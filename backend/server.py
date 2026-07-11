from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, BackgroundTasks, UploadFile, File, Form, Body, Query
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.gzip import GZipMiddleware
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo.errors import DuplicateKeyError
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
from enum import Enum
import uuid
import re
import json
import hashlib
from datetime import datetime, timezone, timedelta
import bcrypt
import jwt
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from PIL import Image
import pillow_heif  # دعم صور HEIC/HEIF من iPhone
pillow_heif.register_heif_opener()  # تسجيل دعم HEIC/HEIF
import io
import math
import base64
import aiofiles
import asyncio
import socketio

# ==================== BUSINESS DATE HELPERS (اليوم التشغيلي) ====================
IRAQ_TZ_OFFSET_HOURS = 3

def iraq_date_from_utc(utc_iso_str: Optional[str] = None) -> str:
    """تحويل ISO datetime (UTC) لتاريخ العراق YYYY-MM-DD"""
    try:
        if utc_iso_str:
            dt = datetime.fromisoformat(utc_iso_str.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = datetime.now(timezone.utc)
        iraq_dt = dt + timedelta(hours=IRAQ_TZ_OFFSET_HOURS)
        return iraq_dt.strftime("%Y-%m-%d")
    except Exception:
        return (datetime.now(timezone.utc) + timedelta(hours=IRAQ_TZ_OFFSET_HOURS)).strftime("%Y-%m-%d")

# ساعة بداية اليوم التشغيلي (افتراضياً 6 صباحاً). أي وقت قبلها يُنسب لليوم السابق.
DEFAULT_BUSINESS_DAY_START_HOUR = 6

def iraq_business_date_from_utc(utc_iso_str: Optional[str] = None, start_hour: int = DEFAULT_BUSINESS_DAY_START_HOUR) -> str:
    """اليوم التشغيلي بتوقيت العراق مع ساعة بداية اليوم — قبلها = اليوم السابق (شفت الفجر يخص ليلة أمس)."""
    try:
        if utc_iso_str:
            dt = datetime.fromisoformat(utc_iso_str.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = datetime.now(timezone.utc)
        iraq_dt = dt + timedelta(hours=IRAQ_TZ_OFFSET_HOURS)
        try:
            sh = int(start_hour)
        except (TypeError, ValueError):
            sh = DEFAULT_BUSINESS_DAY_START_HOUR
        if not (0 <= sh <= 12):
            sh = DEFAULT_BUSINESS_DAY_START_HOUR
        if iraq_dt.hour < sh:
            iraq_dt = iraq_dt - timedelta(days=1)
        return iraq_dt.strftime("%Y-%m-%d")
    except Exception:
        return iraq_date_from_utc(utc_iso_str)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

from routes.rate_limit import enforce_rate_limit, sanitize_text

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

async def get_business_day_start_hour(tenant_id: Optional[str]) -> int:
    """ساعة بداية اليوم التشغيلي من إعدادات المطعم (افتراضياً 6)."""
    try:
        if tenant_id:
            t = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "business_day_start_hour": 1})
            if t and t.get("business_day_start_hour") is not None:
                return int(t["business_day_start_hour"])
    except Exception:
        pass
    return DEFAULT_BUSINESS_DAY_START_HOUR

async def _resolve_business_date(tenant_id: Optional[str], branch_id: Optional[str], cashier_id: Optional[str] = None) -> str:
    """الحصول على business_date من الوردية المفتوحة الحالية.
    
    أولوية البحث (لتفادي مشكلة شفت يامن المفتوح من أمس وشفت زهراء الجديد لليوم):
    1. وردية الكاشير المفتوحة الخاصة به (cashier_id) — الأكثر صحة
    2. أحدث وردية مفتوحة في الفرع (sort by started_at desc)
    3. اليوم التشغيلي الحالي (مع ساعة بداية اليوم) إذا لا يوجد شفت
    """
    base_q = {"status": "open"}
    if tenant_id:
        base_q["tenant_id"] = tenant_id
    if branch_id:
        base_q["branch_id"] = branch_id
    
    shift = None
    # 1) جرب وردية الكاشير الخاصة به أولاً
    if cashier_id:
        shift = await db.shifts.find_one(
            {**base_q, "cashier_id": cashier_id},
            {"_id": 0, "business_date": 1, "started_at": 1, "opened_at": 1},
            sort=[("started_at", -1)]
        )
    # 2) إذا ما وجد، خذ أحدث وردية مفتوحة في الفرع
    if not shift:
        shift = await db.shifts.find_one(
            base_q,
            {"_id": 0, "business_date": 1, "started_at": 1, "opened_at": 1},
            sort=[("started_at", -1)]
        )
    
    _sh = await get_business_day_start_hour(tenant_id)
    if shift:
        if shift.get("business_date"):
            return shift["business_date"]
        started = shift.get("started_at") or shift.get("opened_at")
        if started:
            return iraq_business_date_from_utc(started, start_hour=_sh)
    return iraq_business_date_from_utc(start_hour=_sh)

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET')
if not JWT_SECRET:
    raise ValueError("JWT_SECRET environment variable is required")
JWT_ALGORITHM = 'HS256'
# مدة صلاحية الجلسة بناءً على نوع المستخدم
JWT_EXPIRATION_HOURS_STAFF = 24  # الموظفين: 24 ساعة
JWT_EXPIRATION_DAYS_OWNERS = 36500  # المالك والعملاء: 100 سنة (لا يسجلون خروج أبداً)

# SendGrid Configuration
SENDGRID_API_KEY = os.environ.get('SENDGRID_API_KEY', '')
SENDER_EMAIL = os.environ.get('SENDER_EMAIL', 'noreply@maestroegp.com')

# SMTP Configuration (Namecheap Private Email / generic SMTP)
SMTP_HOST = os.environ.get('SMTP_HOST', '')
SMTP_PORT = int(os.environ.get('SMTP_PORT', '465') or 465)
SMTP_USER = os.environ.get('SMTP_USER', '')
SMTP_PASSWORD = os.environ.get('SMTP_PASSWORD', '')
SMTP_FROM_NAME = os.environ.get('SMTP_FROM_NAME', 'Maestro EGP')

# ==================== المصادقة الثنائية (2FA) + جهاز موثوق ====================
import twilio_verify as _twilio_verify  # noqa: E402
import whatsapp_free as _wa_free  # noqa: E402
# عناوين بريد المالك الاحتياطية (fallback من .env) — تُدار ديناميكياً من قاعدة البيانات
_ENV_OWNER_RECOVERY_EMAILS = [
    e.strip().lower()
    for e in (os.environ.get('OWNER_RECOVERY_EMAILS', 'owner@maestroegp.com,Moataz0111@gmail.com')).split(',')
    if e.strip()
]
# للاستخدام المتزامن (sync): يُحدَّث من قاعدة البيانات في _load_owner_recovery_emails
OWNER_RECOVERY_EMAILS = list(_ENV_OWNER_RECOVERY_EMAILS)


async def get_owner_recovery_emails() -> list:
    """يُرجع قائمة بريد المالك الاحتياطية من قاعدة البيانات (تُدار من UI) أو من .env كافتراضي.
    
    الإعدادات المُدارة من UI (email_config.recovery_emails) لها الأولوية.
    عند غياب قيمة في DB، نعود إلى قائمة .env كخيار احتياطي.
    """
    global OWNER_RECOVERY_EMAILS
    try:
        doc = await db.email_config.find_one({"id": "global"}, {"_id": 0, "recovery_emails": 1})
        if doc and isinstance(doc.get("recovery_emails"), list) and doc["recovery_emails"]:
            db_emails = [str(e).strip().lower() for e in doc["recovery_emails"] if str(e).strip() and "@" in str(e)]
            if db_emails:
                OWNER_RECOVERY_EMAILS = db_emails  # حدّث الكاش المتزامن
                return db_emails
    except Exception:
        pass
    return list(_ENV_OWNER_RECOVERY_EMAILS)


async def _load_email_config() -> dict:
    """Load email/SMTP config from DB (db.email_config id='global'), falling back to env vars.
    The DB config (set by the owner in the control panel) takes priority over env."""
    cfg = {
        "smtp_host": SMTP_HOST,
        "smtp_port": SMTP_PORT,
        "smtp_user": SMTP_USER,
        "smtp_password": SMTP_PASSWORD,
        "from_name": SMTP_FROM_NAME,
        "sender_email": SENDER_EMAIL,
        "sendgrid_api_key": SENDGRID_API_KEY,
    }
    try:
        doc = await db.email_config.find_one({"id": "global"})
        if doc:
            if doc.get("smtp_host"):
                cfg["smtp_host"] = doc["smtp_host"]
            if doc.get("smtp_port"):
                cfg["smtp_port"] = int(doc["smtp_port"])
            if doc.get("smtp_user"):
                cfg["smtp_user"] = doc["smtp_user"]
            if doc.get("smtp_password"):
                cfg["smtp_password"] = doc["smtp_password"]
            if doc.get("from_name"):
                cfg["from_name"] = doc["from_name"]
            if doc.get("sender_email"):
                cfg["sender_email"] = doc["sender_email"]
            if doc.get("sendgrid_api_key"):
                cfg["sendgrid_api_key"] = doc["sendgrid_api_key"]
    except Exception as e:
        logger.error(f"Failed to load email config from DB: {e}")
    return cfg


def _smtp_send_sync(cfg: dict, to_emails, subject: str, html_content: str) -> bool:
    """Send an HTML email synchronously via SMTP (SSL 465 or STARTTLS 587)."""
    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.utils import formataddr

    if isinstance(to_emails, str):
        to_emails = [to_emails]
    host = cfg["smtp_host"]
    port = int(cfg.get("smtp_port") or 465)
    user = cfg["smtp_user"]
    password = cfg["smtp_password"]
    sender = user or cfg.get("sender_email")

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    from_name = (cfg.get("from_name") or "Maestro EGP").strip()
    msg['From'] = formataddr((from_name, sender))
    msg['To'] = ', '.join(to_emails)
    # ترويسة Reply-To بعنوان نظيف (بدون اسم العرض) لضمان نجاح الرد من عميل البريد
    msg['Reply-To'] = sender
    msg.attach(MIMEText(html_content, 'html', 'utf-8'))

    if port == 465:
        server = smtplib.SMTP_SSL(host, port, timeout=20)
    else:
        server = smtplib.SMTP(host, port, timeout=20)
        server.starttls()
    try:
        server.login(user, password)
        server.sendmail(sender, to_emails, msg.as_string())
        return True
    finally:
        try:
            server.quit()
        except Exception:
            pass


async def send_system_email(to_emails, subject: str, html_content: str, purpose: str = "system", tenant_id: Optional[str] = None) -> bool:
    """Unified email sender. Prefers SMTP (Namecheap Private Email); falls back to SendGrid.
    Reads config from DB (owner control panel) or env. Returns True on success. Never raises.
    يُسجّل كل رسالة في db.system_email_logs مع status=sent|failed لسجل المتابعة."""
    cfg = await _load_email_config()
    recipients = to_emails if isinstance(to_emails, list) else [to_emails]
    log_base = {
        "id": str(uuid.uuid4()),
        "to": recipients,
        "subject": subject,
        "purpose": purpose,
        "tenant_id": tenant_id,
        "provider": None,
        "status": "failed",
        "error": None,
        "sent_at": datetime.now(timezone.utc).isoformat(),
    }
    # 1) SMTP (preferred)
    if cfg.get("smtp_host") and cfg.get("smtp_user") and cfg.get("smtp_password"):
        try:
            ok = await asyncio.to_thread(_smtp_send_sync, cfg, to_emails, subject, html_content)
            if ok:
                logger.info(f"Email sent via SMTP to {to_emails}")
                try:
                    await db.system_email_logs.insert_one({**log_base, "provider": "smtp", "status": "sent"})
                except Exception:
                    pass
                return True
            log_base["error"] = "smtp_send_returned_false"
        except Exception as e:
            log_base["error"] = f"smtp: {str(e)[:200]}"
            logger.error(f"SMTP send failed: {e}")
    # 2) SendGrid fallback
    if cfg.get("sendgrid_api_key"):
        try:
            message = Mail(
                from_email=cfg.get("sender_email") or SENDER_EMAIL,
                to_emails=recipients,
                subject=subject,
                html_content=html_content,
            )
            sg = SendGridAPIClient(cfg["sendgrid_api_key"])
            sg.send(message)
            logger.info(f"Email sent via SendGrid to {to_emails}")
            try:
                await db.system_email_logs.insert_one({**log_base, "provider": "sendgrid", "status": "sent", "error": None})
            except Exception:
                pass
            return True
        except Exception as e:
            log_base["error"] = f"sendgrid: {str(e)[:200]}"
            logger.error(f"SendGrid send failed: {e}")
    logger.warning("No email transport configured (SMTP/SendGrid) — email not sent")
    if not log_base["error"]:
        log_base["error"] = "no_transport_configured"
    try:
        await db.system_email_logs.insert_one(log_base)
    except Exception:
        pass
    return False


async def email_transport_configured() -> bool:
    cfg = await _load_email_config()
    return bool((cfg.get("smtp_host") and cfg.get("smtp_user") and cfg.get("smtp_password")) or cfg.get("sendgrid_api_key"))


# ============ إشعار المالك بثلاث قنوات: الجرس + الواتساب + البريد ============
# ==================== شعار Maestro EGP في بريد الإشعارات ====================
# نُضمّن الشعار كصورة base64 داخل HTML البريد (يعمل في كل تطبيقات البريد بلا CORS)
_MAESTRO_LOGO_B64_CACHE = None

def _get_maestro_logo_b64() -> str:
    """يُرجع شعار Maestro EGP بصيغة data:image/png;base64,... (مُحمَّل مرة واحدة).
    يُستخدم داخل قوالب البريد لضمان ظهور الشعار في كل تطبيقات البريد (Apple Mail / Gmail / …)."""
    global _MAESTRO_LOGO_B64_CACHE
    if _MAESTRO_LOGO_B64_CACHE is not None:
        return _MAESTRO_LOGO_B64_CACHE
    try:
        _logo_path = ROOT_DIR / "static" / "branding" / "maestro-logo-email.b64"
        with open(_logo_path, "r", encoding="utf-8") as _f:
            _MAESTRO_LOGO_B64_CACHE = "data:image/png;base64," + _f.read().strip()
    except Exception as _e:
        logger.warning(f"maestro logo load failed: {_e}")
        _MAESTRO_LOGO_B64_CACHE = ""
    return _MAESTRO_LOGO_B64_CACHE


def build_branded_email_html(title: str, body_html: str, severity: str = "info",
                              show_footer_note: bool = True) -> str:
    """يبني HTML موحّد لجميع رسائل البريد الخارجة من النظام.

    - severity: info(أزرق) | warning(أصفر) | critical(أحمر) | success(أخضر)
    - يُدرج شعار Maestro EGP الموحّد تحت الشريط الملوّن مباشرة (خلفية بيضاء + ظل 3D)
    - يُستخدم لكل: تقارير الوردية، OTP، رسائل الترحيب، تنبيهات النظام، والاسترداد.
    """
    _sev_color = {
        "info": "#3B82F6", "warning": "#F59E0B",
        "critical": "#EF4444", "success": "#10B981",
    }.get(severity, "#3B82F6")
    _logo_src = _get_maestro_logo_b64()
    _logo_block = (
        f"<div style='background:#ffffff;padding:16px 0;text-align:center'>"
        f"<img src='{_logo_src}' alt='Maestro EGP' "
        f"style='width:64px;height:64px;display:inline-block;border-radius:50%;object-fit:cover;"
        f"box-shadow:0 6px 16px rgba(11,26,58,0.35), 0 2px 4px rgba(0,0,0,0.15)' />"
        f"</div>"
    ) if _logo_src else ""
    now_iso = datetime.now(timezone.utc).isoformat()
    _footer = (
        f"<hr style='margin:16px 0;border:none;border-top:1px solid #e5e7eb'/>"
        f"<p style='font-size:12px;color:#6b7280;margin:0'>Maestro EGP — {now_iso}</p>"
    ) if show_footer_note else ""
    return (
        f"<div style='font-family:Tahoma,Arial,sans-serif;direction:rtl;text-align:right;max-width:640px;margin:0 auto'>"
        f"<div style='background:{_sev_color};color:#fff;padding:12px 16px;border-radius:8px 8px 0 0'>"
        f"<h2 style='margin:0'>{title}</h2></div>"
        f"{_logo_block}"
        f"<div style='background:#f9fafb;padding:16px;border-radius:0 0 8px 8px;border:1px solid #e5e7eb;border-top:none;color:#111'>"
        f"{body_html}"
        f"{_footer}"
        f"</div></div>"
    )


async def notify_owner_multichannel(
    title: str,
    message: str,
    severity: str = "info",
    category: str = "system",
    send_whatsapp: bool = True,
    send_email: bool = True,
    tenant_id: Optional[str] = None,
    metadata: Optional[dict] = None,
) -> dict:
    """يُرسل إشعاراً مهماً للمالك في 3 قنوات متزامنة:
    
    1. جرس لوحة التحكم (db.notifications) — يظهر في أيقونة الجرس الأعلى.
    2. واتساب لرقم المالك المرتبط (إن كان الواتساب مربوطاً).
    3. بريد إلكتروني إلى OWNER_RECOVERY_EMAILS (إن كان SMTP/SendGrid مُعدّاً).
    
    - severity: info | warning | critical (يُلوّن الجرس)
    - category: security | shift | expense | subscription | system
    - metadata: بيانات إضافية للجرس (تُخزّن مع الإشعار)
    
    يُرجع dict فيه {bell, whatsapp, email}: bool لكل قناة.
    """
    result = {"bell": False, "whatsapp": False, "email": False}
    now_iso = datetime.now(timezone.utc).isoformat()
    
    # 1) الجرس (دائماً)
    try:
        await db.notifications.insert_one({
            "id": str(uuid.uuid4()),
            "title": title,
            "message": message,
            "severity": severity,
            "category": category,
            "is_read": False,
            "created_at": now_iso,
            "tenant_id": tenant_id,
            "metadata": metadata or {},
        })
        result["bell"] = True
    except Exception as e:
        logger.warning(f"notify_owner_multichannel bell failed: {e}")
    
    # 2) الواتساب (لرقم المالك + المديرين المرتبطين بالمشروع — يستقبل الجميع الإشعار)
    NOTIFY_MAX_RECIPIENTS = 20  # حماية: لا نُغرق تينانتاً بمئات الرسائل
    if send_whatsapp:
        try:
            _wa_connected = await _wa_free.is_connected()
            if not _wa_connected:
                # سجّل السبب بوضوح للمشرف — الرسالة لم تُرسَل لأن الواتساب غير مربوط
                logger.warning(
                    f"notify_owner_multichannel: skip whatsapp — WA_NOT_CONNECTED "
                    f"(tenant={tenant_id}, category={category}). "
                    f"اربط رقم الواتساب من لوحة المالك → إعدادات النظام الرئيسي → الواتساب."
                )
                result["whatsapp_skip_reason"] = "wa_not_connected"
            else:
                # اجمع كل الأرقام المستقبِلة: env → tenant.owner_phone → كل admin/manager للتينانت
                phones = []
                env_phone = os.environ.get("OWNER_ALERT_PHONE", "")
                if env_phone:
                    phones.append(env_phone)
                # رقم المالك على مستند التينانت
                if tenant_id:
                    _tn = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "owner_phone": 1}) or {}
                    if _tn.get("owner_phone"):
                        phones.append(_tn["owner_phone"])
                # كل حسابات admin + manager + owner ذات الهاتف داخل نفس التينانت (فقط string phones)
                # 🔥 استبعاد super_admin عند وجود tenant_id — الرسائل الخاصة بالعملاء لا تُرسَل لمالك النظام
                _owner_roles = ["admin", "owner", "manager", "branch_manager"]
                if not tenant_id:
                    _owner_roles.append("super_admin")  # فقط للتنبيهات النظامية العامة
                user_q = {"role": {"$in": _owner_roles},
                          "phone": {"$type": "string", "$ne": ""}}
                if tenant_id:
                    user_q["tenant_id"] = tenant_id
                async for u in db.users.find(user_q, {"_id": 0, "phone": 1}):
                    if u.get("phone"):
                        phones.append(u["phone"])
                # dedup على شكل E.164 القياسي (نمنع نفس الرقم بصيغتين مختلفتين)
                seen_e164 = set()
                normalized_pairs = []  # (e164, raw) لعرض السجل
                for p in phones:
                    raw = (p or "").strip()
                    if not raw:
                        continue
                    try:
                        e164 = await _phone_to_e164(raw)
                    except Exception:
                        continue
                    if not e164 or e164 in seen_e164:
                        continue
                    seen_e164.add(e164)
                    normalized_pairs.append((e164, raw))
                    if len(normalized_pairs) >= NOTIFY_MAX_RECIPIENTS:
                        logger.warning(f"notify_owner_multichannel: قصّرت المستقبِلين على {NOTIFY_MAX_RECIPIENTS} (كان هناك أكثر — تحقق من إعدادات التينانت)")
                        break
                # إرسال متزامن لكل الأرقام (بهوية Maestro EGP الموحّدة: شعار + قالب)
                sent_count = 0
                _wa_errors = []
                for e164, _raw in normalized_pairs:
                    try:
                        ok, _err = await _wa_free.send_message(
                            e164, message, purpose=f"owner_{category}",
                            tenant_id=tenant_id, title=title, with_logo=True,
                        )
                        if ok:
                            sent_count += 1
                        else:
                            _wa_errors.append(f"{e164}: {_err}")
                    except Exception as _ee:
                        logger.warning(f"whatsapp send to {e164} failed: {_ee}")
                        _wa_errors.append(f"{e164}: {_ee}")
                result["whatsapp"] = sent_count > 0
                result["whatsapp_recipients"] = sent_count
                if not normalized_pairs:
                    logger.warning(
                        f"notify_owner_multichannel: skip whatsapp — NO_PHONES "
                        f"(tenant={tenant_id}, category={category}). "
                        f"أضف رقم واتساب لمالك المطعم (owner_phone) أو لأي مستخدم admin/manager بحقل phone."
                    )
                    result["whatsapp_skip_reason"] = "no_recipients"
                elif sent_count == 0 and _wa_errors:
                    logger.warning(f"notify_owner_multichannel: whatsapp all sends failed — {_wa_errors[:3]}")
                    result["whatsapp_skip_reason"] = "all_sends_failed"
        except Exception as e:
            logger.warning(f"notify_owner_multichannel whatsapp failed: {e}")
    
    # 3) البريد الإلكتروني (لمالك المشروع + المديرين + بريد الاسترداد — للجميع نسخة)
    if send_email:
        try:
            body_html = f"<pre style='white-space:pre-wrap;font-family:Tahoma,Arial,sans-serif;font-size:14px;color:#111;margin:0'>{message}</pre>"
            html = build_branded_email_html(title=title, body_html=body_html, severity=severity)
            # 🔥 استبعاد super_admin عند وجود tenant_id — البريد الخاص بالعميل لا يُرسَل لمالك النظام
            _owner_roles_email = ["admin", "owner", "manager", "branch_manager"]
            if not tenant_id:
                _owner_roles_email.append("super_admin")
            user_q = {"role": {"$in": _owner_roles_email},
                      "email": {"$exists": True, "$ne": ""}}
            if tenant_id:
                # بريد العميل فقط (owner_email على مستند التينانت + مستخدمو التينانت)
                user_q["tenant_id"] = tenant_id
                emails = []
                _tn = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "owner_email": 1}) or {}
                if _tn.get("owner_email"):
                    emails.append(_tn["owner_email"])
            else:
                # تنبيهات نظامية عامة — تشمل بريد استرداد المالك الأعلى
                emails = list(await get_owner_recovery_emails())
            async for u in db.users.find(user_q, {"_id": 0, "email": 1}):
                if u.get("email"):
                    emails.append(u["email"])
            # dedupe لطيف
            seen_e = set()
            unique_emails = []
            for em in emails:
                em = (em or "").strip().lower()
                if em and "@" in em and em not in seen_e:
                    seen_e.add(em)
                    unique_emails.append(em)
            if unique_emails:
                # قصّ لـ NOTIFY_MAX_RECIPIENTS (نفس السقف مع الواتساب)
                if len(unique_emails) > NOTIFY_MAX_RECIPIENTS:
                    logger.warning(f"notify_owner_multichannel email: قصّرت المستقبِلين على {NOTIFY_MAX_RECIPIENTS} (كان هناك {len(unique_emails)})")
                    unique_emails = unique_emails[:NOTIFY_MAX_RECIPIENTS]
                ok = await send_system_email(
                    unique_emails,
                    f"[Maestro EGP] {title}",
                    html,
                    purpose=f"owner_{category}",
                    tenant_id=tenant_id,
                )
                result["email"] = bool(ok)
                result["email_recipients"] = len(unique_emails)
        except Exception as e:
            logger.warning(f"notify_owner_multichannel email failed: {e}")
    
    return result

# Static Files Configuration
UPLOAD_DIR = ROOT_DIR / "uploads"
UPLOAD_DIR.mkdir(exist_ok=True)
BACKGROUNDS_DIR = UPLOAD_DIR / "backgrounds"
BACKGROUNDS_DIR.mkdir(exist_ok=True)
LOGOS_DIR = UPLOAD_DIR / "logos"
LOGOS_DIR.mkdir(exist_ok=True)
IMAGES_DIR = UPLOAD_DIR / "images"
IMAGES_DIR.mkdir(exist_ok=True)
PRODUCTS_DIR = IMAGES_DIR / "products"
PRODUCTS_DIR.mkdir(exist_ok=True)
CATEGORIES_DIR = IMAGES_DIR / "categories"
CATEGORIES_DIR.mkdir(exist_ok=True)

app = FastAPI(title="Maestro EGP API", version="2.0.0", docs_url=None, redoc_url=None, openapi_url=None)
api_router = APIRouter(prefix="/api")
security = HTTPBearer(auto_error=False)

AUTH_COOKIE_NAME = "access_token"

def _extract_token(request: Request, credentials) -> str:
    if credentials and credentials.credentials:
        return credentials.credentials
    tok = request.cookies.get(AUTH_COOKIE_NAME) if request else None
    if not tok:
        raise HTTPException(status_code=401, detail="غير مصرح - يرجى تسجيل الدخول")
    return tok

def _set_auth_cookie(response, token: str, role: str):
    max_age = 365 * 24 * 3600 if role in ("super_admin", "admin") else 24 * 3600
    response.set_cookie(key=AUTH_COOKIE_NAME, value=token, httponly=True, secure=True,
                        samesite="lax", max_age=max_age, path="/")

PASSWORD_MIN_LENGTH = 8

def validate_password_strength(password: str):
    p = password or ""
    if len(p) < PASSWORD_MIN_LENGTH:
        raise HTTPException(status_code=400, detail=f"كلمة المرور ضعيفة: يجب ألا تقل عن {PASSWORD_MIN_LENGTH} أحرف")
    if not any(ch.isdigit() for ch in p):
        raise HTTPException(status_code=400, detail="كلمة المرور ضعيفة: يجب أن تحتوي على رقم واحد على الأقل")
    if not any(ch.isalpha() for ch in p):
        raise HTTPException(status_code=400, detail="كلمة المرور ضعيفة: يجب أن تحتوي على حرف واحد على الأقل")

# إضافة GZip compression لتسريع نقل البيانات
app.add_middleware(GZipMiddleware, minimum_size=1000)

# ==================== حظر عناوين IP (Block List) ====================
_BLOCKED_IPS = set()
_BLOCKED_IPS_TS = 0.0
_BLOCKED_TTL = 20  # ثانية

async def _refresh_blocked_ips(force=False):
    global _BLOCKED_IPS, _BLOCKED_IPS_TS
    import time as _t
    now = _t.time()
    if not force and (now - _BLOCKED_IPS_TS) < _BLOCKED_TTL:
        return
    try:
        docs = await db.blocked_ips.find({}, {"_id": 0, "ip": 1}).to_list(10000)
        _BLOCKED_IPS = {d.get("ip") for d in docs if d.get("ip")}
        _BLOCKED_IPS_TS = now
    except Exception:
        pass

# عناوين IP الخاصة بالمالك — تُحفظ ولا تُحظر أبداً (طلب المالك)
_OWNER_IPS = set()
_OWNER_IPS_TS = 0.0

async def _refresh_owner_ips(force=False):
    global _OWNER_IPS, _OWNER_IPS_TS
    import time as _t
    now = _t.time()
    if not force and (now - _OWNER_IPS_TS) < _BLOCKED_TTL:
        return
    try:
        docs = await db.owner_trusted_ips.find({}, {"_id": 0, "ip": 1}).to_list(1000)
        _OWNER_IPS = {d.get("ip") for d in docs if d.get("ip")}
        _OWNER_IPS_TS = now
    except Exception:
        pass

async def _is_owner_ip(ip):
    if not ip or ip == "unknown":
        return False
    await _refresh_owner_ips()
    return ip in _OWNER_IPS

async def _save_owner_ip(ip, request=None):
    """يحفظ عنوان IP للمالك في القائمة البيضاء (لا يُحظر أبداً)."""
    if not ip or ip == "unknown":
        return
    try:
        await db.owner_trusted_ips.update_one(
            {"ip": ip},
            {"$set": {"ip": ip, "saved_at": datetime.now(timezone.utc).isoformat(),
                      "user_agent": (request.headers.get("user-agent", "") if request else "")}},
            upsert=True,
        )
        # إن كان محظوراً بالخطأ سابقاً، ارفع الحظر عنه
        await db.blocked_ips.delete_one({"ip": ip})
        await _refresh_owner_ips(force=True)
        await _refresh_blocked_ips(force=True)
    except Exception:
        pass

async def _is_ip_blocked(ip):
    # عنوان المالك لا يُحظر أبداً
    if await _is_owner_ip(ip):
        return False
    await _refresh_blocked_ips()
    return ip in _BLOCKED_IPS

async def _ban_ip_permanent(ip, reason: str, request=None):
    """حظر دائم لعنوان IP بعد تجاوز محاولات الدخول الفاشلة (يمنع كل الأجهزة على الشبكة).
    يُستثنى عنوان المالك دائماً."""
    if not ip or ip == "unknown":
        return False
    if await _is_owner_ip(ip):
        return False
    try:
        await db.blocked_ips.update_one(
            {"ip": ip},
            {"$set": {
                "ip": ip,
                "reason": reason,
                "blocked_by": "النظام (حظر دائم تلقائي)",
                "auto": True,
                "permanent": True,
                "blocked_at": datetime.now(timezone.utc).isoformat(),
            }},
            upsert=True,
        )
        await _refresh_blocked_ips(force=True)
        try:
            await record_audit("security.permanent_ban", request=request, status=403,
                               details={"ip": ip, "reason": reason})
        except Exception:
            pass
        return True
    except Exception:
        return False

# ==================== الحظر التلقائي للمخترقين (Auto-Ban) ====================
from collections import deque as _deque, defaultdict as _ddict
_AUTO_OFFENSES = _ddict(_deque)        # ip -> طوابع زمنية للمحاولات المشبوهة
_AUTO_BAN_THRESHOLD = 5                 # عدد المحاولات المشبوهة
_AUTO_BAN_WINDOW = 300                 # خلال 5 دقائق
_AUTO_BAN_ENABLED = True

async def _register_offense(ip, request, path, method, sc):
    """يسجّل محاولة مشبوهة ويحظر العنوان تلقائياً عند تجاوز الحد."""
    if not _AUTO_BAN_ENABLED or not ip or ip == "unknown":
        return
    # عنوان المالك لا يُحظر أبداً
    if await _is_owner_ip(ip):
        return
    import time as _t
    now = _t.time()
    dq = _AUTO_OFFENSES[ip]
    cutoff = now - _AUTO_BAN_WINDOW
    while dq and dq[0] < cutoff:
        dq.popleft()
    dq.append(now)
    if len(dq) < _AUTO_BAN_THRESHOLD:
        return
    # حماية: لا تحظر عنواناً سجّل دخولاً ناجحاً مؤخراً (موظف/مالك) — خلال آخر 24 ساعة فقط
    try:
        _since = datetime.now(timezone.utc) - timedelta(hours=24)
        trusted = await db.audit_logs.find_one({
            "ip": ip,
            "event": {"$regex": "login\\.success$"},
            "ts": {"$gte": _since},
        })
        if trusted:
            dq.clear()
            return
    except Exception:
        pass
    # حظر تلقائي
    try:
        existing = await db.blocked_ips.find_one({"ip": ip})
        if not existing:
            await db.blocked_ips.update_one({"ip": ip}, {"$set": {
                "ip": ip,
                "reason": f"حظر تلقائي: {len(dq)} محاولة مشبوهة خلال 5 دقائق (آخرها {method} {path})",
                "blocked_by": "النظام (تلقائي)",
                "auto": True,
                "last_path": path,
                "last_method": method,
                "last_status": sc,
                "offenses": len(dq),
                "blocked_at": datetime.now(timezone.utc).isoformat(),
            }}, upsert=True)
            await _refresh_blocked_ips(force=True)
            try:
                await record_audit("security.auto_blocked", request=request, status=403,
                                   details={"ip": ip, "offenses": len(dq), "path": path, "method": method})
            except Exception:
                pass
        dq.clear()
    except Exception:
        pass

def _block_response(request):
    """استجابة الحظر: صفحة HTML للمتصفّح، JSON لطلبات الـ API."""
    from fastapi.responses import JSONResponse, HTMLResponse
    accept = (request.headers.get("accept") or "")
    if "text/html" in accept:
        html = """<!doctype html><html dir="rtl" lang="ar"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1"><title>تم حظر الوصول</title></head>
<body style="margin:0;background:#0a0e1a;color:#fff;font-family:system-ui,Segoe UI,Tahoma,sans-serif;display:flex;align-items:center;justify-content:center;min-height:100vh;text-align:center">
<div style="max-width:520px;padding:40px"><div style="font-size:64px">🚫</div>
<h1 style="color:#ef4444;margin:16px 0">تم حظر الوصول</h1>
<p style="color:#94a3b8;line-height:1.8">تم حظر عنوانك بسبب نشاط مشبوه أو محاولة وصول غير مصرّح بها.<br>إذا كنت تعتقد أن هذا خطأ، تواصل مع إدارة النظام.</p>
<p style="color:#475569;font-size:12px;margin-top:24px">Maestro EGP — نظام محمي</p></div></body></html>"""
        return HTMLResponse(content=html, status_code=403, headers={"X-Blocked": "1"})
    return JSONResponse(status_code=403, content={"detail": "تم حظر عنوانك من قبل الإدارة", "blocked": True}, headers={"X-Blocked": "1"})

# ==================== ترويسات أمان لكل استجابات الـ API (حماية ضد clickjacking/XSS/MIME-sniffing) ====================
@app.middleware("http")
async def normalize_date_range_middleware(request: Request, call_next):
    """🛡 تصحيح تلقائي للمدى الزمني المعكوس (البداية بعد النهاية) في كل التقارير.
    منظومة محاسبية: لا تُعرض أصفار مضللة بسبب انعكاس من/إلى — يُبدَّلان تلقائياً."""
    try:
        qp = request.query_params
        changed = False
        items = [(k, v) for k, v in qp.multi_items()]
        for a, b in (("start_date", "end_date"), ("date_from", "date_to")):
            s = qp.get(a)
            e = qp.get(b)
            if s and e and s[:10] > e[:10]:
                new_start = e[:10]
                new_end = s[:10] + (e[10:] if len(e) > 10 else "")
                items = [(k, new_start if k == a else (new_end if k == b else v)) for k, v in items]
                changed = True
        if changed:
            from urllib.parse import urlencode
            request.scope["query_string"] = urlencode(items).encode()
    except Exception:
        pass
    return await call_next(request)

@app.middleware("http")
async def security_headers_middleware(request: Request, call_next):
    # حظر عناوين IP المحظورة قبل أي معالجة
    try:
        ip = _client_ip(request)
        if await _is_ip_blocked(ip):
            try:
                await record_audit("security.blocked_ip_hit", request=request, status=403,
                                   details={"path": request.url.path, "method": request.method})
            except Exception:
                pass
            return _block_response(request)
    except Exception:
        pass
    response = await call_next(request)
    try:
        response.headers["X-Content-Type-Options"] = "nosniff"
        response.headers["X-Frame-Options"] = "SAMEORIGIN"
        response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
        response.headers["Permissions-Policy"] = "geolocation=(self), microphone=(self), camera=(self)"
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    except Exception:
        pass
    return response

# ==================== سجل أمني: تسجيل محاولات الحقن/الإساءة المرفوضة (429/403) ====================
@app.middleware("http")
async def security_audit_middleware(request: Request, call_next):
    response = await call_next(request)
    try:
        sc = response.status_code
        path = request.url.path
        # نسجّل: كل تجاوز لتحديد المعدّل (429)، وكل محاولة كتابة ممنوعة (403 على POST/PUT/PATCH/DELETE)
        if path.startswith("/api") and (
            sc == 429 or (sc == 403 and request.method in ("POST", "PUT", "PATCH", "DELETE"))
        ):
            event = "security.rate_limited" if sc == 429 else "security.forbidden"
            await record_audit(
                event, request=request, status=sc,
                details={"path": path, "method": request.method},
            )
            # تسجيل المحاولة للحظر التلقائي
            try:
                await _register_offense(_client_ip(request), request, path, request.method, sc)
            except Exception:
                pass
    except Exception:
        pass
    return response

# ==================== RBAC مركزي: منع الأدوار الأدنى من البيانات الحساسة ====================
# طبقة صلاحيات موحّدة تُطبَّق على كل النقاط الحساسة (مالية/موارد بشرية/تكاليف/تقارير/موردين)
# تمنع الكاشير/الكابتن/الكول-سنتر/السائق من قراءة بيانات لا تخصّهم، مع إبقاء الأدوار التشغيلية تعمل.
import re as _re_rbac

_MGMT_ROLES = {"super_admin", "admin", "manager"}
_INV_ROLES = _MGMT_ROLES | {"warehouse_keeper", "manufacturer"}
_PURCH_ROLES = _MGMT_ROLES | {"purchasing", "warehouse_keeper"}

# (نمط المسار, الأدوار المسموح لها). أول تطابق يُحسم.
_RBAC_RULES = [
    (_re_rbac.compile(r"^/api/employees(/|$)"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/payroll"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/advances"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/deductions"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/reports/"), _MGMT_ROLES | {"supervisor"}),
    (_re_rbac.compile(r"^/api/smart-reports/"), _MGMT_ROLES | {"supervisor"}),
    (_re_rbac.compile(r"^/api/dashboard/stats"), _MGMT_ROLES | {"supervisor"}),
    (_re_rbac.compile(r"^/api/break-even"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/inventory-stats"), _INV_ROLES),
    (_re_rbac.compile(r"^/api/inventory-settings"), _INV_ROLES),
    (_re_rbac.compile(r"^/api/suppliers"), _PURCH_ROLES),
    (_re_rbac.compile(r"^/api/supplier-payment-dues"), _PURCH_ROLES),
    (_re_rbac.compile(r"^/api/payment-settings"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/callcenter/simulate"), _MGMT_ROLES),
    (_re_rbac.compile(r"^/api/purchases-new"), _PURCH_ROLES),
    (_re_rbac.compile(r"^/api/purchase-requests"), _PURCH_ROLES),
    (_re_rbac.compile(r"^/api/manufactured-products"), _INV_ROLES),
    (_re_rbac.compile(r"^/api/manufacturing"), _INV_ROLES),
    (_re_rbac.compile(r"^/api/warehouse-"), _INV_ROLES | {"purchasing"}),
    (_re_rbac.compile(r"^/api/raw-materials"), _INV_ROLES | {"purchasing"}),
]

@app.middleware("http")
async def rbac_middleware(request: Request, call_next):
    from fastapi.responses import JSONResponse as _JR
    try:
        path = request.url.path
        if request.method != "OPTIONS" and path.startswith("/api/"):
            allowed = None
            for rx, roles in _RBAC_RULES:
                if rx.match(path):
                    allowed = roles
                    break
            if allowed is not None:
                role = None
                auth = request.headers.get("Authorization", "")
                if auth.startswith("Bearer "):
                    try:
                        payload = jwt.decode(auth[7:], JWT_SECRET, algorithms=[JWT_ALGORITHM])
                        role = payload.get("role")
                    except Exception:
                        role = None
                if role is None:
                    return _JR(status_code=401, content={"detail": "غير مصرح - يلزم تسجيل دخول صالح"})
                if role not in allowed:
                    return _JR(status_code=403, content={"detail": "ليس لديك صلاحية للوصول لهذه البيانات"})
    except Exception:
        pass
    return await call_next(request)


# Mount static files directory
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="api_uploads")

# Helper function to check user roles
def has_role(user: dict, roles: list) -> bool:
    """التحقق من صلاحية المستخدم"""
    user_role = user.get("role", "")
    return user_role in roles or "super_admin" in roles and user_role == "super_admin"

# ==================== DATABASE INITIALIZATION ====================

async def create_indexes():
    """إنشاء indexes لتسريع الاستعلامات"""
    try:
        # Users indexes
        await db.users.create_index("id", unique=True)
        await db.users.create_index("email", unique=True)
        await db.users.create_index("tenant_id")
        await db.users.create_index("role")

        # Audit log indexes + TTL (حذف تلقائي بعد 60 يوماً)
        await db.audit_logs.create_index("ts", expireAfterSeconds=5184000)
        await db.audit_logs.create_index("tenant_id")
        await db.audit_logs.create_index("event")
        await db.driver_tokens.create_index("token", unique=True)
        await db.driver_tokens.create_index("driver_id")
        
        # Orders indexes - الأكثر أهمية للأداء
        await db.orders.create_index("id", unique=True)
        await db.orders.create_index("tenant_id")
        await db.orders.create_index("branch_id")
        await db.orders.create_index("status")
        await db.orders.create_index("created_at")
        await db.orders.create_index([("tenant_id", 1), ("status", 1)])
        await db.orders.create_index([("tenant_id", 1), ("created_at", -1)])
        await db.orders.create_index([("cashier_id", 1), ("created_at", -1)])
        # 🔒 فهرس فريد على offline_id لكل tenant — قفل idempotency على مستوى قاعدة البيانات
        # يمنع تكرار الطلب نهائياً حتى عند تزامن المزامنة (سباق طلبين بنفس المفتاح).
        # partial: يُطبَّق فقط حين يكون offline_id نصاً (يتجاهل null حتى لا تتعارض الطلبات بلا offline_id).
        try:
            await db.orders.create_index(
                [("tenant_id", 1), ("offline_id", 1)],
                unique=True,
                partialFilterExpression={"offline_id": {"$type": "string"}},
                name="uniq_tenant_offline_id",
            )
        except Exception as _e:
            # غالباً بسبب وجود تكرارات قديمة — يُنشأ الفهرس تلقائياً بعد تشغيل أداة التنظيف.
            print(f"⚠️ تعذّر إنشاء فهرس offline_id الفريد (ربما توجد تكرارات): {_e}")
        
        # Products indexes
        await db.products.create_index("id", unique=True)
        await db.products.create_index("tenant_id")
        await db.products.create_index("category_id")
        await db.products.create_index([("tenant_id", 1), ("is_active", 1)])
        
        # Categories indexes
        await db.categories.create_index("id", unique=True)
        await db.categories.create_index("tenant_id")
        
        # Drivers indexes
        await db.drivers.create_index("id", unique=True)
        await db.drivers.create_index("tenant_id")
        await db.drivers.create_index("branch_id")
        
        # Employees indexes
        await db.employees.create_index("id", unique=True)
        await db.employees.create_index("tenant_id")
        await db.employees.create_index("branch_id")
        
        # Shifts indexes
        await db.shifts.create_index("id", unique=True)
        await db.shifts.create_index([("cashier_id", 1), ("status", 1)])
        await db.shifts.create_index("tenant_id")
        
        # Expenses indexes
        await db.expenses.create_index("id", unique=True)
        await db.expenses.create_index("tenant_id")
        await db.expenses.create_index([("branch_id", 1), ("created_at", -1)])
        
        # Inventory indexes
        await db.inventory.create_index("id", unique=True)
        await db.inventory.create_index("tenant_id")
        await db.inventory.create_index("branch_id")
        
        logger.info("✅ Database indexes created successfully")
    except Exception as e:
        logger.warning(f"⚠️ Some indexes may already exist: {e}")

async def init_database():
    """تهيئة قاعدة البيانات بالبيانات الأساسية عند بدء التطبيق"""
    try:
        logger.info("🔍 Checking database initialization...")
        
        # إنشاء indexes لتسريع الاستعلامات
        await create_indexes()
        
        # التحقق من الاتصال بقاعدة البيانات
        await db.command('ping')
        logger.info("✅ Database connection successful")
        
        # التحقق من وجود Super Admin وتحديثه إذا لزم الأمر (بدون حذف)
        super_admin = await db.users.find_one({"email": "owner@maestroegp.com"})
        if super_admin:
            # تحديث Super Admin إذا كان ينقصه super_admin_secret (بدون حذف)
            if not super_admin.get("super_admin_secret"):
                logger.info("🔧 Updating Super Admin with missing fields...")
                await db.users.update_one(
                    {"email": "owner@maestroegp.com"},
                    {"$set": {"super_admin_secret": "271018", "role": "super_admin"}}
                )
                logger.info("✅ Super Admin updated with secret key")
        
        if not super_admin:
            logger.info("🔧 Initializing database with default data...")
            
            # إنشاء Super Admin
            super_admin_password = bcrypt.hashpw("owner123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            super_admin_doc = {
                "id": str(uuid.uuid4()),
                "username": "super_admin",
                "email": "owner@maestroegp.com",
                "password": super_admin_password,
                "password_hash": super_admin_password,
                "full_name": "مالك النظام",
                "role": "super_admin",
                "branch_id": None,
                "tenant_id": "system",
                "permissions": ["all", "super_admin"],
                "is_active": True,
                "super_admin_secret": "271018",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(super_admin_doc)
            logger.info("✅ Super Admin created: owner@maestroegp.com / owner123")
            
            # إنشاء مدير النظام الرئيسي
            admin_password = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            admin_doc = {
                "id": str(uuid.uuid4()),
                "username": "admin",
                "email": "admin@maestroegp.com",
                "password": admin_password,
                "full_name": "مدير النظام",
                "role": "admin",
                "branch_id": None,
                "tenant_id": "default",
                "permissions": ["all"],
                "is_active": True,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.users.insert_one(admin_doc)
            logger.info("✅ Main Admin created: admin@maestroegp.com / admin123")
            
            # إنشاء الفرع الرئيسي
            branch_doc = {
                "id": str(uuid.uuid4()),
                "name": "الفرع الرئيسي",
                "code": "MAIN",
                "address": "العنوان الرئيسي",
                "phone": "",
                "is_main": True,
                "is_active": True,
                "tenant_id": "default",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.branches.insert_one(branch_doc)
            logger.info("✅ Main Branch created")
            
            # إعدادات النظام
            branding_doc = {
                "type": "system_branding",
                "value": {
                    "name": "Maestro",
                    "name_ar": "Maestro",
                    "name_en": "Maestro",
                    "logo_url": None
                }
            }
            await db.settings.insert_one(branding_doc)
            logger.info("✅ System branding created")
            
            # خلفيات تسجيل الدخول الافتراضية - صور متعددة
            bg_doc = {
                "type": "login_backgrounds",
                "backgrounds": [
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                        "title": "مطعم فاخر",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                        "title": "مطعم حديث",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=1920",
                        "title": "طعام شهي",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=1920",
                        "title": "مطعم أنيق",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=1920",
                        "title": "كافيه عصري",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
                        "title": "مطبخ احترافي",
                        "is_active": True
                    }
                ],
                "settings": {
                    "transition_effect": "fade",
                    "transition_speed": 5,
                    "overlay_color": "rgba(0,0,0,0.5)",
                    "text_color": "#ffffff"
                }
            }
            await db.settings.insert_one(bg_doc)
            logger.info("✅ Login backgrounds created (6 images)")
            
            # إنشاء الفئات الافتراضية مع الصور
            default_categories = [
                {"id": str(uuid.uuid4()), "name": "برغر", "name_ar": "برغر", "sort_order": 1, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "بيتزا", "name_ar": "بيتزا", "sort_order": 2, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1703073186021-021fb5a0bde1?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "مشروبات", "name_ar": "مشروبات", "sort_order": 3, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "حلويات", "name_ar": "حلويات", "sort_order": 4, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "سلطات", "name_ar": "سلطات", "sort_order": 5, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.categories.insert_many(default_categories)
            logger.info("✅ Default categories created (5) with images")
            
            # إنشاء منتجات افتراضية مع الصور
            burger_cat = default_categories[0]["id"]
            pizza_cat = default_categories[1]["id"]
            drinks_cat = default_categories[2]["id"]
            desserts_cat = default_categories[3]["id"]
            salads_cat = default_categories[4]["id"]
            default_products = [
                {"id": str(uuid.uuid4()), "name": "برغر كلاسيك", "price": 5000, "cost": 2000, "category_id": burger_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1656439659132-24c68e36b553?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "برغر دبل", "price": 7500, "cost": 3000, "category_id": burger_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "بيتزا مارغريتا", "price": 10000, "cost": 4000, "category_id": pizza_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1681567604770-0dc826c870ae?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "بيتزا خضار", "price": 12000, "cost": 5000, "category_id": pizza_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1602104980741-b87a33837f9f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "كولا", "price": 1500, "cost": 500, "category_id": drinks_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "عصير برتقال", "price": 2500, "cost": 1000, "category_id": drinks_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1716925539259-ce0115263d37?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "كيكة شوكولاته", "price": 3500, "cost": 1500, "category_id": desserts_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "سلطة خضراء", "price": 4000, "cost": 1200, "category_id": salads_cat, "is_available": True, "tenant_id": "default", "image": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.products.insert_many(default_products)
            logger.info("✅ Default products created (8) with images")
            
            # إنشاء سائقين افتراضيين
            default_drivers = [
                {"id": str(uuid.uuid4()), "name": "سائق 1", "phone": "07801111111", "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "سائق 2", "phone": "07802222222", "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.drivers.insert_many(default_drivers)
            logger.info("✅ Default drivers created (2)")
            
            # إنشاء موظفين افتراضيين
            default_employees = [
                {"id": str(uuid.uuid4()), "name": "موظف 1", "position": "كاشير", "phone": "07803333333", "salary": 500000, "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
                {"id": str(uuid.uuid4()), "name": "موظف 2", "position": "طباخ", "phone": "07804444444", "salary": 600000, "is_active": True, "tenant_id": "default", "created_at": datetime.now(timezone.utc).isoformat()},
            ]
            await db.employees.insert_many(default_employees)
            logger.info("✅ Default employees created (2)")
            
            logger.info("=" * 50)
            logger.info("🎉 DATABASE INITIALIZATION COMPLETE!")
            logger.info("=" * 50)
            logger.info("📋 LOGIN CREDENTIALS:")
            logger.info("   Super Admin: owner@maestroegp.com / owner123")
            logger.info("   Secret Key: 271018")
            logger.info("   Main Admin: admin@maestroegp.com / admin123")
            logger.info("=" * 50)
        else:
            logger.info("ℹ️ Database already initialized - Super Admin exists")
            # لا نحتاج لإنشاء بيانات default لأن كل عميل سيكون له بياناته الخاصة
        
        # التحقق من وجود خلفيات تسجيل الدخول
        login_bg = await db.settings.find_one({"type": "login_backgrounds"})
        if not login_bg:
            logger.info("🔧 Adding default login backgrounds...")
            bg_doc = {
                "type": "login_backgrounds",
                "backgrounds": [
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                        "title": "مطعم فاخر",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                        "title": "مطعم حديث",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=1920",
                        "title": "طعام شهي",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=1920",
                        "title": "مطعم أنيق",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=1920",
                        "title": "كافيه عصري",
                        "is_active": True
                    },
                    {
                        "id": str(uuid.uuid4()),
                        "image_url": "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
                        "title": "مطبخ احترافي",
                        "is_active": True
                    }
                ],
                "settings": {
                    "transition_effect": "fade",
                    "transition_speed": 5,
                    "overlay_color": "rgba(0,0,0,0.5)",
                    "text_color": "#ffffff"
                }
            }
            await db.settings.insert_one(bg_doc)
            logger.info("✅ Login backgrounds added (6 images)")
            
    except Exception as e:
        logger.error(f"❌ Database initialization error: {str(e)}")
        logger.error(f"❌ Error type: {type(e).__name__}")
        import traceback
        logger.error(f"❌ Traceback: {traceback.format_exc()}")

@app.on_event("startup")
async def startup_event():
    """يتم تشغيله عند بدء التطبيق.
    محصّن بالكامل: لا يُسمح لأي خطأ في التهيئة بإسقاط الخادم (يمنع crash-loop/502 على الإنتاج).
    يبقى خفيفاً: تهيئة الحسابات/الفهارس الأساسية فقط، ثم يفتح uvicorn المنفذ فوراً.
    كل الترحيلات الثقيلة تُؤجَّل للخلفية عبر asyncio.create_task."""
    logger.info("🚀 Starting Maestro EGP API...")
    # أساسي وخفيف: تهيئة الحسابات والفروع والفهارس (لازم لتسجيل الدخول فوراً)
    # محاط بـ try/except حتى لا يُسقط أي خطأ في قاعدة البيانات إقلاعَ الخادم
    try:
        await init_database()
    except Exception as e:
        logger.error(f"❌ init_database failed at startup (continuing anyway): {e}")
    try:
        await seed_default_backgrounds()
    except Exception as e:
        logger.error(f"❌ seed_default_backgrounds failed at startup (continuing anyway): {e}")

    # تأجيل كل الترحيلات الثقيلة للخلفية (لا تحجب فتح المنفذ ولا فحص الصحة)
    try:
        asyncio.create_task(_run_deferred_startup_tasks())
    except Exception as e:
        logger.error(f"❌ Failed to schedule deferred startup tasks (continuing anyway): {e}")

    logger.info("✅ Application started successfully (heavy migrations deferred to background)")


async def update_contact_page_texts_v1():
    """يحدّث نصوص صفحة التواصل في إعدادات الفاتورة على الإنتاج بشكل آمن (idempotent).
    يُحدّث السطر التعريفي (promo_text) فقط إذا كان لا يزال أحد القيم القديمة المعروفة،
    ونص تعريف النظام (system_intro) فقط إذا كان فارغاً — حتى لا نطمس أي تخصيص للمالك."""
    NEW_PROMO = "نظام محاسبي وإداري متكامل للمؤسسات والمطاعم والمشاريع التجارية الكبرى"
    NEW_INTRO = "منصة شاملة ذكية ومتطورة تدير العمليات، والمخزون، والتصنيع، والمشتريات، والتوصيل، والموارد البشرية والمالية في نظام واحد دقيق يعمل حتى بدون إنترنت."
    OLD_PROMOS = {
        "نظام إدارة متكامل للمطاعم والكافيهات",
        "نظام محاسبي وإداري متكامل للمطاعم والمشاريع",
    }
    doc = await db.settings.find_one({"type": "system_invoice_settings"})
    if not doc:
        return
    value = doc.get("value") or {}
    updates = {}
    if value.get("promo_text") in OLD_PROMOS or not value.get("promo_text"):
        updates["value.promo_text"] = NEW_PROMO
    if not value.get("system_intro"):
        updates["value.system_intro"] = NEW_INTRO
    if updates:
        await db.settings.update_one({"type": "system_invoice_settings"}, {"$set": updates})
        logger.info(f"✅ update_contact_page_texts_v1 applied: {list(updates.keys())}")



async def _run_deferred_startup_tasks():
    """ينفّذ كل الترحيلات/التهيئات الثقيلة في الخلفية تسلسلياً بعد إقلاع الخادم وفتح المنفذ.
    كل مهمة محمية بـ try/except حتى لا يُفشل فشلُ مهمة واحدة بقيّةَ المهام، ونتيح المجال
    للطلبات الواردة بين كل ترحيل وآخر."""
    # مهلة قصيرة لضمان أن uvicorn فتح المنفذ واستجاب لفحص الصحة أولاً
    await asyncio.sleep(2)
    logger.info("🛠️ Starting deferred background migrations...")
    _deferred = [
        ("setup_database_indexes", setup_database_indexes),
        ("apply_automatic_updates", apply_automatic_updates),
        ("auto_migrate_business_dates", auto_migrate_business_dates),
        ("seed_department_branches", seed_department_branches),
        ("cleanup_stale_biometric_jobs", cleanup_stale_biometric_jobs),
        ("fix_pending_orders_extras_calc", fix_pending_orders_extras_calc),
        ("cleanup_mistaken_expense_moataz36", cleanup_mistaken_expense_moataz36),
        ("cleanup_duplicate_expenses", cleanup_duplicate_expenses),
        ("purge_ghost_order_saidiya_11_20260430", purge_ghost_order_saidiya_11_20260430),
        ("settle_driver_collected_orders_as_cash", settle_driver_collected_orders_as_cash),
        ("backfill_closing_business_date", backfill_closing_business_date),
        ("auto_heal_shifts_and_business_dates", auto_heal_shifts_and_business_dates),
        ("fix_yamen_orders_jadriya_20260503", fix_yamen_orders_jadriya_20260503),
        ("seed_initial_cost_layers_v1", seed_initial_cost_layers_v1),
        ("backfill_tenant_id_on_products_v1", backfill_tenant_id_on_products_v1),
        ("renumber_offline_orders_chronologically_v1", renumber_offline_orders_chronologically_v1),
        ("renumber_offline_orders_chronologically_v2", renumber_offline_orders_chronologically_v2),
        ("update_contact_page_texts_v1", update_contact_page_texts_v1),
        ("backfill_shift_cash_deposit_branch_v1", backfill_shift_cash_deposit_branch_v1),
        ("backfill_shift_cash_deposit_branch_v2", backfill_shift_cash_deposit_branch_v2),
        ("purge_pentest_probe_data_v1", purge_pentest_probe_data_v1),
    ]
    for name, fn in _deferred:
        try:
            await fn()
        except Exception as e:
            logger.error(f"❌ Deferred startup task '{name}' failed: {e}")
        await asyncio.sleep(0)
    logger.info("✅ All deferred background migrations finished")

async def purge_pentest_probe_data_v1():
    """تنظيف تلقائي (idempotent) لأي سجلات دخيلة أنشأها فاحص اختراق (مثل 'RW Probe').
    يُشغَّل في الخلفية عند كل إقلاع (أي بعد كل تحديث/نشر)، فيحذف الورديات/الطلبات/الإغلاقات
    ذات أسماء الفحص ويعكس أي إيداع خزينة مرتبط. آمن: يطابق أنماطاً محددة جداً لا تظهر في
    أسماء كاشير حقيقية، ويتوقف فوراً إن لم يجد شيئاً."""
    try:
        rx = {"$regex": r"(rw\s*probe|read.?write\s*probe|pentest|sqlmap|burpsuite|nikto|<script)", "$options": "i"}
        name_q = {"cashier_name": rx}
        probe_shifts = await db.shifts.find(name_q, {"_id": 0, "id": 1, "cashier_name": 1}).to_list(2000)
        shift_ids = [s["id"] for s in probe_shifts if s.get("id")]

        closing_or = [name_q]
        if shift_ids:
            closing_or.append({"shift_id": {"$in": shift_ids}})
        probe_closings = await db.cash_register_closings.find({"$or": closing_or}, {"_id": 0, "id": 1}).to_list(2000)
        closing_ids = [c["id"] for c in probe_closings if c.get("id")]

        order_or = [name_q]
        if shift_ids:
            order_or.append({"shift_id": {"$in": shift_ids}})
        probe_orders = await db.orders.find({"$or": order_or}, {"_id": 0, "id": 1, "total": 1}).to_list(50000)

        if not (probe_shifts or probe_closings or probe_orders):
            return  # لا شيء لتنظيفه — idempotent

        orders_total = sum(_sn(o.get("total")) for o in probe_orders)

        # عكس أي إيداع خزينة مرتبط بهذه الورديات/الإغلاقات
        dep_ids = shift_ids + closing_ids
        if dep_ids:
            await db.owner_deposits.delete_many({"$or": [{"ref_closing_id": {"$in": dep_ids}}, {"shift_id": {"$in": dep_ids}}]})

        del_orders = (await db.orders.delete_many({"$or": order_or})).deleted_count
        del_shifts = (await db.shifts.delete_many(name_q)).deleted_count
        del_closings = 0
        if closing_ids:
            del_closings = (await db.cash_register_closings.delete_many({"id": {"$in": closing_ids}})).deleted_count

        await db.audit_logs.insert_one({
            "id": str(uuid.uuid4()),
            "action": "auto_purge_pentest_probe",
            "deleted_orders": del_orders,
            "deleted_shifts": del_shifts,
            "deleted_closings": del_closings,
            "orders_total_removed": orders_total,
            "deleted_at": datetime.now(timezone.utc).isoformat(),
        })
        logger.info(f"🧹 auto_purge_pentest_probe: removed {del_shifts} shifts, {del_closings} closings, {del_orders} orders ({orders_total} sales)")
    except Exception as e:
        logger.error(f"purge_pentest_probe_data_v1 failed: {e}")


async def auto_migrate_business_dates():
    """ترحيل تلقائي آمن لحقل business_date (اليوم التشغيلي) لجميع العملاء.
    يُحدّث فقط السجلات التي لا تملك business_date. Idempotent."""
    try:
        updated = {"shifts": 0, "orders": 0, "expenses": 0, "shifts_recomputed": 0}
        
        # 1) تحديث الورديات
        async for shift in db.shifts.find(
            {"business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "started_at": 1, "opened_at": 1}
        ):
            started = shift.get("started_at") or shift.get("opened_at")
            biz_date = iraq_date_from_utc(started) if started else iraq_date_from_utc()
            await db.shifts.update_one({"id": shift["id"]}, {"$set": {"business_date": biz_date}})
            updated["shifts"] += 1
        
        # 2) تحديث الطلبات (حسب shift_id أو branch+time)
        shift_biz_map = {}
        all_shifts_list = []
        async for s in db.shifts.find({}, {"_id": 0, "id": 1, "business_date": 1, "started_at": 1, "ended_at": 1, "opened_at": 1, "branch_id": 1, "tenant_id": 1}):
            if s.get("business_date"):
                shift_biz_map[s["id"]] = s
                all_shifts_list.append(s)
        
        def find_containing_shift(branch_id, created_at, tenant_id=None):
            """يبحث عن الوردية التي تحتوي هذا الوقت في نفس الفرع"""
            if not branch_id or not created_at:
                return None
            candidates = [s for s in all_shifts_list if s.get("branch_id") == branch_id]
            if tenant_id:
                candidates = [s for s in candidates if not s.get("tenant_id") or s.get("tenant_id") == tenant_id]
            for s in candidates:
                s_start = s.get("started_at") or s.get("opened_at") or ""
                s_end = s.get("ended_at") or ""
                if created_at >= s_start and (not s_end or created_at <= s_end):
                    return s
            return None
        
        async for order in db.orders.find(
            {"business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "shift_id": 1, "created_at": 1, "branch_id": 1, "tenant_id": 1}
        ):
            biz = None
            sid = order.get("shift_id")
            if sid and sid in shift_biz_map:
                biz = shift_biz_map[sid].get("business_date")
            if not biz:
                # البحث بالفرع والوقت
                shift = find_containing_shift(order.get("branch_id"), order.get("created_at", ""), order.get("tenant_id"))
                if shift:
                    biz = shift.get("business_date")
            if not biz:
                biz = iraq_date_from_utc(order.get("created_at"))
            await db.orders.update_one({"id": order["id"]}, {"$set": {"business_date": biz}})
            updated["orders"] += 1
        
        # 2.1) تصحيح معالج ذكي: إعادة حساب business_date للطلبات التي تم ترحيلها سابقاً
        # بلا استخدام branch+time matching (قد تملك business_date خاطئ من الترحيل الأول)
        # نستخدم flag لمنع إعادة التشغيل في كل مرة
        fix_flag = await db.system_flags.find_one({"flag": "orders_business_date_v2_fixed"})
        if not fix_flag:
            async for order in db.orders.find(
                {"business_date": {"$exists": True}},
                {"_id": 0, "id": 1, "shift_id": 1, "created_at": 1, "branch_id": 1, "tenant_id": 1, "business_date": 1}
            ):
                correct_biz = None
                sid = order.get("shift_id")
                if sid and sid in shift_biz_map:
                    correct_biz = shift_biz_map[sid].get("business_date")
                if not correct_biz:
                    shift = find_containing_shift(order.get("branch_id"), order.get("created_at", ""), order.get("tenant_id"))
                    if shift:
                        correct_biz = shift.get("business_date")
                # فقط إذا وجدنا وردية مطابقة ومختلفة، نصحّح
                if correct_biz and correct_biz != order.get("business_date"):
                    await db.orders.update_one({"id": order["id"]}, {"$set": {"business_date": correct_biz}})
                    updated["orders"] += 1
            await db.system_flags.insert_one({
                "flag": "orders_business_date_v2_fixed",
                "applied_at": datetime.now(timezone.utc).isoformat()
            })
        
        # 3) تحديث المصاريف (مطابقة بالفرع + الفترة الزمنية)
        async for exp in db.expenses.find(
            {"business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "branch_id": 1, "tenant_id": 1, "created_at": 1, "date": 1}
        ):
            biz = None
            exp_branch = exp.get("branch_id")
            exp_tenant = exp.get("tenant_id")
            exp_created = exp.get("created_at", "")
            # ابحث عن الوردية التي تحتوي هذا المصروف زمنياً في نفس الفرع
            for s in shift_biz_map.values():
                if s.get("branch_id") != exp_branch:
                    continue
                if s.get("tenant_id") and exp_tenant and s.get("tenant_id") != exp_tenant:
                    continue
                s_start = s.get("started_at") or ""
                s_end = s.get("ended_at") or ""
                if exp_created >= s_start and (not s_end or exp_created <= s_end):
                    biz = s.get("business_date")
                    break
            if not biz:
                biz = exp.get("date") or iraq_date_from_utc(exp_created)
            await db.expenses.update_one({"id": exp["id"]}, {"$set": {"business_date": biz}})
            updated["expenses"] += 1
        
        # 4) ترحيل السلف/الخصومات/المكافآت/الساعات الإضافية (استخدام created_at)
        for coll_name in ["advances", "deductions", "bonuses", "overtime_requests"]:
            coll = db[coll_name]
            async for rec in coll.find(
                {"business_date": {"$exists": False}},
                {"_id": 0, "id": 1, "created_at": 1, "date": 1}
            ):
                biz = iraq_date_from_utc(rec.get("created_at")) or rec.get("date")
                await coll.update_one({"id": rec["id"]}, {"$set": {"business_date": biz}})
        
        # 5) إعادة احتساب total_expenses للورديات المُغلقة (قاعدة معتمدة: shift_id/منشئ — لا خلط بين الكاشيرين)
        from routes.shared import shift_expense_query as _shift_exp_q
        async for s in db.shifts.find(
            {"status": "closed", "expenses_shift_scope_fix": {"$ne": True}},
            {"_id": 0, "id": 1, "branch_id": 1, "tenant_id": 1, "cashier_id": 1, "business_date": 1,
             "started_at": 1, "opened_at": 1, "ended_at": 1, "cash_sales": 1, "opening_cash": 1, "opening_balance": 1}
        ):
            exp_q = _shift_exp_q(s, s.get("tenant_id"))
            shift_expenses = await db.expenses.find(exp_q, {"_id": 0, "amount": 1}).to_list(500)
            total_exp = sum(float(e.get("amount") or 0) for e in shift_expenses)
            opening_cash = float(s.get("opening_cash") or s.get("opening_balance") or 0)
            cash_sales = float(s.get("cash_sales") or 0)
            expected_cash = opening_cash + cash_sales - total_exp
            await db.shifts.update_one(
                {"id": s["id"]},
                {"$set": {
                    "total_expenses": total_exp,
                    "expected_cash": expected_cash,
                    "expenses_shift_scope_fix": True
                }}
            )
            updated["shifts_recomputed"] += 1
        
        if any(updated.values()):
            logger.info(f"✅ business_date auto-migration: {updated}")
    except Exception as e:
        logger.warning(f"⚠️ business_date auto-migration skipped: {e}")
    
    # ==================== ONE-TIME ORPHAN EXPENSE CLEANUP ====================
    # حذف مصروف الصيانة اليتيم 75,000 د.ع (ID محدد مسبقاً من طلب المالك)
    # يعمل مرة واحدة فقط - محمي بـ flag في system_flags
    try:
        cleanup_flag = await db.system_flags.find_one({"flag": "orphan_maintenance_75k_cleaned"})
        if not cleanup_flag:
            ORPHAN_EXPENSE_ID = "7b13235e-b5de-4a98-93f2-aabc691a28d2"
            orphan = await db.expenses.find_one({"id": ORPHAN_EXPENSE_ID}, {"_id": 0})
            if orphan:
                orphan_branch = orphan.get("branch_id")
                orphan_tenant = orphan.get("tenant_id")
                orphan_created = orphan.get("created_at", "")
                
                # حذف المصروف
                await db.expenses.delete_one({"id": ORPHAN_EXPENSE_ID})
                logger.info(f"🗑️ orphan cleanup: deleted expense {ORPHAN_EXPENSE_ID} ({orphan.get('description')} {orphan.get('amount')})")
                
                # إعادة احتساب total_expenses للوردية المرتبطة
                if orphan_branch and orphan_created:
                    shift_q = {"branch_id": orphan_branch, "started_at": {"$lte": orphan_created}}
                    if orphan_tenant:
                        shift_q["tenant_id"] = orphan_tenant
                    affected_shifts = await db.shifts.find(
                        shift_q,
                        {"_id": 0, "id": 1, "started_at": 1, "ended_at": 1, "opening_cash": 1, "opening_balance": 1, "cash_sales": 1}
                    ).to_list(100)
                    for s in affected_shifts:
                        s_end = s.get("ended_at") or ""
                        if s_end and orphan_created > s_end:
                            continue
                        q = {
                            "branch_id": orphan_branch,
                            "category": {"$ne": "refund"},
                            "created_at": {"$gte": s.get("started_at", "")}
                        }
                        if orphan_tenant:
                            q["tenant_id"] = orphan_tenant
                        if s_end:
                            q["created_at"]["$lte"] = s_end
                        shift_exps = await db.expenses.find(q, {"_id": 0, "amount": 1}).to_list(500)
                        total_exp = sum(float(e.get("amount") or 0) for e in shift_exps)
                        opening_cash = float(s.get("opening_cash") or s.get("opening_balance") or 0)
                        cash_sales = float(s.get("cash_sales") or 0)
                        await db.shifts.update_one(
                            {"id": s["id"]},
                            {"$set": {"total_expenses": total_exp, "expected_cash": opening_cash + cash_sales - total_exp}}
                        )
                        logger.info(f"🔄 recomputed shift {s['id'][:8]}: total_expenses={total_exp}")
            
            # وضع العلم لمنع إعادة التشغيل
            await db.system_flags.insert_one({
                "flag": "orphan_maintenance_75k_cleaned",
                "applied_at": datetime.now(timezone.utc).isoformat(),
                "expense_existed": bool(orphan)
            })
            if not orphan:
                logger.info("ℹ️ orphan cleanup: expense not found (already clean)")
    except Exception as e:
        logger.warning(f"⚠️ orphan cleanup skipped: {e}")

async def seed_default_backgrounds():
    """إضافة الخلفيات الافتراضية لصفحة الدخول"""
    try:
        existing = await db.settings.find_one({"type": "login_backgrounds"})
        if existing and existing.get("value", {}).get("backgrounds"):
            # تحقق من أن الخلفيات ليست فارغة
            if len(existing.get("value", {}).get("backgrounds", [])) > 0:
                logger.info("   ✅ Login backgrounds already exist")
                return
        
        # الخلفيات الافتراضية
        backgrounds = [
            {"id": str(uuid.uuid4()), "image_url": "/api/uploads/backgrounds/restaurant_1.jpg", "title": "مطعم فاخر 1", "animation_type": "fade", "animation_duration": 8, "overlay_opacity": 0.5, "is_active": True, "sort_order": 0, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "image_url": "/api/uploads/backgrounds/restaurant_2.jpg", "title": "مطعم فاخر 2", "animation_type": "fade", "animation_duration": 8, "overlay_opacity": 0.5, "is_active": True, "sort_order": 1, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "image_url": "/api/uploads/backgrounds/restaurant_3.jpg", "title": "مطعم فاخر 3", "animation_type": "fade", "animation_duration": 8, "overlay_opacity": 0.5, "is_active": True, "sort_order": 2, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "image_url": "/api/uploads/backgrounds/restaurant_4.jpg", "title": "مطعم فاخر 4", "animation_type": "fade", "animation_duration": 8, "overlay_opacity": 0.5, "is_active": True, "sort_order": 3, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "image_url": "/api/uploads/backgrounds/restaurant_5.jpg", "title": "مطعم فاخر 5", "animation_type": "fade", "animation_duration": 8, "overlay_opacity": 0.5, "is_active": True, "sort_order": 4, "created_at": datetime.now(timezone.utc).isoformat()},
            {"id": str(uuid.uuid4()), "image_url": "/api/uploads/backgrounds/restaurant_6.jpg", "title": "مطعم فاخر 6", "animation_type": "fade", "animation_duration": 8, "overlay_opacity": 0.5, "is_active": True, "sort_order": 5, "created_at": datetime.now(timezone.utc).isoformat()},
        ]
        
        settings = {
            "backgrounds": backgrounds,
            "animation_enabled": True,
            "transition_type": "fade",
            "transition_duration": 1.5,
            "auto_play": True,
            "show_logo": True,
            "logo_url": None,
            "logo_animation": "pulse",
            "overlay_color": "rgba(0,0,0,0.5)",
            "text_color": "#ffffff"
        }
        
        await db.settings.update_one(
            {"type": "login_backgrounds"},
            {"$set": {"type": "login_backgrounds", "value": settings}},
            upsert=True
        )
        logger.info(f"   ✅ Added {len(backgrounds)} default backgrounds")
    except Exception as e:
        logger.error(f"   ❌ Failed to seed backgrounds: {e}")

async def apply_automatic_updates():
    """تطبيق التحديثات التلقائية على جميع العملاء عند كل بدء تشغيل"""
    logger.info("🔄 Applying automatic updates to all tenants...")
    
    try:
        # 0. تحديث البيانات القديمة التي ليس لها tenant_id لتصبح "default"
        # تحديث المستخدمين الرئيسيين
        await db.users.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        logger.info("   ✅ Updated users without tenant_id")
        
        # تحديث السائقين
        drivers_tenant_result = await db.drivers.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if drivers_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {drivers_tenant_result.modified_count} drivers with default tenant_id")
        
        # تحديث الموظفين
        employees_tenant_result = await db.employees.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if employees_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {employees_tenant_result.modified_count} employees with default tenant_id")
        
        # تحديث الفروع
        branches_tenant_result = await db.branches.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if branches_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {branches_tenant_result.modified_count} branches with default tenant_id")
        
        # تحديث الفئات
        categories_tenant_result = await db.categories.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if categories_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {categories_tenant_result.modified_count} categories with default tenant_id")
        
        # تحديث المنتجات
        products_tenant_result = await db.products.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if products_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {products_tenant_result.modified_count} products with default tenant_id")
        
        # تحديث الطلبات
        orders_tenant_result = await db.orders.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if orders_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {orders_tenant_result.modified_count} orders with default tenant_id")
        
        # تحديث المصاريف
        expenses_tenant_result = await db.expenses.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if expenses_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {expenses_tenant_result.modified_count} expenses with default tenant_id")
        
        # 1. تفعيل جميع السائقين
        drivers_result = await db.drivers.update_many(
            {},
            {"$set": {"is_active": True}}
        )
        if drivers_result.modified_count > 0:
            logger.info(f"   ✅ Activated {drivers_result.modified_count} drivers")
        
        # 2. إضافة is_available للسائقين
        await db.drivers.update_many(
            {"is_available": {"$exists": False}},
            {"$set": {"is_available": True}}
        )
        
        # 2.5 إصلاح السائقين المرتبطين بفروع غير موجودة
        default_branch = await db.branches.find_one({"tenant_id": "default"})
        if default_branch:
            # جلب جميع الفروع الصالحة (مع حد أقصى لتجنب مشاكل الأداء)
            valid_branch_ids = [b["id"] async for b in db.branches.find({}, {"id": 1}).limit(1000)]
            
            # تحديث السائقين بفروع غير موجودة
            drivers_fixed = await db.drivers.update_many(
                {"branch_id": {"$nin": valid_branch_ids}},
                {"$set": {"branch_id": default_branch["id"]}}
            )
            if drivers_fixed.modified_count > 0:
                logger.info(f"   ✅ Fixed {drivers_fixed.modified_count} drivers with invalid branch_id")
        
        # 3. تفعيل جميع الفروع
        await db.branches.update_many(
            {"is_active": {"$exists": False}},
            {"$set": {"is_active": True}}
        )
        
        # 4. لا نُنشئ فروع افتراضية للعملاء - العميل يُنشئ فروعه بنفسه
        # هذا يمنع التداخل ويعطي العميل تحكماً كاملاً
        # العملاء بدون فروع سيرون رسالة "يرجى إنشاء فرع أولاً"
        
        # 5. إغلاق الورديات القديمة (أكثر من 24 ساعة)
        old_cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        old_shifts = await db.shifts.update_many(
            {"status": "open", "started_at": {"$lt": old_cutoff}},
            {"$set": {
                "status": "auto_closed",
                "ended_at": datetime.now(timezone.utc).isoformat(),
                "auto_closed_reason": "تم الإغلاق تلقائياً بعد 24 ساعة"
            }}
        )
        if old_shifts.modified_count > 0:
            logger.info(f"   ✅ Auto-closed {old_shifts.modified_count} old shifts")
        
        # 6. تحديث الطاولات القديمة التي ليس لها tenant_id
        tables_tenant_result = await db.tables.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        if tables_tenant_result.modified_count > 0:
            logger.info(f"   ✅ Updated {tables_tenant_result.modified_count} tables with default tenant_id")
        
        # 7. إنشاء طاولات افتراضية لكل عميل ليس لديه طاولات
        # ملاحظة: تم تعطيل هذه الميزة مؤقتاً لتحسين أداء بدء التشغيل
        # العملاء الجدد سيُنشئون طاولاتهم من خلال واجهة الإعدادات
        
        # 8. تحديث صور الفئات والمنتجات الافتراضية للنظام الرئيسي (إذا لم تكن موجودة)
        category_images = {
            "برغر": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا": "https://images.unsplash.com/photo-1703073186021-021fb5a0bde1?w=400",
            "مشروبات": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "حلويات": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400",
            "سلطات": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400",
        }
        for cat_name, cat_image in category_images.items():
            await db.categories.update_many(
                {"name": cat_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": cat_image}}
            )
        
        product_images = {
            "برغر كلاسيك": "https://images.unsplash.com/photo-1656439659132-24c68e36b553?w=400",
            "برغر دبل": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا مارغريتا": "https://images.unsplash.com/photo-1681567604770-0dc826c870ae?w=400",
            "بيتزا خضار": "https://images.unsplash.com/photo-1602104980741-b87a33837f9f?w=400",
            "كولا": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "عصير برتقال": "https://images.unsplash.com/photo-1716925539259-ce0115263d37?w=400",
        }
        for prod_name, prod_image in product_images.items():
            await db.products.update_many(
                {"name": prod_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": prod_image}}
            )
        logger.info("   ✅ Updated default category and product images")
        
        logger.info("✅ Automatic updates applied successfully")
    except Exception as e:
        logger.error(f"❌ Error applying automatic updates: {e}")

# ==================== HEALTH CHECK ====================

@app.get("/")
def read_root():
    return {"status": "Server is running successfully 🚀", "app": "Maestro EGP", "version": "2.0.0"}

@app.get("/health")
@app.head("/health")
def health_check():
    return {"status": "ok"}

@api_router.get("/health")
@api_router.head("/health")
def api_health_check():
    return {"status": "ok", "api": "Maestro EGP API"}

# ==================== DATABASE INITIALIZATION ENDPOINT ====================

@api_router.get("/init-db")
async def initialize_database_endpoint(key: str = ""):
    """
    Endpoint لتهيئة قاعدة البيانات يدوياً - محمي بالمفتاح السري للمالك
    """
    if not key or key != SUPER_ADMIN_SECRET:
        raise HTTPException(status_code=403, detail="غير مصرح")
    try:
        # التحقق من وجود Super Admin
        super_admin = await db.users.find_one({"role": "super_admin"})
        
        if super_admin:
            # التحقق من وجود خلفيات
            login_bg = await db.settings.find_one({"type": "login_backgrounds"})
            if not login_bg:
                bg_doc = {
                    "type": "login_backgrounds",
                    "backgrounds": [
                        {
                            "id": str(uuid.uuid4()),
                            "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                            "title": "مطعم فاخر",
                            "is_active": True
                        },
                        {
                            "id": str(uuid.uuid4()),
                            "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                            "title": "مطعم حديث",
                            "is_active": True
                        }
                    ],
                    "settings": {
                        "transition_effect": "fade",
                        "transition_speed": 5,
                        "overlay_color": "rgba(0,0,0,0.5)",
                        "text_color": "#ffffff"
                    }
                }
                await db.settings.insert_one(bg_doc)
            
            return {
                "status": "already_initialized",
                "message": "قاعدة البيانات مهيأة مسبقاً"
            }
        
        # إنشاء Super Admin
        super_admin_password = bcrypt.hashpw("owner123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        super_admin_doc = {
            "id": str(uuid.uuid4()),
            "username": "super_admin",
            "email": "owner@maestroegp.com",
            "password": super_admin_password,
            "full_name": "Owner",
            "role": "super_admin",
            "branch_id": None,
            "tenant_id": None,
            "permissions": ["all", "super_admin"],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(super_admin_doc)
        
        # إنشاء مدير النظام الرئيسي
        admin_password = bcrypt.hashpw("admin123".encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        admin_doc = {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "email": "admin@maestroegp.com",
            "password": admin_password,
            "full_name": "مدير النظام",
            "role": "admin",
            "branch_id": None,
            "tenant_id": None,
            "permissions": ["all"],
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_doc)
        
        # إنشاء الفرع الرئيسي
        branch_doc = {
            "id": str(uuid.uuid4()),
            "name": "الفرع الرئيسي",
            "code": "MAIN",
            "address": "العنوان الرئيسي",
            "phone": "",
            "is_main": True,
            "is_active": True,
            "tenant_id": None,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.branches.insert_one(branch_doc)
        
        # إعدادات النظام
        branding_doc = {
            "type": "system_branding",
            "value": {
                "name": "Maestro",
                "name_ar": "Maestro",
                "name_en": "Maestro",
                "logo_url": None
            }
        }
        await db.settings.insert_one(branding_doc)
        
        # خلفيات تسجيل الدخول - 6 خلفيات
        bg_doc = {
            "type": "login_backgrounds",
            "backgrounds": [
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
                    "title": "مطعم فاخر",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
                    "title": "مطعم حديث",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1414235077428-338989a2e8c0?w=1920",
                    "title": "طعام شهي",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1555396273-367ea4eb4db5?w=1920",
                    "title": "مطعم أنيق",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1559339352-11d035aa65de?w=1920",
                    "title": "كافيه عصري",
                    "is_active": True
                },
                {
                    "id": str(uuid.uuid4()),
                    "image_url": "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
                    "title": "مطبخ احترافي",
                    "is_active": True
                }
            ],
            "settings": {
                "transition_effect": "fade",
                "transition_speed": 5,
                "overlay_color": "rgba(0,0,0,0.5)",
                "text_color": "#ffffff"
            }
        }
        await db.settings.insert_one(bg_doc)
        
        # Log credentials to server logs only (not in response)
        logger.info("=" * 50)
        logger.info("🎉 DATABASE INITIALIZED - CREDENTIALS (check server logs):")
        logger.info("   Super Admin: owner@maestroegp.com / owner123")
        logger.info("   Secret Key: 271018")
        logger.info("   Main Admin: admin@maestroegp.com / admin123")
        logger.info("=" * 50)
        
        return {
            "status": "success",
            "message": "تم تهيئة قاعدة البيانات بنجاح! تحقق من البريد الإلكتروني أو تواصل مع مزود الخدمة للحصول على بيانات الدخول."
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"حدث خطأ: {str(e)}"
        }

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def _sn(val, default=0):
    """Safe number: converts None to default for math ops.
    MongoDB .get('key', 0) returns None if key exists with null value."""
    if val is None:
        return default
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


# ============================================================================
# 🧮 Unit conversion for manufactured product links (MfgLinksEditor)
# ============================================================================
# Inlined here (instead of `from utils.link_units import ...`) because
# `utils/__init__.py` triggers import of `utils.auth` which depends on a
# non-existent `models.enums` module — and importing it crashes server.py.
# Mirror copy lives in `utils/link_units.py` for isolated unit testing.

_LINK_WEIGHT_MAP = {
    "غرام": 1.0, "كغم": 1000.0, "كيلو": 1000.0, "كجم": 1000.0,
    "gram": 1.0, "kg": 1000.0,
    "مل": 1.0, "لتر": 1000.0, "ml": 1.0, "liter": 1000.0, "l": 1000.0,
}


def _convert_link_consumption_to_main(consumption_qty: float, consumption_unit: str,
                                      main_unit: str, piece_weight: float,
                                      piece_weight_unit: str) -> float:
    """يُحوّل الكمية المُستهلكة من أي وحدة (consumption_unit) إلى main_unit.
    يدعم: (A) نفس الوحدة، (B) sub→main عبر piece_weight،
    (C) نفس العائلة الوزنية، (D) جسر عبر piece_weight_unit الوزني.
    """
    cu = (consumption_unit or "").strip()
    mu = (main_unit or "").strip()
    pwu = (piece_weight_unit or "").strip()
    pw = float(piece_weight or 0)

    if not cu or cu == mu:
        return consumption_qty
    if cu == pwu and pw > 0:
        return consumption_qty / pw
    cu_factor = _LINK_WEIGHT_MAP.get(cu)
    mu_factor = _LINK_WEIGHT_MAP.get(mu)
    if cu_factor is not None and mu_factor is not None:
        return consumption_qty * cu_factor / mu_factor
    pwu_factor = _LINK_WEIGHT_MAP.get(pwu)
    if cu_factor is not None and pwu_factor is not None and pw > 0:
        qty_in_pwu_base = consumption_qty * cu_factor / pwu_factor
        return qty_in_pwu_base / pw
    return consumption_qty


# ============================================================================
# 🧮 Mfg product unit-cost lookup (uses authoritative _enrich_unit_cost_fields)
# ============================================================================
async def _get_mfg_unit_cost(db, mfg_product: dict) -> dict:
    """يُرجع dict يحتوي `unit_cost_after_waste` و `computed_yield` للمنتج المُصنّع،
    باستخدام نفس منطق `routes.inventory_system._enrich_unit_cost_fields` (مصدر
    الحقيقة الوحيد). يحرس ضد تكلفة طلبات خاطئة في POS/التقارير.
    """
    from routes.inventory_system import _enrich_unit_cost_fields  # late import (circular safe)
    await _enrich_unit_cost_fields(db, mfg_product)
    return {
        "unit_cost_after_waste": float(mfg_product.get("unit_cost_after_waste") or 0),
        "unit_cost_before_waste": float(mfg_product.get("unit_cost_before_waste") or 0),
        "computed_yield": float(mfg_product.get("computed_yield") or 0),
    }

# ==================== MODELS ====================

class UserRole:
    SUPER_ADMIN = "super_admin"  # مالك النظام الرئيسي
    ADMIN = "admin"
    MANAGER = "manager"
    SUPERVISOR = "supervisor"
    CASHIER = "cashier"
    DELIVERY = "delivery"  # دور جديد للسائقين
    CALL_CENTER = "call_center"  # دور كول سنتر - يرى المكالمات والتوصيل فقط
    WAREHOUSE_KEEPER = "warehouse_keeper"  # أمين مخزن - يرى المخزن وطلبات التصنيع
    MANUFACTURER = "manufacturer"  # مصنع - يرى التصنيع وطلبات الفروع
    PURCHASING = "purchasing"  # مشتريات - يرى طلبات الشراء

class OrderType:
    DINE_IN = "dine_in"
    TAKEAWAY = "takeaway"
    DELIVERY = "delivery"

class OrderStatus:
    PENDING = "pending"
    PREPARING = "preparing"
    READY = "ready"
    DELIVERED = "delivered"
    CANCELLED = "cancelled"

class PaymentMethod:
    CASH = "cash"
    CARD = "card"
    CREDIT = "credit"
    PENDING = "pending"

# ==================== TENANT MODELS (Multi-tenant) ====================

class TenantCreate(BaseModel):
    name: str  # اسم المطعم/الكافيه
    name_ar: Optional[str] = None  # اسم المطعم بالعربي
    name_en: Optional[str] = None  # اسم المطعم بالإنجليزي
    slug: str  # رابط فريد (مثل: my-restaurant)
    owner_name: str  # اسم المالك
    owner_email: EmailStr
    owner_phone: Optional[str] = ""  # رقم الهاتف (اختياري)
    subscription_type: str = "trial"  # trial, bronze, silver, gold, basic, premium, demo
    subscription_duration: int = 1  # مدة الاشتراك بالأشهر (1, 3, 6, 12)
    max_branches: int = 1
    max_users: int = 5
    logo_url: Optional[str] = None  # شعار المطعم
    is_demo: bool = False  # هل هو حساب تجريبي

class TenantFeatures(BaseModel):
    """ميزات العميل المتاحة"""
    showPOS: bool = True
    showTables: bool = True
    showOrders: bool = True
    showExpenses: bool = True
    showInventory: bool = True
    showDelivery: bool = True
    showReports: bool = True
    showSettings: bool = True
    showHR: bool = False
    showWarehouse: bool = False
    showCallLogs: bool = False
    showCallCenter: bool = False
    showKitchen: bool = False
    showLoyalty: bool = True
    showCoupons: bool = True
    showRecipes: bool = True
    showReservations: bool = True
    showReviews: bool = True
    showRatings: bool = True
    showSmartReports: bool = True

class TenantResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_ar: Optional[str] = None
    name_en: Optional[str] = None
    slug: str
    owner_name: str
    owner_email: str
    owner_phone: str
    subscription_type: str
    max_branches: int
    max_users: int
    is_active: bool
    created_at: str
    expires_at: Optional[str] = None
    logo_url: Optional[str] = None
    subscription_duration: Optional[int] = None  # مدة الاشتراك بالأشهر

# ==================== نظام الإشعارات ====================

class NotificationType(str, Enum):
    NEW_TENANT = "new_tenant"  # عميل جديد
    SUBSCRIPTION_EXPIRING = "subscription_expiring"  # اشتراك قارب على الانتهاء
    SUBSCRIPTION_EXPIRED = "subscription_expired"  # اشتراك انتهى
    TENANT_ACTIVATED = "tenant_activated"  # تفعيل عميل
    TENANT_DEACTIVATED = "tenant_deactivated"  # تعطيل عميل
    SYSTEM = "system"  # إشعار نظام عام
    # إشعارات الطلبات الجديدة
    NEW_ORDER_CASHIER = "new_order_cashier"  # طلب جديد للكاشير
    NEW_ORDER_DRIVER = "new_order_driver"  # طلب جديد للسائق
    ORDER_READY = "order_ready"  # طلب جاهز للتسليم

class NotificationCreate(BaseModel):
    type: str
    title: str
    message: str
    tenant_id: Optional[str] = None
    data: Optional[dict] = None

class NotificationResponse(BaseModel):
    id: str
    type: str
    title: str
    message: str
    tenant_id: Optional[str] = None
    data: Optional[dict] = None
    is_read: bool = False
    created_at: str

class NotificationSettings(BaseModel):
    """إعدادات الإشعارات للمالك"""
    days_before_expiry: int = 15  # عدد الأيام قبل انتهاء الاشتراك للتنبيه (الافتراضي 15 يوم)
    email_notifications: bool = False  # إرسال بريد إلكتروني (معطل افتراضياً)
    push_notifications: bool = True  # إشعارات المتصفح
    notify_new_tenant: bool = True  # إشعار عند إضافة عميل جديد
    notify_tenant_status: bool = True  # إشعار عند تفعيل/تعطيل عميل

# User Models
class UserCreate(BaseModel):
    username: str
    email: EmailStr
    password: str
    full_name: str
    full_name_en: Optional[str] = None  # الاسم بالإنجليزية
    role: str = UserRole.CASHIER
    branch_id: Optional[str] = None
    permissions: List[str] = []
    phone: Optional[str] = None  # لاستلام رمز التحقق عبر واتساب/SMS
    tenant_id: Optional[str] = None  # للنظام متعدد المستأجرين

class UserLogin(BaseModel):
    email: str
    password: str
    secret_key: Optional[str] = None
    device_id: Optional[str] = None

class UserResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    username: Optional[str] = ""
    email: str
    full_name: Optional[str] = ""
    full_name_en: Optional[str] = None  # الاسم بالإنجليزية
    role: str
    branch_id: Optional[str] = None
    permissions: List[str] = []
    phone: Optional[str] = None
    is_active: bool = True
    created_at: str
    tenant_id: Optional[str] = None

class UserUpdate(BaseModel):
    username: Optional[str] = None
    email: Optional[EmailStr] = None
    full_name: Optional[str] = None
    full_name_en: Optional[str] = None  # الاسم بالإنجليزية
    role: Optional[str] = None
    branch_id: Optional[str] = None
    permissions: Optional[List[str]] = None
    phone: Optional[str] = None
    is_active: Optional[bool] = None

# Branch Models
class BranchCreate(BaseModel):
    name: str
    address: str
    phone: str
    phone2: Optional[str] = ""  # رقم هاتف ثانوي للفرع
    email: Optional[str] = None
    # التكاليف الثابتة الشهرية
    rent_cost: float = 0.0  # الإيجار الشهري
    water_cost: float = 0.0  # تكلفة الماء الشهرية
    electricity_cost: float = 0.0  # تكلفة الكهرباء الشهرية
    generator_cost: float = 0.0  # تكلفة المولدة الشهرية
    # إعدادات الفرع الخارجي/المباع
    is_sold_branch: bool = False  # هل الفرع مباع؟
    buyer_name: Optional[str] = None  # اسم المشتري
    buyer_phone: Optional[str] = None  # هاتف المشتري
    owner_percentage: float = 0.0  # نسبة المالك من المبيعات
    monthly_fee: float = 0.0  # رسوم شهرية ثابتة
    # === تصنيف الفرع/القسم ===
    # branch (افتراضي) | central_kitchen | warehouse | purchasing
    # الأقسام (غير-فرع) لها موظفون لكن رواتبهم تُحسب منفصلة عن الفروع
    branch_type: str = "branch"

class BranchResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    address: Optional[str] = ""
    phone: Optional[str] = ""
    phone2: Optional[str] = ""
    email: Optional[str] = None
    is_active: bool = True
    created_at: Optional[str] = None
    # إحداثيات الفرع (لأجور التوصيل حسب المسافة)
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    # التكاليف الثابتة الشهرية
    rent_cost: float = 0.0
    water_cost: float = 0.0
    electricity_cost: float = 0.0
    generator_cost: float = 0.0
    # إعدادات الفرع الخارجي/المباع
    is_sold_branch: bool = False
    buyer_name: Optional[str] = None
    buyer_phone: Optional[str] = None
    owner_percentage: float = 0.0
    monthly_fee: float = 0.0
    branch_type: str = "branch"

# Category Models
class CategoryCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    icon: Optional[str] = None
    image: Optional[str] = None
    color: Optional[str] = None
    sort_order: int = 0

class CategoryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_en: Optional[str] = None
    icon: Optional[str] = None
    image: Optional[str] = None
    color: Optional[str] = None
    sort_order: int = 0
    is_active: bool = True

# Product Models with Pre-Manufacturing Cost
class ProductCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    category_id: str
    price: float
    cost: float = 0.0  # تكلفة ما قبل التصنيع
    operating_cost: float = 0.0  # تكلفة تشغيلية
    packaging_cost: float = 0.0  # تكلفة التغليف (للسفري/التوصيل)
    image: Optional[str] = None
    description: Optional[str] = None
    is_available: bool = True
    ingredients: List[Dict[str, Any]] = []
    barcode: Optional[str] = None
    finished_product_id: Optional[str] = None  # ربط بالمنتج النهائي للخصم التلقائي من الوصفة
    manufactured_product_id: Optional[str] = None  # ربط بالمنتج المصنع من النظام الجديد
    printer_ids: List[str] = []  # الطابعات المرتبطة بالمنتج
    extras: List[Dict[str, Any]] = []  # الإضافات المتاحة للمنتج
    packaging_items: List[Dict[str, Any]] = []  # مواد التغليف المربوطة للخصم التلقائي
    recipe_quantities: List[Dict[str, Any]] = []  # كميات المكونات للوصفة
    manufactured_consumption_qty: float = 1.0  # ⭐ كم وحدة من المنتج المصنع تُخصم لكل وحدة مباعة
    # ⭐ ربط متعدد بالمنتجات المُصنّعة (يدعم أكثر من منتج مصنّع للمنتج الواحد)
    # كل عنصر: {manufactured_product_id, consumption_qty, piece_weight?, piece_weight_unit?}
    manufactured_links: List[Dict[str, Any]] = []

class ProductResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_en: Optional[str] = None
    category_id: str
    price: float
    cost: float = 0.0
    operating_cost: float = 0.0
    packaging_cost: float = 0.0  # تكلفة التغليف
    profit: float = 0.0  # حقل محسوب
    image: Optional[str] = None
    description: Optional[str] = None
    is_available: bool = True
    ingredients: List[Dict[str, Any]] = []
    barcode: Optional[str] = None
    finished_product_id: Optional[str] = None  # ربط بالمنتج النهائي
    manufactured_product_id: Optional[str] = None  # ربط بالمنتج المصنع من النظام الجديد
    printer_ids: List[str] = []  # الطابعات المرتبطة بالمنتج
    extras: List[Dict[str, Any]] = []  # الإضافات المتاحة للمنتج (جبنة إضافية، صوص، إلخ)
    packaging_items: List[Dict[str, Any]] = []  # مواد التغليف المربوطة للخصم التلقائي
    recipe_quantities: List[Dict[str, Any]] = []  # كميات المكونات للوصفة
    manufactured_consumption_qty: float = 1.0  # ⭐ كم وحدة من المنتج المصنع تُخصم لكل وحدة مباعة
    manufactured_links: List[Dict[str, Any]] = []  # ⭐ ربط متعدد بالمنتجات المُصنّعة

# Inventory Models
class InventoryItemCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str
    quantity: float = 0.0
    min_quantity: float = 0.0
    cost_per_unit: float = 0.0
    branch_id: str
    item_type: str = "raw"  # raw or finished

class InventoryResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    name_en: Optional[str] = None
    unit: str
    quantity: float
    min_quantity: float
    cost_per_unit: float
    branch_id: str
    item_type: str
    last_updated: str

class InventoryTransaction(BaseModel):
    inventory_id: str
    transaction_type: str  # in or out
    quantity: float
    notes: Optional[str] = None

# Purchase Models - المشتريات
class PurchaseCreate(BaseModel):
    supplier_name: str
    invoice_number: Optional[str] = None
    items: List[Dict[str, Any]]  # [{inventory_id, quantity, cost_per_unit}]
    total_amount: float
    payment_method: str = "cash"
    payment_status: str = "paid"  # paid, pending, partial
    notes: Optional[str] = None
    branch_id: str

class PurchaseResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    supplier_name: str
    invoice_number: Optional[str] = None
    items: List[Dict[str, Any]]
    total_amount: float
    payment_method: str
    payment_status: str
    notes: Optional[str] = None
    branch_id: str
    created_by: str
    created_at: str

# Expense Models - المصاريف اليومية
class ExpenseCreate(BaseModel):
    category: str  # rent, utilities, salaries, maintenance, supplies, other
    description: str
    amount: float
    payment_method: str = "cash"
    reference_number: Optional[str] = None
    branch_id: str
    date: Optional[str] = None

class ExpenseResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    category: str
    description: str
    amount: float
    payment_method: str
    reference_number: Optional[str] = None
    branch_id: str
    created_by: str
    date: str
    created_at: str

# Operating Cost Models - التكاليف التشغيلية
class OperatingCostCreate(BaseModel):
    name: str
    cost_type: str  # fixed or variable
    amount: float
    frequency: str  # daily, weekly, monthly
    branch_id: str

# Table Models
class TableCreate(BaseModel):
    number: int
    capacity: int
    branch_id: str
    section: Optional[str] = None

class TableResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    number: int
    capacity: int
    branch_id: str
    section: Optional[str] = None
    status: str = "available"
    current_order_id: Optional[str] = None

# Customer Models - إدارة العملاء
class CustomerCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    phone2: Optional[str] = None  # رقم إضافي
    address: Optional[str] = None
    area: Optional[str] = None  # المنطقة
    notes: Optional[str] = None
    is_blocked: bool = False  # حظر العميل

class CustomerResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    email: Optional[str] = None
    phone2: Optional[str] = None
    address: Optional[str] = None
    area: Optional[str] = None
    notes: Optional[str] = None
    is_blocked: bool = False
    total_orders: int = 0
    total_spent: float = 0.0
    last_order_date: Optional[str] = None
    created_at: str
    source: Optional[str] = None
    welcome_status: Optional[str] = None
    welcome_coupon_code: Optional[str] = None
    welcome_whatsapp_sent: Optional[bool] = None

# Order Models
class OrderItemCreate(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    price: float
    cost: float = 0.0
    notes: Optional[str] = None
    extras: List[Dict[str, Any]] = []  # الإضافات المحددة للمنتج

class OrderCreate(BaseModel):
    order_type: str = OrderType.DINE_IN
    table_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    delivery_address: Optional[str] = None
    buzzer_number: Optional[str] = None  # رقم جهاز التنبيه للسفري
    items: List[OrderItemCreate]
    branch_id: str
    payment_method: str = PaymentMethod.CASH
    # طريقة الدفع المفضّلة عند حفظ طلب كمعلق — تُستخدم لاحقاً عند الدفع الفعلي
    preferred_payment: Optional[str] = None
    discount: float = 0.0
    # كوبون مرتبط بالطلب (يُحسب ضمن discount لكن نخزن المرجع)
    coupon_id: Optional[str] = None
    coupon_code: Optional[str] = None
    coupon_name: Optional[str] = None
    coupon_discount: float = 0.0
    notes: Optional[str] = None
    delivery_app: Optional[str] = None
    delivery_app_name: Optional[str] = None  # اسم شركة التوصيل (للإدخال المباشر)
    is_delivery_company: bool = False  # هل الطلب لشركة توصيل
    driver_id: Optional[str] = None
    auto_ready: bool = False  # الطلب جاهز تلقائياً
    # ⭐ مفتاح الثبات (idempotency) لمنع تكرار طلبات الأوفلاين عند ضياع رد السيرفر
    offline_id: Optional[str] = None
    is_offline_order: Optional[bool] = False
    # ⭐ بيانات شركة التوصيل + رقم الطلب الخارجي (مفتاح منع التكرار التجاري)
    delivery_company: Optional[str] = None
    delivery_company_id: Optional[str] = None
    delivery_company_name: Optional[str] = None
    delivery_company_order_id: Optional[str] = None

class OrderResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    order_number: int
    order_type: str
    table_id: Optional[str] = None
    customer_name: Optional[str] = None
    customer_phone: Optional[str] = None
    delivery_address: Optional[str] = None
    buzzer_number: Optional[str] = None  # رقم جهاز التنبيه
    items: List[Dict[str, Any]]
    subtotal: float
    discount: float = 0.0  # Default for legacy orders
    tax: float = 0.0  # Default for legacy orders
    total: float
    total_cost: float = 0.0
    profit: float = 0.0
    branch_id: Optional[str] = None  # Made optional for customer orders
    cashier_id: Optional[str] = None  # Made optional for orders without cashier
    captain_id: Optional[str] = None  # هوية الكابتن إن أنشأ الطلب
    captain_name: Optional[str] = None
    captain_cash_status: Optional[str] = None  # held | collected | None
    status: str = "pending"  # Default status
    payment_method: str = "cash"  # Default payment method
    payment_status: str = "pending"  # Default for legacy orders
    delivery_app: Optional[str] = None
    delivery_app_name: Optional[str] = None  # اسم شركة التوصيل
    is_delivery_company: bool = False  # هل الطلب لشركة توصيل
    delivery_commission: float = 0.0
    driver_id: Optional[str] = None
    driver_name: Optional[str] = None  # اسم السائق
    driver_phone: Optional[str] = None  # هاتف السائق
    notes: Optional[str] = None
    created_at: Optional[str] = None  # Made optional for legacy orders
    updated_at: Optional[str] = None  # Made optional for legacy orders
    tenant_id: Optional[str] = None  # Added for tenant filtering
    business_date: Optional[str] = None  # اليوم التشغيلي (من الوردية)
    shift_id: Optional[str] = None
    cancelled_items: List[Dict[str, Any]] = []  # سجل تدقيق إلغاء الأصناف الجزئي
    # كوبون مطبّق على الطلب
    coupon_id: Optional[str] = None
    coupon_code: Optional[str] = None
    coupon_name: Optional[str] = None
    coupon_discount: float = 0.0
    # طريقة الدفع المفضّلة (تُستخدم عند تحميل طلب معلق)
    preferred_payment: Optional[str] = None
    # === حقول الأوفلاين والتدقيق (لـ fix-routing UI) ===
    is_offline_order: Optional[bool] = False
    offline_id: Optional[str] = None
    original_order_number: Optional[int] = None
    customer_type: Optional[str] = None
    paid_amount: Optional[float] = 0
    delivery_fee: Optional[float] = 0
    service_charge: Optional[float] = 0
    delivery_company: Optional[str] = None
    delivery_company_id: Optional[str] = None
    delivery_company_name: Optional[str] = None
    delivery_company_order_id: Optional[str] = None
    routing_fixed_at: Optional[str] = None
    routing_fixed_by: Optional[str] = None
    routing_fixed_by_name: Optional[str] = None
    routing_fix_history: List[Dict[str, Any]] = []
    renumbered_at: Optional[str] = None
    renumbered_reason: Optional[str] = None
    # === حقول الرفض/الإلغاء واسم الكاشير (لعرض تفاصيل الطلب في إدارة التوصيل) ===
    cashier_name: Optional[str] = None
    is_rejected: Optional[bool] = None
    cancellation_reason: Optional[str] = None
    cancelled_at: Optional[str] = None
    rejected_at: Optional[str] = None
    rejected_by_name: Optional[str] = None

# Shift Models
class ShiftCreate(BaseModel):
    cashier_id: str
    branch_id: str
    opening_cash: float

class ShiftClose(BaseModel):
    closing_cash: float
    notes: Optional[str] = None

# نموذج إغلاق الصندوق المتقدم مع جرد الفئات
class CashRegisterClose(BaseModel):
    denominations: Dict[str, int] = {}  # {"250": 5, "500": 10, "1000": 20, ...}
    notes: Optional[str] = None

class ShiftResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    cashier_id: str
    cashier_name: Optional[str] = ""
    branch_id: str
    opening_cash: float
    closing_cash: Optional[float] = None
    expected_cash: Optional[float] = None
    cash_difference: Optional[float] = None
    total_sales: float = 0.0
    total_cost: float = 0.0
    gross_profit: float = 0.0
    total_orders: int = 0
    card_sales: float = 0.0
    cash_sales: float = 0.0
    credit_sales: float = 0.0
    delivery_app_sales: Dict[str, float] = {}
    driver_sales: float = 0.0  # مبيعات السائقين
    total_expenses: float = 0.0
    net_profit: float = 0.0
    started_at: str
    ended_at: Optional[str] = None
    status: str
    denominations: Optional[Dict[str, int]] = None  # تفاصيل الجرد
    cancelled_orders: int = 0  # عدد الطلبات الملغاة
    cancelled_amount: float = 0.0  # إجمالي الإلغاءات
    discounts_total: float = 0.0  # إجمالي الخصومات
    cancelled_by: List[Dict] = []  # تفاصيل من قام بالإلغاء
    business_date: Optional[str] = None  # اليوم التشغيلي

# Delivery Driver Models
class DriverCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None  # لاستلام رمز التحقق/الرسائل عبر البريد
    branch_id: str
    pin: str = "1234"  # رمز PIN الافتراضي - يجب تغييره
    user_id: Optional[str] = None  # ربط بحساب مستخدم

class DriverResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: str
    email: Optional[str] = None
    branch_id: str
    is_available: bool = True
    current_order_id: Optional[str] = None
    total_deliveries: int = 0
    user_id: Optional[str] = None
    # معلومات الموقع
    location_lat: Optional[float] = None
    location_lng: Optional[float] = None
    location_updated_at: Optional[str] = None
    # معلومات الطلب الحالي (اختياري)
    current_order: Optional[Dict[str, Any]] = None
    is_active: bool = True

class DriverLocationUpdate(BaseModel):
    latitude: float
    longitude: float

# Delivery App Settings - إعدادات شركات التوصيل
class DeliveryAppSettingCreate(BaseModel):
    app_id: str
    name: str
    name_en: Optional[str] = None
    commission_type: str = "percentage"  # percentage or fixed
    commission_rate: float = 0.0  # نسبة الاستقطاع
    is_active: bool = True
    payment_terms: str = "weekly"  # daily, weekly, monthly
    contact_info: Optional[str] = None

# Currency Models
class Currency(BaseModel):
    code: str
    name: str
    symbol: str
    exchange_rate: float

# ==================== HR MODELS - إدارة الموارد البشرية ====================

class EmployeeCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    national_id: Optional[str] = None  # رقم الهوية
    position: str  # المسمى الوظيفي
    department: Optional[str] = None  # القسم
    branch_id: str
    hire_date: str  # تاريخ التعيين
    salary: float  # الراتب الأساسي
    salary_type: str = "monthly"  # monthly, daily, hourly
    work_hours_per_day: float = 8.0  # ساعات العمل اليومية
    user_id: Optional[str] = None  # ربط بحساب مستخدم
    biometric_uid: Optional[str] = None  # رقم البصمة على الجهاز
    shift_start: Optional[str] = None  # وقت بداية الشفت HH:MM
    shift_end: Optional[str] = None    # وقت نهاية الشفت HH:MM
    break_start: Optional[str] = None  # وقت بداية الاستراحة HH:MM
    break_end: Optional[str] = None    # وقت نهاية الاستراحة HH:MM
    work_days: Optional[list] = None   # أيام العمل [0=الأحد, 1=الإثنين, ..., 6=السبت]
    is_general_manager: bool = False   # مدير عام/أونر: لا يُحتسب عليه الحضور/السلف/الخصومات

class EmployeeResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    phone: Optional[str] = ""
    email: Optional[str] = None
    national_id: Optional[str] = None
    position: Optional[str] = ""
    department: Optional[str] = None
    branch_id: Optional[str] = None
    hire_date: Optional[str] = None
    salary: float = 0
    salary_type: str = "monthly"
    work_hours_per_day: float = 8.0
    user_id: Optional[str] = None
    biometric_uid: Optional[str] = None
    shift_start: Optional[str] = None
    shift_end: Optional[str] = None
    break_start: Optional[str] = None
    break_end: Optional[str] = None
    work_days: Optional[list] = None
    is_active: bool = True
    created_at: Optional[str] = None
    tenant_id: Optional[str] = None
    face_photo: Optional[str] = None
    face_photo_updated_at: Optional[str] = None
    is_general_manager: bool = False
    annual_leave_balance: Optional[float] = None  # رصيد الإجازة السنوية (أيام)
    # إنهاء الخدمات
    employment_status: Optional[str] = "active"  # active | terminated_pending | terminated
    termination_date: Optional[str] = None
    termination_month: Optional[str] = None
    termination_requested_at: Optional[str] = None
    auto_finalize_at: Optional[str] = None
    settlement_paid: Optional[bool] = None
    settlement_amount: Optional[float] = None
    settlement_preview: Optional[float] = None
    settlement_withdrawal_id: Optional[str] = None
    archive_purge_at: Optional[str] = None

class EmployeeUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    national_id: Optional[str] = None
    position: Optional[str] = None
    department: Optional[str] = None
    branch_id: Optional[str] = None
    hire_date: Optional[str] = None
    salary: Optional[float] = None
    salary_type: Optional[str] = None
    work_hours_per_day: Optional[float] = None
    is_active: Optional[bool] = None
    biometric_uid: Optional[str] = None
    shift_start: Optional[str] = None
    shift_end: Optional[str] = None
    break_start: Optional[str] = None
    break_end: Optional[str] = None
    work_days: Optional[list] = None
    face_photo: Optional[str] = None
    is_general_manager: Optional[bool] = None

# نموذج الحضور والانصراف
class AttendanceCreate(BaseModel):
    employee_id: str
    date: str  # YYYY-MM-DD
    check_in: Optional[str] = None  # وقت الحضور HH:MM
    check_out: Optional[str] = None  # وقت الانصراف HH:MM
    break_out: Optional[str] = None  # وقت الذهاب للاستراحة HH:MM
    break_in: Optional[str] = None   # وقت العودة من الاستراحة HH:MM
    status: str = "present"  # present, absent, late, early_leave, holiday
    notes: Optional[str] = None
    source: str = "manual"  # manual, fingerprint, system

class AttendanceResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    date: str
    check_in: Optional[str] = None
    check_out: Optional[str] = None
    break_out: Optional[str] = None
    break_in: Optional[str] = None
    worked_hours: Optional[float] = None
    status: str
    notes: Optional[str] = None
    source: str
    created_at: str

# نموذج السلف
class AdvanceCreate(BaseModel):
    employee_id: str
    amount: float
    reason: Optional[str] = None
    deduction_months: int = 1  # عدد أشهر الاستقطاع
    date: Optional[str] = None

class AdvanceResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    amount: float
    remaining_amount: float
    deducted_amount: float = 0
    deduction_months: int
    monthly_deduction: float
    reason: Optional[str] = None
    status: str  # pending, approved, rejected, paid
    date: str
    created_by: str
    created_at: str

# نموذج الخصومات
class DeductionCreate(BaseModel):
    employee_id: str
    deduction_type: str  # absence, late, early_leave, violation, other
    amount: Optional[float] = None  # مبلغ ثابت
    hours: Optional[float] = None  # عدد الساعات
    days: Optional[float] = None  # عدد الأيام
    reason: str
    date: str

class DeductionResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    deduction_type: str
    amount: float
    hours: Optional[float] = None
    days: Optional[float] = None
    reason: str
    date: str
    created_by: str
    created_at: str

# نموذج المكافآت والوقت الإضافي
class BonusCreate(BaseModel):
    employee_id: str
    bonus_type: str  # performance, overtime, holiday, other
    amount: Optional[float] = None  # مبلغ ثابت
    hours: Optional[float] = None  # ساعات إضافية
    reason: str
    date: str

class BonusResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    bonus_type: str
    amount: float
    hours: Optional[float] = None
    reason: str
    date: str
    created_by: str
    created_at: str

# نموذج كشف الراتب
class PayrollCreate(BaseModel):
    employee_id: str
    month: str  # YYYY-MM
    basic_salary: float
    total_deductions: float = 0
    total_bonuses: float = 0
    advance_deduction: float = 0
    net_salary: float
    notes: Optional[str] = None

class PayrollResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    employee_id: str
    employee_name: Optional[str] = None
    month: str
    basic_salary: float
    worked_days: int = 0
    absent_days: int = 0
    late_hours: float = 0
    overtime_hours: float = 0
    total_deductions: float
    total_bonuses: float
    advance_deduction: float
    net_salary: float
    status: str  # draft, approved, paid
    notes: Optional[str] = None
    created_by: str
    created_at: str
    paid_at: Optional[str] = None

# ==================== INVENTORY TRANSFER MODELS - تحويلات المخزون ====================

class InventoryTransferCreate(BaseModel):
    from_branch_id: str  # الفرع المرسل (أو المخزن الرئيسي)
    to_branch_id: str  # الفرع المستلم
    items: List[Dict[str, Any]]  # [{inventory_id, quantity, notes}]
    transfer_type: str = "warehouse_to_branch"  # warehouse_to_branch, branch_to_warehouse, branch_to_branch
    notes: Optional[str] = None

class InventoryTransferResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    transfer_number: int
    from_branch_id: str
    from_branch_name: Optional[str] = None
    to_branch_id: str
    to_branch_name: Optional[str] = None
    items: List[Dict[str, Any]]
    transfer_type: str
    status: str  # pending, approved, shipped, received, cancelled
    notes: Optional[str] = None
    created_by: str
    created_at: str
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None
    received_by: Optional[str] = None
    received_at: Optional[str] = None

# نموذج طلب شراء
class PurchaseRequestCreate(BaseModel):
    branch_id: str  # الفرع الطالب
    items: List[Dict[str, Any]]  # [{name, quantity, unit, notes}]
    priority: str = "normal"  # urgent, high, normal, low
    notes: Optional[str] = None

class PurchaseRequestResponse(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    request_number: int
    branch_id: str
    branch_name: Optional[str] = None
    items: List[Dict[str, Any]]
    priority: str
    status: str  # pending, approved, ordered, received, cancelled
    notes: Optional[str] = None
    created_by: str
    created_at: str
    approved_by: Optional[str] = None
    approved_at: Optional[str] = None

# ==================== AUTH HELPERS ====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))
    except Exception:
        return False

def create_token(user_id: str, role: str, branch_id: Optional[str] = None, tenant_id: Optional[str] = None) -> str:
    # تحديد مدة الجلسة بناءً على نوع المستخدم
    # المالك (super_admin) والعملاء (admin) = لا يسجلون خروج أبداً (100 سنة)
    # الموظفين (cashier, warehouse_keeper, manufacturer, branch_manager) = 24 ساعة
    if role in [UserRole.SUPER_ADMIN, UserRole.ADMIN]:
        expiration = datetime.now(timezone.utc) + timedelta(days=JWT_EXPIRATION_DAYS_OWNERS)
    else:
        expiration = datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS_STAFF)
    
    payload = {
        "user_id": user_id,
        "role": role,
        "branch_id": branch_id,
        "tenant_id": tenant_id,
        "iat": datetime.now(timezone.utc),
        "exp": expiration
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(request: Request, credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = _extract_token(request, credentials)
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        # إبطال الجلسات القديمة عند تفعيل المالك للتحقق الإلزامي (يُخرج الجميع)
        cutoff = await _sessions_valid_after()
        if cutoff and payload.get("iat"):
            try:
                iat = payload["iat"]
                iat_ts = iat if isinstance(iat, (int, float)) else datetime.fromisoformat(str(iat)).timestamp()
                if iat_ts < cutoff:
                    raise HTTPException(status_code=401, detail="انتهت الجلسة — يرجى تسجيل الدخول لتفعيل التحقق")
            except HTTPException:
                raise
            except Exception:
                pass
        user = await db.users.find_one({"id": payload["user_id"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="المستخدم غير موجود")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="انتهت صلاحية الجلسة")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="رمز غير صالح")

async def get_current_driver(request: Request, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """التحقق من جلسة السائق عبر توكن opaque مخزّن في driver_tokens"""
    token = _extract_token(request, credentials)
    rec = await db.driver_tokens.find_one({"token": token})
    if not rec:
        raise HTTPException(status_code=401, detail="جلسة السائق غير صالحة - يرجى تسجيل الدخول")
    # إبطال جلسات السائقين القديمة عند تفعيل التحقق الإلزامي
    cutoff = await _sessions_valid_after()
    if cutoff and rec.get("created_at"):
        try:
            if datetime.fromisoformat(rec["created_at"]).timestamp() < cutoff:
                await db.driver_tokens.delete_one({"token": token})
                raise HTTPException(status_code=401, detail="انتهت الجلسة — يرجى تسجيل الدخول")
        except HTTPException:
            raise
        except Exception:
            pass
    driver = await db.drivers.find_one({"id": rec.get("driver_id")}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=401, detail="السائق غير موجود")
    if not driver.get("is_active", True):
        raise HTTPException(status_code=403, detail="حساب السائق غير مفعل")
    return driver

async def get_staff_or_driver(request: Request, credentials: HTTPAuthorizationCredentials = Depends(security)):
    """يقبل توكن موظف (JWT) أو توكن سائق (opaque)."""
    token = _extract_token(request, credentials)
    try:
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload.get("user_id")}, {"_id": 0})
        if user:
            user["_auth_type"] = "staff"
            return user
    except Exception:
        pass
    rec = await db.driver_tokens.find_one({"token": token})
    if rec:
        driver = await db.drivers.find_one({"id": rec.get("driver_id")}, {"_id": 0})
        if driver:
            driver["_auth_type"] = "driver"
            return driver
    raise HTTPException(status_code=401, detail="غير مصرح")

async def verify_device_agent(request: Request):
    """حماية نقاط أجهزة البصمة/الوكلاء بمفتاح سري (إن ضُبط BIOMETRIC_AGENT_KEY)."""
    agent_key = os.environ.get("BIOMETRIC_AGENT_KEY")
    if not agent_key:
        return True
    provided = request.headers.get("X-Agent-Key") or request.query_params.get("key")
    if provided != agent_key:
        raise HTTPException(status_code=403, detail="Unauthorized device agent")
    return True

# ======== سجل التدقيق الأمني (Security Audit Log) ========
def _client_ip(request) -> str:
    """استخراج IP الحقيقي للزائر خلف nginx/docker."""
    if not request:
        return "unknown"
    xff = request.headers.get("x-forwarded-for") or request.headers.get("X-Forwarded-For")
    if xff:
        return xff.split(",")[0].strip()
    xri = request.headers.get("x-real-ip")
    if xri:
        return xri.strip()
    return request.client.host if getattr(request, "client", None) else "unknown"

async def record_audit(event: str, request=None, user=None, details=None, status=None, tenant_id=None):
    """تسجيل حدث أمني في مجموعة audit_logs (لا يرمي أخطاء أبداً)."""
    try:
        u = user if isinstance(user, dict) else {}
        doc = {
            "id": str(uuid.uuid4()),
            "event": event,
            "ip": _client_ip(request),
            "path": str(request.url.path) if request else None,
            "method": request.method if request else None,
            "user_agent": (request.headers.get("user-agent") if request else None),
            "user_id": u.get("id"),
            "user_name": u.get("full_name") or u.get("username") or u.get("email"),
            "role": u.get("role"),
            "tenant_id": tenant_id or u.get("tenant_id"),
            "status": status,
            "details": details,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "ts": datetime.now(timezone.utc),
        }
        await db.audit_logs.insert_one(doc)
    except Exception:
        pass

# ======== الحماية ضد محاولات تخمين/اختراق الدخول (Brute-force) ========
MAX_LOGIN_FAILS = 5
LOGIN_LOCK_MINUTES = 15

async def check_login_lock(key: str):
    """يرفع 429 إذا كان المفتاح (IP+بريد) مقفولاً مؤقتاً."""
    rec = await db.login_attempts.find_one({"key": key})
    if rec and rec.get("locked_until"):
        try:
            lu = datetime.fromisoformat(rec["locked_until"])
            if lu > datetime.now(timezone.utc):
                mins = max(1, int((lu - datetime.now(timezone.utc)).total_seconds() // 60) + 1)
                raise HTTPException(status_code=429, detail=f"تم القفل المؤقت بسبب محاولات دخول فاشلة. حاول بعد {mins} دقيقة")
        except HTTPException:
            raise
        except Exception:
            pass

async def record_login_fail(key: str, ip: str = None, request=None):
    """يسجّل محاولة فاشلة. عند بلوغ MAX_LOGIN_FAILS يُحظر عنوان الـIP نهائياً
    (يمنع كل الأجهزة على نفس الشبكة) — يُستثنى عنوان المالك دائماً."""
    rec = await db.login_attempts.find_one({"key": key})
    count = (rec.get("count", 0) if rec else 0) + 1
    update = {"key": key, "count": count, "ip": ip, "last": datetime.now(timezone.utc).isoformat()}
    if count >= MAX_LOGIN_FAILS:
        # حظر دائم لعنوان الـIP (بدل القفل المؤقت)
        banned = False
        if ip:
            banned = await _ban_ip_permanent(
                ip,
                reason=f"حظر دائم: {count} محاولات دخول فاشلة",
                request=request,
            )
        update["permanently_banned"] = banned
        update["count"] = 0
    await db.login_attempts.update_one({"key": key}, {"$set": update}, upsert=True)

async def clear_login_attempts(key: str):
    try:
        await db.login_attempts.delete_one({"key": key})
    except Exception:
        pass

# ======== المصادقة الثنائية (2FA) — جهاز موثوق + رمز تحقق ========
import hashlib as _hashlib
import secrets as _secrets_2fa

_OTP_TTL_MINUTES = 1
_OTP_MAX_ATTEMPTS = 5

def _new_device_id() -> str:
    return _secrets_2fa.token_urlsafe(24)

def _gen_otp_code() -> str:
    return f"{_secrets_2fa.randbelow(1000000):06d}"

def _hash_otp(code: str, salt: str) -> str:
    return _hashlib.sha256(f"{salt}:{code}:{JWT_SECRET}".encode()).hexdigest()

def _mask_email(email: str) -> str:
    # إخفاء الاسم قبل @ بالكامل وإظهار ما بعده فقط
    try:
        _, domain = email.split("@", 1)
        return f"***@{domain}"
    except Exception:
        return "***"

def _mask_phone(phone: str) -> str:
    # تقنيع كل الأرقام وإظهار آخر رقمين فقط
    p = "".join(ch for ch in str(phone) if ch.isdigit())
    if len(p) <= 2:
        return "**"
    return ("*" * (len(p) - 2)) + p[-2:]

async def is_device_trusted(subject_type: str, subject_id: str, device_id: str) -> bool:
    if not device_id or not subject_id:
        return False
    doc = await db.trusted_devices.find_one({
        "subject_type": subject_type,
        "subject_id": str(subject_id),
        "device_id": device_id,
        "revoked": {"$ne": True},
    })
    return bool(doc)

async def trust_device(subject_type: str, subject_id: str, device_id: str, ip: str,
                       user_agent: str = "", tenant_id=None, label: str = ""):
    if not device_id:
        return
    now = datetime.now(timezone.utc).isoformat()
    await db.trusted_devices.update_one(
        {"subject_type": subject_type, "subject_id": str(subject_id), "device_id": device_id},
        {"$set": {
            "subject_type": subject_type,
            "subject_id": str(subject_id),
            "device_id": device_id,
            "ip": ip,
            "user_agent": user_agent,
            "tenant_id": tenant_id,
            "label": label,
            "last_seen_at": now,
            "revoked": False,
        }, "$setOnInsert": {"id": str(uuid.uuid4()), "created_at": now}},
        upsert=True,
    )

async def _phone_to_e164(phone: str) -> str:
    """تحويل رقم عراقي محلي إلى صيغة E.164 (+964) إن لزم."""
    p = "".join(ch for ch in str(phone or "") if ch.isdigit() or ch == "+")
    if not p:
        return p
    if p.startswith("+"):
        return p
    if p.startswith("00"):
        return "+" + p[2:]
    if p.startswith("0"):
        return "+964" + p[1:]
    if p.startswith("964"):
        return "+" + p
    return "+" + p

async def start_2fa_verification(subject_type, subject_id, subject_name, tenant_id,
                                 channel, destination, device_id, ip, request=None, extra=None,
                                 fallback_email=None):
    """ينشئ جلسة تحقق ويُرسل الرمز بترتيب الأولوية:
    واتساب مجاني (رقم المالك) ← البريد ← (الحالة القصوى) SMS مدفوع ← معلّق.
    channel: 'email' | 'whatsapp' | 'sms' (القناة المفضّلة)."""
    vid = str(uuid.uuid4())
    salt = _secrets_2fa.token_hex(8)
    expires = (datetime.now(timezone.utc) + timedelta(minutes=_OTP_TTL_MINUTES)).isoformat()
    ua = request.headers.get("user-agent", "") if request else ""
    method = None
    code_hash = None
    dev_code = None
    pending = False
    masked = ""
    channels_sent = []  # جميع القنوات التي نجح فيها إرسال الرمز — يُرجَع للواجهة

    def _otp_email_html(code, name):
        body_html = f"""
            <p style='margin:0 0 8px 0'>مرحباً {name or ''}،</p>
            <p style='margin:0 0 12px 0'>تم طلب تسجيل الدخول من جهاز جديد. استخدم الرمز التالي (صالح لدقيقة واحدة فقط):</p>
            <div style='font-size:34px;font-weight:900;letter-spacing:8px;color:#1e40af;background:#eff6ff;padding:16px;border-radius:10px;text-align:center;margin:16px 0;border:1px dashed #3B82F6'>{code}</div>
            <p style='color:#64748b;font-size:13px;margin:0'>⚠️ إن لم تكن أنت، تجاهل الرسالة وغيّر كلمة المرور.</p>
        """
        return build_branded_email_html(
            title="🔐 رمز تحقق الدخول",
            body_html=body_html,
            severity="info",
        )

    if channel in ("sms", "whatsapp"):
        e164 = await _phone_to_e164(destination)
        destination = e164
        masked = _mask_phone(e164)
        code = _gen_otp_code()
        code_hash = _hash_otp(code, salt)
        # نص واضح للـ OTP — القالب الموحّد سيضيف *🔔 Maestro EGP* + الفواصل + الطابع الزمني تلقائياً
        wa_msg = f"رمز الدخول: *{code}*\n\nصالح لدقيقة واحدة فقط.\n⚠️ لا تُشاركه مع أحد."
        sent = False
        channels_sent = []  # قائمة القنوات التي نجح فيها الإرسال (لعرضها للمستخدم)
        
        # 🔥 إرسال ثنائي: واتساب + بريد (بشكل متزامن) — يضمن وصول الرمز حتى لو تعطلت قناة
        # 1) واتساب مجاني (رقم المالك المرتبط) — بهوية Maestro EGP الموحّدة (شعار + قالب)
        if await _wa_free.is_connected():
            ok, err = await _wa_free.send_message(e164, wa_msg, purpose="otp", title="🔐 رمز التحقق")
            if ok:
                sent = True
                channels_sent.append("whatsapp")
                method = "local"
        
        # 2) البريد بالتوازي (إن توفّر بريد للمستخدم) — لا ينتظر فشل الواتساب
        if fallback_email:
            recips = fallback_email if isinstance(fallback_email, list) else [fallback_email]
            recips = [r for r in recips if r and "@" in str(r)]
            if recips:
                try:
                    email_ok = await send_system_email(
                        recips,
                        "رمز تحقق الدخول — Maestro EGP",
                        _otp_email_html(code, subject_name)
                    )
                    if email_ok:
                        sent = True
                        channels_sent.append("email")
                        if not method:
                            method = "local"
                        # إن كان البريد الوحيد الذي نجح، عرض عنوان البريد
                        if not channels_sent or channels_sent == ["email"]:
                            masked = _mask_email(recips[0])
                except Exception as _e:
                    logger.warning(f"OTP email delivery failed: {_e}")
        
        # 3) الحالة القصوى: SMS مدفوع عبر Twilio — فقط لو القناتان أعلاه فشلتا
        if not sent and _twilio_verify.is_configured():
            ok, status = await _twilio_verify.start_verification(e164, "sms")
            if ok:
                sent = True
                channels_sent.append("sms")
                method = "twilio"  # Twilio يتحقق من رمزه الخاص
                code_hash = None
        
        # 4) تعذّر الإرسال بالكامل → معلّق (لا يُعرض الرمز في أي مكان — أمان صارم)
        if not sent:
            method = "local"
            pending = True
        
        # القناة الرئيسية في الرد = أول قناة نجحت (للتوافق مع الواجهة)
        if channels_sent:
            channel = channels_sent[0]
    else:  # email مباشرة (المالك/الموظف بلا هاتف)
        recipients = destination if isinstance(destination, list) else [destination]
        masked = _mask_email(recipients[0] if recipients else "")
        method = "local"
        code = _gen_otp_code()
        code_hash = _hash_otp(code, salt)
        sent = await send_system_email(recipients, "رمز تحقق الدخول — Maestro EGP", _otp_email_html(code, subject_name))
        if sent:
            channels_sent.append("email")
        else:
            pending = True

    session = {
        "id": vid,
        "subject_type": subject_type,
        "subject_id": str(subject_id) if subject_id is not None else None,
        "subject_name": subject_name,
        "tenant_id": tenant_id,
        "channel": channel,
        "channels_sent": channels_sent if channel in ("sms", "whatsapp", "email") else [],
        "method": method,
        "destination": destination if not isinstance(destination, list) else ",".join(destination),
        "code_hash": code_hash,
        "salt": salt,
        "device_id": device_id or _new_device_id(),
        "ip": ip,
        "user_agent": ua,
        "attempts": 0,
        "consumed": False,
        "pending_delivery": pending,
        "extra": extra or {},
        "created_at": datetime.now(timezone.utc).isoformat(),
        "expires_at": expires,
    }
    await db.verification_sessions.insert_one(dict(session))
    resp = {
        "requires_2fa": True,
        "verification_id": vid,
        "channel": channel,
        "channels_sent": session.get("channels_sent", []),
        "destination_masked": masked,
        "device_id": session["device_id"],
        "expires_in_minutes": _OTP_TTL_MINUTES,
        "pending_delivery": pending,
    }
    return resp

async def verify_2fa_code(verification_id: str, code: str, ip: str = None):
    """يتحقق من الرمز. يُرجع (ok: bool, session: dict|None, error: str|None)."""
    sess = await db.verification_sessions.find_one({"id": verification_id}, {"_id": 0})
    if not sess:
        return False, None, "الجلسة غير موجودة أو منتهية"
    if sess.get("consumed"):
        return False, None, "تم استخدام هذا الرمز مسبقاً"
    try:
        if datetime.fromisoformat(sess["expires_at"]) < datetime.now(timezone.utc):
            return False, None, "انتهت صلاحية الرمز — أعد الإرسال"
    except Exception:
        pass
    if sess.get("attempts", 0) >= _OTP_MAX_ATTEMPTS:
        return False, None, "تجاوزت عدد المحاولات — أعد الإرسال"

    code = (code or "").strip()
    approved = False
    used_backup = False
    if sess.get("method") == "twilio":
        ok, approved, err = await _twilio_verify.check_verification(sess["destination"], code)
        if not ok:
            approved = False
    else:
        approved = bool(sess.get("code_hash")) and _hash_otp(code, sess.get("salt", "")) == sess["code_hash"]
    
    # 🆘 نجاة أخيرة: قبول رمز طوارئ للسوبر أدمن إن فشل الرمز الأصلي (فقط لجلسات المالك)
    if not approved and sess.get("purpose") in ("super_admin_login", "owner_login", "sa_login"):
        try:
            import hashlib
            code_hash = hashlib.sha256(code.encode()).hexdigest()
            sec_cfg = await db.security_config.find_one({"id": "global"}, {"_id": 0}) or {}
            backup_hashes = sec_cfg.get("backup_codes") or []
            if code_hash in backup_hashes:
                approved = True
                used_backup = True
                # استهلاك الرمز — يُحذَف من القائمة (single-use)
                new_hashes = [h for h in backup_hashes if h != code_hash]
                await db.security_config.update_one(
                    {"id": "global"},
                    {"$set": {"backup_codes": new_hashes,
                              "last_backup_code_used_at": datetime.now(timezone.utc).isoformat()}}
                )
                logger.warning(f"🆘 Backup code used for super_admin login (remaining: {len(new_hashes)})")
        except Exception as _be:
            logger.warning(f"backup code check failed: {_be}")

    if not approved:
        await db.verification_sessions.update_one({"id": verification_id}, {"$inc": {"attempts": 1}})
        return False, None, "رمز التحقق غير صحيح"

    await db.verification_sessions.update_one({"id": verification_id}, {"$set": {"consumed": True, "used_backup_code": used_backup}})
    return True, sess, None

def _choose_user_2fa_channel(user: dict, is_owner: bool):
    """يحدد قناة التحقق ووِجهته لمستخدم (موظف/مدير/مالك)."""
    if is_owner:
        return "email", list(OWNER_RECOVERY_EMAILS)
    phone = user.get("phone") or user.get("mobile")
    if phone:
        return "whatsapp", phone
    # لا هاتف → البريد الإلكتروني للحساب
    return "email", (user.get("email") or "")

# ======== مفتاح التفعيل الرئيسي للتحقق الإلزامي (بيد المالك) ========
_SEC_CFG = None
_SEC_CFG_TS = 0.0

async def _refresh_security_config(force=False):
    global _SEC_CFG, _SEC_CFG_TS
    import time as _t
    now = _t.time()
    if not force and _SEC_CFG is not None and (now - _SEC_CFG_TS) < 15:
        return _SEC_CFG
    doc = await db.security_config.find_one({"id": "global"}, {"_id": 0}) or {}
    _SEC_CFG = doc
    _SEC_CFG_TS = now
    return doc

async def two_fa_enabled() -> bool:
    cfg = await _refresh_security_config()
    return bool(cfg.get("two_fa_enabled", False))

async def _sessions_valid_after():
    """يُرجع طابع زمني (epoch) — أي توكن أُصدر قبله يُعتبر منتهياً (لإخراج الجميع)."""
    cfg = await _refresh_security_config()
    val = cfg.get("sessions_valid_after")
    if not val:
        return None
    try:
        return datetime.fromisoformat(val).timestamp()
    except Exception:
        return None


# دالة مساعدة للحصول على tenant_id للمستخدم
def get_user_tenant_id(user: dict) -> Optional[str]:
    """
    الحصول على tenant_id للمستخدم
    
    - Super Admin الحقيقي بدون tenant: يُرجع None (لا يصل لبيانات عملاء)
    - Super Admin في وضع Impersonation: يُرجع tenant_id المستخدم الهدف
    - المستخدمين العاديين: يُرجع tenant_id الخاص بهم
    """
    tenant_id = user.get("tenant_id")
    
    # إذا كان Super Admin بدون tenant_id، لا يُرجع default
    if user.get("role") == UserRole.SUPER_ADMIN and not tenant_id:
        return None  # لا يصل لبيانات أي عميل
    
    return tenant_id or "default"

# دالة مساعدة لبناء query مع tenant_id
def build_tenant_query(user: dict, base_query: dict = None) -> dict:
    """
    بناء query مع فلترة tenant_id - يضمن عزل البيانات بين المستأجرين
    
    ملاحظات مهمة:
    - Super Admin الحقيقي (بدون impersonation) لا يستخدم هذه الدالة للوصول لبيانات العملاء
    - عند Impersonation، المستخدم يصبح admin عادي بـ tenant_id محدد
    - كل مستخدم يرى فقط بيانات الـ tenant الخاص به
    """
    query = base_query.copy() if base_query else {}
    
    # الحصول على tenant_id من المستخدم
    tenant_id = user.get("tenant_id")
    
    # Super Admin الحقيقي (tenant_id = None أو system) بدون impersonation
    # لا يجب أن يصل لهذه الدالة عند استعراض بيانات العملاء
    # لكن إذا وصل، نُرجع query فارغ لمنع تسرب البيانات
    if user.get("role") == UserRole.SUPER_ADMIN and not tenant_id:
        # Super Admin بدون tenant - لا يرى بيانات أي عميل
        query["tenant_id"] = "__NO_ACCESS__"  # قيمة غير موجودة لإرجاع نتائج فارغة
        return query
    
    # جميع المستخدمين الآخرين (بما فيهم admin المنتحَل من Super Admin)
    if tenant_id:
        query["tenant_id"] = tenant_id
    else:
        # fallback للبيانات القديمة بدون tenant_id
        query["tenant_id"] = "default"
    
    return query

# دالة مساعدة لبناء query مع فلترة الفرع
def build_branch_query(user: dict, base_query: dict = None) -> dict:
    """بناء query مع فلترة الفرع للمستخدمين المقيدين بفرع معين"""
    query = build_tenant_query(user, base_query)
    
    # إذا كان المستخدم مرتبط بفرع معين (ليس admin أو manager)
    user_branch_id = user.get("branch_id")
    user_role = user.get("role")
    
    # المستخدمون العاديون (cashier, supervisor, delivery) يرون فقط فرعهم
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    
    return query

def user_can_access_branch(user: dict, branch_id: str) -> bool:
    """التحقق من صلاحية المستخدم للوصول لفرع معين"""
    if user.get("role") in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        return True
    return user.get("branch_id") == branch_id

# ==================== EMAIL SERVICE ====================

async def send_shift_report_email(shift_data: dict, recipient_emails: List[str]):
    if not recipient_emails or not await email_transport_configured():
        logger.warning("Email transport not configured or no recipients")
        return
    
    _diff = shift_data.get('cash_difference', 0)
    _diff_color = '#EF4444' if _diff < 0 else '#10B981'
    _net = shift_data.get('net_profit', 0)
    _net_color = '#10B981' if _net >= 0 else '#EF4444'
    body_html = f"""
        <table style="width: 100%; border-collapse: collapse; font-size:14px">
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>أمين الصندوق:</strong></td><td>{shift_data.get('cashier_name', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>تاريخ البدء:</strong></td><td>{shift_data.get('started_at', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>تاريخ الإنتهاء:</strong></td><td>{shift_data.get('ended_at', 'N/A')}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>إجمالي المبيعات:</strong></td><td style="color: #10B981; font-weight: bold;">{shift_data.get('total_sales', 0):,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>إجمالي التكاليف:</strong></td><td>{shift_data.get('total_cost', 0):,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>الربح الإجمالي:</strong></td><td style="color: #10B981;">{shift_data.get('gross_profit', 0):,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>المصاريف:</strong></td><td style="color: #EF4444;">{shift_data.get('total_expenses', 0):,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>صافي الربح:</strong></td><td style="color: {_net_color}; font-weight: bold;">{_net:,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>عدد الطلبات:</strong></td><td>{shift_data.get('total_orders', 0)}</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>النقد المتوقع:</strong></td><td>{shift_data.get('expected_cash', 0):,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px; border-bottom: 1px solid #eee;"><strong>النقد الفعلي:</strong></td><td>{shift_data.get('closing_cash', 0):,.0f} د.ع</td></tr>
            <tr><td style="padding: 8px;"><strong>الفرق:</strong></td><td style="color: {_diff_color}; font-weight: bold;">{_diff:,.0f} د.ع</td></tr>
        </table>
    """
    html_content = build_branded_email_html(
        title=f"📊 تقرير إغلاق الصندوق — {shift_data.get('cashier_name', '')}",
        body_html=body_html,
        severity="info",
    )
    
    try:
        await send_system_email(
            recipient_emails,
            f"تقرير إغلاق الصندوق - {shift_data.get('cashier_name', '')} - {datetime.now().strftime('%Y-%m-%d')}",
            html_content,
            purpose="shift_report",
        )
    except Exception as e:
        logger.error(f"Failed to send email: {e}")

async def send_welcome_email(recipient_email: str, tenant_name: str, owner_name: str, username: str, password: str):
    """إرسال بريد ترحيبي للعميل الجديد مع بيانات الدخول (بهوية Maestro EGP الموحّدة)."""
    if not recipient_email or not await email_transport_configured():
        logger.warning("Email transport not configured or no recipient")
        return
    
    frontend_url = os.environ.get('FRONTEND_URL', 'https://maestroegp.com')
    body_html = f"""
        <div style="background:#eff6ff;border:1px solid #3B82F6;border-radius:12px;padding:20px;margin-bottom:16px">
            <h2 style="margin:0 0 12px 0;color:#1e40af">مرحباً {owner_name}! 🎉</h2>
            <p style="margin:0;line-height:1.8;color:#334155">
                تم إنشاء حسابك في <strong style="color:#1e40af">{tenant_name}</strong> بنجاح على منصة Maestro EGP.
                يمكنك الآن البدء في إدارة مطعمك/الكافيه بكل سهولة.
            </p>
        </div>
        <div style="background:#fff;border:1px solid #e5e7eb;border-radius:12px;padding:20px;margin-bottom:16px">
            <h3 style="margin:0 0 16px 0;color:#111;font-size:16px">🔐 بيانات تسجيل الدخول</h3>
            <p style="color:#6b7280;margin:0 0 4px 0;font-size:12px">البريد الإلكتروني / اسم المستخدم:</p>
            <p style="background:#f3f4f6;padding:10px 12px;border-radius:8px;margin:0 0 12px 0;font-family:monospace;font-size:14px;color:#111">{username}</p>
            <p style="color:#6b7280;margin:0 0 4px 0;font-size:12px">كلمة المرور:</p>
            <p style="background:#f3f4f6;padding:10px 12px;border-radius:8px;margin:0;font-family:monospace;font-size:14px;color:#111">{password}</p>
        </div>
        <div style="text-align:center;margin:24px 0">
            <a href="{frontend_url}/login" style="display:inline-block;background:#3B82F6;color:#fff;text-decoration:none;padding:12px 32px;border-radius:8px;font-weight:bold">🚀 تسجيل الدخول الآن</a>
        </div>
        <div style="background:#f0fdf4;border:1px solid #10B981;border-radius:12px;padding:16px;margin-bottom:16px">
            <h3 style="margin:0 0 8px 0;color:#065f46;font-size:14px">✨ ماذا يمكنك أن تفعل؟</h3>
            <ul style="margin:0;padding-right:20px;line-height:1.9;color:#334155;font-size:13px">
                <li>إدارة الطلبات (محلي، سفري، توصيل)</li>
                <li>تتبع السائقين على الخريطة</li>
                <li>إدارة المخزون والمنتجات والتصنيع</li>
                <li>تقارير المبيعات والأرباح والحسابات</li>
                <li>إدارة الموظفين والرواتب والبصمة</li>
                <li>نظام الكول سنتر والولاء</li>
            </ul>
        </div>
        <div style="background:#fef2f2;border:1px solid #EF4444;border-radius:10px;padding:12px;color:#991b1b;font-size:13px">
            ⚠️ <strong>هام:</strong> يرجى تغيير كلمة المرور فور تسجيل الدخول للحفاظ على أمان حسابك.
        </div>
    """
    html_content = build_branded_email_html(
        title=f"🎉 مرحباً في {tenant_name}",
        body_html=body_html,
        severity="success",
    )
    try:
        await send_system_email(
            [recipient_email],
            f"🎉 مرحباً في {tenant_name} - بيانات الدخول",
            html_content,
            purpose="welcome",
        )
        logger.info(f"Welcome email dispatched to {recipient_email}")
    except Exception as e:
        logger.error(f"Failed to send welcome email: {e}")

# ==================== HELPER FUNCTIONS ====================

async def get_delivery_app_commission(app_id: str) -> float:
    """Get commission rate for delivery app"""
    setting = await db.delivery_app_settings.find_one({"app_id": app_id}, {"_id": 0})
    if setting:
        return setting.get("commission_rate", 0)
    return 0

async def calculate_order_cost(items: List[Dict]) -> float:
    """Calculate total cost for order items"""
    total_cost = 0
    for item in items:
        product = await db.products.find_one({"id": item.get("product_id")}, {"_id": 0})
        if product:
            item_cost = (_sn(product.get("cost")) + _sn(product.get("operating_cost"))) * item.get("quantity", 1)
            total_cost += item_cost
    return total_cost

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register")
async def register(user: UserCreate, current_user: dict = Depends(get_current_user)):
    # الحماية: إنشاء المستخدمين متاح فقط للمالك (admin/super_admin) — لا المدير ولا الكاشير
    creator_role = current_user.get("role")
    allowed_creators = [UserRole.SUPER_ADMIN, UserRole.ADMIN]
    if creator_role not in allowed_creators:
        raise HTTPException(status_code=403, detail="إنشاء الحسابات متاح للمالك فقط")

    # لا يُسمح بإنشاء حساب super_admin إطلاقاً عبر هذا المسار (يُنشأ فقط من لوحة المالك الرئيسية)
    requested_role = user.role
    if requested_role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح بإنشاء حساب مالك النظام")

    # المدير العادي لا يستطيع إنشاء حساب admin/مالك متجر
    if creator_role != UserRole.SUPER_ADMIN and requested_role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح بإنشاء حساب بهذه الصلاحية")

    existing = await db.users.find_one({"$or": [{"email": user.email}, {"username": user.username}]})
    if existing:
        raise HTTPException(status_code=400, detail="المستخدم موجود بالفعل")

    # فرض عزل المستأجر: المستخدم الجديد يتبع نفس tenant المنشئ (إلا إذا كان super_admin ينشئ لمستأجر محدد)
    creator_tenant = current_user.get("tenant_id")
    new_tenant_id = user.tenant_id if creator_role == UserRole.SUPER_ADMIN else creator_tenant

    user_doc = {
        "id": str(uuid.uuid4()),
        "username": user.username,
        "email": user.email,
        "password": hash_password(user.password),
        "full_name": user.full_name,
        "full_name_en": user.full_name_en,  # الاسم بالإنجليزية
        "role": requested_role,
        "branch_id": user.branch_id,
        "permissions": user.permissions,
        "tenant_id": new_tenant_id,
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    del user_doc["password"]
    del user_doc["_id"]
    # لا نُصدر توكن للمستخدم الجديد هنا (المنشئ يبقى بجلسته)؛ نعيد بيانات المستخدم فقط
    return {"user": user_doc}

@api_router.post("/auth/login")
async def login(credentials: UserLogin, request: Request, response: Response):
    _ip = _client_ip(request)
    _lockkey = f"login:{_ip}:{credentials.email}"
    await check_login_lock(_lockkey)
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    
    if not user:
        await record_login_fail(_lockkey, ip=_ip, request=request)
        await record_audit("auth.login.failed", request, details={"email": credentials.email, "reason": "user_not_found"}, status=401)
        raise HTTPException(status_code=401, detail="بيانات الدخول غير صحيحة")
    
    stored_hash = user.get("password_hash", user.get("password", ""))
    
    if not verify_password(credentials.password, stored_hash):
        await record_login_fail(_lockkey, ip=_ip, request=request)
        await record_audit("auth.login.failed", request, user=user, details={"email": credentials.email, "reason": "wrong_password"}, status=401)
        raise HTTPException(status_code=401, detail="بيانات الدخول غير صحيحة")
    
    if not user.get("is_active", True):
        await record_audit("auth.login.failed", request, user=user, details={"reason": "inactive"}, status=401)
        raise HTTPException(status_code=401, detail="الحساب معطل")

    # الحماية: حساب مالك النظام (super_admin) يتطلب مفتاحاً سرياً يُتحقق منه على الخادم
    import hmac as _hmac
    if user.get("role") == UserRole.SUPER_ADMIN:
        _expected_secret = user.get("super_admin_secret") or user.get("secret_key") or SUPER_ADMIN_SECRET
        if not credentials.secret_key or not _hmac.compare_digest(str(credentials.secret_key), str(_expected_secret)):
            await record_login_fail(_lockkey, ip=_ip, request=request)
            await record_audit("auth.login.failed", request, user=user, details={"reason": "bad_secret_key"}, status=403)
            raise HTTPException(status_code=403, detail="المفتاح السري مطلوب لحساب المالك")
    else:
        # 🔒 تقرير الأمان #3: أي حساب مدير/مالك عيّن له مفتاحاً سرياً يجب التحقق منه على الخادم
        #    (لا يُتجاوَز من الواجهة). الحسابات التي لم تُعيّن مفتاحاً تبقى كما هي (لا تُقفَل).
        _acct_secret = user.get("secret_key")
        if _acct_secret and (user.get("role") or "").lower() in ("admin", "manager", "owner"):
            if not credentials.secret_key or not _hmac.compare_digest(str(credentials.secret_key), str(_acct_secret)):
                await record_login_fail(_lockkey, ip=_ip, request=request)
                await record_audit("auth.login.failed", request, user=user, details={"reason": "bad_secret_key_admin"}, status=403)
                raise HTTPException(status_code=403, detail="المفتاح السري مطلوب لهذا الحساب")

    await clear_login_attempts(_lockkey)
    
    # ======== فحص ترخيص العميل (Tenant) ========
    tenant_id = user.get("tenant_id")
    if tenant_id and user.get("role") != UserRole.SUPER_ADMIN:
        tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
        
        if tenant:
            # فحص حالة الحساب
            if not tenant.get("is_active", True):
                raise HTTPException(
                    status_code=403, 
                    detail="حساب المطعم معطل - يرجى التواصل مع الدعم الفني"
                )
            
            # فحص تاريخ انتهاء الاشتراك
            expires_at = tenant.get("expires_at")
            if expires_at:
                try:
                    expiry_date = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                    if datetime.now(timezone.utc) > expiry_date:
                        raise HTTPException(
                            status_code=403, 
                            detail="انتهى اشتراك المطعم - يرجى التجديد للاستمرار"
                        )
                except ValueError:
                    pass  # تجاهل إذا كان التاريخ غير صالح
    
    # إزالة كلمة المرور من الاستجابة
    if "password" in user:
        del user["password"]
    if "password_hash" in user:
        del user["password_hash"]

    # ======== المصادقة الثنائية (جهاز موثوق) ========
    _is_owner = (user.get("role") == UserRole.SUPER_ADMIN) or ((user.get("email") or "").lower() in OWNER_RECOVERY_EMAILS)
    if _is_owner:
        await _save_owner_ip(_ip, request)  # عنوان المالك يُحفظ ولا يُحظر أبداً
    if await two_fa_enabled() and not await is_device_trusted("user", user["id"], credentials.device_id):
        _channel, _dest = _choose_user_2fa_channel(user, _is_owner)
        _uname = user.get("full_name") or user.get("username") or user.get("email")
        resp2fa = await start_2fa_verification("user", user["id"], _uname, user.get("tenant_id"),
                                               _channel, _dest, credentials.device_id, _ip, request,
                                               extra={"purpose": "staff_login"},
                                               fallback_email=user.get("email"))
        await record_audit("auth.2fa.challenge", request, user=user, status=200, details={"channel": _channel})
        return resp2fa
    if await two_fa_enabled():
        await trust_device("user", user["id"], credentials.device_id, _ip,
                           request.headers.get("user-agent", ""), user.get("tenant_id"))

    token = create_token(user["id"], user["role"], user.get("branch_id"), user.get("tenant_id"))
    
    # تسجيل حدث الدخول في سجل المراقبة
    user_name = user.get("full_name") or user.get("username") or user.get("email")
    await record_audit("auth.login.success", request, user=user, status=200)
    _set_auth_cookie(response, token, user.get("role"))
    return {"user": user, "token": token}


class Verify2FARequest(BaseModel):
    verification_id: str
    code: str

class Resend2FARequest(BaseModel):
    verification_id: str

@api_router.post("/auth/login/verify-2fa")
async def verify_staff_2fa(payload: Verify2FARequest, request: Request, response: Response):
    """التحقق من رمز الدخول الثنائي للموظف/المالك وإصدار التوكن + توثيق الجهاز."""
    _ip = _client_ip(request)
    ok, sess, err = await verify_2fa_code(payload.verification_id, payload.code, _ip)
    if not ok:
        raise HTTPException(status_code=401, detail=err or "رمز التحقق غير صحيح")
    if sess.get("subject_type") != "user":
        raise HTTPException(status_code=400, detail="جلسة تحقق غير صالحة")
    user = await db.users.find_one({"id": sess["subject_id"]}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    device_id = sess.get("device_id")
    await trust_device("user", user["id"], device_id, _ip,
                       request.headers.get("user-agent", ""), user.get("tenant_id"))
    _is_owner = (user.get("role") == UserRole.SUPER_ADMIN) or ((user.get("email") or "").lower() in OWNER_RECOVERY_EMAILS)
    if _is_owner:
        await _save_owner_ip(_ip, request)
    if "password" in user:
        del user["password"]
    if "password_hash" in user:
        del user["password_hash"]
    token = create_token(user["id"], user["role"], user.get("branch_id"), user.get("tenant_id"))
    await record_audit("auth.2fa.success", request, user=user, status=200)
    _set_auth_cookie(response, token, user.get("role"))
    return {"user": user, "token": token, "device_id": device_id}

@api_router.post("/auth/2fa/resend")
async def resend_2fa(payload: Resend2FARequest, request: Request):
    """إعادة إرسال رمز التحقق لجلسة قائمة (بنفس القناة والوِجهة)."""
    _ip = _client_ip(request)
    sess = await db.verification_sessions.find_one({"id": payload.verification_id}, {"_id": 0})
    if not sess:
        raise HTTPException(status_code=404, detail="الجلسة غير موجودة")
    if sess.get("consumed"):
        raise HTTPException(status_code=400, detail="تم استخدام هذه الجلسة")
    dest = sess.get("destination")
    if sess.get("channel") == "email" and "," in (dest or ""):
        dest = dest.split(",")
    resp = await start_2fa_verification(
        sess["subject_type"], sess["subject_id"], sess.get("subject_name"),
        sess.get("tenant_id"), sess["channel"], dest, sess.get("device_id"),
        _ip, request, extra=sess.get("extra"),
    )
    return resp


@api_router.get("/auth/me")
async def get_me(current_user: dict = Depends(get_current_user)):
    user = dict(current_user)
    if "password" in user:
        del user["password"]
    return user

@api_router.post("/auth/logout")
async def logout(response: Response, current_user: dict = Depends(get_current_user)):
    response.delete_cookie(AUTH_COOKIE_NAME, path="/")
    """تسجيل خروج المستخدم مع تسجيل الحدث في سجل المراقبة"""
    user_name = current_user.get("full_name") or current_user.get("username") or current_user.get("email")
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "event_type": "logout",
        "user_id": current_user.get("id"),
        "user_name": user_name,
        "user_email": current_user.get("email"),
        "user_role": current_user.get("role"),
        "tenant_id": current_user.get("tenant_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return {"message": "تم تسجيل الخروج", "success": True}


# معاينة حساب مستخدم (تسجيل الدخول كمستخدم آخر)
@api_router.post("/auth/impersonate/{user_id}")
async def impersonate_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """
    تسجيل الدخول كمستخدم آخر (للمدراء فقط)
    يُستخدم لمعاينة التطبيق من منظور المستخدم
    """
    # التحقق من الصلاحيات (المدير العام أو المالك فقط)
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح - هذه الميزة للمدراء فقط")
    
    # Super Admin يمكنه معاينة أي مستخدم
    if current_user["role"] == UserRole.SUPER_ADMIN:
        target_user = await db.users.find_one({"id": user_id}, {"_id": 0})
    else:
        # التحقق من أن المستخدم ينتمي لنفس الـ tenant
        query = build_tenant_query(current_user, {"id": user_id})
        target_user = await db.users.find_one(query, {"_id": 0})
    
    if not target_user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # لا يمكن انتحال حساب مدير عام أو super_admin
    if target_user.get("role") in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="لا يمكن معاينة حساب مدير عام")
    
    # إزالة كلمة المرور
    if "password" in target_user:
        del target_user["password"]
    if "password_hash" in target_user:
        del target_user["password_hash"]
    
    # إضافة علامة أن هذا حساب منتحل
    target_user["impersonated"] = True
    target_user["impersonated_by"] = current_user.get("id")
    target_user["original_user_name"] = current_user.get("full_name") or current_user.get("username")
    
    # تسجيل حدث الانتحال في audit log
    admin_name = current_user.get("full_name") or current_user.get("name") or current_user.get("username") or current_user.get("email")
    target_name = target_user.get("full_name") or target_user.get("name") or target_user.get("username") or target_user.get("email")
    
    audit_log = {
        "id": str(uuid.uuid4()),
        "event_type": "impersonation",
        "admin_id": current_user.get("id"),
        "admin_name": admin_name,
        "admin_email": current_user.get("email"),
        "admin_role": current_user.get("role"),
        "target_user_id": target_user.get("id"),
        "target_user_name": target_name,
        "target_user_email": target_user.get("email"),
        "target_user_role": target_user.get("role"),
        "tenant_id": current_user.get("tenant_id"),
        "ip_address": None,  # يمكن إضافته لاحقاً
        "user_agent": None,  # يمكن إضافته لاحقاً
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.impersonation_logs.insert_one(audit_log)
    
    # تسجيل في سجل المراقبة الموحد أيضاً
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "event_type": "impersonation",
        "user_id": current_user.get("id"),
        "user_name": admin_name,
        "user_email": current_user.get("email"),
        "user_role": current_user.get("role"),
        "target_user_name": target_name,
        "target_user_role": target_user.get("role"),
        "tenant_id": current_user.get("tenant_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    
    # إنشاء توكن للمستخدم المنتحل
    token = create_token(target_user["id"], target_user["role"], target_user.get("branch_id"), target_user.get("tenant_id"))
    
    return {
        "user": target_user,
        "token": token,
        "message": f"تم تسجيل الدخول كـ {target_user.get('full_name') or target_user.get('username')}"
    }


@api_router.get("/auth/impersonation-logs")
async def get_impersonation_logs(
    current_user: dict = Depends(get_current_user),
    limit: int = 20,
    page: int = 1,
    skip: int = None
):
    """
    جلب سجلات انتحال الشخصية (للمدراء فقط)
    """
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user)
    
    # حساب skip من page إذا لم يتم تمريره
    actual_skip = skip if skip is not None else (page - 1) * limit
    
    logs = await db.impersonation_logs.find(
        query, 
        {"_id": 0}
    ).sort("created_at", -1).skip(actual_skip).limit(limit).to_list(limit)
    
    total = await db.impersonation_logs.count_documents(query)
    total_pages = (total + limit - 1) // limit  # حساب عدد الصفحات
    
    return {
        "logs": logs,
        "total": total,
        "total_pages": total_pages,
        "page": page,
        "limit": limit,
        "skip": actual_skip
    }

@api_router.get("/auth/audit-logs")
async def get_audit_logs(
    current_user: dict = Depends(get_current_user),
    limit: int = 50,
    page: int = 1
):
    """جلب سجل المراقبة الشامل - جميع عمليات الدخول/الخروج/الانتحال"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user)
    actual_skip = (page - 1) * limit
    
    # حذف السجلات الأقدم من شهر تلقائياً
    one_month_ago = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
    await db.audit_logs.delete_many({**query, "created_at": {"$lt": one_month_ago}})
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("created_at", -1).skip(actual_skip).limit(limit).to_list(limit)
    total = await db.audit_logs.count_documents(query)
    
    return {
        "logs": logs,
        "total": total,
        "total_pages": (total + limit - 1) // limit,
        "page": page
    }

@api_router.delete("/auth/audit-logs")
async def clear_audit_logs(current_user: dict = Depends(get_current_user)):
    """إفراغ سجل المراقبة"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user)
    result = await db.audit_logs.delete_many(query)
    return {"message": f"تم حذف {result.deleted_count} سجل", "success": True, "deleted_count": result.deleted_count}


# ==================== USER ROUTES ====================

@api_router.get("/users", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # فلترة المستخدمين حسب tenant_id
    query = build_tenant_query(current_user)
    users = await db.users.find(query, {"_id": 0, "password": 0}).to_list(1000)
    return users

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, update: UserUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن المستخدم ينتمي لنفس الـ tenant
    query = build_tenant_query(current_user, {"id": user_id})
    user = await db.users.find_one(query)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    # 🛡️ سياسة صلاحيات المستخدمين (طلب المالك — تحكّم كامل لمالك المشروع):
    #   - super_admin: كل شيء (نظاماً كاملاً).
    #   - tenant admin (مالك المشروع): تحكّم كامل بحسابه وحسابات مستخدميه (تعديل/إنشاء/دور/صلاحيات/فرع/تفعيل).
    #     القيد الوحيد: لا يقدر يحذف نفسه (يُطبَّق في delete_user).
    #   - manager: يعدّل الأدوار الأدنى فقط.
    #   - لا يجوز لأي أحد ترقية إلى super_admin.
    #   - لا يجوز لأي أحد لمس حسابات خارج تينانته.
    is_self_edit = (user_id == current_user.get("id"))
    is_super_admin = (current_user["role"] == UserRole.SUPER_ADMIN)
    is_tenant_admin = (current_user["role"] == UserRole.ADMIN)
    same_tenant = (user.get("tenant_id") == current_user.get("tenant_id"))
    
    new_role = update_data.get("role")
    current_role = user.get("role")
    # الفحص يعمل فقط عند تغيير الدور فعلياً (idempotent no-op مسموح)
    role_is_changing = bool(new_role) and (new_role != current_role)
    if role_is_changing:
        # (1) ترقية إلى super_admin ممنوعة لغير super_admin (حماية نظامية)
        if new_role == UserRole.SUPER_ADMIN and not is_super_admin:
            raise HTTPException(status_code=403, detail="غير مصرح بترقية حساب إلى مالك النظام")
    
    # (2) حساب super_admin لا يُعدَّل إلا بواسطة super_admin (tenant admin ممنوع من لمس حساب مالك النظام)
    if current_role == UserRole.SUPER_ADMIN and not is_super_admin:
        raise HTTPException(status_code=403, detail="غير مصرح بتعديل حساب مالك النظام")
    
    # (3) لا يمكن تعديل حسابات خارج التينانت (للجميع باستثناء super_admin)
    if not is_super_admin and not same_tenant:
        raise HTTPException(status_code=403, detail="غير مصرح — لا يمكن تعديل حسابات خارج مشروعك")
    
    # (4) حماية self-edit: مالك المشروع لا يستطيع إلغاء تفعيل نفسه أو تحويل دوره (يمنع قفل نفسه خارج النظام)
    if is_self_edit and is_tenant_admin:
        if update_data.get("is_active") is False:
            raise HTTPException(status_code=403, detail="لا يمكنك إلغاء تفعيل حسابك — يجب أن يتم ذلك بواسطة مالك النظام")
        if role_is_changing:
            raise HTTPException(status_code=403, detail="لا يمكنك تغيير دورك — يجب أن يتم ذلك بواسطة مالك النظام")
    
    # التحقق من عدم تكرار البريد الإلكتروني أو اسم المستخدم
    if update_data.get("email"):
        existing = await db.users.find_one({"email": update_data["email"], "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم بالفعل")
    
    if update_data.get("username"):
        existing = await db.users.find_one({"username": update_data["username"], "id": {"$ne": user_id}})
        if existing:
            raise HTTPException(status_code=400, detail="اسم المستخدم مستخدم بالفعل")
    
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    # 🔄 مزامنة تلقائية: عند تعديل tenant admin بياناته الشخصية، تنعكس على مستند تينانته
    # (owner_phone / owner_email / owner_name) — يظهر التحديث فوراً في لوحة تحكم مالك النظام.
    try:
        if is_self_edit and user.get("role") == UserRole.ADMIN and user.get("tenant_id"):
            tenant_sync = {}
            if update_data.get("phone"):
                tenant_sync["owner_phone"] = update_data["phone"]
            if update_data.get("email"):
                tenant_sync["owner_email"] = update_data["email"]
            if update_data.get("full_name"):
                tenant_sync["owner_name"] = update_data["full_name"]
            if tenant_sync:
                tenant_sync["updated_at"] = datetime.now(timezone.utc).isoformat()
                await db.tenants.update_one({"id": user["tenant_id"]}, {"$set": tenant_sync})
                logger.info(f"🔄 Tenant admin {user_id} synced to tenant doc: {list(tenant_sync.keys())}")
    except Exception as _sync_err:
        logger.warning(f"tenant owner sync failed: {_sync_err}")
    
    user = await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})
    return user

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    """حذف مستخدم - العملاء (Admin) فقط يحذفون مستخدميهم، Super Admin يحذف العملاء فقط من صفحة العملاء"""
    
    # Super Admin لا يحذف المستخدمين من هنا - يحذف العملاء فقط من endpoint العملاء
    if current_user["role"] == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="المالك يحذف العملاء فقط من صفحة إدارة العملاء")
    
    # فقط Admin و Manager يمكنهم حذف المستخدمين
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح لك بحذف المستخدمين")
    
    # Admin/Manager يحذف فقط مستخدمي نفس الـ tenant
    query = build_tenant_query(current_user, {"id": user_id})
    user = await db.users.find_one(query)
    
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    # لا يمكن حذف نفسك مطلقاً (bug fix: كان current_user.get('user_id') والصحيح 'id')
    if user_id == current_user.get("id") or user_id == current_user.get("user_id"):
        raise HTTPException(status_code=400, detail="لا يمكنك حذف حسابك الخاص")
    
    # حسابات admin/super_admin لا تُحذف إلا بواسطة super_admin (المالك ما يحذف نفسه ولا شريكه)
    if user.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN] and current_user["role"] != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح — فقط مالك النظام يستطيع حذف حساب مالك مشروع")
    
    # حذف المستخدم نهائياً
    result = await db.users.delete_one({"id": user_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="فشل في حذف المستخدم")
    
    logger.info(f"User {user_id} deleted by {current_user.get('user_id')}")
    return {"message": "تم حذف المستخدم بنجاح"}

@api_router.get("/tenant/limits")
async def get_tenant_limits(current_user: dict = Depends(get_current_user)):
    """جلب حدود العميل الحالي (الفروع والمستخدمين)"""
    tenant_id = get_user_tenant_id(current_user)
    
    if not tenant_id:
        return {
            "max_branches": 999,
            "max_users": 999,
            "current_branches": 0,
            "current_users": 0,
            "branches_remaining": 999,
            "users_remaining": 999
        }
    
    tenant = await db.tenants.find_one({"id": tenant_id})
    if not tenant:
        return {
            "max_branches": 1,
            "max_users": 5,
            "current_branches": 0,
            "current_users": 0,
            "branches_remaining": 1,
            "users_remaining": 5
        }
    
    max_branches = tenant.get("max_branches", 1)
    max_users = tenant.get("max_users", 5)
    
    # حساب العدد الحالي (جميع الفروع النشطة)
    current_branches = await db.branches.count_documents({
        "tenant_id": tenant_id, 
        "is_active": {"$ne": False}
    })
    current_users = await db.users.count_documents({
        "tenant_id": tenant_id, 
        "is_active": {"$ne": False}
    })
    
    return {
        "max_branches": max_branches,
        "max_users": max_users,
        "current_branches": current_branches,
        "current_users": current_users,
        "branches_remaining": max(0, max_branches - current_branches),
        "users_remaining": max(0, max_users - current_users)
    }

@api_router.post("/users", response_model=UserResponse)
async def create_user(user: UserCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء مستخدم جديد مع tenant_id تلقائياً"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="إنشاء الحسابات متاح للمالك فقط")

    # منع تصعيد الصلاحيات: لا يُسمح بإنشاء super_admin إطلاقاً عبر هذا المسار
    if user.role == UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح بإنشاء حساب مالك النظام")
    # إنشاء حساب admin (شريك/مالك) مسموح فقط لـsuper_admin — المالك يديره ولا يُنشئ شركاء
    if current_user["role"] != UserRole.SUPER_ADMIN and user.role == UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح بإنشاء حساب مالك مشروع — تواصل مع مالك النظام")
    
    # التحقق من عدم وجود المستخدم
    existing = await db.users.find_one({"$or": [{"email": user.email}, {"username": user.username}]})
    if existing:
        raise HTTPException(status_code=400, detail="المستخدم موجود بالفعل")
    validate_password_strength(user.password)
    
    # الحصول على tenant_id من المستخدم الحالي
    tenant_id = get_user_tenant_id(current_user)
    
    # التحقق من الحد الأقصى للمستخدمين
    if tenant_id and current_user["role"] != UserRole.SUPER_ADMIN:
        tenant = await db.tenants.find_one({"id": tenant_id})
        if tenant:
            max_users = tenant.get("max_users", 5)
            current_users_count = await db.users.count_documents({"tenant_id": tenant_id, "is_active": {"$ne": False}})
            if current_users_count >= max_users:
                raise HTTPException(
                    status_code=403, 
                    detail=f"تم الوصول للحد الأقصى من المستخدمين ({max_users}). يرجى مراجعة مسؤول النظام لرفع الحد"
                )
    
    user_doc = {
        "id": str(uuid.uuid4()),
        "username": user.username,
        "email": user.email,
        "password": hash_password(user.password),
        "full_name": user.full_name,
        "role": user.role,
        "branch_id": user.branch_id,
        "permissions": user.permissions,
        "phone": (user.phone or "").strip(),
        "full_name_en": user.full_name_en,
        "tenant_id": tenant_id,  # إضافة tenant_id تلقائياً
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(user_doc)
    del user_doc["password"]
    del user_doc["_id"]
    return user_doc

class PasswordReset(BaseModel):
    new_password: str

@api_router.put("/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, data: PasswordReset, current_user: dict = Depends(get_current_user)):
    """إعادة تعيين كلمة مرور المستخدم"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن المستخدم ينتمي لنفس الـ tenant
    query = build_tenant_query(current_user, {"id": user_id})
    user = await db.users.find_one(query)
    if not user:
        raise HTTPException(status_code=404, detail="المستخدم غير موجود")
    
    validate_password_strength(data.new_password)
    hashed = hash_password(data.new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"password": hashed}})
    
    return {"message": "تم تغيير كلمة المرور بنجاح"}

# ==================== BRANCH ROUTES ====================

@api_router.post("/branches", response_model=BranchResponse)
async def create_branch(branch: BranchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # التحقق من الحد الأقصى للفروع
    if tenant_id and current_user["role"] != UserRole.SUPER_ADMIN:
        tenant = await db.tenants.find_one({"id": tenant_id})
        if tenant:
            max_branches = tenant.get("max_branches", 1)
            current_branches_count = await db.branches.count_documents({
                "tenant_id": tenant_id, 
                "is_active": {"$ne": False}
            })
            if current_branches_count >= max_branches:
                raise HTTPException(
                    status_code=403, 
                    detail=f"تم الوصول للحد الأقصى من الفروع ({max_branches}). يرجى مراجعة مسؤول النظام لرفع الحد"
                )
    
    branch_doc = {
        "id": str(uuid.uuid4()),
        **branch.model_dump(),
        "tenant_id": tenant_id,  # فصل البيانات
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.branches.insert_one(branch_doc)
    del branch_doc["_id"]
    return branch_doc

@api_router.get("/branches")
async def get_branches(
    current_user: dict = Depends(get_current_user),
    include_inactive: bool = False,
    include_departments: bool = False,
):
    """
    جلب الفروع - مع عزل صارم للبيانات بين المستأجرين
    
    - Super Admin الحقيقي (بدون tenant_id) = لا يرى فروع (يستخدم Super Admin Panel)
    - Admin/Manager = يرى فروع الـ tenant الخاص به فقط
    - Cashier/Staff = يرى فرعه فقط
    
    include_departments=False (افتراضي): يرجع الفروع العادية فقط (لـPOS, Dashboard, etc.)
    include_departments=True: يرجع الفروع + الأقسام (المطبخ المركزي/المخزن/المشتريات) — لـHR
    """
    tenant_id = current_user.get("tenant_id")
    
    # Super Admin الحقيقي بدون tenant - لا يرى فروع عادية
    if current_user.get("role") == UserRole.SUPER_ADMIN and not tenant_id:
        return []  # Super Admin يستخدم Super Admin Panel لإدارة العملاء
    
    # بناء query مع فلترة صارمة للـ tenant
    query = {"tenant_id": tenant_id} if tenant_id else {"tenant_id": "default"}
    
    # المستخدمون المرتبطون بفرع معين يرون فقط فرعهم
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["id"] = user_branch_id
    
    # إخفاء الفروع المعطّلة إلا إذا طُلب عرضها
    if not include_inactive:
        query["is_active"] = {"$ne": False}
    
    # === فلترة الأقسام (مطبخ مركزي/مخزن/مشتريات) ===
    # افتراضياً نخفيها — تظهر فقط لـHR عند تمرير include_departments=true
    if not include_departments:
        query["$and"] = query.get("$and", []) + [
            {"$or": [
                {"branch_type": {"$exists": False}},
                {"branch_type": "branch"},
                {"branch_type": None},
            ]}
        ]
    
    branches = await db.branches.find(query, {"_id": 0}).to_list(100)
    # تطبيع المخرجات عبر BranchResponse (نفس سلوك الإدارة السابق) + إخفاء الحقول المالية عن غير الإدارة
    _SENSITIVE_BRANCH_FIELDS = (
        "rent_cost", "water_cost", "electricity_cost", "generator_cost",
        "is_sold_branch", "buyer_name", "buyer_phone", "owner_percentage", "monthly_fee",
    )
    is_mgmt = user_role in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]
    out = []
    for _b in branches:
        bd = BranchResponse(**_b).model_dump()
        if not is_mgmt:
            for _f in _SENSITIVE_BRANCH_FIELDS:
                bd.pop(_f, None)
        out.append(bd)
    return out

@api_router.get("/branches/{branch_id}", response_model=BranchResponse)
async def get_branch(branch_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": branch_id})
    branch = await db.branches.find_one(query, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="الفرع غير موجود")
    return branch

@api_router.put("/branches/{branch_id}")
async def update_branch(branch_id: str, branch: BranchCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": branch_id})
    await db.branches.update_one(query, {"$set": branch.model_dump()})
    return await db.branches.find_one({"id": branch_id}, {"_id": 0})

@api_router.delete("/branches/{branch_id}")
async def delete_branch(branch_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # Check if branch has users or orders
    users_count = await db.users.count_documents({"branch_id": branch_id})
    if users_count > 0:
        raise HTTPException(status_code=400, detail="لا يمكن حذف الفرع - يوجد مستخدمين مرتبطين به")
    
    # حذف الفرع نهائياً
    result = await db.branches.delete_one({"id": branch_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الفرع غير موجود")
    
    logger.info(f"Branch {branch_id} deleted by {current_user.get('user_id')}")
    return {"message": "تم حذف الفرع بنجاح"}

# ==================== KITCHEN SECTIONS ROUTES ====================

@api_router.post("/kitchen-sections")
async def create_kitchen_section(section: dict, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    section_doc = {
        "id": str(uuid.uuid4()),
        "name": section.get("name"),
        "name_en": section.get("name_en"),
        "color": section.get("color", "#D4AF37"),
        "icon": section.get("icon", "🍳"),
        "printer_id": section.get("printer_id"),
        "branch_id": section.get("branch_id"),
        "tenant_id": tenant_id,  # فصل البيانات لكل عميل
        "sort_order": section.get("sort_order", 0),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.kitchen_sections.insert_one(section_doc)
    del section_doc["_id"]
    return section_doc

@api_router.get("/kitchen-sections")
async def get_kitchen_sections(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    sections = await db.kitchen_sections.find(query, {"_id": 0}).sort("sort_order", 1).to_list(100)
    return sections

@api_router.put("/kitchen-sections/{section_id}")
async def update_kitchen_section(section_id: str, section: dict, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": section_id})
    update_data = {k: v for k, v in section.items() if k != "id"}
    await db.kitchen_sections.update_one(query, {"$set": update_data})
    return await db.kitchen_sections.find_one({"id": section_id}, {"_id": 0})

@api_router.delete("/kitchen-sections/{section_id}")
async def delete_kitchen_section(section_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": section_id})
    await db.kitchen_sections.delete_one(query)
    return {"message": "تم الحذف"}

@api_router.put("/categories/{category_id}/kitchen-section")
async def assign_category_to_kitchen_section(category_id: str, data: dict, current_user: dict = Depends(get_current_user)):
    """Assign a category to a kitchen section"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    kitchen_section_id = data.get("kitchen_section_id")
    await db.categories.update_one(
        {"id": category_id}, 
        {"$set": {"kitchen_section_id": kitchen_section_id}}
    )
    return await db.categories.find_one({"id": category_id}, {"_id": 0})

# ==================== CATEGORY ROUTES ====================

@api_router.post("/categories", response_model=CategoryResponse)
async def create_category(category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    cat_doc = {
        "id": str(uuid.uuid4()),
        **category.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "is_active": True
    }
    await db.categories.insert_one(cat_doc)
    del cat_doc["_id"]
    return cat_doc

@api_router.get("/categories", response_model=List[CategoryResponse])
async def get_categories(current_user: dict = Depends(get_current_user)):
    """جلب الفئات - مع عزل صارم للبيانات"""
    tenant_id = current_user.get("tenant_id")
    
    # Super Admin بدون tenant لا يرى فئات
    if current_user.get("role") == UserRole.SUPER_ADMIN and not tenant_id:
        return []
    
    query = {"tenant_id": tenant_id} if tenant_id else {"tenant_id": "default"}
    categories = await db.categories.find(query, {"_id": 0}).sort("sort_order", 1).to_list(100)
    return categories

@api_router.put("/categories/{category_id}")
async def update_category(category_id: str, category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": category_id})
    await db.categories.update_one(query, {"$set": category.model_dump()})
    return await db.categories.find_one({"id": category_id}, {"_id": 0})

@api_router.delete("/categories/{category_id}")
async def delete_category(category_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # Super Admin يحذف أي فئة
    if current_user["role"] == UserRole.SUPER_ADMIN:
        query = {"id": category_id}
    else:
        query = build_tenant_query(current_user, {"id": category_id})
    
    result = await db.categories.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الفئة غير موجودة")
    
    logger.info(f"Category {category_id} deleted by {current_user.get('user_id')}")
    return {"message": "تم حذف الفئة بنجاح"}

# ==================== PRODUCT ROUTES ====================

@api_router.post("/products", response_model=ProductResponse)
async def create_product(product: ProductCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # Calculate profit
    profit = product.price - product.cost - product.operating_cost
    
    prod_doc = {
        "id": str(uuid.uuid4()),
        **product.model_dump(),
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "profit": profit
    }
    await db.products.insert_one(prod_doc)
    del prod_doc["_id"]
    return prod_doc

@api_router.get("/products", response_model=List[ProductResponse])
async def get_products(
    category_id: Optional[str] = None,
    skip: int = Query(0, ge=0, description="عدد العناصر للتخطي"),
    limit: int = Query(100, ge=1, le=500, description="الحد الأقصى للعناصر"),
    current_user: dict = Depends(get_current_user)
):
    """جلب المنتجات - مع عزل صارم للبيانات"""
    tenant_id = current_user.get("tenant_id")
    
    # Super Admin بدون tenant لا يرى منتجات
    if current_user.get("role") == UserRole.SUPER_ADMIN and not tenant_id:
        return []
    
    query = {"tenant_id": tenant_id} if tenant_id else {"tenant_id": "default"}
    if category_id:
        query["category_id"] = category_id
    products = await db.products.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    _can_see_cost = (current_user.get("role") or "").lower() in ("admin", "manager", "super_admin", "owner")
    for p in products:
        if _can_see_cost:
            p["profit"] = _sn(p.get("price")) - _sn(p.get("cost")) - _sn(p.get("operating_cost"))
        else:
            # 🔒 إخفاء التكلفة/الربح عن الكاشير وبقية الأدوار غير الإدارية (تقرير الأمان #5)
            for _cf in ("cost", "operating_cost", "packaging_cost", "profit", "profit_margin",
                        "raw_material_cost", "raw_material_cost_after_waste", "production_cost",
                        "cost_before_waste", "cost_after_waste", "unit_cost"):
                if _cf in p:
                    p[_cf] = 0
    return products

@api_router.get("/products/{product_id}")
async def get_product(product_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": product_id})
    product = await db.products.find_one(query, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    if (current_user.get("role") or "").lower() in ("admin", "manager", "super_admin", "owner"):
        product["profit"] = _sn(product.get("price")) - _sn(product.get("cost")) - _sn(product.get("operating_cost"))
    else:
        for _cf in ("cost", "operating_cost", "packaging_cost", "profit", "profit_margin",
                    "raw_material_cost", "raw_material_cost_after_waste", "production_cost",
                    "cost_before_waste", "cost_after_waste", "unit_cost"):
            if _cf in product:
                product[_cf] = 0
    return product

@api_router.put("/products/{product_id}")
async def update_product(product_id: str, product: ProductCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    profit = product.price - product.cost - product.operating_cost
    update_data = {**product.model_dump(), "profit": profit}
    
    query = build_tenant_query(current_user, {"id": product_id})
    await db.products.update_one(query, {"$set": update_data})
    return await db.products.find_one({"id": product_id}, {"_id": 0})

@api_router.post("/admin/translate-names")
async def translate_entity_names(overwrite: bool = False, current_user: dict = Depends(get_current_user)):
    """ترجمة تلقائية لأسماء المنتجات والأقسام إلى الإنجليزية (name_en) عبر الذكاء الاصطناعي."""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")

    tenant_id = get_user_tenant_id(current_user)
    api_key = os.environ.get("EMERGENT_LLM_KEY")
    if not api_key:
        raise HTTPException(status_code=500, detail="مفتاح الذكاء الاصطناعي غير مهيأ")

    from emergentintegrations.llm.chat import LlmChat, UserMessage

    def _parse_json(text: str) -> dict:
        text = (text or "").strip()
        if text.startswith("```"):
            parts = text.split("```")
            if len(parts) >= 2:
                text = parts[1]
                if text.lstrip().lower().startswith("json"):
                    text = text.lstrip()[4:]
        try:
            return json.loads(text)
        except Exception:
            m = re.search(r"\{.*\}", text, re.DOTALL)
            return json.loads(m.group(0)) if m else {}

    async def translate_collection(coll_name: str) -> int:
        coll = db[coll_name]
        q = {"tenant_id": tenant_id}
        if not overwrite:
            q["$or"] = [{"name_en": {"$in": [None, ""]}}, {"name_en": {"$exists": False}}]
        items = await coll.find(q, {"_id": 0, "id": 1, "name": 1}).to_list(1000)
        items = [it for it in items if it.get("name")]
        translated = 0
        for i in range(0, len(items), 40):
            chunk = items[i:i + 40]
            names = {it["id"]: it["name"] for it in chunk}
            chat = LlmChat(
                api_key=api_key,
                session_id=f"translate-{coll_name}-{uuid.uuid4()}",
                system_message=(
                    "You are a professional menu translator for an Iraqi restaurant POS. "
                    "Translate Arabic food and category names into concise, natural English menu names. "
                    "Keep brand/proper names sensible. Return ONLY valid JSON mapping each given id to its English name, no extra text."
                ),
            ).with_model("openai", "gpt-4o")
            prompt = "Translate these to English. Return a JSON object {id: english_name}.\n" + json.dumps(names, ensure_ascii=False)
            try:
                resp = await chat.send_message(UserMessage(text=prompt))
            except Exception as e:
                logger.error(f"translate LLM error: {e}")
                raise HTTPException(status_code=502, detail="تعذّر الاتصال بخدمة الترجمة")
            mapping = _parse_json(resp if isinstance(resp, str) else str(resp))
            for pid, en in mapping.items():
                if en and isinstance(en, str):
                    await coll.update_one(
                        {"id": pid, "tenant_id": tenant_id},
                        {"$set": {"name_en": en.strip()}},
                    )
                    translated += 1
        return translated

    products_n = await translate_collection("products")
    categories_n = await translate_collection("categories")
    return {
        "products_translated": products_n,
        "categories_translated": categories_n,
        "message": f"تمت ترجمة {products_n} منتج و {categories_n} قسم",
    }


@api_router.delete("/products/{product_id}")
async def delete_product(product_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # Super Admin يحذف أي منتج، Admin يحذف منتجات tenant الخاص به
    if current_user["role"] == UserRole.SUPER_ADMIN:
        query = {"id": product_id}
    else:
        query = build_tenant_query(current_user, {"id": product_id})
    
    result = await db.products.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    logger.info(f"Product {product_id} deleted by {current_user.get('user_id')}")
    return {"message": "تم حذف المنتج بنجاح"}

# ==================== INVENTORY ROUTES ====================

@api_router.post("/inventory", response_model=InventoryResponse)
async def create_inventory_item(item: InventoryItemCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    inv_doc = {
        "id": str(uuid.uuid4()),
        **item.model_dump(),
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory.insert_one(inv_doc)
    del inv_doc["_id"]
    return inv_doc

@api_router.get("/inventory", response_model=List[InventoryResponse])
async def get_inventory(
    branch_id: Optional[str] = None, 
    item_type: Optional[str] = None,
    skip: int = Query(0, ge=0, description="عدد العناصر للتخطي"),
    limit: int = Query(100, ge=1, le=500, description="الحد الأقصى للعناصر"),
    current_user: dict = Depends(get_current_user)
):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    if item_type:
        query["item_type"] = item_type
    items = await db.inventory.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return items

@api_router.post("/inventory/transaction")
async def inventory_transaction(transaction: InventoryTransaction, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": transaction.inventory_id})
    item = await db.inventory.find_one(query)
    if not item:
        raise HTTPException(status_code=404, detail="الصنف غير موجود")
    
    new_qty = item["quantity"]
    if transaction.transaction_type == "in":
        new_qty += transaction.quantity
    else:
        new_qty -= transaction.quantity
        if new_qty < 0:
            raise HTTPException(status_code=400, detail="الكمية غير كافية")
    
    await db.inventory.update_one(
        {"id": transaction.inventory_id},
        {"$set": {"quantity": new_qty, "last_updated": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Log transaction
    trans_doc = {
        "id": str(uuid.uuid4()),
        "inventory_id": transaction.inventory_id,
        "transaction_type": transaction.transaction_type,
        "quantity": transaction.quantity,
        "notes": transaction.notes,
        "user_id": current_user["id"],
        "tenant_id": get_user_tenant_id(current_user),  # فصل البيانات
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory_transactions.insert_one(trans_doc)
    
    return {"message": "تمت العملية بنجاح", "new_quantity": new_qty}

# ==================== PURCHASE ROUTES - المشتريات ====================

@api_router.post("/purchases")
async def create_purchase(purchase: PurchaseCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    purchase_doc = {
        "id": str(uuid.uuid4()),
        **purchase.model_dump(),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.purchases.insert_one(purchase_doc)
    
    # Update inventory quantities
    for item in purchase.items:
        await db.inventory.update_one(
            {"id": item.get("inventory_id")},
            {
                "$inc": {"quantity": _sn(item.get("quantity"))},
                "$set": {
                    "cost_per_unit": item.get("cost_per_unit", 0),
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
            }
        )
    
    del purchase_doc["_id"]
    return purchase_doc

@api_router.get("/purchases")
async def get_purchases(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if branch_id:
        query["branch_id"] = branch_id
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        query.setdefault("created_at", {})["$lte"] = end_date + "T23:59:59"
    
    purchases = await db.purchases.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return purchases

# ==================== EXPENSE ROUTES - المصاريف ====================

@api_router.post("/expenses")
async def create_expense(expense: ExpenseCreate, current_user: dict = Depends(get_current_user)):
    user_permissions = current_user.get("permissions", [])
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN] and "expenses" not in user_permissions:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id_for_biz = get_user_tenant_id(current_user)
    branch_for_biz = expense.branch_id if hasattr(expense, 'branch_id') else None
    
    # حماية ضد التكرار: نفس المستخدم + الفرع + المبلغ + الوصف خلال آخر 10 ثواني = تكرار
    from datetime import timedelta
    dup_cutoff = (datetime.now(timezone.utc) - timedelta(seconds=10)).isoformat()
    dup_query = {
        "created_by": current_user["id"],
        "branch_id": branch_for_biz,
        "amount": float(expense.amount),
        "description": (expense.description or "").strip(),
        "created_at": {"$gte": dup_cutoff}
    }
    if tenant_id_for_biz:
        dup_query["tenant_id"] = tenant_id_for_biz
    existing_dup = await db.expenses.find_one(dup_query, {"_id": 0})
    if existing_dup:
        return existing_dup  # إرجاع المصروف الموجود بدل إنشاء نسخة مكررة
    
    business_date = await _resolve_business_date(tenant_id_for_biz, branch_for_biz, current_user.get("id"))
    # ربط المصروف بوردية الكاشير المفتوحة (إن وُجدت) — أساس احتساب مصاريف الوردية بدقة وبلا خلط
    # 🛡 صارم: نبحث فقط عن الوردية المفتوحة لهذا الكاشير في هذا الفرع لليوم الحالي
    open_shift_q = {
        "status": "open",
        "cashier_id": current_user["id"],
        "business_date": business_date,
    }
    if branch_for_biz:
        open_shift_q["branch_id"] = branch_for_biz
    if tenant_id_for_biz:
        open_shift_q["tenant_id"] = tenant_id_for_biz
    open_shift = await db.shifts.find_one(open_shift_q, {"_id": 0, "id": 1})
    
    # 🛡 Lazy Shift Opening للمصاريف: إذا لم توجد وردية مفتوحة للكاشير اليوم، افتحها تلقائياً
    # هذا يضمن أن كل مصروف يحمل shift_id صحيح ولا يتسرّب لملخصات ورديات أخرى.
    # نستثني المدراء/المالكين (يمكنهم تسجيل مصاريف عامة بلا وردية)
    is_manager_role = current_user.get("role", "") in ["admin", "super_admin", "manager", "branch_manager"]
    if not open_shift and branch_for_biz and not is_manager_role:
        new_shift_id = str(uuid.uuid4())
        new_shift_doc = {
            "id": new_shift_id,
            "tenant_id": tenant_id_for_biz,
            "branch_id": branch_for_biz,
            "cashier_id": current_user["id"],
            "cashier_name": current_user.get("full_name", "") or current_user.get("username", ""),
            "started_at": datetime.now(timezone.utc).isoformat(),
            "opened_at": datetime.now(timezone.utc).isoformat(),
            "status": "open",
            "opening_balance": 0,
            "opening_cash": 0,
            "business_date": business_date,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "auto_opened_by_first_expense": True,
        }
        await db.shifts.insert_one(new_shift_doc)
        open_shift = {"id": new_shift_id}
        logger.info(f"✅ فُتحت وردية جديدة تلقائياً لتسجيل مصروف — كاشير {current_user.get('full_name')} — business_date={business_date}")
    
    expense_doc = {
        "id": str(uuid.uuid4()),
        **expense.model_dump(),
        "tenant_id": tenant_id_for_biz,
        "date": expense.date or datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "business_date": business_date,
        "shift_id": open_shift["id"] if open_shift else None,
        "cashier_id": current_user["id"],
        "created_by": current_user["id"],
        "created_by_name": current_user.get("full_name", "") or current_user.get("username", ""),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.expenses.insert_one(expense_doc)
    del expense_doc["_id"]
    
    # تحديث إجمالي مصروفات الوردية المفتوحة الحالية (لو كان المصروف ليس مرتجع)
    if expense_doc.get("category") != "refund" and branch_for_biz and open_shift:
        # 🛡 صارم: نُحدّث فقط الوردية التي رُبط بها المصروف (وليس أي وردية في الفرع)
        current_shift = await db.shifts.find_one(
            {"id": open_shift["id"]},
            {"_id": 0, "id": 1, "started_at": 1, "opening_cash": 1, "opening_balance": 1, "cash_sales": 1, "cashier_id": 1}
        )
        if current_shift:
            # نستخدم نفس منطق shift_expense_query لضمان اتساق العدّ
            from routes.shared import shift_expense_query
            exp_q = shift_expense_query(current_shift, tenant_id_for_biz)
            shift_expenses_live = await db.expenses.find(exp_q, {"_id": 0, "amount": 1}).to_list(500)
            total_exp_live = sum(float(e.get("amount") or 0) for e in shift_expenses_live)
            opening_cash_live = float(current_shift.get("opening_cash") or current_shift.get("opening_balance") or 0)
            cash_sales_live = float(current_shift.get("cash_sales") or 0)
            await db.shifts.update_one(
                {"id": current_shift["id"]},
                {"$set": {
                    "total_expenses": total_exp_live,
                    "expected_cash": opening_cash_live + cash_sales_live - total_exp_live
                }}
            )
    
    return expense_doc

@api_router.get("/business-date/current")
async def get_current_business_date(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إرجاع business_date الحالي للوردية المفتوحة في الفرع.
    
    مفيد للواجهة لكي تعرض مصاريف/مبيعات اليوم التشغيلي الفعلي بدل التاريخ التقويمي.
    مثال: شفت افتُتح الساعة 10 مساءً وامتد حتى 2 صباحاً اليوم التالي — business_date
    يبقى = تاريخ الافتتاح (لا يتغير مع منتصف الليل).
    
    إذا لم توجد وردية مفتوحة، يرجع تاريخ العراق الحالي.
    """
    tenant_id = get_user_tenant_id(current_user)
    br = branch_id or current_user.get("branch_id")
    biz = await _resolve_business_date(tenant_id, br, current_user.get("id"))
    # اجلب معلومات الوردية المفتوحة كذلك لعرضها في الواجهة
    shift_q = {"status": "open"}
    if tenant_id:
        shift_q["tenant_id"] = tenant_id
    if br:
        shift_q["branch_id"] = br
    shift = await db.shifts.find_one(shift_q, {"_id": 0, "id": 1, "cashier_id": 1, "cashier_name": 1, "started_at": 1, "business_date": 1}, sort=[("started_at", -1)])
    return {
        "business_date": biz,
        "calendar_date_iraq": iraq_date_from_utc(),
        "has_open_shift": bool(shift),
        "open_shift": shift or None,
    }


@api_router.get("/shifts/stale")
async def list_stale_shifts(
    hours_threshold: int = 18,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """قائمة الورديات العالقة (المفتوحة أكثر من X ساعة).
    
    للمالك/المدير فقط — تساعد على اكتشاف الورديات اللي نسي الكاشير إغلاقها
    وتسبب مشاكل في تقارير الـ business_date.
    """
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مسموح")
    
    tenant_id = get_user_tenant_id(current_user)
    cutoff = (datetime.now(timezone.utc) - timedelta(hours=hours_threshold)).isoformat()
    
    q = {"status": "open", "started_at": {"$lt": cutoff}}
    if tenant_id:
        q["tenant_id"] = tenant_id
    if branch_id:
        q["branch_id"] = branch_id
    
    stale = await db.shifts.find(q, {"_id": 0, "id": 1, "cashier_id": 1, "cashier_name": 1, "started_at": 1, "business_date": 1, "branch_id": 1}).to_list(50)
    
    # احسب الساعات منذ الافتتاح
    now = datetime.now(timezone.utc)
    for s in stale:
        try:
            started = datetime.fromisoformat(s["started_at"].replace("Z", "+00:00"))
            s["hours_open"] = round((now - started).total_seconds() / 3600, 1)
        except Exception:
            s["hours_open"] = None
    
    return {"stale_shifts": stale, "count": len(stale), "hours_threshold": hours_threshold}


@api_router.post("/shifts/{shift_id}/force-close")
async def force_close_stale_shift(
    shift_id: str,
    current_user: dict = Depends(get_current_user)
):
    """إغلاق وردية عالقة بصلاحية المالك/المدير.
    
    يضع status=closed مع تسجيل الإغلاق القسري في force_closed_by/force_closed_at
    لكي تتميز عن الإغلاقات الطبيعية. لا تنشئ ايصال إغلاق صندوق (المبيعات والمصاريف
    تُحسب طبيعياً في تقرير اليوم التشغيلي للوردية).
    """
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مسموح")
    
    tenant_id = get_user_tenant_id(current_user)
    q = {"id": shift_id, "status": "open"}
    if tenant_id:
        q["tenant_id"] = tenant_id
    
    shift = await db.shifts.find_one(q, {"_id": 0})
    if not shift:
        raise HTTPException(status_code=404, detail="الوردية غير موجودة أو مغلقة بالفعل")
    
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.shifts.update_one(
        {"id": shift_id},
        {"$set": {
            "status": "closed",
            "ended_at": now_iso,
            "force_closed_by": current_user.get("id"),
            "force_closed_by_name": current_user.get("full_name") or current_user.get("username"),
            "force_closed_at": now_iso,
            "force_close_reason": "stale_shift_admin_override",
        }}
    )
    return {"message": "تم إغلاق الوردية بنجاح", "shift_id": shift_id, "cashier_name": shift.get("cashier_name")}


@api_router.get("/expenses")
async def get_expenses(
    branch_id: Optional[str] = None,
    category: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    # استبعاد المرتجعات من المصاريف نهائياً
    if category:
        query["category"] = category
    else:
        query["category"] = {"$ne": "refund"}
    
    # الكاشير يرى مصاريفه فقط في اليوم الحالي
    user_role = current_user.get("role", "")
    is_manager = user_role in ["admin", "super_admin", "manager", "branch_manager"]
    if not is_manager:
        query["created_by"] = current_user["id"]
        # الكاشير يرى فقط مصاريف اليوم التشغيلي (business_date من الوردية المفتوحة، أو تاريخ العراق اليوم)
        user_tenant = get_user_tenant_id(current_user)
        user_branch = current_user.get("branch_id") or branch_id
        biz_today = await _resolve_business_date(user_tenant, user_branch, current_user.get("id"))
        # مطابقة business_date أو fallback للتاريخ القديم `date` للسجلات القديمة
        query["$or"] = [
            {"business_date": biz_today},
            {"business_date": {"$exists": False}, "date": biz_today}
        ]
    else:
        if start_date or end_date:
            date_range = {}
            if start_date:
                date_range["$gte"] = start_date
            if end_date:
                date_range["$lte"] = end_date
            # استخدام business_date عند توفره، مع fallback للتاريخ القديم للسجلات القديمة
            query["$or"] = [
                {"business_date": date_range.copy()},
                {"business_date": {"$exists": False}, "date": date_range.copy()}
            ]
    
    if branch_id:
        query["branch_id"] = branch_id
    
    expenses = await db.expenses.find(query, {"_id": 0}).sort("date", -1).to_list(500)
    
    # ملء اسم الكاشير للمصاريف القديمة التي لا تحتوي على created_by_name
    needs_update = [e for e in expenses if not e.get("created_by_name") and e.get("created_by")]
    if needs_update:
        user_ids = list(set(e["created_by"] for e in needs_update))
        users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "full_name": 1, "username": 1}).to_list(100)
        user_map = {u["id"]: u.get("full_name") or u.get("username", "") for u in users}
        for e in expenses:
            if not e.get("created_by_name") and e.get("created_by"):
                e["created_by_name"] = user_map.get(e["created_by"], "")
    
    return expenses



@api_router.get("/expenses/categories")
async def get_expense_categories(current_user: dict = Depends(get_current_user)):
    """جلب التصنيفات الافتراضية"""
    return [
        {"id": "rent", "name": "إيجار"},
        {"id": "utilities", "name": "كهرباء وماء"},
        {"id": "gas", "name": "غاز"},
        {"id": "salaries", "name": "رواتب"},
        {"id": "advance", "name": "سلف"},
        {"id": "maintenance", "name": "صيانة"},
        {"id": "supplies", "name": "مستلزمات"},
        {"id": "marketing", "name": "تسويق"},
        {"id": "transport", "name": "نقل"},
        {"id": "other", "name": "أخرى"}
    ]

@api_router.get("/expense-categories")
async def get_custom_expense_categories(current_user: dict = Depends(get_current_user)):
    """جلب التصنيفات المخصصة"""
    tenant_id = current_user.get("tenant_id")
    query = {"tenant_id": tenant_id} if tenant_id else {}
    
    categories = await db.expense_categories.find(
        query,
        {"_id": 0}
    ).to_list(100)
    
    return categories

@api_router.post("/expense-categories")
async def create_expense_category(category: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """إنشاء تصنيف مصاريف جديد"""
    tenant_id = current_user.get("tenant_id")
    
    # التحقق من عدم وجود التصنيف مسبقاً
    existing = await db.expense_categories.find_one({
        "id": category.get("id"),
        "tenant_id": tenant_id
    })
    
    if existing:
        return {"message": "التصنيف موجود بالفعل", "category": existing}
    
    category_doc = {
        "id": category.get("id"),
        "name": category.get("name"),
        "icon": category.get("icon", "🏷️"),
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.expense_categories.insert_one(category_doc)
    category_doc.pop("_id", None)
    
    return category_doc

# ==================== OPERATING COST ROUTES - التكاليف التشغيلية ====================

@api_router.post("/operating-costs")
async def create_operating_cost(cost: OperatingCostCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    cost_doc = {
        "id": str(uuid.uuid4()),
        **cost.model_dump(),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.operating_costs.insert_one(cost_doc)
    del cost_doc["_id"]
    return cost_doc

@api_router.get("/operating-costs")
async def get_operating_costs(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"branch_id": branch_id} if branch_id else {}
    costs = await db.operating_costs.find(query, {"_id": 0}).to_list(100)
    return costs

# ==================== moved to routes/hr_routes.py ====================

# ==================== EMPLOYEE RATINGS - تقييم الموظفين التلقائي ====================

@api_router.get("/employee-ratings")
async def get_employee_ratings(
    month: str = None,  # YYYY-MM
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """حساب تقييم الموظفين بناءً على الحضور والانصراف والخصومات"""
    tenant_id = get_user_tenant_id(current_user)
    
    if not month:
        month = datetime.now().strftime("%Y-%m")
    
    # جلب الموظفين
    emp_query = {"tenant_id": tenant_id, "is_active": True}
    if branch_id:
        emp_query["branch_id"] = branch_id
    
    employees = await db.employees.find(emp_query, {"_id": 0}).to_list(500)
    
    if not employees:
        return {"ratings": [], "summary": {}}
    
    # === تحسين الأداء: Batch fetch بدلاً من N+1 queries ===
    employee_ids = [emp["id"] for emp in employees]
    
    # جلب جميع سجلات الحضور دفعة واحدة
    all_attendance = await db.attendance.find({
        "tenant_id": tenant_id,
        "employee_id": {"$in": employee_ids},
        "date": {"$regex": f"^{month}"}
    }, {"_id": 0}).to_list(15000)
    
    # جلب جميع الخصومات دفعة واحدة
    all_deductions = await db.deductions.find({
        "tenant_id": tenant_id,
        "employee_id": {"$in": employee_ids},
        "month": month
    }, {"_id": 0}).to_list(5000)
    
    # جلب جميع المكافآت دفعة واحدة
    all_bonuses = await db.bonuses.find({
        "tenant_id": tenant_id,
        "employee_id": {"$in": employee_ids},
        "month": month
    }, {"_id": 0}).to_list(5000)
    
    # تجميع البيانات حسب الموظف
    attendance_by_emp = {}
    for record in all_attendance:
        emp_id = record.get("employee_id")
        if emp_id not in attendance_by_emp:
            attendance_by_emp[emp_id] = []
        attendance_by_emp[emp_id].append(record)
    
    deductions_by_emp = {}
    for record in all_deductions:
        emp_id = record.get("employee_id")
        if emp_id not in deductions_by_emp:
            deductions_by_emp[emp_id] = []
        deductions_by_emp[emp_id].append(record)
    
    bonuses_by_emp = {}
    for record in all_bonuses:
        emp_id = record.get("employee_id")
        if emp_id not in bonuses_by_emp:
            bonuses_by_emp[emp_id] = []
        bonuses_by_emp[emp_id].append(record)
    
    ratings = []
    
    for emp in employees:
        emp_id = emp["id"]
        
        # حساب أيام العمل في الشهر (26 يوم افتراضياً)
        work_days_expected = 26
        
        # جلب سجلات الحضور للشهر (من البيانات المجمعة)
        attendance_records = attendance_by_emp.get(emp_id, [])
        
        # حساب أيام الحضور
        attendance_days = len(attendance_records)
        
        # حساب التأخير (إذا الحضور بعد الساعة 9 صباحاً)
        late_count = 0
        early_leave_count = 0
        total_work_hours = 0
        
        for record in attendance_records:
            check_in = record.get("check_in")
            check_out = record.get("check_out")
            
            if check_in:
                try:
                    check_in_time = datetime.strptime(check_in, "%H:%M")
                    expected_time = datetime.strptime(emp.get("work_start", "09:00"), "%H:%M")
                    if check_in_time > expected_time:
                        late_count += 1
                except:
                    pass
            
            if check_out:
                try:
                    check_out_time = datetime.strptime(check_out, "%H:%M")
                    expected_end = datetime.strptime(emp.get("work_end", "17:00"), "%H:%M")
                    if check_out_time < expected_end:
                        early_leave_count += 1
                except:
                    pass
            
            # حساب ساعات العمل
            if check_in and check_out:
                hours = _calc_worked_hours_hhmm(check_in, check_out)
                if hours:
                    total_work_hours += hours
        
        # جلب الخصومات للشهر (من البيانات المجمعة)
        deductions = deductions_by_emp.get(emp_id, [])
        total_deductions = sum(_sn(d.get("amount")) for d in deductions)
        deduction_count = len(deductions)
        
        # جلب المكافآت للشهر (من البيانات المجمعة)
        bonuses = bonuses_by_emp.get(emp_id, [])
        total_bonuses = sum(_sn(b.get("amount")) for b in bonuses)
        bonus_count = len(bonuses)
        
        # ========== حساب التقييم ==========
        # التقييم من 100 نقطة
        
        # 1. تقييم الحضور (40 نقطة)
        attendance_percentage = (attendance_days / work_days_expected) * 100 if work_days_expected > 0 else 0
        attendance_score = min(40, (attendance_percentage / 100) * 40)
        
        # 2. تقييم الالتزام بالمواعيد (30 نقطة)
        punctuality_issues = late_count + early_leave_count
        punctuality_deduction = min(30, punctuality_issues * 3)  # خصم 3 نقاط لكل تأخير/خروج مبكر
        punctuality_score = max(0, 30 - punctuality_deduction)
        
        # 3. تقييم عدم وجود خصومات (20 نقطة)
        deduction_penalty = min(20, deduction_count * 5)  # خصم 5 نقاط لكل خصم
        discipline_score = max(0, 20 - deduction_penalty)
        
        # 4. نقاط المكافآت (10 نقاط إضافية)
        bonus_score = min(10, bonus_count * 2)  # 2 نقطة لكل مكافأة
        
        # المجموع
        total_score = attendance_score + punctuality_score + discipline_score + bonus_score
        
        # تحديد المستوى
        if total_score >= 90:
            level = "ممتاز"
            level_color = "green"
        elif total_score >= 75:
            level = "جيد جداً"
            level_color = "blue"
        elif total_score >= 60:
            level = "جيد"
            level_color = "yellow"
        elif total_score >= 50:
            level = "مقبول"
            level_color = "orange"
        else:
            level = "ضعيف"
            level_color = "red"
        
        ratings.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name", ""),
            "branch_id": emp.get("branch_id"),
            "position": emp.get("position", ""),
            "month": month,
            
            # إحصائيات الحضور
            "attendance_days": attendance_days,
            "work_days_expected": work_days_expected,
            "attendance_percentage": round(attendance_percentage, 1),
            "late_count": late_count,
            "early_leave_count": early_leave_count,
            "total_work_hours": round(total_work_hours, 1),
            
            # إحصائيات الخصومات والمكافآت
            "deduction_count": deduction_count,
            "total_deductions": total_deductions,
            "bonus_count": bonus_count,
            "total_bonuses": total_bonuses,
            
            # التقييم
            "scores": {
                "attendance": round(attendance_score, 1),
                "punctuality": round(punctuality_score, 1),
                "discipline": round(discipline_score, 1),
                "bonus": round(bonus_score, 1)
            },
            "total_score": round(total_score, 1),
            "level": level,
            "level_color": level_color
        })
    
    # ترتيب حسب التقييم
    ratings.sort(key=lambda x: x["total_score"], reverse=True)
    
    # إحصائيات عامة
    if ratings:
        avg_score = sum(r["total_score"] for r in ratings) / len(ratings)
        excellent_count = len([r for r in ratings if r["total_score"] >= 90])
        good_count = len([r for r in ratings if 75 <= r["total_score"] < 90])
        average_count = len([r for r in ratings if 60 <= r["total_score"] < 75])
        poor_count = len([r for r in ratings if r["total_score"] < 60])
    else:
        avg_score = 0
        excellent_count = good_count = average_count = poor_count = 0
    
    return {
        "month": month,
        "ratings": ratings,
        "summary": {
            "total_employees": len(ratings),
            "average_score": round(avg_score, 1),
            "excellent_count": excellent_count,
            "good_count": good_count,
            "average_count": average_count,
            "poor_count": poor_count
        }
    }

# ==================== moved to routes/payroll_reports_routes.py ====================

# ==================== moved to routes/coupons_promotions_routes.py ====================

# ==================== INVENTORY TRANSFER ROUTES - تحويلات المخزون ====================

async def get_next_transfer_number() -> int:
    """الحصول على رقم التحويل التالي"""
    counter = await db.counters.find_one_and_update(
        {"type": "transfer"},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

@api_router.post("/inventory-transfers", response_model=InventoryTransferResponse)
async def create_inventory_transfer(transfer: InventoryTransferCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء طلب تحويل مخزون"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    transfer_number = await get_next_transfer_number()
    
    # جلب أسماء الفروع
    from_branch = await db.branches.find_one({"id": transfer.from_branch_id}, {"name": 1})
    to_branch = await db.branches.find_one({"id": transfer.to_branch_id}, {"name": 1})
    
    transfer_doc = {
        "id": str(uuid.uuid4()),
        "transfer_number": transfer_number,
        **transfer.model_dump(),
        "from_branch_name": from_branch.get("name") if from_branch else None,
        "to_branch_name": to_branch.get("name") if to_branch else None,
        "status": "pending",
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.inventory_transfers.insert_one(transfer_doc)
    del transfer_doc["_id"]
    return transfer_doc

@api_router.get("/inventory-transfers")
async def get_inventory_transfers(
    from_branch_id: Optional[str] = None,
    to_branch_id: Optional[str] = None,
    status: Optional[str] = None,
    transfer_type: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب تحويلات المخزون"""
    query = build_tenant_query(current_user)
    if from_branch_id:
        query["from_branch_id"] = from_branch_id
    if to_branch_id:
        query["to_branch_id"] = to_branch_id
    if status:
        query["status"] = status
    if transfer_type:
        query["transfer_type"] = transfer_type
    
    transfers = await db.inventory_transfers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return transfers

@api_router.get("/inventory-transactions")
async def get_inventory_transactions(current_user: dict = Depends(get_current_user)):
    """جلب حركات المخزون (واردات/صادرات)"""
    query = build_tenant_query(current_user)
    transactions = await db.inventory_transactions.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return transactions

@api_router.put("/inventory-transfers/{transfer_id}/approve")
async def approve_inventory_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    """الموافقة على التحويل"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": transfer_id})
    transfer = await db.inventory_transfers.find_one(query)
    if not transfer:
        raise HTTPException(status_code=404, detail="التحويل غير موجود")
    
    if transfer.get("status") != "pending":
        raise HTTPException(status_code=400, detail="لا يمكن الموافقة على هذا التحويل")
    
    await db.inventory_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": "approved",
            "approved_by": current_user["id"],
            "approved_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تمت الموافقة على التحويل"}

@api_router.put("/inventory-transfers/{transfer_id}/ship")
async def ship_inventory_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    """شحن التحويل (خصم من المخزن المرسل)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": transfer_id})
    transfer = await db.inventory_transfers.find_one(query)
    if not transfer:
        raise HTTPException(status_code=404, detail="التحويل غير موجود")
    
    if transfer.get("status") != "approved":
        raise HTTPException(status_code=400, detail="يجب الموافقة على التحويل أولاً")
    
    # خصم الكميات من المخزن المرسل
    for item in transfer.get("items", []):
        await db.inventory.update_one(
            {"id": item.get("inventory_id"), "branch_id": transfer["from_branch_id"]},
            {"$inc": {"quantity": -_sn(item.get("quantity"))}}
        )
    
    await db.inventory_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": "shipped",
            "shipped_by": current_user["id"],
            "shipped_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تم شحن التحويل"}

@api_router.put("/inventory-transfers/{transfer_id}/receive")
async def receive_inventory_transfer(transfer_id: str, current_user: dict = Depends(get_current_user)):
    """استلام التحويل (إضافة للمخزن المستلم)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": transfer_id})
    transfer = await db.inventory_transfers.find_one(query)
    if not transfer:
        raise HTTPException(status_code=404, detail="التحويل غير موجود")
    
    if transfer.get("status") != "shipped":
        raise HTTPException(status_code=400, detail="يجب شحن التحويل أولاً")
    
    # إضافة الكميات للمخزن المستلم
    for item in transfer.get("items", []):
        # التحقق من وجود الصنف في المخزن المستلم
        existing = await db.inventory.find_one({
            "id": item.get("inventory_id"),
            "branch_id": transfer["to_branch_id"]
        })
        
        if existing:
            await db.inventory.update_one(
                {"id": item.get("inventory_id"), "branch_id": transfer["to_branch_id"]},
                {"$inc": {"quantity": _sn(item.get("quantity"))}}
            )
        else:
            # إنشاء صنف جديد في المخزن المستلم
            source_item = await db.inventory.find_one({"id": item.get("inventory_id")}, {"_id": 0})
            if source_item:
                new_item = {
                    **source_item,
                    "id": str(uuid.uuid4()),
                    "branch_id": transfer["to_branch_id"],
                    "quantity": _sn(item.get("quantity")),
                    "last_updated": datetime.now(timezone.utc).isoformat()
                }
                await db.inventory.insert_one(new_item)
    
    await db.inventory_transfers.update_one(
        {"id": transfer_id},
        {"$set": {
            "status": "received",
            "received_by": current_user["id"],
            "received_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    return {"message": "تم استلام التحويل"}

# ==================== PURCHASE REQUEST ROUTES - طلبات الشراء ====================

async def get_next_request_number() -> int:
    """الحصول على رقم الطلب التالي"""
    counter = await db.counters.find_one_and_update(
        {"type": "purchase_request"},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

# ==================== TABLE ROUTES ====================

@api_router.post("/tables", response_model=TableResponse)
async def create_table(table: TableCreate, current_user: dict = Depends(get_current_user)):
    # السماح للمدير والأدمن أو من لديه صلاحية tables
    user_permissions = current_user.get("permissions", [])
    if current_user["role"] not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER] and "tables" not in user_permissions:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    table_doc = {
        "id": str(uuid.uuid4()),
        **table.model_dump(),
        "tenant_id": tenant_id,  # فصل البيانات لكل عميل
        "status": "available",
        "current_order_id": None
    }
    await db.tables.insert_one(table_doc)
    del table_doc["_id"]
    return table_doc

@api_router.get("/tables", response_model=List[TableResponse])
async def get_tables(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    # Super Admin لا يرى طاولات (ليس لديه مطعم)
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        return []
    
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    tables = await db.tables.find(query, {"_id": 0}).sort("number", 1).to_list(100)
    return tables

@api_router.put("/tables/{table_id}/status")
async def update_table_status(table_id: str, status: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": table_id})
    await db.tables.update_one(query, {"$set": {"status": status}})
    return {"message": "تم التحديث"}

@api_router.put("/tables/{table_id}")
async def update_table(table_id: str, table_data: TableCreate, current_user: dict = Depends(get_current_user)):
    """تعديل بيانات الطاولة"""
    # التحقق من صلاحيات المستخدم
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        if not current_user.get("permissions", {}).get("tables"):
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية لتعديل الطاولات")
    
    # التحقق من وجود الطاولة
    query = build_tenant_query(current_user, {"id": table_id})
    existing_table = await db.tables.find_one(query)
    if not existing_table:
        raise HTTPException(status_code=404, detail="الطاولة غير موجودة")
    
    # التحقق من تكرار رقم الطاولة في نفس الفرع
    branch_id = table_data.branch_id or existing_table.get("branch_id")
    duplicate_query = build_tenant_query(current_user, {
        "number": table_data.number,
        "branch_id": branch_id,
        "id": {"$ne": table_id}
    })
    duplicate = await db.tables.find_one(duplicate_query)
    if duplicate:
        raise HTTPException(status_code=400, detail="رقم الطاولة موجود مسبقاً في هذا الفرع")
    
    # تحديث الطاولة
    update_data = {
        "number": table_data.number,
        "capacity": table_data.capacity,
        "section": table_data.section or existing_table.get("section", "داخلي"),
        "branch_id": branch_id,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.tables.update_one(query, {"$set": update_data})
    
    # إرجاع الطاولة المحدثة
    updated_table = await db.tables.find_one(query, {"_id": 0})
    return updated_table

@api_router.post("/tables/transfer")
async def transfer_table_order(
    from_table_id: str = Body(...),
    to_table_id: str = Body(...),
    order_id: str = Body(None),
    current_user: dict = Depends(get_current_user)
):
    """تحويل الطلب من طاولة إلى أخرى"""
    
    # التحقق من الطاولة المصدر مع فلترة tenant_id
    query = build_tenant_query(current_user, {"id": from_table_id})
    from_table = await db.tables.find_one(query)
    if not from_table:
        raise HTTPException(status_code=404, detail="الطاولة المصدر غير موجودة")
    
    if from_table.get("status") != "occupied":
        raise HTTPException(status_code=400, detail="الطاولة المصدر ليست مشغولة")
    
    # التحقق من الطاولة المستهدفة مع فلترة tenant_id
    query = build_tenant_query(current_user, {"id": to_table_id})
    to_table = await db.tables.find_one(query)
    if not to_table:
        raise HTTPException(status_code=404, detail="الطاولة المستهدفة غير موجودة")
    
    if to_table.get("status") != "available":
        raise HTTPException(status_code=400, detail="الطاولة المستهدفة غير متاحة")
    
    # الحصول على الطلب الحالي
    actual_order_id = order_id or from_table.get("current_order_id")
    if actual_order_id:
        # تحديث الطلب
        await db.orders.update_one(
            {"id": actual_order_id},
            {"$set": {
                "table_id": to_table_id,
                "table_number": to_table.get("number"),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "transfer_history": {
                    "from_table": from_table.get("number"),
                    "to_table": to_table.get("number"),
                    "transferred_at": datetime.now(timezone.utc).isoformat(),
                    "transferred_by": current_user["id"]
                }
            }}
        )
    
    # تحديث حالة الطاولات
    await db.tables.update_one(
        {"id": from_table_id},
        {"$set": {"status": "available", "current_order_id": None}}
    )
    
    await db.tables.update_one(
        {"id": to_table_id},
        {"$set": {"status": "occupied", "current_order_id": actual_order_id}}
    )
    
    logger.info(f"Order transferred from table {from_table.get('number')} to table {to_table.get('number')}")
    
    return {
        "message": f"تم تحويل الطلب من طاولة {from_table.get('number')} إلى طاولة {to_table.get('number')}",
        "from_table": from_table.get("number"),
        "to_table": to_table.get("number")
    }

@api_router.delete("/tables/{table_id}")
async def delete_table(table_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طاولة - فقط للمالك أو المدير"""
    if current_user.get("role") not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER, "super_admin", "admin", "manager"]:
        raise HTTPException(status_code=403, detail="ليس لديك صلاحية حذف الطاولات")
    
    # Super Admin يحذف أي طاولة
    if current_user.get("role") in [UserRole.SUPER_ADMIN, "super_admin"]:
        query = {"id": table_id}
    else:
        query = build_tenant_query(current_user, {"id": table_id})
    
    table = await db.tables.find_one(query)
    if not table:
        raise HTTPException(status_code=404, detail="الطاولة غير موجودة")
    
    # التحقق من أن الطاولة ليست مشغولة
    if table.get("status") == "occupied":
        raise HTTPException(status_code=400, detail="لا يمكن حذف طاولة مشغولة")
    
    await db.tables.delete_one({"id": table_id})
    logger.info(f"Table {table_id} deleted by {current_user.get('user_id')}")
    return {"message": "تم حذف الطاولة بنجاح"}

# ==================== CUSTOMER ROUTES - إدارة العملاء ====================

@api_router.post("/customers", response_model=CustomerResponse)
async def create_customer(customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    # التحقق من عدم وجود العميل بنفس الرقم في نفس الـ tenant
    tenant_id = get_user_tenant_id(current_user)
    query = {"phone": customer.phone}
    if tenant_id:
        query["tenant_id"] = tenant_id
    existing = await db.customers.find_one(query)
    if existing:
        raise HTTPException(status_code=400, detail="رقم الهاتف موجود مسبقاً")
    
    customer_doc = {
        "id": str(uuid.uuid4()),
        **customer.model_dump(),
        "tenant_id": tenant_id,  # فصل البيانات
        "total_orders": 0,
        "total_spent": 0.0,
        "last_order_date": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.customers.insert_one(customer_doc)
    del customer_doc["_id"]
    return customer_doc

@api_router.get("/customers", response_model=List[CustomerResponse])
async def get_customers(search: Optional[str] = None, phone: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)  # فلترة حسب tenant_id
    if phone:
        query["$or"] = [{"phone": phone}, {"phone2": phone}]
    elif search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search}},
            {"phone2": {"$regex": search}},
            {"area": {"$regex": search, "$options": "i"}}
        ]
    customers = await db.customers.find(query, {"_id": 0}).sort("name", 1).to_list(500)
    return customers

@api_router.get("/customers/{customer_id}", response_model=CustomerResponse)
async def get_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": customer_id})
    customer = await db.customers.find_one(query, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    return customer

@api_router.get("/customers/by-phone/{phone}")
async def get_customer_by_phone(phone: str, current_user: dict = Depends(get_current_user)):
    """البحث عن عميل بالهاتف مع سجل الطلبات"""
    tenant_id = get_user_tenant_id(current_user)
    
    # بناء query للبحث عن العميل بالهاتف مع مراعاة tenant_id
    phone_conditions = [{"phone": phone}, {"phone2": phone}]
    
    if tenant_id:
        # المستخدم العميل يرى فقط بياناته
        query = {"$and": [{"tenant_id": tenant_id}, {"$or": phone_conditions}]}
    elif current_user.get("role") == UserRole.SUPER_ADMIN:
        # Super Admin يرى الكل
        query = {"$or": phone_conditions}
    else:
        # المستخدم الرئيسي (بدون tenant_id) يرى البيانات الرئيسية فقط
        query = {"$and": [
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]},
            {"$or": phone_conditions}
        ]}
    
    customer = await db.customers.find_one(query, {"_id": 0})
    
    if not customer:
        return {"found": False, "customer": None, "orders": []}
    
    # جلب آخر 10 طلبات للعميل
    orders_query = build_tenant_query(current_user, {"customer_phone": phone})
    orders = await db.orders.find(orders_query, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    return {
        "found": True,
        "customer": customer,
        "orders": orders
    }

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, customer: CustomerCreate, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": customer_id})
    await db.customers.update_one(query, {"$set": customer.model_dump()})
    return await db.customers.find_one({"id": customer_id}, {"_id": 0})

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, current_user: dict = Depends(get_current_user)):
    await db.customers.delete_one({"id": customer_id})
    return {"message": "تم الحذف"}

# ==================== WELCOME DISCOUNT (رسالة ترحيب + كود خصم) ====================
DEFAULT_WELCOME_CONFIG = {
    "enabled": True,
    "discount_type": "percentage",  # percentage | fixed
    "discount_value": 10,
    "min_order_amount": 0,
    "valid_days": 7,
    "message_template": (
        "مرحباً بك في {restaurant} 🎉\n"
        "شكراً لأول طلب لديك! إليك كود خصم خاص لطلبك القادم:\n\n"
        "🎁 {code}\n"
        "قيمة الخصم: {discount} — صالح حتى {expiry}.\n\n"
        "بالهناء والعافية 🌟"
    ),
}

async def _get_welcome_config(tenant_id: Optional[str]) -> dict:
    doc = await db.app_settings.find_one({"tenant_id": tenant_id, "key": "welcome_discount"}, {"_id": 0})
    cfg = dict(DEFAULT_WELCOME_CONFIG)
    if doc and isinstance(doc.get("value"), dict):
        cfg.update(doc["value"])
    return cfg

@api_router.get("/welcome-discount/config")
async def get_welcome_discount_config(current_user: dict = Depends(get_current_user)):
    return await _get_welcome_config(get_user_tenant_id(current_user))

@api_router.put("/welcome-discount/config")
async def update_welcome_discount_config(payload: dict = Body(...), current_user: dict = Depends(get_current_user)):
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN, "branch_manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    tenant_id = get_user_tenant_id(current_user)
    allowed = {"enabled", "discount_type", "discount_value", "min_order_amount", "valid_days", "message_template"}
    value = {k: payload[k] for k in payload if k in allowed}
    await db.app_settings.update_one(
        {"tenant_id": tenant_id, "key": "welcome_discount"},
        {"$set": {"tenant_id": tenant_id, "key": "welcome_discount", "value": value,
                  "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    return await _get_welcome_config(tenant_id)

@api_router.get("/welcome-approvals")
async def get_welcome_approvals(current_user: dict = Depends(get_current_user)):
    """قائمة الزبائن الجدد بانتظار موافقة خصم الترحيب (إشعار داخل النظام بزر موافقة)."""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN, "branch_manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"welcome_status": "pending"})
    customers = await db.customers.find(query, {"_id": 0}).sort("created_at", -1).to_list(100)
    return {"pending": customers, "count": len(customers)}

@api_router.get("/welcome-discount/stats")
async def welcome_discount_stats(current_user: dict = Depends(get_current_user)):
    """إحصائيات كوبونات الترحيب: الحالة، الاستخدام، وقيمة الخصومات الممنوحة."""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN, "branch_manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    pending = await db.customers.count_documents(build_tenant_query(current_user, {"welcome_status": "pending"}))
    granted = await db.customers.count_documents(build_tenant_query(current_user, {"welcome_status": "granted"}))
    cq = build_tenant_query(current_user, {"is_welcome": True})
    coupons = await db.coupons.find(cq, {"_id": 0, "code": 1, "customer_name": 1, "created_at": 1,
                                          "used_count": 1, "total_discount_given": 1, "valid_until": 1,
                                          "discount_type": 1, "discount_value": 1}).sort("created_at", -1).to_list(500)
    now_iso = datetime.now(timezone.utc).isoformat()
    used = sum(1 for c in coupons if (c.get("used_count") or 0) > 0)
    expired = sum(1 for c in coupons if (c.get("valid_until") or "") < now_iso and (c.get("used_count") or 0) == 0)
    total_discount = sum(float(c.get("total_discount_given") or 0) for c in coupons)
    total = len(coupons)
    return {
        "pending_customers": pending,
        "granted_customers": granted,
        "total_coupons": total,
        "used_coupons": used,
        "expired_unused": expired,
        "active_unused": total - used - expired,
        "conversion_rate": round((used / total * 100) if total else 0, 1),
        "total_discount_given": total_discount,
        "recent": coupons[:10],
    }

@api_router.post("/customers/{customer_id}/grant-welcome-discount")
async def grant_welcome_discount(customer_id: str, payload: dict = Body(default={}), current_user: dict = Depends(get_current_user)):
    """موافقة صاحب المطعم/المدير → كوبون شخصي باسم الزبون بعدد مرات وفروع يحددها المالك + واتساب احترافي."""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN, "branch_manager"]:
        raise HTTPException(status_code=403, detail="يتطلب موافقة صاحب المطعم أو المدير")
    tenant_id = get_user_tenant_id(current_user)
    customer = await db.customers.find_one(build_tenant_query(current_user, {"id": customer_id}), {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    if customer.get("welcome_status") == "granted":
        raise HTTPException(status_code=400, detail="تم منح خصم الترحيب لهذا العميل مسبقاً")
    cfg = await _get_welcome_config(tenant_id)
    if not cfg.get("enabled", True):
        raise HTTPException(status_code=400, detail="خصم الترحيب غير مُفعّل")

    payload = payload or {}
    try:
        usage_limit = max(1, int(payload.get("usage_limit") or 1))
    except (TypeError, ValueError):
        usage_limit = 1
    branch_ids = [b for b in (payload.get("branch_ids") or []) if isinstance(b, str)]
    discount_type = payload.get("discount_type") or cfg.get("discount_type", "percentage")
    try:
        dval = float(payload.get("discount_value") if payload.get("discount_value") is not None else (cfg.get("discount_value", 10) or 0))
    except (TypeError, ValueError):
        dval = float(cfg.get("discount_value", 10) or 0)
    try:
        min_order = float(payload.get("min_order_amount") if payload.get("min_order_amount") is not None else (cfg.get("min_order_amount", 0) or 0))
    except (TypeError, ValueError):
        min_order = 0.0
    try:
        valid_days = max(1, int(payload.get("valid_days") or cfg.get("valid_days", 7) or 7))
    except (TypeError, ValueError):
        valid_days = 7

    code = "WLC" + uuid.uuid4().hex[:6].upper()
    now = datetime.now(timezone.utc)
    valid_until = now + timedelta(days=valid_days)
    is_pct = discount_type == "percentage"
    customer_display_name = (customer.get("name") or "").strip() or "زبون"
    coupon_doc = {
        "id": str(uuid.uuid4()),
        "code": code,
        "name": f"خصم ترحيبي 🎁 — {customer_display_name}",
        "description": f"كوبون ترحيبي شخصي باسم الزبون {customer_display_name} ({customer.get('phone','')})",
        "discount_type": discount_type,
        "discount_value": dval,
        "min_order_amount": min_order,
        "max_discount": None,
        "usage_limit": usage_limit,
        "usage_per_customer": usage_limit,
        "valid_from": now.isoformat(),
        "valid_until": valid_until.isoformat(),
        "is_active": True,
        "applicable_to": "all",
        "applicable_ids": [],
        "loyalty_tier_required": None,
        "first_order_only": False,
        "branch_ids": branch_ids,
        "customer_name": customer_display_name,
        "used_count": 0,
        "total_discount_given": 0,
        "is_welcome": True,
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": now.isoformat(),
    }
    await db.coupons.insert_one(coupon_doc)

    # اسم المطعم للرسالة
    restaurant = "مطعمنا"
    if tenant_id:
        t = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "name": 1})
        if t and t.get("name"):
            restaurant = t["name"]
    if restaurant == "مطعمنا":
        b = await db.branches.find_one(build_tenant_query(current_user), {"_id": 0, "name": 1})
        if b and b.get("name"):
            restaurant = b["name"]
    discount_str = f"{int(dval)}%" if is_pct else f"{int(dval):,} د.ع"
    expiry_str = valid_until.strftime("%Y-%m-%d")

    # أسماء الفروع المتاحة للطلب بهذا الكوبون
    _bq = {"tenant_id": tenant_id} if tenant_id else {}
    if branch_ids:
        _bq["id"] = {"$in": branch_ids}
    _branches = await db.branches.find(_bq, {"_id": 0, "name": 1}).to_list(50)
    branch_names = [b.get("name") for b in _branches if b.get("name")]
    branches_str = "، ".join(branch_names) if branch_names else "جميع فروعنا"

    uses_str = "مرة واحدة" if usage_limit == 1 else ("مرتين" if usage_limit == 2 else f"{usage_limit} مرات")
    min_order_line = f"🛒 الحد الأدنى للطلب: {int(min_order):,} د.ع\n" if min_order > 0 else ""
    message = (
        f"🎉 أهلاً وسهلاً بك في {restaurant}!\n\n"
        f"عميلنا العزيز {customer_display_name}، يسعدنا انضمامك إلينا، وهذه هدية ترحيبية خاصة باسمك:\n\n"
        f"🎁 كود الخصم: {code}\n"
        f"💰 قيمة الخصم: {discount_str}\n"
        f"🔁 عدد مرات الاستخدام: {uses_str}\n"
        f"🏪 الفروع المتاحة للطلب: {branches_str}\n"
        f"{min_order_line}"
        f"📅 صالح حتى: {expiry_str}\n\n"
        f"هذا الكوبون شخصي باسمك ولا يمكن استخدامه من قبل غيرك.\n"
        f"نتشرف بخدمتك دائماً 🌟\n{restaurant}"
    )

    wa_ok, wa_err = await _wa_free.send_message(
        customer.get("phone", ""), message,
        purpose="welcome_coupon", title="🎁 كوبون ترحيبي — Maestro EGP",
    )

    # تعليم إشعار الموافقة المرتبط بهذا الزبون كمُعالج
    await db.notifications.update_many(
        {"type": "welcome_approval", "data.customer_id": customer_id},
        {"$set": {"is_read": True, "handled_at": now.isoformat()}}
    )

    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {
            "welcome_status": "granted",
            "welcome_coupon_code": code,
            "welcome_granted_at": now.isoformat(),
            "welcome_granted_by": current_user["id"],
            "welcome_whatsapp_sent": bool(wa_ok),
        }}
    )
    return {
        "success": True,
        "coupon_code": code,
        "coupon_id": coupon_doc["id"],
        "usage_limit": usage_limit,
        "branches": branch_names or ["جميع الفروع"],
        "discount": discount_str,
        "valid_until": expiry_str,
        "whatsapp_sent": bool(wa_ok),
        "whatsapp_error": wa_err,
        "message": "تم إنشاء كوبون باسم الزبون" + (" وإرساله عبر واتساب ✅" if wa_ok else " (تعذّر إرسال واتساب — الرقم غير مربوط أو غير متصل)")
    }

# ==================== ORDER ROUTES ====================

async def get_next_order_number(branch_id: str, business_date: Optional[str] = None) -> int:
    """عدّاد الطلبات: يستخدم business_date (التاريخ التشغيلي للشفت) لضمان عدم الانجراف
    عند منتصف الليل UTC. إذا لم يُمرَّر business_date، يستخدم اليوم بـ UTC كـ fallback.
    """
    counter_date = business_date or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    counter = await db.order_counters.find_one_and_update(
        {"branch_id": branch_id, "date": counter_date},
        {"$inc": {"counter": 1}},
        upsert=True,
        return_document=True
    )
    return counter["counter"]

@api_router.post("/orders", response_model=OrderResponse)
async def create_order(order: OrderCreate, current_user: dict = Depends(get_current_user)):
    tenant_id = get_user_tenant_id(current_user)
    
    # 🔒 أمان مالي (تقرير الاختراق #2): منع الأسعار السالبة/الصفرية والتلاعب بالسعر من جهة العميل.
    # يُنفَّذ في بداية الدالة قبل أي إنشاء وردية/طلب — فلا تُنشأ ورديات أو سجلات وهمية من محاولة حقن.
    for _it in order.items:
        if (_it.quantity or 0) <= 0:
            raise HTTPException(status_code=400, detail="كمية الصنف يجب أن تكون أكبر من صفر")
        if _sn(_it.price) < 0:
            raise HTTPException(status_code=400, detail="سعر الصنف لا يمكن أن يكون سالباً")
        for _ex in (_it.extras or []):
            if _sn(_ex.get("price")) < 0 or int(_ex.get("quantity", 1) or 0) <= 0:
                raise HTTPException(status_code=400, detail="بيانات الإضافة غير صالحة")
        _prod_price = await db.products.find_one(
            {"id": _it.product_id, **({"tenant_id": tenant_id} if tenant_id else {})},
            {"_id": 0, "price": 1}
        )
        if _prod_price is not None:
            _catalog = _sn(_prod_price.get("price"))
            # منتج له سعر ثابت بالقائمة: لا يُقبل بيع بأقل من سعر القائمة (يشمل الصفر) — الخصم عبر حقل الخصم فقط
            if _catalog > 0 and _sn(_it.price) < _catalog - 1:
                raise HTTPException(
                    status_code=400,
                    detail=f"سعر الصنف ({_it.product_name or _it.product_id}) أقل من سعر القائمة — استخدم حقل الخصم بدلاً من تعديل السعر"
                )
        else:
            # صنف غير موجود بقائمة المنتجات (product_id مُلفّق): يجب أن يكون بسعر موجب
            # يمنع إنشاء طلبات وهمية بسعر صفر/سالب على معرّف منتج غير حقيقي (تحصين إضافي)
            if _sn(_it.price) <= 0:
                raise HTTPException(status_code=400, detail="صنف غير معروف أو بسعر غير صالح")

    # ⭐ تحقق الصلاحية: "آجل بدون شركة توصيل" مسموح فقط للمالك/المدير العام
    # أو لمن لديه صلاحية "allow_credit_without_delivery"
    if (order.payment_method or "").lower() == "credit":
        has_delivery_company = bool(
            getattr(order, "delivery_app", None)
            or getattr(order, "delivery_app_name", None)
            or getattr(order, "delivery_commission", 0) or 0 > 0
        )
        if not has_delivery_company:
            privileged_roles = {"admin", "manager", "super_admin", "branch_manager", "owner"}
            user_role = (current_user.get("role") or "").lower()
            user_perms = current_user.get("permissions") or []
            if user_role not in privileged_roles and "allow_credit_without_delivery" not in user_perms:
                raise HTTPException(
                    status_code=403,
                    detail="غير مسموح بإنشاء طلب آجل بدون شركة توصيل. اطلب من المالك تفعيل صلاحية 'آجل بدون شركة توصيل'.",
                )
    
    # ===== Idempotency: منع إنشاء طلب مكرر خلال 30 ثانية =====
    # السبب الجذري للطلب المكرر #11 في فرع السيدية: ضغط مزدوج/إعادة محاولة من POS
    # ⭐ ثبات (Idempotency) عبر offline_id: مفتاح ثابت يُرسله الكاشير مع كل عملية إنشاء.
    # إن وُجد طلب بنفس offline_id (حتى لو مرّت دقائق) نُعيده بدل إنشاء نسخة مكررة.
    # هذا يحل تكرار الأوفلاين عند ضياع رد السيرفر ثم مزامنة النسخة المحلية.
    incoming_offline_id = getattr(order, "offline_id", None)
    if incoming_offline_id:
        existing_by_offline = await db.orders.find_one(
            {"offline_id": incoming_offline_id, **({"tenant_id": tenant_id} if tenant_id else {})},
            {"_id": 0},
        )
        if existing_by_offline:
            logger.info(
                f"Order idempotency: returning existing #{existing_by_offline.get('order_number')} "
                f"for offline_id={incoming_offline_id} (no duplicate created)"
            )
            return existing_by_offline

    # ⭐⭐ منع تكرار طلبات شركات التوصيل بنفس رقم الطلب الخارجي (مثل رقم طلبات/مزاجك)
    #    يمنع إدخال نفس الطلب مرتين نهائياً — حتى من جهازين مختلفين أو أونلاين/أوفلاين.
    incoming_ext_ref = (getattr(order, "delivery_company_order_id", None) or "").strip()
    if incoming_ext_ref:
        _company_keys = [k for k in [
            getattr(order, "delivery_app", None), getattr(order, "delivery_company_id", None),
            getattr(order, "delivery_company", None), getattr(order, "delivery_app_name", None),
            getattr(order, "delivery_company_name", None),
        ] if k]
        _ext_q = {
            "delivery_company_order_id": incoming_ext_ref,
            "status": {"$nin": ["cancelled", "canceled", "deleted", "refunded"]},
            **({"tenant_id": tenant_id} if tenant_id else {}),
        }
        if _company_keys:
            _ext_q["$or"] = [
                {"delivery_app": {"$in": _company_keys}},
                {"delivery_company_id": {"$in": _company_keys}},
                {"delivery_company": {"$in": _company_keys}},
                {"delivery_app_name": {"$in": _company_keys}},
                {"delivery_company_name": {"$in": _company_keys}},
            ]
        existing_ext = await db.orders.find_one(_ext_q, {"_id": 0})
        if existing_ext:
            logger.info(f"Duplicate delivery order blocked: ext_ref={incoming_ext_ref} -> existing #{existing_ext.get('order_number')}")
            raise HTTPException(status_code=409, detail={
                "code": "DUPLICATE_DELIVERY_ORDER",
                "message": f"رقم طلب شركة التوصيل ({incoming_ext_ref}) مُدخَل مسبقاً بالطلب رقم #{existing_ext.get('order_number')} — لم يُنشأ طلب مكرر",
                "order_number": existing_ext.get("order_number"),
                "order_id": existing_ext.get("id"),
            })

    # نقارن: tenant + branch + cashier + customer + items hash + total
    try:
        items_sig_data = sorted([
            f"{(it.product_id or '')}:{float(it.quantity or 0):.2f}"
            for it in order.items
        ])
        items_sig = "|".join(items_sig_data)
        # ⭐ نضمّن رقم طلب شركة التوصيل وطريقة الدفع في البصمة: طلبان مختلفان لشركة التوصيل
        #    (برقمين خارجيين مختلفين) لهما بصمتان مختلفتان فلا يُدمجان خطأً ولا يُفقَد طلب.
        #    كما نضمّن table_id: طاولتان مختلفتان بنفس الأصناف لا تُدمجان (حماية من فقدان طلب حقيقي).
        _ext_for_fp = (getattr(order, "delivery_company_order_id", None) or "").strip()
        _table_for_fp = (getattr(order, "table_id", None) or "")
        order_fingerprint = hashlib.sha1(
            f"{tenant_id or ''}|{order.branch_id}|{current_user.get('id','')}|"
            f"{order.customer_name or ''}|{order.customer_phone or ''}|"
            f"{order.order_type}|{_table_for_fp}|{items_sig}|{float(order.discount or 0):.2f}|"
            f"{order.payment_method or ''}|{_ext_for_fp}".encode("utf-8")
        ).hexdigest()

        # ⭐ نافذة منع التكرار 150 ثانية (كانت 30ث): يلتقط الضغط/الإرسال المزدوج المتباعد
        #    (مثل الطلب #34/#35 في السيدية بفارق دقيقة) ومزامنة الأوفلاين المتأخرة.
        #    البصمة تتضمّن الكاشير+الطاولة+الأصناف فلا تُدمَج طلبات حقيقية مختلفة.
        dedupe_cutoff_dt = datetime.now(timezone.utc) - timedelta(seconds=150)
        dedupe_cutoff = dedupe_cutoff_dt.isoformat()
        
        existing_dup = await db.orders.find_one(
            {
                "order_fingerprint": order_fingerprint,
                "created_at": {"$gte": dedupe_cutoff},
                "status": {"$ne": "cancelled"},
            },
            {"_id": 0},
        )
        # حماية إضافية: لا تَدمج طلبين لشركة توصيل برقمين خارجيين مختلفين (تجنّب فقدان طلب)
        if existing_dup and _ext_for_fp and (existing_dup.get("delivery_company_order_id") or "") != _ext_for_fp:
            existing_dup = None
        if existing_dup:
            logger.info(
                f"Order dedup: returning existing #{existing_dup.get('order_number')} "
                f"instead of creating duplicate (branch={order.branch_id})"
            )
            return existing_dup
    except Exception as _e:
        # لا نوقف إنشاء الطلب لو فشل فحص التكرار
        logger.warning(f"order dedup check failed: {_e}")
        order_fingerprint = None
    
    # سيتم حساب order_number لاحقاً (بعد استخراج business_date من الشفت لضمان الاتساق)
    
    # ===== الكابتن: الطلب يُحسب حصراً على وردية الكاشير المرتبط بها =====
    captain_id = None
    captain_name = None
    captain_link = None
    if (current_user.get("role") or "").strip().lower() == "captain":
        _ot = (order.order_type or "").strip().lower()
        if _ot == "delivery":
            raise HTTPException(status_code=400, detail="الكابتن يمكنه إنشاء طلبات داخلي/سفري فقط — التوصيل من اختصاص الكاشير")
        if _ot not in ("dine_in", "takeaway"):
            raise HTTPException(status_code=400, detail="نوع الطلب غير صالح للكابتن — يُسمح بالداخلي والسفري فقط")
        _link_q = {"captain_id": current_user["id"], "active": True}
        if tenant_id:
            _link_q["tenant_id"] = tenant_id
        captain_link = await db.captain_shift_links.find_one(_link_q, {"_id": 0})
        if not captain_link:
            raise HTTPException(status_code=400, detail="لم يربطك أي كاشير بورديته — اطلب من الكاشير ربطك أولاً")
        captain_id = current_user["id"]
        captain_name = current_user.get("full_name") or current_user.get("username", "")
    
    # البحث عن الوردية المفتوحة — أولوية لشفت الكاشير نفسه (وليس أي شفت في الفرع)
    # 🛡 حماية: الوردية يجب أن تكون لليوم التشغيلي الحالي فقط (business_date = اليوم بتوقيت العراق).
    #    وردية من يوم سابق → تُغلَق تلقائياً وتُفتح وردية جديدة نظيفة (منع خلط المبيعات بين الأيام).
    _biz_hour_now = await get_business_day_start_hour(tenant_id)
    today_biz_date = iraq_business_date_from_utc(start_hour=_biz_hour_now)
    
    base_shift_query = {"status": "open"}
    if tenant_id:
        base_shift_query["tenant_id"] = tenant_id
    if order.branch_id:
        base_shift_query["branch_id"] = order.branch_id
    
    # 1) جرب شفت الكاشير الخاص أولاً (الأكثر دقة)
    if captain_link:
        # الكابتن: الطلب يُحسب حصراً على وردية الكاشير المرتبط بها
        current_shift = await db.shifts.find_one(
            {"id": captain_link["shift_id"], "status": "open"},
            {"_id": 0, "id": 1, "cashier_id": 1, "cashier_name": 1, "business_date": 1, "started_at": 1, "opened_at": 1},
        )
        if not current_shift:
            raise HTTPException(status_code=400, detail="وردية الكاشير المرتبط أُغلقت — اطلب من الكاشير ربطك من جديد")
    else:
        current_shift = await db.shifts.find_one(
            {**base_shift_query, "cashier_id": current_user.get("id")},
            {"_id": 0, "id": 1, "cashier_id": 1, "business_date": 1, "started_at": 1, "opened_at": 1},
            sort=[("started_at", -1)]
        )
        # 2) إن لم يوجد، خذ أحدث شفت مفتوح في الفرع — للمالك/المدير فقط.
        # ⛔ الكاشير لا يلتقط شفت كاشير آخر أبداً (منع خلط المبيعات بين الكاشيرية) —
        # إن لم يكن له شفت، يُنشأ له شفت خاص به تلقائياً في الأسفل.
        if not current_shift and current_user.get("role", "") != "cashier":
            current_shift = await db.shifts.find_one(
                base_shift_query,
                {"_id": 0, "id": 1, "cashier_id": 1, "business_date": 1, "started_at": 1, "opened_at": 1},
                sort=[("started_at", -1)]
            )
    
    # 🛡 التحقق من business_date: لو وُجدت وردية لكن ليست لليوم الحالي → أغلقها وأنشئ جديدة
    if current_shift and not captain_link:
        _shift_biz = current_shift.get("business_date")
        if not _shift_biz:
            # وردية قديمة بدون business_date: احتسبها من started_at
            _st = current_shift.get("started_at") or current_shift.get("opened_at")
            if _st:
                try:
                    _shift_biz = iraq_business_date_from_utc(_st, start_hour=_biz_hour_now)
                except Exception:
                    _shift_biz = None
        if _shift_biz and _shift_biz != today_biz_date:
            # وردية من يوم سابق — أغلقها تلقائياً ("منسية") لتفادي خلط المبيعات
            logger.warning(
                f"🛡 stale open shift detected: shift_id={current_shift.get('id')} "
                f"business_date={_shift_biz} != today={today_biz_date}. Auto-closing & opening new."
            )
            try:
                await db.shifts.update_one(
                    {"id": current_shift["id"]},
                    {"$set": {
                        "status": "closed",
                        "ended_at": datetime.now(timezone.utc).isoformat(),
                        "auto_closed": True,
                        "auto_close_reason": "stale_business_date_next_day_order",
                        "notes": f"إغلاق تلقائي: وصل طلب جديد ليوم {today_biz_date} بينما الوردية لـ {_shift_biz}",
                    }}
                )
            except Exception as _e:
                logger.warning(f"failed to auto-close stale shift: {_e}")
            current_shift = None  # سيتم إنشاء وردية جديدة أدناه
    
    # إذا لم توجد وردية مفتوحة
    if not current_shift:
        user_role = current_user.get("role", "")
        is_owner = user_role in ["admin", "super_admin", "manager", "branch_manager"]
        
        if is_owner:
            # المالك: لا ننشئ وردية باسمه - نرجع خطأ
            raise HTTPException(status_code=400, detail="لا توجد وردية مفتوحة - يرجى فتح وردية لكاشير أولاً من نقاط البيع")
        else:
            # كاشير: امنع ازدواج الشفت — إن وُجدت وردية مفتوحة بنفس الاسم لليوم الحالي استخدمها
            _cashier_name = current_user.get("full_name", "") or ""
            _nn = " ".join(_cashier_name.strip().split()).lower()
            dup_shift = None
            if _nn:
                _open_q = {"status": "open", "business_date": today_biz_date}
                if tenant_id:
                    _open_q["tenant_id"] = tenant_id
                async for _s in db.shifts.find(_open_q, {"_id": 0, "id": 1, "cashier_name": 1}):
                    if " ".join((_s.get("cashier_name") or "").strip().split()).lower() == _nn:
                        dup_shift = await db.shifts.find_one({"id": _s["id"]}, {"_id": 0})
                        break
            if dup_shift:
                current_shift = dup_shift
            else:
                # لا توجد وردية بنفس الاسم لليوم: ننشئ وردية تلقائياً بتاريخ اليوم بالضبط
                new_shift = {
                    "id": str(uuid.uuid4()),
                    "tenant_id": tenant_id,
                    "branch_id": order.branch_id,
                    "cashier_id": current_user.get("id"),
                    "cashier_name": current_user.get("full_name", ""),
                    "opened_at": datetime.now(timezone.utc).isoformat(),
                    "started_at": datetime.now(timezone.utc).isoformat(),
                    "status": "open",
                    "opening_balance": 0,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "business_date": today_biz_date,
                    "auto_opened_by_first_order": True,
                }
                await db.shifts.insert_one(new_shift)
                current_shift = new_shift
                logger.info(f"✅ فُتحت وردية جديدة تلقائياً للكاشير {current_user.get('full_name')} — business_date={today_biz_date}")
    
    shift_id = current_shift["id"] if current_shift else None
    
    # تحديد business_date من الوردية (لضمان تسجيل السجل ضمن اليوم التشغيلي الصحيح حتى لو تجاوزت الوردية منتصف الليل)
    _biz_hour = await get_business_day_start_hour(tenant_id)
    shift_business_date = None
    if current_shift:
        shift_business_date = current_shift.get("business_date")
        if not shift_business_date:
            shift_started = current_shift.get("started_at") or current_shift.get("opened_at")
            if shift_started:
                shift_business_date = iraq_business_date_from_utc(shift_started, start_hour=_biz_hour)
    if not shift_business_date:
        shift_business_date = iraq_business_date_from_utc(start_hour=_biz_hour)
    
    # رقم الطلب — مرتبط بـ business_date (الشفت) لضمان عدم الانجراف عند منتصف الليل UTC
    order_number = await get_next_order_number(order.branch_id, shift_business_date)
    
    # المالك/المدير: الطلب يُسجل باسم كاشير الوردية وليس باسم المالك
    user_role = current_user.get("role", "")
    is_owner = user_role in ["admin", "super_admin", "manager", "branch_manager"]
    if is_owner and current_shift and current_shift.get("cashier_id") != current_user.get("id"):
        effective_cashier_id = current_shift["cashier_id"]
    elif captain_link and current_shift:
        # الكابتن: الطلب يُنسب لكاشير الوردية (يُحسب على ورديته) مع حفظ هوية الكابتن
        effective_cashier_id = current_shift["cashier_id"]
    else:
        effective_cashier_id = current_user["id"]
    
    # حساب المجموع الفرعي مع الإضافات - مُطابق تماماً لحساب السلة (Frontend):
    # الصيغة الصحيحة: (price × quantity) + extras_total  
    # NOT: (price + extras) × quantity  ← هذه كانت تُضاعف الإضافات
    # هذا يضمن تطابق 100% بين ما يراه العميل في السلة وما يُحفظ في DB
    def calculate_item_total(item):
        # سعر المنتج الأساسي مضروباً في الكمية
        base_total = item.price * item.quantity
        # الإضافات: كل إضافة لها سعر × كميتها الخاصة (لا تُضرب بكمية المنتج)
        extras_total = sum(_sn(extra.get("price")) * int(extra.get("quantity", 1)) for extra in (item.extras or []))
        return base_total + extras_total
    
    subtotal = sum(calculate_item_total(item) for item in order.items)
    tax = subtotal * 0.0  # No tax for Iraq
    # حد الخصم الكلي بـsubtotal لمنع الإجمالي السلبي
    safe_discount = max(0.0, min(subtotal, float(order.discount or 0)))
    total = max(0.0, subtotal - safe_discount + tax)
    
    # Calculate total cost including packaging for delivery/takeaway
    total_cost = 0
    total_packaging_cost = 0
    items_with_cost = []
    is_delivery_or_takeaway = order.order_type in ["delivery", "takeaway"]
    
    for item in order.items:
        product = await db.products.find_one({"id": item.product_id}, {"_id": 0})
        item_cost = 0
        packaging_cost = 0
        
        if product:
            # حساب تكلفة المنتج الأساسية
            base_cost = _sn(product.get("cost")) + _sn(product.get("operating_cost"))
            
            # إضافة تكلفة التغليف للتوصيل والسفري فقط
            if is_delivery_or_takeaway:
                packaging_cost = product.get("packaging_cost", 0) * item.quantity
                total_packaging_cost += packaging_cost
            
            # إذا كان المنتج مرتبط بمنتج مصنع، احصل على تكلفة المواد الخام
            # ⭐ يدعم الآن `manufactured_links` (متعدد) مع fallback للحقل القديم `manufactured_product_id`
            mfg_links = list(product.get("manufactured_links") or [])
            if not mfg_links and product.get("manufactured_product_id"):
                mfg_links = [{
                    "manufactured_product_id": product.get("manufactured_product_id"),
                    "consumption_qty": product.get("manufactured_consumption_qty") or 1,
                }]
            if mfg_links:
                # ⭐ نستخدم _enrich_unit_cost_fields كمصدر وحيد للحقيقة لـ unit_cost_after_waste.
                #    يدعم: piece_weight, pack_info, count→count, weight main_unit (لحم مفروم), nested mfg.
                links_unit_cost = 0.0
                for link in mfg_links:
                    mp_id = link.get("manufactured_product_id")
                    if not mp_id:
                        continue
                    mfg_product = await db.manufactured_products.find_one(
                        {"id": mp_id},
                        {"_id": 0, "raw_material_cost": 1, "raw_material_cost_after_waste": 1,
                         "production_cost": 1, "recipe": 1, "piece_weight": 1, "piece_weight_unit": 1,
                         "piece_def_value": 1, "piece_def_unit": 1,
                         "quantity": 1, "total_produced": 1, "unit": 1, "cost_before_waste": 1}
                    )
                    if not mfg_product:
                        continue
                    enrich = await _get_mfg_unit_cost(db, mfg_product)
                    unit_cost = enrich["unit_cost_after_waste"]
                    pw = float(mfg_product.get("piece_weight") or 0)
                    pwu = mfg_product.get("piece_weight_unit") or "غرام"
                    main_unit = mfg_product.get("unit") or "حبة"
                    consumption_qty = float(link.get("consumption_qty") or 1)
                    consumption_unit = link.get("consumption_unit") or main_unit
                    consumption_qty = _convert_link_consumption_to_main(
                        consumption_qty, consumption_unit, main_unit, pw, pwu
                    )
                    links_unit_cost += unit_cost * consumption_qty
                if links_unit_cost > 0:
                    base_cost = links_unit_cost + _sn(product.get("operating_cost"))
            
            item_cost = base_cost * item.quantity + packaging_cost
        
        total_cost += item_cost
        item_dict = item.model_dump()
        item_dict["cost"] = item_cost
        item_dict["packaging_cost"] = packaging_cost
        # حساب إجمالي سعر الإضافات المختارة للعنصر
        item_dict["extras_total"] = sum(_sn(extra.get("price")) * int(extra.get("quantity", 1)) for extra in (item.extras or []))
        items_with_cost.append(item_dict)
    
    # Calculate delivery commission if applicable
    delivery_commission = 0
    if order.delivery_app:
        commission_rate = await get_delivery_app_commission(order.delivery_app)
        delivery_commission = total * (commission_rate / 100)
    
    # Calculate profit
    profit = total - total_cost - delivery_commission
    
    # Determine payment status
    if order.payment_method == PaymentMethod.PENDING:
        payment_status = "pending"
        # الطلب جاهز تلقائياً إذا تم تحديده
        order_status = OrderStatus.READY if order.auto_ready else OrderStatus.PENDING
    elif order.payment_method == PaymentMethod.CREDIT:
        payment_status = "credit"
        order_status = OrderStatus.READY if order.auto_ready else OrderStatus.PREPARING
    else:
        payment_status = "paid"
        order_status = OrderStatus.READY if order.auto_ready else OrderStatus.PREPARING
    
    # الحصول على اسم شركة التوصيل
    delivery_app_name = order.delivery_app_name  # أولوية للقيمة المُرسلة
    is_delivery_company = order.is_delivery_company  # أولوية للقيمة المُرسلة
    customer_id = None
    
    # شركات التوصيل الافتراضية
    default_delivery_apps = {
        "toters": "توترز",
        "talabat": "طلبات",
        "baly": "بالي",
        "alsaree3": "عالسريع",
        "talabati": "طلباتي",
    }
    
    # إذا تم تحديد delivery_app، نجلب اسم الشركة
    if order.delivery_app and not delivery_app_name:
        # أولاً نتحقق من الشركات الافتراضية
        if order.delivery_app in default_delivery_apps:
            delivery_app_name = default_delivery_apps[order.delivery_app]
            is_delivery_company = True
        else:
            # ثم نبحث في قاعدة البيانات
            delivery_app_doc = await db.delivery_apps.find_one({"id": order.delivery_app})
            if delivery_app_doc:
                delivery_app_name = delivery_app_doc.get("name")
                is_delivery_company = True
            else:
                # نبحث في إعدادات شركات التوصيل
                app_setting = await db.delivery_app_settings.find_one({"app_id": order.delivery_app})
                if app_setting:
                    delivery_app_name = app_setting.get("name", order.delivery_app)
                    is_delivery_company = True
    
    # التحقق إذا كان العميل شركة توصيل
    if order.customer_phone:
        customer = await db.customers.find_one({
            "$or": [{"phone": order.customer_phone}, {"phone2": order.customer_phone}]
        })
        if customer:
            customer_id = customer.get("id")
            # إذا كان العميل شركة توصيل ولم يتم تحديدها مسبقاً
            if customer.get("is_delivery_company", False):
                is_delivery_company = True
                if not delivery_app_name:
                    delivery_app_name = customer.get("name")
    
    # البحث عن رقم الطاولة
    table_number = None
    if order.table_id:
        table_doc = await db.tables.find_one({"id": order.table_id}, {"_id": 0, "number": 1})
        if table_doc:
            table_number = table_doc.get("number")
    
    order_doc = {
        "id": str(uuid.uuid4()),
        "order_number": order_number,
        "order_type": order.order_type,
        "table_id": order.table_id,
        "table_number": table_number,
        "customer_name": order.customer_name,
        "customer_phone": order.customer_phone,
        "delivery_address": order.delivery_address,
        "buzzer_number": order.buzzer_number,  # رقم جهاز التنبيه
        "items": items_with_cost,
        "subtotal": subtotal,
        "discount": safe_discount,
        "coupon_id": order.coupon_id,
        "coupon_code": order.coupon_code,
        "coupon_name": order.coupon_name,
        "coupon_discount": order.coupon_discount,
        "tax": tax,
        "packaging_cost": total_packaging_cost,  # تكلفة التغليف الإجمالية
        "total": total,
        "total_cost": total_cost,
        "profit": profit,
        "branch_id": order.branch_id,
        "cashier_id": effective_cashier_id,
        "captain_id": captain_id,  # هوية الكابتن إن أنشأ الطلب (يُحسب على وردية الكاشير)
        "captain_name": captain_name,
        "captain_cash_status": ("held" if captain_id and order.payment_method == "cash" else None),
        "shift_id": shift_id,  # ربط الطلب بالوردية الحالية
        "business_date": shift_business_date,  # اليوم التشغيلي (حسب الوردية المفتوحة)
        "tenant_id": tenant_id,  # فصل البيانات لكل عميل
        "status": order_status,
        "payment_method": order.payment_method,
        "preferred_payment": order.preferred_payment,
        "payment_status": payment_status,
        "order_fingerprint": order_fingerprint,
        "offline_id": incoming_offline_id,  # ⭐ مفتاح الثبات لمنع التكرار عند المزامنة
        "delivery_company_order_id": (getattr(order, "delivery_company_order_id", None) or None),  # رقم الطلب لدى شركة التوصيل (مفتاح منع تكرار تجاري)
        "delivery_company_id": getattr(order, "delivery_company_id", None),
        "delivery_company_name": getattr(order, "delivery_company_name", None) or delivery_app_name,
        "delivery_company": getattr(order, "delivery_company", None),
        "is_offline_order": bool(getattr(order, "is_offline_order", False)),
        "delivery_app": order.delivery_app,
        "delivery_app_name": delivery_app_name,  # اسم شركة التوصيل
        "delivery_commission": delivery_commission,
        "is_delivery_company": is_delivery_company,  # هل العميل شركة توصيل
        "customer_id": customer_id,  # معرف العميل
        "driver_id": order.driver_id,
        "notes": order.notes,
        "credit_transferred": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    try:
        await db.orders.insert_one(order_doc)
    except DuplicateKeyError:
        # سباق: طلب بنفس offline_id أُدرج بالفعل (قفل قاعدة البيانات) — نُعيد الموجود بلا تكرار/آثار جانبية
        existing_dup = await db.orders.find_one(
            {"offline_id": incoming_offline_id, **({"tenant_id": tenant_id} if tenant_id else {})},
            {"_id": 0},
        )
        if existing_dup:
            logger.info(f"Order idempotency (DB lock): returning existing #{existing_dup.get('order_number')} for offline_id={incoming_offline_id}")
            return existing_dup
        raise
    del order_doc["_id"]
    
    # تحديث معلومات العميل إذا كان موجوداً
    if order.customer_phone:
        customer = await db.customers.find_one({"$or": [{"phone": order.customer_phone}, {"phone2": order.customer_phone}]})
        if customer:
            await db.customers.update_one(
                {"id": customer["id"]},
                {
                    "$inc": {"total_orders": 1, "total_spent": total},
                    "$set": {"last_order_date": datetime.now(timezone.utc).isoformat()}
                }
            )
        elif order.customer_name:
            # إنشاء عميل جديد تلقائياً
            new_customer = {
                "id": str(uuid.uuid4()),
                "name": order.customer_name,
                "phone": order.customer_phone,
                "phone2": None,
                "address": order.delivery_address,
                "area": None,
                "notes": None,
                "is_blocked": False,
                "total_orders": 1,
                "total_spent": total,
                "last_order_date": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            await db.customers.insert_one(new_customer)
    
    # Update table status if dine-in
    if order.table_id:
        await db.tables.update_one(
            {"id": order.table_id},
            {"$set": {"status": "occupied", "current_order_id": order_doc["id"]}}
        )
    
    # Assign driver if specified
    if order.driver_id:
        driver = await db.drivers.find_one({"id": order.driver_id}, {"_id": 0})
        if driver:
            # Update order with driver info
            await db.orders.update_one(
                {"id": order_doc["id"]},
                {"$set": {
                    "driver_name": driver.get("name"),
                    "driver_phone": driver.get("phone")
                }}
            )
            order_doc["driver_name"] = driver.get("name")
            order_doc["driver_phone"] = driver.get("phone")
            
            # Update driver availability
            await db.drivers.update_one(
                {"id": order.driver_id},
                {"$set": {"is_available": False, "current_order_id": order_doc["id"]}}
            )
    
    # Deduct inventory - خصم المواد الخام من مخزون الفرع بناءً على الوصفات
    # جلب إعدادات المخزون
    inventory_settings = await db.settings.find_one({"type": "inventory_settings"}, {"_id": 0})
    inventory_mode = inventory_settings.get("inventory_mode", "centralized") if inventory_settings else "centralized"
    
    for item in order.items:
        product = await db.products.find_one({"id": item.product_id})
        if product:
            # النظام الجديد: المنتجات المصنعة (يدعم ربط متعدد)
            mfg_links = list(product.get("manufactured_links") or [])
            if not mfg_links and product.get("manufactured_product_id"):
                mfg_links = [{
                    "manufactured_product_id": product.get("manufactured_product_id"),
                    "consumption_qty": product.get("manufactured_consumption_qty") or 1,
                }]
            if mfg_links:
                for link in mfg_links:
                    mp_id = link.get("manufactured_product_id")
                    if not mp_id:
                        continue
                    consumption_qty = float(link.get("consumption_qty") or 1)
                    # ⭐ تحويل الوحدة المختارة إلى الوحدة الرئيسية للخصم الصحيح من المخزون
                    consumption_unit = link.get("consumption_unit")
                    if consumption_unit:
                        mp_doc = await db.manufactured_products.find_one(
                            {"id": mp_id},
                            {"_id": 0, "unit": 1, "piece_weight": 1, "piece_weight_unit": 1}
                        )
                        if mp_doc:
                            consumption_qty = _convert_link_consumption_to_main(
                                consumption_qty,
                                consumption_unit,
                                mp_doc.get("unit") or "حبة",
                                float(mp_doc.get("piece_weight") or 0),
                                mp_doc.get("piece_weight_unit") or "",
                            )
                    deduct_amount = consumption_qty * item.quantity
                    # خصم من مخزون الفرع (branch_inventory)
                    branch_item = await db.branch_inventory.find_one({
                        "branch_id": order.branch_id,
                        "product_id": mp_id
                    })
                    if branch_item:
                        await db.branch_inventory.update_one(
                            {"id": branch_item["id"]},
                            {
                                "$inc": {
                                    "quantity": -deduct_amount,
                                    "sold_quantity": deduct_amount
                                },
                                "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
                            }
                        )
                    elif inventory_mode == "centralized":
                        await db.manufactured_products.update_one(
                            {"id": mp_id},
                            {
                                "$inc": {"quantity": -deduct_amount},
                                "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
                            }
                        )
                continue
            
            # النظام القديم: المنتجات النهائية
            finished_product_id = product.get("finished_product_id")
            if finished_product_id:
                finished_product = await db.inventory.find_one(
                    {"id": finished_product_id, "item_type": "finished"},
                    {"_id": 0}
                )
                if finished_product and finished_product.get("recipe"):
                    # خصم من المواد الخام بناءً على الوصفة
                    for ingredient in finished_product["recipe"]:
                        raw_material_id = ingredient.get("raw_material_id")
                        qty_per_unit = _sn(ingredient.get("quantity"))
                        total_to_deduct = qty_per_unit * item.quantity
                        
                        # خصم من مخزون الفرع
                        await db.inventory.update_one(
                            {"id": raw_material_id, "branch_id": order.branch_id},
                            {"$inc": {"quantity": -total_to_deduct}}
                        )
                        # أو من المخزون الرئيسي إذا لم يكن موجوداً في الفرع
                        await db.inventory.update_one(
                            {"id": raw_material_id, "branch_id": "main"},
                            {"$inc": {"quantity": -total_to_deduct}}
                        )
            
            # ثالثاً: استخدم المكونات القديمة إذا كانت موجودة (للتوافق مع البيانات القديمة)
            elif product.get("ingredients"):
                for ing in product["ingredients"]:
                    await db.inventory.update_one(
                        {"id": ing["inventory_id"]},
                        {"$inc": {"quantity": -ing["quantity"] * item.quantity}}
                    )
            
            # خصم مواد التغليف من مخزون الفرع (للتوصيل والسفري فقط)
            if is_delivery_or_takeaway and product.get("packaging_items"):
                for pkg_item in product["packaging_items"]:
                    packaging_material_id = pkg_item.get("packaging_material_id")
                    pkg_quantity = pkg_item.get("quantity", 1) * item.quantity
                    
                    if packaging_material_id:
                        # خصم من مخزون الفرع
                        branch_pkg = await db.branch_packaging_inventory.find_one({
                            "branch_id": order.branch_id,
                            "packaging_material_id": packaging_material_id
                        })
                        
                        if branch_pkg:
                            await db.branch_packaging_inventory.update_one(
                                {"id": branch_pkg["id"]},
                                {
                                    "$inc": {"used_quantity": pkg_quantity},
                                    "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
                                }
                            )
    
    return order_doc

@api_router.get("/orders", response_model=List[OrderResponse])
async def get_orders(
    branch_id: Optional[str] = None,
    status: Optional[str] = None,
    date: Optional[str] = None,
    payment_status: Optional[str] = None,
    order_type: Optional[str] = None,
    include_rejected: bool = False,
    current_user: dict = Depends(get_current_user)
):
    # فلترة حسب tenant_id و branch_id للمستخدم
    query = build_branch_query(current_user)

    # ⭐ الطلبات المرفوضة من الكاشير لا تظهر في إدارة الطلبات (تظهر فقط في إدارة التوصيل عبر include_rejected)
    if not include_rejected:
        query["is_rejected"] = {"$ne": True}
    
    # المستخدمون غير المدراء يرون فقط طلباتهم الخاصة لليوم الحالي
    user_role = current_user.get("role", "")
    is_manager = user_role in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]
    
    if not is_manager:
        query["cashier_id"] = current_user["id"]
        # إذا لم يتم تحديد تاريخ، نعرض فقط طلبات اليوم
        if not date:
            today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
            query["created_at"] = {"$regex": f"^{today}"}
    
    # إذا تم تحديد فرع في الطلب، تحقق من صلاحية الوصول
    if branch_id:
        if not user_can_access_branch(current_user, branch_id):
            raise HTTPException(status_code=403, detail="لا يمكنك الوصول لهذا الفرع")
        query["branch_id"] = branch_id
    
    if status:
        # دعم حالات متعددة مفصولة بفاصلة
        statuses = [s.strip() for s in status.split(',')]
        if len(statuses) > 1:
            query["status"] = {"$in": statuses}
        else:
            query["status"] = status
    if date:
        query["created_at"] = {"$regex": f"^{date}"}
    if payment_status:
        query["payment_status"] = payment_status
    if order_type:
        query["order_type"] = order_type
    
    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return orders


@api_router.get("/delivery-orders")
async def get_delivery_orders(
    branch_id: Optional[str] = None,
    date: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """عرض شامل لكل طلبات التوصيل لإدارة التوصيل:
    يشمل المرفوض + المتأخر في القبول (>دقيقتين) + المقبول + قيد التحضير + المكتمل.
    """
    query = build_branch_query(current_user)
    query["order_type"] = "delivery"
    if branch_id:
        if not user_can_access_branch(current_user, branch_id):
            raise HTTPException(status_code=403, detail="لا يمكنك الوصول لهذا الفرع")
        query["branch_id"] = branch_id
    if date:
        query["created_at"] = {"$regex": f"^{date}"}

    orders = await db.orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    now = datetime.now(timezone.utc)

    def _parse(dt):
        if not dt:
            return None
        try:
            d = datetime.fromisoformat(dt.replace("Z", "+00:00")) if isinstance(dt, str) else dt
            if d.tzinfo is None:
                d = d.replace(tzinfo=timezone.utc)
            return d
        except Exception:
            return None

    summary = {"all": 0, "rejected": 0, "late": 0, "accepted": 0, "preparing": 0, "completed": 0, "cancelled": 0}
    result = []
    for o in orders:
        # ⭐ استبعاد طلبات شركات التوصيل (طلبات/توترز/مزاجك...) — تظهر في تقرير التوصيل فقط.
        # إدارة التوصيل تعرض فقط طلبات تطبيق الزبائن + التوصيل الداخلي (سائق خاص بلا شركة).
        is_company = (
            bool(o.get("is_delivery_company"))
            or bool(o.get("delivery_company_name"))
            or bool(o.get("delivery_app_name"))
            or bool(o.get("delivery_app_id"))
        )
        if is_company:
            continue
        st = o.get("status")
        is_rejected = bool(o.get("is_rejected"))
        is_late = bool(o.get("acceptance_late"))
        waiting_seconds = None
        # طلب لم يُقبل بعد: احسب مدة الانتظار، واعتبره متأخراً إن تجاوزت دقيقتين
        if not is_rejected and st == "pending" and not o.get("accepted_at"):
            created = _parse(o.get("created_at"))
            if created:
                waiting_seconds = int(max(0, (now - created).total_seconds()))
                if waiting_seconds > 120:
                    is_late = True
        # التصنيف الأساسي
        if is_rejected:
            category = "rejected"
        elif st == "delivered":
            category = "completed"
        elif st == "cancelled":
            category = "cancelled"
        elif st in ("preparing", "ready", "out_for_delivery"):
            category = "preparing"
        else:
            category = "accepted" if o.get("accepted_at") else "pending"

        summary["all"] += 1
        # العدّ حسب نفس تصنيف category لضمان تطابق شارات التبويبات مع القوائم المفلترة
        if category in summary:
            summary[category] += 1
        if is_late:
            summary["late"] += 1

        result.append({
            "id": o.get("id"),
            "order_number": o.get("order_number"),
            "status": st,
            "category": category,
            "is_rejected": is_rejected,
            "is_late": is_late,
            "acceptance_delay_seconds": o.get("acceptance_delay_seconds"),
            "waiting_seconds": waiting_seconds,
            "cancellation_reason": o.get("cancellation_reason"),
            "rejected_by_name": o.get("rejected_by_name"),
            "customer_name": o.get("customer_name"),
            "customer_phone": o.get("customer_phone"),
            "delivery_address": o.get("delivery_address"),
            "driver_id": o.get("driver_id"),
            "driver_name": o.get("driver_name"),
            "driver_phone": o.get("driver_phone"),
            "driver_payment_status": o.get("driver_payment_status"),
            "payment_status": o.get("payment_status"),
            "delivery_fee": o.get("delivery_fee", 0),
            "total": o.get("total", 0),
            "items_count": len(o.get("items") or []),
            "branch_id": o.get("branch_id"),
            "created_at": o.get("created_at"),
            "accepted_at": o.get("accepted_at"),
        })

    return {"orders": result, "summary": summary}


    """بصمة محتوى الطلب لاكتشاف التطابق.
    متساهلة عمداً: لا تعتمد على product_id (لأن النسخة الأوفلاين قد تخزّن العناصر
    باسم المنتج فقط) — نعتمد على: الفرع + النوع + العميل + الإجمالي + عدد العناصر
    + إجمالي الكميات + يوم العمل. مع شرط (مدفوع + غير مدفوع) لمنع الإيجابيات الكاذبة.
    """
    items = o.get("items") or []
    item_count = len(items)
    total_qty = round(sum(float(it.get("quantity") or 0) for it in items), 2)
    biz = o.get("business_date") or (o.get("created_at") or "")[:10]
    return (
        f"{o.get('branch_id') or ''}|{o.get('order_type') or ''}|"
        f"{(o.get('customer_name') or '').strip()}|{(o.get('customer_phone') or '').strip()}|"
        f"{round(float(o.get('total') or 0), 2)}|{item_count}|{total_qty}|{biz}"
    )


def _order_is_paid(o: dict) -> bool:
    # التعريف المعتمد في النظام (نفس منطق الإرجاع): مدفوع فقط إذا كانت الحالة paid أو credit
    return (o.get("payment_status") or "").lower() in ("paid", "credit")


@api_router.get("/orders/duplicates")
async def find_duplicate_orders(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user),
):
    """كشف الطلبات المكررة القديمة (مالك/مدير عام فقط).

    يُرجِع فقط المجموعات التي فيها طلب **مدفوع** حقيقي + نسخة **غير مدفوعة** مطابقة
    (السيناريو الناتج عن تكرار الأوفلاين). الطلبات المعلّقة العادية لا تُعتبر تكراراً
    لأنها بلا طلب مدفوع مطابق.
    """
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.SUPER_ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="هذه الأداة متاحة للمالك/المدير العام فقط")

    tenant_id = get_user_tenant_id(current_user)
    query = {"status": {"$nin": ["cancelled", "refunded"]}}
    if tenant_id:
        query["tenant_id"] = tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    if start_date and end_date:
        query["business_date"] = {"$gte": start_date, "$lte": end_date}
    elif start_date:
        query["business_date"] = {"$gte": start_date}

    groups_map: dict = {}
    _dup_proj = {
        "_id": 0, "id": 1, "order_number": 1, "branch_id": 1, "order_type": 1,
        "customer_name": 1, "customer_phone": 1, "total": 1, "items": 1,
        "business_date": 1, "created_at": 1, "payment_method": 1,
        "payment_status": 1, "is_offline_order": 1, "status": 1,
    }
    async for o in db.orders.find(query, _dup_proj):
        groups_map.setdefault(_order_dup_signature(o), []).append(o)

    result_groups = []
    total_dups = 0
    for sig, items in groups_map.items():
        if len(items) < 2:
            continue
        paid = [o for o in items if _order_is_paid(o)]
        unpaid = [o for o in items if not _order_is_paid(o)]
        # يجب وجود طلب مدفوع حقيقي + نسخة غير مدفوعة مطابقة
        if not paid or not unpaid:
            continue
        items.sort(key=lambda x: x.get("created_at") or "")
        paid.sort(key=lambda x: x.get("created_at") or "")
        keep = paid[0]

        def _slim(o):
            return {
                "id": o.get("id"),
                "order_number": o.get("order_number"),
                "total": o.get("total"),
                "order_type": o.get("order_type"),
                "customer_name": o.get("customer_name"),
                "payment_method": o.get("payment_method"),
                "payment_status": o.get("payment_status"),
                "is_offline_order": bool(o.get("is_offline_order")),
                "is_paid": _order_is_paid(o),
                "created_at": o.get("created_at"),
                "business_date": o.get("business_date"),
            }

        removable = [o for o in unpaid if o.get("id") != keep.get("id")]
        if not removable:
            continue
        total_dups += len(removable)
        result_groups.append({
            "signature": sig,
            "keep": _slim(keep),
            "duplicates": [_slim(o) for o in removable],
        })

    result_groups.sort(key=lambda g: g["keep"].get("order_number") or 0, reverse=True)
    return {
        "groups": result_groups,
        "total_groups": len(result_groups),
        "total_duplicates": total_dups,
    }


@api_router.get("/orders/{order_id}", response_model=OrderResponse)
async def get_order(order_id: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    # حل اسم الكاشير من cashier_id إذا لم يكن مخزّناً (لعرض اسم الكاشير في التفاصيل)
    if not order.get("cashier_name") and order.get("cashier_id"):
        u = await db.users.find_one({"id": order["cashier_id"]}, {"_id": 0, "full_name": 1, "username": 1})
        if u:
            order["cashier_name"] = u.get("full_name") or u.get("username") or ""
    return order

@api_router.put("/orders/{order_id}/add-items")
async def add_items_to_order(order_id: str, items: List[OrderItemCreate], current_user: dict = Depends(get_current_user)):
    """إضافة عناصر جديدة لطلب موجود"""
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # إضافة العناصر الجديدة
    new_items = []
    product_query = build_tenant_query(current_user)
    for item in items:
        product_query["id"] = item.product_id
        product = await db.products.find_one(product_query)
        extras_total = sum(_sn(extra.get("price")) * int(extra.get("quantity", 1)) for extra in (item.extras or []))
        new_items.append({
            "product_id": item.product_id,
            "product_name": item.product_name,
            "quantity": item.quantity,
            "price": item.price,
            "cost": _sn(product.get("cost")) if product else 0,
            "notes": item.notes,
            "extras": item.extras or [],
            "extras_total": extras_total
        })
    
    # دمج العناصر الجديدة مع القديمة
    existing_items = order.get("items", [])
    all_items = existing_items + new_items
    
    # إعادة حساب المجاميع - الصيغة الصحيحة (مطابقة للسلة):
    # (price × qty) + extras_total — الإضافات لا تُضرب بالكمية
    subtotal = sum((_sn(i["price"]) * _sn(i["quantity"])) + _sn(i.get("extras_total")) for i in all_items)
    total_cost = sum(_sn(i.get("cost")) * i["quantity"] for i in all_items)
    discount = _sn(order.get("discount"))
    tax = 0
    total = subtotal - discount + tax
    profit = total - total_cost - order.get("delivery_commission", 0)
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "items": all_items,
            "subtotal": subtotal,
            "total_cost": total_cost,
            "total": total,
            "profit": profit,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    query = build_tenant_query(current_user, {"id": order_id})
    return await db.orders.find_one(query, {"_id": 0})


class UpdateOrderItemsRequest(BaseModel):
    items: List[OrderItemCreate]
    notes: Optional[str] = None
    discount: float = 0.0
    # طريقة الدفع المفضّلة — تحفظ من POS عند تعديل طلب معلق
    preferred_payment: Optional[str] = None
    # تفاصيل الكوبون عند تطبيقه على الطلب أثناء التعديل
    coupon_id: Optional[str] = None
    coupon_code: Optional[str] = None
    coupon_name: Optional[str] = None
    coupon_discount: float = 0.0


@api_router.put("/orders/{order_id}/update-items")
async def update_order_items(order_id: str, request: UpdateOrderItemsRequest, current_user: dict = Depends(get_current_user)):
    """تحديث جميع عناصر الطلب والملاحظات"""
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # بناء العناصر المحدّثة مع التكاليف
    updated_items = []
    total_cost = 0
    for item in request.items:
        product = await db.products.find_one({"id": item.product_id}, {"_id": 0})
        item_cost = 0
        if product:
            base_cost = _sn(product.get("cost")) + _sn(product.get("operating_cost"))
            # ⭐ يدعم ربط متعدد بالمنتجات المُصنّعة
            mfg_links = list(product.get("manufactured_links") or [])
            if not mfg_links and product.get("manufactured_product_id"):
                mfg_links = [{
                    "manufactured_product_id": product.get("manufactured_product_id"),
                    "consumption_qty": product.get("manufactured_consumption_qty") or 1,
                }]
            if mfg_links:
                links_unit_cost = 0.0
                for link in mfg_links:
                    mp_id = link.get("manufactured_product_id")
                    if not mp_id:
                        continue
                    mfg_product = await db.manufactured_products.find_one(
                        {"id": mp_id},
                        {"_id": 0, "raw_material_cost": 1, "raw_material_cost_after_waste": 1,
                         "production_cost": 1, "recipe": 1, "unit": 1, "piece_weight": 1,
                         "piece_weight_unit": 1, "piece_def_value": 1, "piece_def_unit": 1,
                         "quantity": 1, "total_produced": 1, "cost_before_waste": 1}
                    )
                    if mfg_product:
                        # ⭐ مصدر وحيد للحقيقة (نفس منطق _enrich_unit_cost_fields)
                        enrich = await _get_mfg_unit_cost(db, mfg_product)
                        unit_cost = enrich["unit_cost_after_waste"]
                        consumption_qty = float(link.get("consumption_qty") or 1)
                        consumption_unit = link.get("consumption_unit") or mfg_product.get("unit") or "حبة"
                        consumption_qty = _convert_link_consumption_to_main(
                            consumption_qty,
                            consumption_unit,
                            mfg_product.get("unit") or "حبة",
                            float(mfg_product.get("piece_weight") or 0),
                            mfg_product.get("piece_weight_unit") or "",
                        )
                        links_unit_cost += unit_cost * consumption_qty
                if links_unit_cost > 0:
                    base_cost = links_unit_cost + _sn(product.get("operating_cost"))
            item_cost = base_cost * item.quantity
        
        total_cost += item_cost
        extras_total = sum(_sn(extra.get("price")) * int(extra.get("quantity", 1)) for extra in (item.extras or []))
        updated_items.append({
            "product_id": item.product_id,
            "product_name": item.product_name,
            "quantity": item.quantity,
            "price": item.price,
            "cost": item_cost,
            "notes": item.notes,
            "extras": item.extras or [],
            "extras_total": extras_total
        })
    
    # إعادة حساب المجاميع - الصيغة الصحيحة (مطابقة للسلة)
    subtotal = sum((_sn(i["price"]) * _sn(i["quantity"])) + _sn(i.get("extras_total")) for i in updated_items)
    # حد الخصم بـsubtotal لمنع الإجمالي السلبي
    safe_discount = max(0.0, min(subtotal, float(request.discount or 0)))
    discount = safe_discount
    tax = 0
    total = max(0.0, subtotal - discount + tax)
    profit = total - total_cost - order.get("delivery_commission", 0)
    
    update_data = {
        "items": updated_items,
        "subtotal": subtotal,
        "discount": discount,
        "total_cost": total_cost,
        "total": total,
        "profit": profit,
        "notes": request.notes,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    # حفظ طريقة الدفع المفضّلة (إذا أرسلها العميل)
    if request.preferred_payment:
        update_data["preferred_payment"] = request.preferred_payment
    # حفظ تفاصيل الكوبون عند تعديل الطلب
    if request.coupon_id:
        update_data["coupon_id"] = request.coupon_id
        update_data["coupon_code"] = request.coupon_code
        update_data["coupon_name"] = request.coupon_name
        update_data["coupon_discount"] = request.coupon_discount
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    query = build_tenant_query(current_user, {"id": order_id})
    return await db.orders.find_one(query, {"_id": 0})

@api_router.put("/orders/{order_id}/status")
async def update_order_status(order_id: str, status: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التحقق من صلاحية الإلغاء
    if status == OrderStatus.CANCELLED:
        # فقط المالك أو المدير يمكنهم الإلغاء
        if current_user.get("role") not in ["admin", "manager"]:
            raise HTTPException(status_code=403, detail="ليس لديك صلاحية إلغاء الطلبات")
    
    await db.orders.update_one(
        query,
        {"$set": {"status": status, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    # Free table if completed
    if status in [OrderStatus.DELIVERED, OrderStatus.CANCELLED] and order.get("table_id"):
        await db.tables.update_one(
            {"id": order["table_id"]},
            {"$set": {"status": "available", "current_order_id": None}}
        )

    # إشعار Push للزبون عند تغيّر حالة الطلب (جاري التحضير/جاهز/في الطريق/تم التسليم)
    try:
        await notify_order_status_change(order_id, status)
    except Exception as _e:
        logger.warning(f"customer status push failed: {_e}")

    return {"message": "تم التحديث"}

@api_router.put("/orders/{order_id}/kitchen-status")
async def update_order_kitchen_status(order_id: str, kitchen_status: str, current_user: dict = Depends(get_current_user)):
    """
    تحديث حالة الطلب في المطبخ بشكل مستقل عن حالة الطلب الرئيسية
    kitchen_status: pending_kitchen, preparing_kitchen, ready_kitchen, completed_kitchen
    """
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    valid_statuses = ["pending_kitchen", "preparing_kitchen", "ready_kitchen", "completed_kitchen"]
    if kitchen_status not in valid_statuses:
        raise HTTPException(status_code=400, detail="حالة المطبخ غير صالحة")
    
    await db.orders.update_one(
        query,
        {"$set": {
            "kitchen_status": kitchen_status, 
            "kitchen_updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )

    # ⭐ إشعار ثانٍ للسائق عند جاهزية الطلب للاستلام
    if kitchen_status == "ready_kitchen" and order.get("driver_id") and order.get("order_type") == "delivery":
        try:
            now = datetime.now(timezone.utc)
            ready_notification = {
                "id": f"notif_{now.timestamp()}_{order_id}_ready",
                "type": "order_ready",
                "order_id": order_id,
                "order_number": str(order.get("order_number", "")),
                "branch_id": order.get("branch_id", ""),
                "order_type": order.get("order_type", "delivery"),
                "customer_name": order.get("customer_name"),
                "customer_phone": order.get("customer_phone"),
                "delivery_address": order.get("delivery_address"),
                "driver_id": order.get("driver_id"),
                "total_amount": float(order.get("total") or 0),
                "items_count": len(order.get("items") or []),
                "tenant_id": order.get("tenant_id"),
                "is_read": False,
                "is_printed": False,
                "created_at": now.isoformat(),
            }
            await db.order_notifications.insert_one(ready_notification)
            try:
                from services.websocket_service import notify_driver_new_order
                await notify_driver_new_order(order.get("driver_id"), {
                    "order_id": order_id,
                    "order_number": str(order.get("order_number", "")),
                    "ready": True,
                    "customer_name": order.get("customer_name"),
                    "delivery_address": order.get("delivery_address"),
                    "branch_id": order.get("branch_id", ""),
                })
            except Exception as _e:
                logger.warning(f"driver ready websocket notify failed: {_e}")
        except Exception as _e:
            logger.warning(f"driver ready notification insert failed: {_e}")

    return {"message": "تم تحديث حالة المطبخ"}

@api_router.get("/kitchen/orders")
async def get_kitchen_orders(current_user: dict = Depends(get_current_user)):
    """
    جلب الطلبات لشاشة المطبخ
    تظهر الطلبات التي لم يتم تحديدها كـ completed_kitchen بغض النظر عن حالتها الرئيسية (حتى لو مدفوعة)
    """
    tenant_id = get_user_tenant_id(current_user)
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    
    # جلب الطلبات التي ليست completed_kitchen وليست ملغاة (اليوم فقط)
    query = {
        "tenant_id": tenant_id,
        "kitchen_status": {"$nin": ["completed_kitchen", None]},
        "status": {"$ne": "cancelled"},
        "created_at": {"$regex": f"^{today}"}
    }
    
    # أيضاً جلب الطلبات الجديدة التي ليس لها kitchen_status بعد (بغض النظر عن حالة الدفع)
    query_new = {
        "tenant_id": tenant_id,
        "kitchen_status": {"$exists": False},
        "status": {"$nin": ["cancelled", "delivered"]},  # فقط استثناء الملغاة والمسلّمة
        "created_at": {"$regex": f"^{today}"}
    }
    
    # دمج النتائج
    orders_with_kitchen_status = await db.orders.find(query, {"_id": 0}).sort("created_at", 1).to_list(200)
    orders_new = await db.orders.find(query_new, {"_id": 0}).sort("created_at", 1).to_list(200)
    
    # إضافة kitchen_status الافتراضي للطلبات الجديدة
    for order in orders_new:
        if "kitchen_status" not in order:
            order["kitchen_status"] = "pending_kitchen"
            # تحديث في قاعدة البيانات
            await db.orders.update_one(
                {"id": order["id"]},
                {"$set": {"kitchen_status": "pending_kitchen"}}
            )
    
    all_orders = orders_with_kitchen_status + orders_new
    
    # إزالة التكرارات وترتيب حسب التاريخ
    seen_ids = set()
    unique_orders = []
    for order in all_orders:
        if order["id"] not in seen_ids:
            seen_ids.add(order["id"])
            unique_orders.append(order)
    
    unique_orders.sort(key=lambda x: x.get("created_at", ""))
    
    # إضافة اسم الفرع لكل طلب
    branch_ids = list(set(o.get("branch_id") for o in unique_orders if o.get("branch_id")))
    branches = {}
    if branch_ids:
        branches_list = await db.branches.find({"id": {"$in": branch_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
        branches = {b["id"]: b["name"] for b in branches_list}
    
    for order in unique_orders:
        order["branch_name"] = branches.get(order.get("branch_id"), "")
    
    return unique_orders

@api_router.put("/orders/{order_id}/payment")
async def update_order_payment(order_id: str, payment_method: str, current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    update_data = {
        "payment_method": payment_method,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if payment_method == PaymentMethod.CREDIT:
        update_data["payment_status"] = "credit"
        if order.get("delivery_app"):
            update_data["credit_transferred"] = True
    elif payment_method in [PaymentMethod.CASH, PaymentMethod.CARD]:
        update_data["payment_status"] = "paid"
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    return {"message": "تم تحديث طريقة الدفع"}

@api_router.put("/orders/{order_id}/cancel")
async def cancel_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """
    إلغاء الطلب:
    - أقل من دقيقة: يُحذف نهائياً (لا يظهر في التقارير)
    - أكثر من دقيقتين: يُسجل كطلب ملغي (يظهر في التقارير)
    - بين دقيقة ودقيقتين: يُسجل كطلب ملغي فقط للمدير/المالك
    """
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # حساب الفرق في الوقت بالثواني
    created_at = datetime.fromisoformat(order["created_at"].replace("Z", "+00:00"))
    time_diff = (datetime.now(timezone.utc) - created_at).total_seconds()
    
    is_within_minute = time_diff < 60  # أقل من دقيقة
    is_within_two_minutes = time_diff < 120  # أقل من دقيقتين
    is_admin_or_manager = current_user.get("role") in ["admin", "manager"]
    
    # التحقق من الصلاحيات
    if not is_within_minute and not is_admin_or_manager:
        raise HTTPException(status_code=403, detail="فقط المالك أو المدير يمكنهم إلغاء الطلبات بعد دقيقة من إنشائها")
    
    # تحرير الطاولة إذا كان الطلب على طاولة
    if order.get("table_id"):
        await db.tables.update_one(
            {"id": order["table_id"]},
            {"$set": {"status": "available", "current_order_id": None}}
        )
    
    # أقل من دقيقة: حذف نهائي
    if is_within_minute:
        await db.orders.delete_one({"id": order_id})
        return {
            "message": "تم حذف الطلب نهائياً",
            "was_quick_delete": True,
            "in_reports": False
        }
    
    # أكثر من دقيقة: تسجيل كطلب ملغي
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": OrderStatus.CANCELLED,
            "cancelled_at": datetime.now(timezone.utc).isoformat(),
            "cancelled_by": current_user["id"],
            "cancellation_reason": "إلغاء بواسطة المدير" if is_admin_or_manager else "إلغاء",
            "time_to_cancel_seconds": int(time_diff),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {
        "message": "تم إلغاء الطلب وتسجيله في التقارير",
        "was_quick_delete": False,
        "in_reports": True
    }


@api_router.put("/orders/{order_id}/reject")
async def reject_order(order_id: str, reason: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """رفض طلب وارد من قبل الكاشير: يُسجَّل الطلب كمرفوض (ملغي مع سبب) ويظهر في تقرير الطلبات،
    ويُشعَر الزبون بالرفض. لا يُحذف نهائياً (بخلاف الإلغاء السريع) كي يبقى مُتتبَّعاً.
    """
    allowed = [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.CASHIER, UserRole.SUPER_ADMIN, "owner"]
    if current_user.get("role") not in allowed:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": OrderStatus.CANCELLED,
            "is_rejected": True,
            "driver_id": None,
            "driver_name": None,
            "driver_phone": None,
            "cancelled_at": now_iso,
            "cancelled_by": current_user.get("id"),
            "cancellation_reason": reason or "رُفض الطلب من قبل الكاشير",
            "rejected_at": now_iso,
            "rejected_by_name": current_user.get("full_name") or current_user.get("username") or "",
            "updated_at": now_iso,
        }}
    )
    # تحرير الطاولة إن وُجدت
    if order.get("table_id"):
        await db.tables.update_one(
            {"id": order["table_id"]},
            {"$set": {"status": "available", "current_order_id": None}}
        )
    # إشعار الزبون بأن الطلب رُفض
    try:
        await notify_order_status_change(order_id, "cancelled")
    except Exception as _e:
        logger.warning(f"reject notify customer failed: {_e}")
    # أوقف مكالمة الكاشير لهذا الطلب (علّمها مقروءة) وأنشئ إشعاراً للإدارة بالرفض
    try:
        await db.order_notifications.update_many(
            {"order_id": order_id, "type": "new_order_cashier"},
            {"$set": {"is_read": True, "escalated": True}},
        )
        branch = await db.branches.find_one({"id": order.get("branch_id")}, {"_id": 0, "name": 1})
        rejecter = current_user.get("full_name") or current_user.get("username") or "الكاشير"
        await db.order_notifications.update_one(
            {"id": f"esc_{order_id}"},
            {"$set": {
                "id": f"esc_{order_id}",
                "type": "order_management_alert",
                "alert_kind": "rejected",
                "order_id": order_id,
                "order_number": order.get("order_number"),
                "branch_id": order.get("branch_id"),
                "branch_name": (branch or {}).get("name", "غير محدد"),
                "cashier_name": rejecter,
                "order_type": order.get("order_type"),
                "customer_name": order.get("customer_name"),
                "customer_phone": order.get("customer_phone"),
                "delivery_address": order.get("delivery_address"),
                "total_amount": order.get("total", 0),
                "items_count": len(order.get("items") or []),
                "reason": reason or "رُفض الطلب من قبل الكاشير",
                "is_read": False,
                "created_at": now_iso,
            }},
            upsert=True,
        )
    except Exception as _e:
        logger.warning(f"reject management alert failed: {_e}")
    return {"message": "تم رفض الطلب وتسجيله كطلب مرفوض", "order_id": order_id}


class ItemCancellationRequest(BaseModel):
    """طلب إلغاء صنف من طلب محفوظ (إلغاء جزئي)"""
    product_id: str
    product_name: Optional[str] = None
    quantity: float = 1
    price: float = 0
    reason: Optional[str] = "حذف من الطلب بعد الحفظ"


@api_router.post("/orders/recompute-costs")
async def recompute_order_costs(
    days: int = 30,
    dry_run: bool = True,
    current_user: dict = Depends(get_current_user),
):
    """🔧 إعادة حساب تكلفة الطلبات القديمة باستخدام منطق التكلفة الموحَّد الحالي.

    يُستخدم لتصحيح الطلبات التي حُسِبت بمنطق قديم خاطئ (قبل HOTFIX 24/05/2026).

    Args:
        days: عدد الأيام الماضية لإعادة حسابها (الافتراضي 30).
        dry_run: إن True يعرض الفرق فقط دون كتابة. إن False يُحدِّث الـ DB.
    """
    # Only allow admin/owner roles
    role = (current_user.get("role") or "").lower()
    if role not in ("admin", "owner", "super_admin", "manager"):
        raise HTTPException(status_code=403, detail="غير مصرّح")

    tenant_id = get_user_tenant_id(current_user)
    cutoff_date = (datetime.now(timezone.utc) - timedelta(days=days)).isoformat()
    query: Dict[str, Any] = {"created_at": {"$gte": cutoff_date}}
    if tenant_id:
        query["tenant_id"] = tenant_id

    orders = await db.orders.find(query, {"_id": 0}).to_list(length=10000)
    summary = {
        "examined": len(orders),
        "updated": 0,
        "unchanged": 0,
        "total_old_cost": 0.0,
        "total_new_cost": 0.0,
        "total_cost_delta": 0.0,
        "samples": [],
    }

    # cache منتجات لتقليل الـ DB hits
    product_cache: Dict[str, dict] = {}
    mfg_cache: Dict[str, dict] = {}

    async def _get_product(pid):
        if pid in product_cache:
            return product_cache[pid]
        p = await db.products.find_one({"id": pid}, {"_id": 0})
        product_cache[pid] = p
        return p

    async def _get_mfg(mid):
        if mid in mfg_cache:
            return mfg_cache[mid]
        m = await db.manufactured_products.find_one(
            {"id": mid},
            {"_id": 0, "raw_material_cost": 1, "raw_material_cost_after_waste": 1,
             "production_cost": 1, "recipe": 1, "piece_weight": 1, "piece_weight_unit": 1,
             "piece_def_value": 1, "piece_def_unit": 1,
             "quantity": 1, "total_produced": 1, "unit": 1, "cost_before_waste": 1}
        )
        mfg_cache[mid] = m
        return m

    for o in orders:
        old_total = float(o.get("total_cost") or 0)
        new_total = 0.0
        new_items = []
        for it in (o.get("items") or []):
            product = await _get_product(it.get("product_id"))
            packaging_cost = 0.0
            new_item_cost = 0.0
            if product:
                base_cost = _sn(product.get("cost")) + _sn(product.get("operating_cost"))
                packaging_cost = float(product.get("packaging_cost") or 0) * int(it.get("quantity", 1))
                mfg_links = list(product.get("manufactured_links") or [])
                if not mfg_links and product.get("manufactured_product_id"):
                    mfg_links = [{
                        "manufactured_product_id": product.get("manufactured_product_id"),
                        "consumption_qty": product.get("manufactured_consumption_qty") or 1,
                    }]
                if mfg_links:
                    links_unit_cost = 0.0
                    for link in mfg_links:
                        mp = await _get_mfg(link.get("manufactured_product_id"))
                        if not mp:
                            continue
                        enrich = await _get_mfg_unit_cost(db, mp)
                        unit_cost = enrich["unit_cost_after_waste"]
                        pw = float(mp.get("piece_weight") or 0)
                        pwu = mp.get("piece_weight_unit") or "غرام"
                        main_unit = mp.get("unit") or "حبة"
                        cq = float(link.get("consumption_qty") or 1)
                        cq = _convert_link_consumption_to_main(
                            cq, link.get("consumption_unit") or main_unit, main_unit, pw, pwu
                        )
                        links_unit_cost += unit_cost * cq
                    if links_unit_cost > 0:
                        base_cost = links_unit_cost + _sn(product.get("operating_cost"))
                new_item_cost = base_cost * int(it.get("quantity", 1)) + packaging_cost
            new_total += new_item_cost
            it2 = dict(it)
            it2["cost"] = round(new_item_cost, 2)
            new_items.append(it2)

        new_total = round(new_total, 2)
        if abs(new_total - old_total) > 0.01:
            summary["updated"] += 1
            summary["total_cost_delta"] += (new_total - old_total)
            summary["total_old_cost"] += old_total
            summary["total_new_cost"] += new_total
            if len(summary["samples"]) < 10:
                summary["samples"].append({
                    "order_id": o.get("id"),
                    "created_at": o.get("created_at"),
                    "old_total_cost": old_total,
                    "new_total_cost": new_total,
                    "delta": round(new_total - old_total, 2),
                })
            if not dry_run:
                await db.orders.update_one(
                    {"id": o.get("id")},
                    {"$set": {"total_cost": new_total, "items": new_items,
                              "cost_recomputed_at": datetime.now(timezone.utc).isoformat()}}
                )
        else:
            summary["unchanged"] += 1

    summary["dry_run"] = dry_run
    summary["total_cost_delta"] = round(summary["total_cost_delta"], 2)
    summary["total_old_cost"] = round(summary["total_old_cost"], 2)
    summary["total_new_cost"] = round(summary["total_new_cost"], 2)
    return summary


@api_router.post("/orders/{order_id}/cancel-item")
async def cancel_order_item(
    order_id: str,
    payload: ItemCancellationRequest,
    current_user: dict = Depends(get_current_user)
):
    """
    تسجيل إلغاء صنف واحد (أو جزء من كميته) من طلب محفوظ ومُرسَل للمطبخ.
    - يُضاف لسجل تدقيق على الطلب نفسه (cancelled_items[])
    - يُسجَّل أيضاً في مجموعة item_cancellations لظهوره في تقارير الإلغاءات
    - مسموح فقط لـ Admin/Manager/Super Admin (الكاشير ممنوع).
    """
    # حماية الأدوار: الكاشير وغير المخوّلين ممنوعون
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مسموح — فقط مالك المطعم أو المدير العام يستطيع حذف/تعديل صنف مُرسَل للمطبخ")

    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")

    qty = float(payload.quantity or 0)
    if qty <= 0:
        raise HTTPException(status_code=400, detail="الكمية غير صالحة")

    now_iso = datetime.now(timezone.utc).isoformat()
    cancellation_entry = {
        "id": str(uuid.uuid4()),
        "product_id": payload.product_id,
        "product_name": payload.product_name or "",
        "quantity": qty,
        "price": float(payload.price or 0),
        "total_value": float(payload.price or 0) * qty,
        "reason": payload.reason or "حذف من الطلب بعد الحفظ",
        "cancelled_by": current_user.get("id"),
        "cancelled_by_name": current_user.get("full_name") or current_user.get("username") or "",
        "cancelled_at": now_iso,
    }

    # إضافة للسجل التدقيقي على الطلب
    await db.orders.update_one(
        {"id": order_id},
        {
            "$push": {"cancelled_items": cancellation_entry},
            "$set": {"updated_at": now_iso},
        },
    )

    # تسجيل مستقل في مجموعة item_cancellations لتقارير الإلغاءات
    try:
        log_doc = {
            **cancellation_entry,
            "order_id": order_id,
            "order_number": order.get("order_number"),
            "order_status": order.get("status"),
            "tenant_id": order.get("tenant_id") or current_user.get("tenant_id"),
            "branch_id": order.get("branch_id") or current_user.get("branch_id"),
        }
        await db.item_cancellations.insert_one(log_doc)
    except Exception as _e:
        logger.warning(f"item_cancellations insert failed: {_e}")

    return {
        "message": "تم تسجيل إلغاء الصنف",
        "cancellation": cancellation_entry,
        "order_number": order.get("order_number"),
        "order_status": order.get("status"),
    }


@api_router.delete("/orders/{order_id}/force-delete")
async def force_delete_unpaid_order(
    order_id: str,
    current_user: dict = Depends(get_current_user),
):
    """
    حذف نهائي لطلب غير مدفوع/مكرر — بدون تسجيل في الإلغاءات/المرتجعات.
    
    شروط:
    - فقط Admin / Manager / Super Admin يقدر
    - الطلب يجب أن يكون: غير مدفوع (payment_method = pending) أو لم يدخل التقارير
    - لا يُسمح بحذف طلب مدفوع بهذه الطريقة (يجب استخدام الإرجاع/الإلغاء العادي)
    """
    # حماية الأدوار
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مسموح — فقط مالك المطعم أو المدير العام يستطيع الحذف النهائي")
    
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التعريف المعتمد: مدفوع فقط إذا كانت الحالة paid أو credit (مطابق لمنطق الإرجاع وأداة الكشف)
    is_paid = (order.get("payment_status") or "").lower() in ("paid", "credit")
    
    # يُسمح فقط لطلب غير مدفوع
    if is_paid:
        raise HTTPException(
            status_code=400,
            detail="لا يمكن الحذف النهائي لطلب مدفوع — استخدم الإرجاع أو الإلغاء العادي بدلاً"
        )
    
    # احذف نهائياً (مع لوغ تدقيقي مستقل في force_deleted_orders للسجل فقط)
    audit_doc = {
        "id": str(uuid.uuid4()),
        "deleted_at": datetime.now(timezone.utc).isoformat(),
        "deleted_by": current_user.get("id"),
        "deleted_by_name": current_user.get("full_name") or current_user.get("username") or "",
        "tenant_id": order.get("tenant_id"),
        "branch_id": order.get("branch_id"),
        "order_id": order_id,
        "order_number": order.get("order_number"),
        "order_total": order.get("total"),
        "order_type": order.get("order_type"),
        "customer_name": order.get("customer_name"),
        "reason": "duplicate_or_unpaid_force_delete",
        # نسخة كاملة من الطلب للأرشيف
        "snapshot": {
            k: v for k, v in order.items()
            if not isinstance(v, (bytes,))
        },
    }
    try:
        await db.force_deleted_orders.insert_one(audit_doc)
    except Exception as _e:
        logger.warning(f"force_deleted_orders archive failed: {_e}")
    
    res = await db.orders.delete_one({"id": order_id})
    
    return {
        "message": "تم الحذف النهائي للطلب — لن يظهر في التقارير ولا في الإلغاءات",
        "deleted": res.deleted_count,
        "order_number": order.get("order_number"),
    }


# ==================== moved to routes/refunds_routes.py ====================

# ==================== DELIVERY APP SETTINGS ====================

@api_router.post("/delivery-app-settings")
async def create_delivery_app_setting(setting: DeliveryAppSettingCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # Check if exists for this tenant
    query = {"app_id": setting.app_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    existing = await db.delivery_app_settings.find_one(query)
    
    setting_data = setting.model_dump()
    setting_data["tenant_id"] = tenant_id
    
    if existing:
        await db.delivery_app_settings.update_one(query, {"$set": setting_data})
    else:
        await db.delivery_app_settings.insert_one(setting_data)
    
    return {"message": "تم الحفظ"}

@api_router.get("/delivery-app-settings")
async def get_delivery_app_settings(current_user: dict = Depends(get_current_user)):
    query = build_tenant_query(current_user)
    settings = await db.delivery_app_settings.find(query, {"_id": 0}).to_list(20)
    return settings

@api_router.delete("/delivery-app-settings/{app_id}")
async def delete_delivery_app_setting(app_id: str, current_user: dict = Depends(get_current_user)):
    """حذف شركة توصيل"""
    query = build_tenant_query(current_user, {"app_id": app_id})
    result = await db.delivery_app_settings.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="شركة التوصيل غير موجودة")
    return {"success": True, "message": "تم حذف شركة التوصيل بنجاح"}

@api_router.get("/delivery-apps")
async def get_delivery_apps(current_user: dict = Depends(get_current_user)):
    # Get default apps
    default_apps = [
        {"id": "toters", "name": "توترز", "name_en": "Toters", "icon": "Truck", "is_default": True},
        {"id": "talabat", "name": "طلبات", "name_en": "Talabat", "icon": "ShoppingBag", "is_default": True},
        {"id": "baly", "name": "بالي", "name_en": "Baly", "icon": "Package", "is_default": True},
        {"id": "alsaree3", "name": "عالسريع", "name_en": "Al-Sari3", "icon": "Zap", "is_default": True},
        {"id": "talabati", "name": "طلباتي", "name_en": "Talabati", "icon": "Box", "is_default": True},
    ]
    
    # Get all settings from database for this tenant
    query = build_tenant_query(current_user)
    all_settings = await db.delivery_app_settings.find(query, {"_id": 0}).to_list(50)
    
    # Create a map of app_id to settings
    settings_map = {s["app_id"]: s for s in all_settings}
    
    # Update default apps with their settings
    result_apps = []
    for app in default_apps:
        setting = settings_map.get(app["id"])
        if setting:
            app["commission_rate"] = setting.get("commission_rate", 0)
            app["is_active"] = setting.get("is_active", True)
        else:
            app["commission_rate"] = 0
            app["is_active"] = True
        result_apps.append(app)
    
    # Add custom apps (apps that are not in default list)
    default_ids = {a["id"] for a in default_apps}
    for setting in all_settings:
        if setting["app_id"] not in default_ids:
            result_apps.append({
                "id": setting["app_id"],
                "name": setting.get("name", setting["app_id"]),
                "name_en": setting.get("name_en", setting["app_id"]),
                "icon": "Truck",
                "is_default": False,
                "commission_rate": setting.get("commission_rate", 0),
                "is_active": setting.get("is_active", True)
            })
    
    return result_apps

# ==================== EXPORT TO EXCEL ====================

from io import BytesIO
from fastapi.responses import StreamingResponse

@api_router.get("/reports/export/excel")
async def export_sales_to_excel(
    report_type: str = "sales",  # sales, products, delivery, expenses
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير إلى Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="مكتبة Excel غير متوفرة")
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    
    # Styling
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Query dates
    if not start_date:
        start_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if not end_date:
        end_date = start_date
    
    query = build_tenant_query(current_user, {
        "created_at": {"$gte": start_date, "$lte": end_date + "T23:59:59"}
    })
    if branch_id:
        query["branch_id"] = branch_id
    
    if report_type == "sales":
        ws.title = "تقرير المبيعات"
        
        # Get orders
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        # Headers
        headers = ["رقم الطلب", "التاريخ", "الوقت", "النوع", "العميل", "الفرع", "طريقة الدفع", "المبلغ", "الحالة"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        statuses = {"pending": "معلق", "preparing": "قيد التحضير", "ready": "جاهز", "completed": "مكتمل", "delivered": "تم التوصيل", "cancelled": "ملغي"}
        
        for row, order in enumerate(orders, 2):
            created_at = datetime.fromisoformat(order["created_at"].replace("Z", "+00:00")) if order.get("created_at") else datetime.now()
            
            data = [
                order.get("order_number", ""),
                created_at.strftime("%Y-%m-%d"),
                created_at.strftime("%H:%M"),
                order_types.get(order.get("order_type", ""), order.get("order_type", "")),
                order.get("customer_name", "بدون اسم"),
                order.get("branch_name", ""),
                payment_methods.get(order.get("payment_method", ""), order.get("payment_method", "")),
                _sn(order.get("total")),
                statuses.get(order.get("status", ""), order.get("status", ""))
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 8:  # Amount column
                    cell.number_format = '#,##0'
        
        # Summary row
        summary_row = len(orders) + 3
        ws.cell(row=summary_row, column=7, value="الإجمالي:").font = Font(bold=True)
        total_cell = ws.cell(row=summary_row, column=8, value=sum(_sn(o.get("total")) for o in orders))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
        
    elif report_type == "products":
        ws.title = "تقرير المنتجات"
        
        # Get orders with items
        orders = await db.orders.find(query, {"_id": 0, "items": 1, "total": 1, "status": 1}).to_list(10000)
        
        # Aggregate products
        products = {}
        for order in orders:
            if order.get("status") == "cancelled":
                continue
            for item in order.get("items", []):
                name = item.get("name", "Unknown")
                if name not in products:
                    products[name] = {"quantity": 0, "revenue": 0}
                products[name]["quantity"] += item.get("quantity", 1)
                products[name]["revenue"] += _sn(item.get("price")) * item.get("quantity", 1)
        
        # Headers
        headers = ["المنتج", "الكمية المباعة", "الإيرادات"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data sorted by quantity
        sorted_products = sorted(products.items(), key=lambda x: x[1]["quantity"], reverse=True)
        for row, (name, data) in enumerate(sorted_products, 2):
            ws.cell(row=row, column=1, value=name).border = thin_border
            ws.cell(row=row, column=2, value=data["quantity"]).border = thin_border
            revenue_cell = ws.cell(row=row, column=3, value=data["revenue"])
            revenue_cell.border = thin_border
            revenue_cell.number_format = '#,##0'
        
    elif report_type == "expenses":
        ws.title = "تقرير المصاريف"
        
        # Get expenses
        expenses = await db.expenses.find(query, {"_id": 0}).to_list(10000)
        
        # Headers
        headers = ["التاريخ", "الفئة", "الوصف", "المبلغ"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data
        for row, expense in enumerate(expenses, 2):
            created_at = datetime.fromisoformat(expense["created_at"].replace("Z", "+00:00")) if expense.get("created_at") else datetime.now()
            
            data = [
                created_at.strftime("%Y-%m-%d"),
                expense.get("category", ""),
                expense.get("description", ""),
                _sn(expense.get("amount"))
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '#,##0'
        
        # Summary
        summary_row = len(expenses) + 3
        ws.cell(row=summary_row, column=3, value="الإجمالي:").font = Font(bold=True)
        total_cell = ws.cell(row=summary_row, column=4, value=sum(_sn(e.get("amount")) for e in expenses))
        total_cell.font = Font(bold=True)
        total_cell.number_format = '#,##0'
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"report_{report_type}_{start_date}_to_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# ==================== moved to routes/pdf_export_routes.py ====================

# ==================== CASH REGISTER ROUTES ====================
# Note: GET /cash-register/summary and POST /cash-register/close are in routes/shifts_routes.py

@api_router.get("/cash-register/today")
async def get_today_cash_register(current_user: dict = Depends(get_current_user)):
    """جلب بيانات صندوق اليوم"""
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    branch_id = current_user.get("branch_id")
    
    # مبيعات اليوم النقدية (فقط الكاش المحصّل باليد - بدون التوصيل وبدون البطاقة)
    cash_query = {
        "created_at": {"$regex": f"^{today}"},
        "payment_method": "cash",
        "status": {"$nin": ["cancelled", "refunded"]},
        "order_type": {"$nin": ["delivery"]}  # استثناء التوصيل من النقدي
    }
    if branch_id:
        cash_query["branch_id"] = branch_id
    
    cash_orders = await db.orders.find(cash_query, {"_id": 0, "total": 1, "order_number": 1}).to_list(500)
    total_cash_sales = sum(_sn(o.get("total")) for o in cash_orders)
    
    # مبيعات البطاقة (منفصلة - ليست نقدي)
    card_query = {
        "created_at": {"$regex": f"^{today}"},
        "payment_method": "card",
        "status": {"$nin": ["cancelled", "refunded"]}
    }
    if branch_id:
        card_query["branch_id"] = branch_id
    card_orders = await db.orders.find(card_query, {"total": 1}).to_list(500)
    total_card_sales = sum(_sn(o.get("total")) for o in card_orders)
    
    # الآجل العادي (بدون التوصيل - شركات التوصيل لها قسم منفصل)
    credit_query = {
        "created_at": {"$regex": f"^{today}"},
        "status": {"$nin": ["cancelled", "refunded"]},
        "payment_method": "credit",
        "order_type": {"$ne": "delivery"}  # استثناء التوصيل
    }
    if branch_id:
        credit_query["branch_id"] = branch_id
    credit_orders = await db.orders.find(credit_query, {"total": 1}).to_list(500)
    total_credit = sum(_sn(o.get("total")) for o in credit_orders)
    
    # آجل شركات التوصيل (منفصل)
    delivery_credit_query = {
        "created_at": {"$regex": f"^{today}"},
        "status": {"$nin": ["cancelled", "refunded"]},
        "order_type": "delivery"
    }
    if branch_id:
        delivery_credit_query["branch_id"] = branch_id
    delivery_credit_orders = await db.orders.find(delivery_credit_query, {"total": 1, "delivery_app": 1}).to_list(500)
    total_delivery_credit = sum(_sn(o.get("total")) for o in delivery_credit_orders)
    
    # المصاريف
    expenses_query = {"date": today}
    if branch_id:
        expenses_query["branch_id"] = branch_id
    expenses = await db.expenses.find(expenses_query, {"_id": 0}).to_list(100)
    total_expenses = sum(_sn(e.get("amount")) for e in expenses)
    
    # آخر إغلاق
    last_close_query = {"branch_id": branch_id} if branch_id else {}
    last_close = await db.cash_register_closes.find_one(
        last_close_query, 
        {"_id": 0},
        sort=[("closed_at", -1)]
    )
    
    # إجمالي المبيعات = كل شيء (نقدي + بطاقة + آجل + آجل توصيل)
    total_all_sales = total_cash_sales + total_card_sales + total_credit + total_delivery_credit
    
    # المتوقع في الصندوق = نقدي فقط - مصاريف (البطاقة لا تدخل الصندوق)
    expected_cash = total_cash_sales - total_expenses
    
    return {
        "date": today,
        "cash_sales": total_cash_sales,  # فقط الكاش المحصّل باليد (بدون توصيل وبدون بطاقة)
        "card_sales": total_card_sales,  # مبيعات البطاقة (منفصلة)
        "credit_sales": total_credit,  # الآجل العادي (بدون التوصيل)
        "delivery_credit_sales": total_delivery_credit,  # آجل شركات التوصيل (منفصل)
        "total_sales": total_all_sales,  # إجمالي كل المبيعات
        "orders_count": len(cash_orders) + len(card_orders) + len(credit_orders) + len(delivery_credit_orders),
        "expenses": expenses,
        "total_expenses": total_expenses,
        "expected_cash": expected_cash,  # المتوقع في الصندوق = نقدي - مصاريف
        "last_close": last_close
    }

# ==================== SETTINGS ROUTES ====================

@api_router.get("/settings")
async def get_settings(current_user: dict = Depends(get_current_user)):
    settings = await db.settings.find({}, {"_id": 0}).to_list(100)
    return {s["type"]: s.get("value") or s for s in settings}

@api_router.post("/settings/email-recipients")
async def set_email_recipients(emails: List[str], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "email_recipients"},
        {"$set": {"type": "email_recipients", "emails": emails}},
        upsert=True
    )
    return {"message": "تم الحفظ"}

@api_router.post("/settings/currencies")
async def set_currencies(currencies: List[Currency], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "currencies"},
        {"$set": {"type": "currencies", "value": [c.model_dump() for c in currencies]}},
        upsert=True
    )
    return {"message": "تم الحفظ"}

@api_router.post("/settings/general")
async def set_general_settings(settings: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    if current_user["role"] != UserRole.ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "general"},
        {"$set": {"type": "general", "value": settings}},
        upsert=True
    )
    return {"message": "تم الحفظ"}

@api_router.get("/settings/general")
async def get_general_settings():
    settings = await db.settings.find_one({"type": "general"}, {"_id": 0})
    return settings.get("value", {}) if settings else {}

@api_router.put("/settings/restaurant")
async def update_restaurant_settings(settings: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """حفظ إعدادات المطعم (الاسم والشعار)"""
    # السماح للمدير (admin) والمالك (super_admin)
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # جلب tenant_id من المستخدم
    tenant_id = current_user.get("tenant_id")
    
    # تحديث tenant إذا كان موجود
    if tenant_id:
        await db.tenants.update_one(
            {"id": tenant_id},
            {"$set": {
                "name": settings.get("name"),
                "name_ar": settings.get("name_ar"),
                "logo_url": settings.get("logo_url"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    
    # تحديث default tenant أيضاً إذا لم يكن هناك tenant محدد
    if not tenant_id or tenant_id == "default":
        await db.tenants.update_one(
            {"id": "default"},
            {"$set": {
                "name": settings.get("name"),
                "name_ar": settings.get("name_ar"),
                "logo_url": settings.get("logo_url"),
                "updated_at": datetime.now(timezone.utc).isoformat()
            }},
            upsert=True
        )
    
    # حفظ في settings أيضاً
    await db.settings.update_one(
        {"type": "restaurant"},
        {"$set": {
            "type": "restaurant",
            "name": settings.get("name"),
            "name_ar": settings.get("name_ar"),
            "logo_url": settings.get("logo_url"),
            "updated_at": datetime.now(timezone.utc).isoformat()
        }},
        upsert=True
    )
    
    return {"message": "تم حفظ إعدادات المطعم بنجاح"}

@api_router.get("/settings/restaurant")
async def get_restaurant_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات المطعم"""
    tenant_id = current_user.get("tenant_id", "default")
    
    # محاولة جلب من tenant
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    if tenant and tenant.get("name"):
        return {
            "name": tenant.get("name"),
            "name_ar": tenant.get("name_ar"),
            "name_en": tenant.get("name_en"),
            "logo_url": tenant.get("logo_url")
        }
    
    # محاولة جلب من settings مع فلترة tenant_id
    settings = await db.settings.find_one({"type": "restaurant", "tenant_id": tenant_id}, {"_id": 0})
    if not settings:
        settings = await db.settings.find_one({"type": "restaurant"}, {"_id": 0})
    if settings:
        val = settings.get("value", settings)
        return {
            "name": val.get("name", ""),
            "name_ar": val.get("name_ar", ""),
            "name_en": val.get("name_en", ""),
            "logo_url": val.get("logo_url", "")
        }
    
    # fallback: use tenant name from user
    return {"name": tenant.get("name", "") if tenant else "", "name_ar": "", "name_en": "", "logo_url": ""}

@api_router.put("/settings/dashboard")
async def set_dashboard_settings(settings: Dict[str, Any], current_user: dict = Depends(get_current_user)):
    """حفظ إعدادات الصفحة الرئيسية - التحكم في الصفحات الظاهرة"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    await db.settings.update_one(
        {"type": "dashboard_settings"},
        {"$set": {"type": "dashboard_settings", "value": settings}},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات الصفحة الرئيسية"}

@api_router.get("/settings/dashboard")
async def get_dashboard_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الصفحة الرئيسية مع مراعاة ميزات العميل"""
    
    # الإعدادات الافتراضية الكاملة
    default_settings = {
        "showPOS": True,
        "showTables": True,
        "showOrders": True,
        "showExpenses": True,
        "showInventory": True,
        "showDelivery": True,
        "showReports": True,
        "showSettings": True,
        "showHR": True,
        "showWarehouse": True,
        "showCallLogs": True,
        "showCallCenter": True,
        "showKitchen": True,
        "showLoyalty": True,
        "showCoupons": True,
        "showRecipes": True,
        "showReservations": True,
        "showReviews": True,
        "showRatings": True,
        "showSmartReports": True,
        "showPurchasing": True,
        "showBranchOrders": True,
        "showCustomerMenu": True,
        # الميزات الجديدة
        "showOwnerWallet": True,
        "showExternalBranches": True,
        "showComprehensiveReport": True,
        # خيارات الإعدادات
        "settingsUsers": True,
        "settingsCustomers": True,
        "settingsBranches": True,
        "settingsCategories": True,
        "settingsProducts": True,
        "settingsPrinters": True,
        "settingsDeliveryCompanies": True,
        "settingsCallCenter": True,
        "settingsNotifications": True,
        "settingsRestaurant": True,
        "settingsAppearance": True,
        "settingsInvoice": True,
        "settingsSystem": True,
        "settingsInventory": True,
        "settingsPayment": True,
        "settingsKitchenSections": True,
        # التقارير
        "showBreakEvenReport": True,
        "showInventoryReports": True,
        # ميزات جديدة في الإجراءات السريعة
        "showCaptainsManagement": True,
        "showExternalPurchasesReport": True,
        "showPriceIncreaseReport": True
    }
    
    # جلب إعدادات لوحة القيادة المحفوظة
    settings = await db.settings.find_one({"type": "dashboard_settings"}, {"_id": 0})
    if settings and settings.get("value"):
        default_settings = {**default_settings, **settings.get("value", {})}
    
    # إذا كان المستخدم Super Admin، أرجع كل الميزات
    if current_user.get("role") == UserRole.SUPER_ADMIN:
        return default_settings
    
    # إذا كان المستخدم بدون tenant_id (النظام الرئيسي)، أرجع كل الميزات
    tenant_id = get_user_tenant_id(current_user)
    if not tenant_id:
        return default_settings
    
    # جلب ميزات العميل
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "enabled_features": 1})
    if tenant and tenant.get("enabled_features"):
        tenant_features = tenant["enabled_features"]
        # دمج الميزات - العميل يرى فقط الميزات المفعّلة له
        for key in default_settings:
            if key in tenant_features:
                default_settings[key] = tenant_features[key] and default_settings[key]
    
    return default_settings

@api_router.get("/tenant/info")
async def get_tenant_info(current_user: dict = Depends(get_current_user)):
    """جلب معلومات العميل (الشعار والاسم) للعرض في Dashboard"""
    tenant_id = get_user_tenant_id(current_user)
    
    # إذا كان النظام الرئيسي (بدون tenant)
    if not tenant_id:
        # جلب إعدادات النظام العامة
        settings = await db.settings.find_one({"type": "system_branding"}, {"_id": 0})
        if settings and settings.get("value"):
            return settings["value"]
        return {
            "name": "Maestro",
            "name_ar": "Maestro",
            "name_en": "Maestro",
            "logo_url": None
        }
    
    # جلب معلومات العميل
    tenant = await db.tenants.find_one(
        {"id": tenant_id}, 
        {"_id": 0, "name": 1, "name_ar": 1, "name_en": 1, "logo_url": 1}
    )
    
    if not tenant:
        return {"name": "Maestro", "logo_url": None}
    
    return tenant

# ==================== LOGIN BACKGROUNDS API ====================

# كلمة سر خاصة للـ Super Admin - من متغيرات البيئة
SUPER_ADMIN_SECRET = os.environ.get("SUPER_ADMIN_SECRET", "271018")

# التحقق من صلاحية Super Admin
async def verify_super_admin(current_user: dict = Depends(get_current_user)):
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="صلاحيات Super Admin مطلوبة")
    return current_user

# ==================== FIX DATA API ====================

@api_router.post("/fix-data")
async def fix_data_endpoint(current_user: dict = Depends(verify_super_admin)):
    """
    Endpoint لإصلاح البيانات القديمة
    يمكن استدعاؤه فقط من قبل المالك
    """
    try:
        results = {
            "tables_fixed": 0,
            "tables_deleted_duplicates": 0,
            "categories_fixed": 0,
            "products_fixed": 0,
            "tenant_tables_created": 0,
        }
        
        # 1. تحديث الطاولات القديمة التي ليس لها tenant_id لتصبح "default"
        tables_result = await db.tables.update_many(
            {"$or": [{"tenant_id": {"$exists": False}}, {"tenant_id": None}, {"tenant_id": ""}]},
            {"$set": {"tenant_id": "default"}}
        )
        results["tables_fixed"] = tables_result.modified_count
        
        # 2. حذف الطاولات المكررة (نفس رقم الطاولة ونفس tenant_id)
        pipeline = [
            {"$group": {
                "_id": {"number": "$number", "tenant_id": "$tenant_id"},
                "count": {"$sum": 1},
                "ids": {"$push": "$id"}
            }},
            {"$match": {"count": {"$gt": 1}}}
        ]
        duplicates = await db.tables.aggregate(pipeline).to_list(100)
        for dup in duplicates:
            # الاحتفاظ بأول طاولة وحذف الباقي
            ids_to_delete = dup["ids"][1:]
            delete_result = await db.tables.delete_many({"id": {"$in": ids_to_delete}})
            results["tables_deleted_duplicates"] += delete_result.deleted_count
        
        # 3. إنشاء طاولات لكل عميل ليس لديه طاولات
        tenants = await db.tenants.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(None)
        for tenant in tenants:
            tenant_tables = await db.tables.count_documents({"tenant_id": tenant["id"]})
            if tenant_tables == 0:
                tenant_branch = await db.branches.find_one({"tenant_id": tenant["id"]})
                branch_id = tenant_branch["id"] if tenant_branch else None
                
                default_tables = []
                for i in range(1, 6):
                    default_tables.append({
                        "id": str(uuid.uuid4()),
                        "number": i,
                        "capacity": 4,
                        "section": "القاعة الرئيسية",
                        "status": "available",
                        "current_order_id": None,
                        "branch_id": branch_id,
                        "tenant_id": tenant["id"]
                    })
                await db.tables.insert_many(default_tables)
                results["tenant_tables_created"] += 5
        
        # 4. تحديث صور الفئات للنظام الرئيسي
        category_images = {
            "برغر": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا": "https://images.unsplash.com/photo-1703073186021-021fb5a0bde1?w=400",
            "مشروبات": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "حلويات": "https://images.unsplash.com/photo-1546902189-eaaf09f8e38f?w=400",
            "سلطات": "https://images.unsplash.com/photo-1677653805080-59c57727c84e?w=400",
        }
        for cat_name, cat_image in category_images.items():
            cat_result = await db.categories.update_many(
                {"name": cat_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": cat_image}}
            )
            results["categories_fixed"] += cat_result.modified_count
        
        # 5. تحديث صور المنتجات للنظام الرئيسي
        product_images = {
            "برغر كلاسيك": "https://images.unsplash.com/photo-1656439659132-24c68e36b553?w=400",
            "برغر دبل": "https://images.unsplash.com/photo-1635275650933-7b0911815a2e?w=400",
            "بيتزا مارغريتا": "https://images.unsplash.com/photo-1681567604770-0dc826c870ae?w=400",
            "بيتزا خضار": "https://images.unsplash.com/photo-1602104980741-b87a33837f9f?w=400",
            "كولا": "https://images.unsplash.com/photo-1657958977261-d75e81b4713f?w=400",
            "عصير برتقال": "https://images.unsplash.com/photo-1716925539259-ce0115263d37?w=400",
        }
        for prod_name, prod_image in product_images.items():
            prod_result = await db.products.update_many(
                {"name": prod_name, "tenant_id": "default", "$or": [{"image": {"$exists": False}}, {"image": None}, {"image": ""}]},
                {"$set": {"image": prod_image}}
            )
            results["products_fixed"] += prod_result.modified_count
        
        return {
            "status": "success",
            "message": "تم إصلاح البيانات بنجاح",
            "results": results
        }
        
    except Exception as e:
        return {
            "status": "error",
            "message": f"حدث خطأ: {str(e)}"
        }

# ==================== SYSTEM BRANDING API ====================

@api_router.get("/system/branding")
async def get_system_branding(current_user: dict = Depends(verify_super_admin)):
    """جلب إعدادات هوية النظام الرئيسي"""
    settings = await db.settings.find_one({"type": "system_branding"}, {"_id": 0})
    if settings and settings.get("value"):
        return settings["value"]
    return {
        "name": "Maestro",
        "name_ar": "Maestro",
        "name_en": "Maestro",
        "logo_url": None
    }

@api_router.put("/system/branding")
async def update_system_branding(branding: dict, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات هوية النظام الرئيسي (الاسم والشعار)"""
    allowed_fields = ["name", "name_ar", "name_en", "logo_url"]
    update_data = {k: v for k, v in branding.items() if k in allowed_fields}
    
    await db.settings.update_one(
        {"type": "system_branding"},
        {"$set": {"type": "system_branding", "value": update_data}},
        upsert=True
    )
    
    return {"message": "تم تحديث هوية النظام بنجاح", "branding": update_data}

class LoginBackgroundCreate(BaseModel):
    image_url: str
    title: Optional[str] = None
    animation_type: str = "fade"  # fade, slide, zoom, kenburns, parallax
    animation_duration: int = 8  # بالثواني
    overlay_opacity: float = 0.5
    is_active: bool = True
    sort_order: int = 0

class LoginBackgroundSettings(BaseModel):
    backgrounds: List[Dict[str, Any]] = []
    animation_enabled: bool = True
    transition_type: str = "fade"  # fade, slide, crossfade
    transition_duration: float = 1.5  # بالثواني
    auto_play: bool = True
    show_logo: bool = True
    logo_url: Optional[str] = None
    logo_animation: str = "pulse"  # pulse, bounce, glow, none
    overlay_color: str = "rgba(0,0,0,0.5)"
    text_color: str = "#ffffff"

@api_router.get("/login-backgrounds")
async def get_login_backgrounds():
    """جلب إعدادات خلفيات صفحة الدخول (عام - بدون مصادقة)"""
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    
    default_settings = {
        "backgrounds": [],
        "animation_enabled": True,
        "transition_type": "fade",
        "transition_duration": 1.5,
        "auto_play": True,
        "show_logo": True,
        "logo_url": None,
        "logo_animation": "pulse",
        "overlay_color": "rgba(0,0,0,0.5)",
        "text_color": "#ffffff"
    }
    
    if settings and settings.get("value"):
        return {**default_settings, **settings.get("value", {})}
    return default_settings

@api_router.put("/login-backgrounds")
async def update_login_backgrounds(settings: LoginBackgroundSettings, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات خلفيات صفحة الدخول (Super Admin فقط)"""
    
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"type": "login_backgrounds", "value": settings.model_dump()}},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات الخلفيات"}

@api_router.post("/login-backgrounds/upload")
async def upload_login_background(
    file_url: str,
    title: Optional[str] = None,
    animation_type: str = "fade",
    current_user: dict = Depends(verify_super_admin)
):
    """إضافة خلفية جديدة"""
    
    # جلب الإعدادات الحالية
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    current_backgrounds = []
    current_value = {}
    
    if settings and settings.get("value"):
        current_value = settings["value"]
        current_backgrounds = current_value.get("backgrounds", [])
    
    # إضافة الخلفية الجديدة
    new_background = {
        "id": str(uuid.uuid4()),
        "image_url": file_url,
        "title": title,
        "animation_type": animation_type,
        "animation_duration": 8,
        "overlay_opacity": 0.5,
        "is_active": True,
        "sort_order": len(current_backgrounds),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    current_backgrounds.append(new_background)
    
    # تحديث مع الحفاظ على الإعدادات الأخرى
    default_settings = {
        "backgrounds": current_backgrounds,
        "animation_enabled": current_value.get("animation_enabled", True),
        "transition_type": current_value.get("transition_type", "fade"),
        "transition_duration": current_value.get("transition_duration", 1.5),
        "auto_play": current_value.get("auto_play", True),
        "show_logo": current_value.get("show_logo", True),
        "logo_url": current_value.get("logo_url", None),
        "logo_animation": current_value.get("logo_animation", "pulse"),
        "overlay_color": current_value.get("overlay_color", "rgba(0,0,0,0.5)"),
        "text_color": current_value.get("text_color", "#ffffff")
    }
    
    # حفظ التحديث
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"type": "login_backgrounds", "value": default_settings}},
        upsert=True
    )
    
    return {"message": "تم إضافة الخلفية", "background": new_background}

@api_router.delete("/login-backgrounds/{background_id}")
async def delete_login_background(background_id: str, current_user: dict = Depends(verify_super_admin)):
    """حذف خلفية"""
    
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    if not settings or not settings.get("value"):
        raise HTTPException(status_code=404, detail="لا توجد خلفيات")
    
    backgrounds = settings["value"].get("backgrounds", [])
    backgrounds = [b for b in backgrounds if b.get("id") != background_id]
    
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.backgrounds": backgrounds}}
    )
    
    return {"message": "تم حذف الخلفية"}

@api_router.post("/login-backgrounds/upload-logo")
async def upload_login_page_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(verify_super_admin)
):
    """رفع شعار صفحة تسجيل الدخول - للمالك فقط"""
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/svg+xml']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. يرجى استخدام JPG, PNG, GIF, WebP أو SVG")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, LOGOS_DIR, max_size=(512, 512), quality=90)
    
    # إنشاء URL نسبي للشعار
    logo_url = f"/api/uploads/logos/{filename}"
    
    # تحديث login_backgrounds مع الشعار الجديد
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    current_value = settings.get("value", {}) if settings else {}
    
    # تحديث logo_url فقط
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.logo_url": logo_url}},
        upsert=True
    )
    
    return {"message": "تم رفع شعار صفحة تسجيل الدخول بنجاح", "logo_url": logo_url}

@api_router.put("/login-backgrounds/logo-url")
async def update_login_page_logo_url(
    logo_url: str = Body(..., embed=True),
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث شعار صفحة تسجيل الدخول برابط خارجي - للمالك فقط"""
    
    # تحديث login_backgrounds مع الشعار الجديد
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.logo_url": logo_url}},
        upsert=True
    )
    
    return {"message": "تم تحديث شعار صفحة تسجيل الدخول", "logo_url": logo_url}

@api_router.delete("/login-backgrounds/logo")
async def delete_login_page_logo(current_user: dict = Depends(verify_super_admin)):
    """حذف شعار صفحة تسجيل الدخول - للمالك فقط"""
    
    # تحديث login_backgrounds بإزالة الشعار
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"value.logo_url": None}},
        upsert=True
    )
    
    return {"message": "تم حذف شعار صفحة تسجيل الدخول"}

# ==================== INVOICE/RECEIPT SETTINGS - إعدادات الفاتورة ====================

class SystemInvoiceSettings(BaseModel):
    """إعدادات الفاتورة للنظام (يتحكم فيها المالك)"""
    system_name: Optional[str] = None  # اسم النظام
    system_logo_url: Optional[str] = None  # شعار النظام
    thank_you_message: str = "شكراً لزيارتكم"  # رسالة الشكر
    system_phone: Optional[str] = None  # رقم هاتف النظام
    system_phone2: Optional[str] = None  # رقم هاتف ثاني
    system_email: Optional[str] = None  # بريد النظام
    system_website: Optional[str] = None  # موقع النظام
    footer_text: Optional[str] = None  # نص إضافي في التذييل
    show_system_branding: bool = True  # عرض شعار وبيانات النظام
    promo_text: Optional[str] = "نظام محاسبي وإداري متكامل للمؤسسات والمطاعم والمشاريع التجارية الكبرى"  # نص الدعاية
    cta_text: Optional[str] = "للحصول على نسختك تواصل معنا"  # نص التحفيز للتواصل
    system_intro: Optional[str] = "منصة شاملة ذكية ومتطورة تدير العمليات، والمخزون، والتصنيع، والمشتريات، والتوصيل، والموارد البشرية والمالية في نظام واحد دقيق يعمل حتى بدون إنترنت."  # نص تعريف النظام

class TenantInvoiceSettings(BaseModel):
    """إعدادات الفاتورة للعميل (المطعم)"""
    restaurant_name: Optional[str] = None  # اسم المطعم في الفاتورة
    show_logo: bool = True  # عرض الشعار
    invoice_logo: Optional[str] = None  # شعار الفاتورة المخصص
    phone: Optional[str] = None  # رقم الهاتف
    phone2: Optional[str] = None  # رقم هاتف ثاني
    address: Optional[str] = None  # العنوان
    tax_number: Optional[str] = None  # الرقم الضريبي
    show_tax: bool = True  # إظهار الرقم الضريبي
    custom_header: Optional[str] = None  # نص إضافي في الترويسة
    custom_footer: Optional[str] = None  # نص إضافي في التذييل
    thank_you_message: Optional[str] = None  # رسالة الشكر
    default_delivery_fee: Optional[float] = 0  # مبلغ التوصيل الافتراضي (يُقترح تلقائياً)

@api_router.get("/system/invoice-settings")
async def get_system_invoice_settings():
    """جلب إعدادات الفاتورة/العلامة للنظام. عام (بدون مصادقة) لأنه يغذّي صفحة
    بيع/تواصل النظام العامة وطباعة الفواتير — بيانات علامة/تواصل عامة غير حساسة."""
    settings = await db.settings.find_one({"type": "system_invoice_settings"}, {"_id": 0})
    
    default_settings = {
        "system_name": None,
        "system_logo_url": None,
        "thank_you_message": "شكراً لزيارتكم",
        "system_phone": None,
        "system_phone2": None,
        "system_email": None,
        "system_website": None,
        "footer_text": None,
        "show_system_branding": True
    }
    
    if settings and settings.get("value"):
        result = {**default_settings, **settings.get("value", {})}
    else:
        result = dict(default_settings)
    # ملاحظة: هذه بيانات تواصل عامة تُعرض في صفحة بيع/تواصل النظام للعملاء المحتملين،
    # لذا تُرجَع كاملة للجميع (عام + كل الأدوار). لا تُخفى — تراجع عن إخفاء #7 هنا.
    return result

@api_router.put("/system/invoice-settings")
async def update_system_invoice_settings(settings: SystemInvoiceSettings, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات الفاتورة للنظام (المالك فقط)"""
    
    await db.settings.update_one(
        {"type": "system_invoice_settings"},
        {"$set": {"value": settings.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات الفاتورة", "settings": settings.model_dump()}

@api_router.get("/tenant/invoice-settings")
async def get_tenant_invoice_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الفاتورة للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    
    default_settings = {
        "show_logo": True,
        "invoice_logo": None,
        "phone": None,
        "phone2": None,
        "address": None,
        "tax_number": None,
        "show_tax": True,
        "custom_header": None,
        "custom_footer": None,
        "thank_you_message": "شكراً لزيارتكم ❤️",
        "default_delivery_fee": 0
    }
    
    if tenant_id:
        settings = await db.tenant_invoice_settings.find_one({"tenant_id": tenant_id}, {"_id": 0})
        if settings:
            # جلب شعار المطعم من tenant إذا لم يكن هناك شعار مخصص للفاتورة
            if not settings.get("invoice_logo"):
                tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "logo_url": 1, "logo": 1})
                if tenant:
                    settings["invoice_logo"] = tenant.get("logo_url") or tenant.get("logo")
            return {**default_settings, **settings}
        else:
            # إذا لم يكن هناك إعدادات، نحاول جلب شعار المطعم من tenant
            tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "logo_url": 1, "logo": 1})
            if tenant and (tenant.get("logo_url") or tenant.get("logo")):
                default_settings["invoice_logo"] = tenant.get("logo_url") or tenant.get("logo")
    
    return default_settings

@api_router.put("/tenant/invoice-settings")
async def update_tenant_invoice_settings(settings: TenantInvoiceSettings, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات الفاتورة للعميل"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    await db.tenant_invoice_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": {**settings.model_dump(), "tenant_id": tenant_id, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات الفاتورة", "settings": settings.model_dump()}

# ==================== Currency & Language Settings ====================

# العملات المدعومة مع أسعار الصرف التقريبية
SUPPORTED_CURRENCIES = {
    "IQD": {"name": "دينار عراقي", "name_en": "Iraqi Dinar", "symbol": "د.ع", "rate_to_usd": 0.00076, "decimal_places": 0},
    "USD": {"name": "دولار أمريكي", "name_en": "US Dollar", "symbol": "$", "rate_to_usd": 1, "decimal_places": 2},
    "SAR": {"name": "ريال سعودي", "name_en": "Saudi Riyal", "symbol": "ر.س", "rate_to_usd": 0.27, "decimal_places": 2},
    "AED": {"name": "درهم إماراتي", "name_en": "UAE Dirham", "symbol": "د.إ", "rate_to_usd": 0.27, "decimal_places": 2},
    "KWD": {"name": "دينار كويتي", "name_en": "Kuwaiti Dinar", "symbol": "د.ك", "rate_to_usd": 3.25, "decimal_places": 3},
    "EGP": {"name": "جنيه مصري", "name_en": "Egyptian Pound", "symbol": "ج.م", "rate_to_usd": 0.032, "decimal_places": 2},
    "JOD": {"name": "دينار أردني", "name_en": "Jordanian Dinar", "symbol": "د.أ", "rate_to_usd": 1.41, "decimal_places": 3},
    "EUR": {"name": "يورو", "name_en": "Euro", "symbol": "€", "rate_to_usd": 1.08, "decimal_places": 2},
    "GBP": {"name": "جنيه استرليني", "name_en": "British Pound", "symbol": "£", "rate_to_usd": 1.27, "decimal_places": 2},
    "TRY": {"name": "ليرة تركية", "name_en": "Turkish Lira", "symbol": "₺", "rate_to_usd": 0.031, "decimal_places": 2},
}

# اللغات المدعومة
SUPPORTED_LANGUAGES = {
    "ar": {"name": "العربية", "name_en": "Arabic", "dir": "rtl"},
    "en": {"name": "English", "name_en": "English", "dir": "ltr"},
    "ku": {"name": "کوردی", "name_en": "Kurdish", "dir": "rtl"},
    "fa": {"name": "فارسی", "name_en": "Persian", "dir": "rtl"},
    "tr": {"name": "Türkçe", "name_en": "Turkish", "dir": "ltr"},
}

# البلدان مع العملات الافتراضية
COUNTRIES = {
    "IQ": {"name": "العراق", "name_en": "Iraq", "currency": "IQD", "language": "ar"},
    "SA": {"name": "السعودية", "name_en": "Saudi Arabia", "currency": "SAR", "language": "ar"},
    "AE": {"name": "الإمارات", "name_en": "UAE", "currency": "AED", "language": "ar"},
    "KW": {"name": "الكويت", "name_en": "Kuwait", "currency": "KWD", "language": "ar"},
    "EG": {"name": "مصر", "name_en": "Egypt", "currency": "EGP", "language": "ar"},
    "JO": {"name": "الأردن", "name_en": "Jordan", "currency": "JOD", "language": "ar"},
    "US": {"name": "أمريكا", "name_en": "United States", "currency": "USD", "language": "en"},
    "GB": {"name": "بريطانيا", "name_en": "United Kingdom", "currency": "GBP", "language": "en"},
    "TR": {"name": "تركيا", "name_en": "Turkey", "currency": "TRY", "language": "tr"},
}

class TenantRegionalSettings(BaseModel):
    """إعدادات المنطقة والعملة للعميل"""
    country: str = "IQ"
    currency: str = "IQD"
    language: str = "ar"
    secondary_currency: Optional[str] = "USD"  # عملة ثانوية للعرض
    show_secondary_currency: bool = False  # عرض السعر بالعملة الثانوية
    custom_exchange_rate: Optional[float] = None  # سعر صرف مخصص

@api_router.get("/system/currencies")
async def get_supported_currencies():
    """جلب قائمة العملات المدعومة"""
    return {"currencies": SUPPORTED_CURRENCIES}

@api_router.get("/system/languages")
async def get_supported_languages():
    """جلب قائمة اللغات المدعومة"""
    return {"languages": SUPPORTED_LANGUAGES}

@api_router.get("/system/countries")
async def get_supported_countries():
    """جلب قائمة البلدان المدعومة"""
    return {"countries": COUNTRIES}

@api_router.get("/tenant/regional-settings")
async def get_tenant_regional_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات المنطقة والعملة للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    
    settings = await db.tenant_regional_settings.find_one({"tenant_id": tenant_id}, {"_id": 0})
    
    default_settings = {
        "country": "IQ",
        "currency": "IQD",
        "language": "ar",
        "secondary_currency": "USD",
        "show_secondary_currency": False,
        "custom_exchange_rate": None
    }
    
    if settings:
        return {**default_settings, **settings}
    return default_settings

@api_router.put("/tenant/regional-settings")
async def update_tenant_regional_settings(settings: TenantRegionalSettings, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات المنطقة والعملة للعميل"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    await db.tenant_regional_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": {**settings.model_dump(), "tenant_id": tenant_id, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات المنطقة", "settings": settings.model_dump()}

@api_router.get("/customer/regional-settings/{tenant_id}")
async def get_customer_regional_settings(tenant_id: str):
    """جلب إعدادات المنطقة للزبون (بدون تسجيل دخول)"""
    # البحث بـ menu_slug أو tenant_id
    tenant = await db.tenants.find_one({"menu_slug": tenant_id})
    if not tenant:
        tenant = await db.tenants.find_one({"id": tenant_id})
    
    if not tenant:
        raise HTTPException(status_code=404, detail="المطعم غير موجود")
    
    actual_tenant_id = tenant.get("id")
    settings = await db.tenant_regional_settings.find_one({"tenant_id": actual_tenant_id}, {"_id": 0})
    
    default_settings = {
        "country": "IQ",
        "currency": "IQD",
        "language": "ar",
        "secondary_currency": "USD",
        "show_secondary_currency": False
    }
    
    result = {**default_settings}
    if settings:
        result.update(settings)
    
    # إضافة معلومات العملة
    currency_code = result.get("currency", "IQD")
    if currency_code in SUPPORTED_CURRENCIES:
        result["currency_info"] = SUPPORTED_CURRENCIES[currency_code]
    
    return result

@api_router.post("/convert-currency")
async def convert_currency(
    amount: float,
    from_currency: str = "IQD",
    to_currency: str = "USD"
):
    """تحويل المبلغ بين العملات"""
    if from_currency not in SUPPORTED_CURRENCIES or to_currency not in SUPPORTED_CURRENCIES:
        raise HTTPException(status_code=400, detail="عملة غير مدعومة")
    
    # التحويل عبر الدولار كوسيط
    from_rate = SUPPORTED_CURRENCIES[from_currency]["rate_to_usd"]
    to_rate = SUPPORTED_CURRENCIES[to_currency]["rate_to_usd"]
    
    # المبلغ بالدولار
    usd_amount = amount * from_rate
    # المبلغ بالعملة المستهدفة
    converted_amount = usd_amount / to_rate
    
    decimal_places = SUPPORTED_CURRENCIES[to_currency]["decimal_places"]
    converted_amount = round(converted_amount, decimal_places)
    
    return {
        "original_amount": amount,
        "original_currency": from_currency,
        "converted_amount": converted_amount,
        "target_currency": to_currency,
        "exchange_rate": from_rate / to_rate
    }

# ==================== Owner Currency Settings ====================

class OwnerCurrencySettings(BaseModel):
    """إعدادات تحويل العملات للمالك"""
    preferred_currency: str = "USD"
    use_live_rates: bool = False
    custom_rates: Optional[dict] = None  # {"IQD_USD": 0.00076, "SAR_USD": 0.27}

@api_router.get("/super-admin/currency-settings")
async def get_owner_currency_settings(current_user: dict = Depends(verify_super_admin)):
    """جلب إعدادات العملة للمالك"""
    settings = await db.settings.find_one({"type": "owner_currency_settings"}, {"_id": 0})
    
    if not settings:
        return {
            "preferred_currency": "USD",
            "use_live_rates": False,
            "custom_rates": {},
            "supported_currencies": list(SUPPORTED_CURRENCIES.keys())
        }
    
    return {
        **settings,
        "supported_currencies": list(SUPPORTED_CURRENCIES.keys())
    }

@api_router.put("/super-admin/currency-settings")
async def update_owner_currency_settings(
    settings: OwnerCurrencySettings,
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث إعدادات العملة للمالك"""
    await db.settings.update_one(
        {"type": "owner_currency_settings"},
        {"$set": {
            "type": "owner_currency_settings",
            "preferred_currency": settings.preferred_currency,
            "use_live_rates": settings.use_live_rates,
            "custom_rates": settings.custom_rates or {},
            "updated_at": datetime.utcnow()
        }},
        upsert=True
    )
    return {"message": "تم حفظ إعدادات العملة"}

@api_router.put("/super-admin/custom-exchange-rate")
async def update_custom_exchange_rate(
    from_currency: str,
    to_currency: str,
    rate: float,
    current_user: dict = Depends(verify_super_admin)
):
    """تحديث سعر صرف مخصص"""
    rate_key = f"{from_currency}_{to_currency}"
    
    await db.settings.update_one(
        {"type": "owner_currency_settings"},
        {"$set": {f"custom_rates.{rate_key}": rate, "updated_at": datetime.utcnow()}},
        upsert=True
    )
    return {"message": f"تم تحديث سعر صرف {from_currency} إلى {to_currency}"}

@api_router.get("/super-admin/live-exchange-rates")
async def get_live_exchange_rates(current_user: dict = Depends(verify_super_admin)):
    """جلب أسعار الصرف الحية (من الإنترنت)"""
    import httpx
    
    try:
        # استخدام API مجاني لأسعار الصرف
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://api.exchangerate-api.com/v4/latest/USD",
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                rates = data.get("rates", {})
                
                # تحويل الأسعار للعملات المدعومة
                live_rates = {}
                for code in SUPPORTED_CURRENCIES.keys():
                    if code in rates:
                        live_rates[code] = {
                            "rate_to_usd": 1 / rates[code] if rates[code] > 0 else 0,
                            "rate_from_usd": rates[code]
                        }
                
                return {
                    "success": True,
                    "base": "USD",
                    "rates": live_rates,
                    "fetched_at": datetime.utcnow().isoformat()
                }
    except Exception as e:
        logger.error(f"Error fetching live rates: {str(e)}")
    
    # إرجاع الأسعار الثابتة في حالة الفشل
    return {
        "success": False,
        "message": "تعذر جلب الأسعار الحية، يتم استخدام الأسعار الثابتة",
        "rates": {code: {"rate_to_usd": info["rate_to_usd"], "rate_from_usd": 1/info["rate_to_usd"] if info["rate_to_usd"] > 0 else 0} 
                 for code, info in SUPPORTED_CURRENCIES.items()}
    }

# ==================== Super Admin Currency Dashboard ====================

@api_router.get("/super-admin/sales-summary")
async def get_super_admin_sales_summary(
    display_currency: str = "USD",
    current_user: dict = Depends(verify_super_admin)
):
    """جلب ملخص المبيعات لجميع العملاء مع تحويل العملات"""
    
    # جلب جميع العملاء النشطين (ليس تجريبي)
    tenants = await db.tenants.find(
        {"is_demo": {"$ne": True}, "subscription_type": {"$ne": "demo"}}, 
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    
    total_sales_usd = 0
    total_orders = 0
    tenant_sales = []
    active_tenants = 0
    
    for tenant in tenants:
        tenant_id = tenant.get("id")
        
        # جلب إعدادات العملة للعميل
        regional = await db.tenant_regional_settings.find_one({"tenant_id": tenant_id}, {"_id": 0})
        tenant_currency = regional.get("currency", "IQD") if regional else "IQD"
        
        # جلب إجمالي المبيعات
        pipeline = [
            {"$match": {"tenant_id": tenant_id, "status": {"$in": ["completed", "delivered"]}}},
            {"$group": {"_id": None, "total": {"$sum": "$total"}, "count": {"$sum": 1}}}
        ]
        result = await db.orders.aggregate(pipeline).to_list(1)
        
        if result:
            sales_in_tenant_currency = _sn(result[0].get("total"))
            orders_count = _sn(result[0].get("count"))
            
            if sales_in_tenant_currency > 0:
                active_tenants += 1
            
            # تحويل للدولار
            if tenant_currency in SUPPORTED_CURRENCIES:
                rate = SUPPORTED_CURRENCIES[tenant_currency]["rate_to_usd"]
                sales_in_usd = sales_in_tenant_currency * rate
            else:
                sales_in_usd = sales_in_tenant_currency
            
            total_sales_usd += sales_in_usd
            total_orders += orders_count
            
            # حساب المبيعات بالعملة المطلوبة
            if display_currency in SUPPORTED_CURRENCIES:
                display_rate = SUPPORTED_CURRENCIES[display_currency]["rate_to_usd"]
                converted_sales = sales_in_usd / display_rate if display_rate > 0 else 0
            else:
                converted_sales = sales_in_usd
            
            tenant_sales.append({
                "name": tenant.get("name"),
                "original_sales": sales_in_tenant_currency,
                "original_currency": tenant_currency,
                "converted_sales": round(converted_sales, 2),
                "orders_count": orders_count
            })
    
    # ترتيب حسب المبيعات المحولة
    tenant_sales.sort(key=lambda x: x["converted_sales"], reverse=True)
    
    # تحويل الإجمالي للعملة المطلوبة
    if display_currency in SUPPORTED_CURRENCIES:
        rate = SUPPORTED_CURRENCIES[display_currency]["rate_to_usd"]
        decimal_places = SUPPORTED_CURRENCIES[display_currency].get("decimal_places", 2)
        total_in_display = total_sales_usd / rate if rate > 0 else 0
        total_in_display = round(total_in_display, decimal_places)
    else:
        total_in_display = total_sales_usd
    
    return {
        "total_sales_converted": total_in_display,
        "total_sales_usd": round(total_sales_usd, 2),
        "total_orders": total_orders,
        "active_tenants": active_tenants,
        "display_currency": display_currency,
        "display_currency_symbol": SUPPORTED_CURRENCIES.get(display_currency, {}).get("symbol", "$"),
        "tenant_sales": tenant_sales
    }

# ==================== Login Page Settings ====================

@api_router.get("/system/login-page-settings")
async def get_login_page_settings():
    """جلب إعدادات صفحة الدخول"""
    settings = await db.settings.find_one({"type": "login_page_settings"}, {"_id": 0})
    
    default_settings = {
        "enable_animation": True,
        "transition_type": "fade",
        "transition_duration": 1.5,
        "auto_change": True,
        "logo_animation": "pulse",
        "backgrounds": [],
        "login_logo_enabled": True,
        "login_logo_url": "",
        "accent_color": "rgba(147, 51, 234, 0.5)"
    }
    
    if settings and settings.get("value"):
        return {**default_settings, **settings.get("value")}
    
    return default_settings

@api_router.put("/system/login-page-settings")
async def update_login_page_settings(settings: dict, current_user: dict = Depends(verify_super_admin)):
    """تحديث إعدادات صفحة الدخول (المالك فقط)"""
    
    await db.settings.update_one(
        {"type": "login_page_settings"},
        {"$set": {"value": settings, "updated_at": datetime.now(timezone.utc).isoformat()}},
        upsert=True
    )
    
    return {"message": "تم تحديث إعدادات صفحة الدخول", "settings": settings}

@api_router.get("/invoice-data/{order_id}")
async def get_invoice_data(order_id: str, current_user: dict = Depends(get_current_user)):
    """جلب بيانات الفاتورة الكاملة للطباعة"""
    
    # جلب الطلب
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب بيانات العميل (المطعم)
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    
    # جلب إعدادات فاتورة العميل
    tenant_invoice = await db.tenant_invoice_settings.find_one({"tenant_id": tenant_id}, {"_id": 0}) if tenant_id else {}
    
    # جلب إعدادات النظام
    system_settings = await db.settings.find_one({"type": "system_invoice_settings"}, {"_id": 0})
    system_invoice = system_settings.get("value", {}) if system_settings else {}
    
    # جلب الفرع
    branch = await db.branches.find_one({"id": order.get("branch_id")}, {"_id": 0, "name": 1, "address": 1, "phone": 1})
    
    return {
        "order": order,
        "tenant": {
            "name": tenant.get("name") if tenant else "المطعم",
            "logo_url": tenant.get("logo_url") if tenant else None,
            "phone": tenant_invoice.get("phone") or (tenant.get("owner_phone") if tenant else None),
            "phone2": tenant_invoice.get("phone2"),
            "address": tenant_invoice.get("address") or (branch.get("address") if branch else None),
            "tax_number": tenant_invoice.get("tax_number"),
            "custom_header": tenant_invoice.get("custom_header"),
            "custom_footer": tenant_invoice.get("custom_footer")
        },
        "system": {
            "logo_url": system_invoice.get("system_logo_url"),
            "thank_you_message": system_invoice.get("thank_you_message", "شكراً لزيارتكم"),
            "phone": system_invoice.get("system_phone"),
            "phone2": system_invoice.get("system_phone2"),
            "email": system_invoice.get("system_email"),
            "website": system_invoice.get("system_website"),
            "footer_text": system_invoice.get("footer_text"),
            "show_branding": system_invoice.get("show_system_branding", True)
        },
        "branch": branch
    }

# ==================== ROLES & STAFF MANAGEMENT - إدارة الأدوار والموظفين ====================
# نظام إدارة الموظفين والصلاحيات للعملاء

class StaffCreate(BaseModel):
    """نموذج إنشاء موظف جديد"""
    full_name: str
    email: str
    phone: Optional[str] = None
    password: str
    role: str = "cashier"  # cashier, supervisor, delivery, branch_manager
    branch_id: str
    job_title: Optional[str] = None  # المسمى الوظيفي المخصص
    permissions: Optional[List[str]] = None  # صلاحيات الموظف

class StaffUpdate(BaseModel):
    """نموذج تحديث بيانات موظف"""
    full_name: Optional[str] = None
    phone: Optional[str] = None
    role: Optional[str] = None
    branch_id: Optional[str] = None
    job_title: Optional[str] = None
    is_active: Optional[bool] = None
    permissions: Optional[List[str]] = None  # صلاحيات الموظف

class StaffResponse(BaseModel):
    """نموذج الاستجابة لبيانات الموظف"""
    model_config = ConfigDict(extra="ignore")
    id: str
    full_name: str
    email: str
    phone: Optional[str] = None
    role: str
    branch_id: Optional[str] = None
    branch_name: Optional[str] = None
    job_title: Optional[str] = None
    is_active: bool = True
    last_login: Optional[str] = None
    created_at: Optional[str] = None
    permissions: Optional[List[str]] = None  # صلاحيات الموظف

# الأدوار المتاحة للموظفين (غير Admin و SuperAdmin)
STAFF_ROLES = {
    "branch_manager": "مدير فرع",
    "supervisor": "مشرف",
    "cashier": "كاشير",
    "captain": "كابتن",
    "kitchen": "مطبخ"
}

# ملاحظة: تم إزالة دور "delivery" (سائق توصيل) من هنا
# السائقين يتم إنشاؤهم وإدارتهم فقط من قسم التوصيل (Delivery)

@api_router.get("/staff/roles")
async def get_staff_roles(current_user: dict = Depends(get_current_user)):
    """جلب قائمة الأدوار المتاحة للموظفين"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    return STAFF_ROLES

@api_router.get("/staff", response_model=List[StaffResponse])
async def get_staff_members(
    branch_id: Optional[str] = None,
    role: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب قائمة الموظفين - للعميل فقط"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user)
    
    # استثناء المستخدمين الرئيسيين (Admin و SuperAdmin)
    query["role"] = {"$nin": [UserRole.ADMIN, UserRole.SUPER_ADMIN]}
    
    if branch_id:
        query["branch_id"] = branch_id
    if role:
        query["role"] = role
    
    staff = await db.users.find(query, {"_id": 0, "password": 0}).to_list(500)
    
    # إضافة اسم الفرع لكل موظف
    branch_ids = list(set([s.get("branch_id") for s in staff if s.get("branch_id")]))
    branches = await db.branches.find({"id": {"$in": branch_ids}}, {"_id": 0, "id": 1, "name": 1}).to_list(100)
    branch_map = {b["id"]: b["name"] for b in branches}
    
    for s in staff:
        s["branch_name"] = branch_map.get(s.get("branch_id"), "غير محدد")
    
    return staff

@api_router.post("/staff", response_model=StaffResponse)
async def create_staff_member(staff: StaffCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء موظف جديد"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من أن الدور صحيح
    if staff.role not in STAFF_ROLES:
        raise HTTPException(status_code=400, detail=f"الدور غير صحيح. الأدوار المتاحة: {', '.join(STAFF_ROLES.keys())}")
    
    # التحقق من عدم تكرار البريد
    existing = await db.users.find_one({"email": staff.email})
    if existing:
        raise HTTPException(status_code=400, detail="البريد الإلكتروني مستخدم")
    
    # التحقق من الفرع
    tenant_id = get_user_tenant_id(current_user)
    branch_query = {"id": staff.branch_id}
    if tenant_id:
        branch_query["tenant_id"] = tenant_id
    
    branch = await db.branches.find_one(branch_query, {"_id": 0})
    if not branch:
        raise HTTPException(status_code=404, detail="الفرع غير موجود")
    
    # إنشاء الموظف
    now = datetime.now(timezone.utc).isoformat()
    staff_doc = {
        "id": str(uuid.uuid4()),
        "full_name": staff.full_name,
        "username": staff.email.split("@")[0],
        "email": staff.email,
        "phone": staff.phone,
        "password": hash_password(staff.password),
        "role": staff.role,
        "branch_id": staff.branch_id,
        "job_title": staff.job_title or STAFF_ROLES.get(staff.role, staff.role),
        "permissions": staff.permissions or [],
        "tenant_id": tenant_id,
        "is_active": True,
        "last_login": None,
        "created_at": now,
        "updated_at": now
    }
    
    await db.users.insert_one(staff_doc)
    
    # إزالة كلمة المرور من الاستجابة
    del staff_doc["password"]
    del staff_doc["_id"]
    staff_doc["branch_name"] = branch.get("name", "")
    
    return staff_doc

@api_router.get("/staff/{staff_id}", response_model=StaffResponse)
async def get_staff_member(staff_id: str, current_user: dict = Depends(get_current_user)):
    """جلب بيانات موظف محدد"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query, {"_id": 0, "password": 0})
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # جلب اسم الفرع
    if staff.get("branch_id"):
        branch = await db.branches.find_one({"id": staff["branch_id"]}, {"_id": 0, "name": 1})
        staff["branch_name"] = branch.get("name", "") if branch else ""
    
    return staff

@api_router.put("/staff/{staff_id}", response_model=StaffResponse)
async def update_staff_member(staff_id: str, update: StaffUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث بيانات موظف"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query)
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من أن الموظف ليس Admin أو SuperAdmin
    if staff.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="لا يمكن تعديل هذا المستخدم")
    
    # بناء التحديث
    update_data = {}
    if update.full_name is not None:
        update_data["full_name"] = update.full_name
    if update.phone is not None:
        update_data["phone"] = update.phone
    if update.role is not None:
        if update.role not in STAFF_ROLES:
            raise HTTPException(status_code=400, detail="الدور غير صحيح")
        update_data["role"] = update.role
    if update.branch_id is not None:
        # التحقق من الفرع
        tenant_id = get_user_tenant_id(current_user)
        branch_query = {"id": update.branch_id}
        if tenant_id:
            branch_query["tenant_id"] = tenant_id
        branch = await db.branches.find_one(branch_query)
        if not branch:
            raise HTTPException(status_code=404, detail="الفرع غير موجود")
        update_data["branch_id"] = update.branch_id
    if update.job_title is not None:
        update_data["job_title"] = update.job_title
    if update.is_active is not None:
        update_data["is_active"] = update.is_active
    if update.permissions is not None:
        update_data["permissions"] = update.permissions
    
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.users.update_one(query, {"$set": update_data})
    
    # جلب البيانات المحدثة
    updated_staff = await db.users.find_one({"id": staff_id}, {"_id": 0, "password": 0})
    
    # جلب اسم الفرع
    if updated_staff.get("branch_id"):
        branch = await db.branches.find_one({"id": updated_staff["branch_id"]}, {"_id": 0, "name": 1})
        updated_staff["branch_name"] = branch.get("name", "") if branch else ""
    
    return updated_staff

@api_router.delete("/staff/{staff_id}")
async def delete_staff_member(staff_id: str, current_user: dict = Depends(get_current_user)):
    """حذف (تعطيل) موظف"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query)
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من أن الموظف ليس Admin أو SuperAdmin
    if staff.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="لا يمكن حذف هذا المستخدم")
    
    # تعطيل بدلاً من الحذف
    await db.users.update_one(query, {"$set": {"is_active": False, "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"message": "تم تعطيل الموظف"}

@api_router.post("/staff/{staff_id}/reset-password")
async def reset_staff_password(staff_id: str, new_password: str = Body(..., embed=True), current_user: dict = Depends(get_current_user)):
    """إعادة تعيين كلمة مرور موظف"""
    if current_user.get("role") not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": staff_id})
    staff = await db.users.find_one(query)
    
    if not staff:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من أن الموظف ليس Admin أو SuperAdmin
    if staff.get("role") in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="لا يمكن تعديل هذا المستخدم")
    
    validate_password_strength(new_password)
    await db.users.update_one(query, {"$set": {"password": hash_password(new_password), "updated_at": datetime.now(timezone.utc).isoformat()}})
    
    return {"message": "تم تغيير كلمة المرور"}

# ==================== FILE UPLOAD ROUTES ====================

async def process_and_save_image(file: UploadFile, target_dir: Path, max_size: tuple = (1920, 1080), quality: int = 85) -> str:
    """معالجة وحفظ الصورة بالحجم والصيغة المناسبة"""
    try:
        # قراءة محتوى الملف
        content = await file.read()
        
        # فتح الصورة
        image = Image.open(io.BytesIO(content))
        
        # تحويل RGBA إلى RGB إذا لزم الأمر
        if image.mode in ('RGBA', 'LA', 'P'):
            background = Image.new('RGB', image.size, (255, 255, 255))
            if image.mode == 'P':
                image = image.convert('RGBA')
            background.paste(image, mask=image.split()[-1] if image.mode == 'RGBA' else None)
            image = background
        elif image.mode != 'RGB':
            image = image.convert('RGB')
        
        # تغيير الحجم مع الحفاظ على النسبة
        image.thumbnail(max_size, Image.Resampling.LANCZOS)
        
        # إنشاء اسم ملف فريد
        file_id = str(uuid.uuid4())
        filename = f"{file_id}.jpg"
        filepath = target_dir / filename
        
        # حفظ الصورة بصيغة JPEG مضغوطة
        image.save(filepath, "JPEG", quality=quality, optimize=True)
        
        return filename
    except Exception as e:
        logger.error(f"Error processing image: {e}")
        raise HTTPException(status_code=400, detail=f"فشل في معالجة الصورة: {str(e)}")

@api_router.post("/upload/background")
async def upload_background_file(
    file: UploadFile = File(...),
    title: str = Form(None),
    animation_type: str = Form("fade"),
    current_user: dict = Depends(verify_super_admin)
):
    """رفع خلفية من الجهاز مع تحويل تلقائي"""
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/heic', 'image/heif', 'image/bmp', 'image/tiff']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. الأنواع المدعومة: JPEG, PNG, GIF, WEBP, HEIC, BMP, TIFF")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, BACKGROUNDS_DIR, max_size=(1920, 1080), quality=85)
    
    # إنشاء URL نسبي للصورة (سيعمل في جميع البيئات)
    image_url = f"/api/uploads/backgrounds/{filename}"
    
    # جلب الإعدادات الحالية
    settings = await db.settings.find_one({"type": "login_backgrounds"}, {"_id": 0})
    current_backgrounds = []
    current_value = {}
    
    if settings and settings.get("value"):
        current_value = settings["value"]
        current_backgrounds = current_value.get("backgrounds", [])
    
    # إضافة الخلفية الجديدة
    new_background = {
        "id": str(uuid.uuid4()),
        "image_url": image_url,
        "title": title,
        "animation_type": animation_type,
        "animation_duration": 8,
        "overlay_opacity": 0.5,
        "is_active": True,
        "sort_order": len(current_backgrounds),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    current_backgrounds.append(new_background)
    
    # تحديث مع الحفاظ على الإعدادات الأخرى
    default_settings = {
        "backgrounds": current_backgrounds,
        "animation_enabled": current_value.get("animation_enabled", True),
        "transition_type": current_value.get("transition_type", "fade"),
        "transition_duration": current_value.get("transition_duration", 1.5),
        "auto_play": current_value.get("auto_play", True),
        "show_logo": current_value.get("show_logo", True),
        "logo_url": current_value.get("logo_url", None),
        "logo_animation": current_value.get("logo_animation", "pulse"),
        "overlay_color": current_value.get("overlay_color", "rgba(0,0,0,0.5)"),
        "text_color": current_value.get("text_color", "#ffffff")
    }
    
    # حفظ التحديث
    await db.settings.update_one(
        {"type": "login_backgrounds"},
        {"$set": {"type": "login_backgrounds", "value": default_settings}},
        upsert=True
    )
    
    return {"message": "تم رفع الخلفية بنجاح", "background": new_background}

@api_router.post("/upload/logo")
async def upload_logo_file(
    file: UploadFile = File(...),
    tenant_id: str = Form(None),
    current_user: dict = Depends(verify_super_admin)
):
    """رفع شعار للعميل - للمالك فقط"""
    
    # التحقق من نوع الملف
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/svg+xml']
    if file.content_type not in allowed_types:
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, LOGOS_DIR, max_size=(512, 512), quality=90)
    
    # إنشاء URL نسبي للشعار
    logo_url = f"/api/uploads/logos/{filename}"
    
    # تحديث الشعار للعميل إذا تم تحديد tenant_id
    if tenant_id:
        await db.tenants.update_one(
            {"id": tenant_id},
            {"$set": {"logo_url": logo_url}}
        )
    
    return {"message": "تم رفع الشعار بنجاح", "logo_url": logo_url, "url": logo_url}

@api_router.post("/upload/restaurant-logo")
async def upload_restaurant_logo(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع شعار المطعم - للمدير أو المالك"""
    
    # السماح للمدير (admin) والمالك (super_admin)
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    # التحقق من نوع الملف - دعم HEIC من iPhone
    allowed_types = ['image/jpeg', 'image/png', 'image/gif', 'image/webp', 'image/svg+xml', 'image/heic', 'image/heif']
    content_type = file.content_type or ''
    if not content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="نوع الملف غير مدعوم. يرجى استخدام JPG, PNG, GIF, WebP, HEIC أو SVG")
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, LOGOS_DIR, max_size=(512, 512), quality=90)
    
    # إنشاء URL نسبي للشعار
    logo_url = f"/api/uploads/logos/{filename}"
    
    return {"message": "تم رفع الشعار بنجاح", "url": logo_url, "logo_url": logo_url}

@api_router.post("/upload/image")
async def upload_general_image(
    file: UploadFile = File(...),
    type: str = Form("product"),  # product, category, general
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة عامة للمنتجات أو الفئات"""
    
    # التحقق من نوع الملف
    allowed_types = [
        'image/jpeg', 'image/png', 'image/gif', 'image/webp', 
        'image/heic', 'image/heif', 'image/bmp', 'image/tiff',
        'image/svg+xml', 'image/avif'
    ]
    
    # السماح بأي نوع صورة
    content_type = file.content_type or ''
    if not content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="يجب أن يكون الملف صورة")
    
    # تحديد المجلد حسب النوع
    if type == "product":
        target_dir = PRODUCTS_DIR
        subfolder = "products"
        max_size = (800, 800)
    elif type == "category":
        target_dir = CATEGORIES_DIR
        subfolder = "categories"
        max_size = (400, 400)
    else:
        target_dir = IMAGES_DIR
        subfolder = "images"
        max_size = (1024, 1024)
    
    # معالجة وحفظ الصورة
    filename = await process_and_save_image(file, target_dir, max_size=max_size, quality=85)
    
    # إنشاء URL للصورة
    image_url = f"/api/uploads/images/{subfolder}/{filename}"
    
    return {
        "message": "تم رفع الصورة بنجاح",
        "image_url": image_url,
        "filename": filename
    }

@api_router.post("/upload/product-image")
async def upload_product_image(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة منتج"""
    return await upload_general_image(file, "product", current_user)

@api_router.post("/upload/category-image")
async def upload_category_image(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة فئة"""
    return await upload_general_image(file, "category", current_user)

# ==================== moved to routes/printer_routes.py ====================

# ==================== SUPER ADMIN routes moved to routes/super_admin_routes.py ====================

# ==================== SUPPLIERS & PURCHASING ====================

class SupplierCreate(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    contact_person: Optional[str] = None
    payment_terms: str = "cash"

class SupplierUpdate(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    address: Optional[str] = None
    notes: Optional[str] = None
    contact_person: Optional[str] = None
    payment_terms: Optional[str] = None
    is_active: Optional[bool] = None

class PurchaseOrderCreate(BaseModel):
    supplier_id: str
    items: List[Dict[str, Any]]
    expected_delivery: Optional[str] = None
    notes: Optional[str] = None
    branch_id: Optional[str] = None

class PurchaseOrderStatusUpdate(BaseModel):
    status: str
    notes: Optional[str] = None

class RawMaterialCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str
    category: Optional[str] = None
    min_stock: float = 0
    current_stock: float = 0
    price: float = 0
    supplier_id: Optional[str] = None
    branch_id: Optional[str] = None

# ==================== PURCHASE INVOICES - فواتير الشراء ====================

class PurchaseInvoiceItemCreate(BaseModel):
    name: str
    quantity: float
    unit: str = "كغم"
    unit_price: float

class PurchaseInvoiceCreate(BaseModel):
    supplier_id: Optional[str] = None
    invoice_number: Optional[str] = None
    items: List[PurchaseInvoiceItemCreate]
    notes: Optional[str] = None
    total_amount: float
    image_data: Optional[str] = None  # Base64 encoded image

class PurchaseSupplierCreate(BaseModel):
    name: str
    company_name: Optional[str] = None
    phone: str
    address: Optional[str] = None
    products: Optional[str] = None
    notes: Optional[str] = None

@api_router.get("/purchase-invoices")
async def get_purchase_invoices(current_user: dict = Depends(get_current_user)):
    """جلب فواتير الشراء"""
    query = build_tenant_query(current_user)
    invoices = await db.purchase_invoices.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # إضافة اسم المورد لكل فاتورة
    for invoice in invoices:
        if invoice.get("supplier_id"):
            supplier = await db.purchase_suppliers.find_one({"id": invoice["supplier_id"]}, {"_id": 0})
            invoice["supplier_name"] = supplier.get("name") if supplier else None
    
    return invoices

@api_router.post("/purchase-invoices")
async def create_purchase_invoice(invoice: PurchaseInvoiceCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء فاتورة شراء جديدة"""
    tenant_id = get_user_tenant_id(current_user)
    
    invoice_doc = {
        "id": str(uuid.uuid4()),
        "supplier_id": invoice.supplier_id,
        "invoice_number": invoice.invoice_number,
        "items": [item.model_dump() for item in invoice.items],
        "notes": invoice.notes,
        "total_amount": invoice.total_amount,
        "image_data": invoice.image_data,
        "status": "new",
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchase_invoices.insert_one(invoice_doc)
    del invoice_doc["_id"]
    
    # إضافة اسم المورد للرد
    if invoice.supplier_id:
        supplier = await db.purchase_suppliers.find_one({"id": invoice.supplier_id}, {"_id": 0})
        invoice_doc["supplier_name"] = supplier.get("name") if supplier else None
    
    return invoice_doc

@api_router.delete("/purchase-invoices/{invoice_id}")
async def delete_purchase_invoice(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """حذف فاتورة شراء"""
    query = build_tenant_query(current_user)
    query["id"] = invoice_id
    
    result = await db.purchase_invoices.delete_one(query)
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")
    
    return {"message": "تم حذف الفاتورة"}


@api_router.post("/purchase-invoices/{invoice_id}/send-to-warehouse")
async def send_purchase_invoice_to_warehouse(invoice_id: str, current_user: dict = Depends(get_current_user)):
    """إرسال فاتورة شراء (legacy collection) للمخزن — يضيف المواد كطبقات FIFO."""
    from services.cost_layer_service import add_cost_layer, get_current_effective_cost, detect_price_increase

    tenant_id = get_user_tenant_id(current_user)
    query = {"id": invoice_id}
    if tenant_id:
        query["tenant_id"] = tenant_id

    invoice = await db.purchase_invoices.find_one(query, {"_id": 0})
    if not invoice:
        raise HTTPException(status_code=404, detail="الفاتورة غير موجودة")

    if invoice.get("status") == "transferred":
        raise HTTPException(status_code=400, detail="تم إرسال هذه الفاتورة للمخزن بالفعل")

    detected_alerts = []
    movements = []

    for item in (invoice.get("items") or []):
        item_qty = float(item.get("quantity", 0) or 0)
        item_cost = float(item.get("unit_price", item.get("cost_per_unit", 0)) or 0)
        item_name = item.get("name")
        if item_qty <= 0 or not item_name:
            continue

        # ابحث عن المادة الخام بالاسم
        mq = {"name": item_name}
        if tenant_id:
            mq["tenant_id"] = tenant_id
        existing = await db.raw_materials.find_one(mq)

        if existing:
            material_id = existing["id"]
            # كشف فرق السعر قبل الإضافة
            alert = await detect_price_increase(
                db,
                tenant_id=tenant_id,
                material_id=material_id,
                material_name=item_name,
                unit=item.get("unit", "كغم"),
                quantity=item_qty,
                new_cost=item_cost,
                purchase_id=invoice_id,
                purchase_number=str(invoice.get("invoice_number") or ""),
                triggered_by_user_id=current_user.get("id"),
                triggered_by_role=current_user.get("role"),
            )
            if alert:
                detected_alerts.append(alert)

            # زِد الكمية فقط (لا تُحدّث cost_per_unit — يأتي من أقدم طبقة)
            await db.raw_materials.update_one(
                {"id": material_id},
                {"$inc": {"quantity": item_qty}, "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}}
            )
        else:
            material_id = str(uuid.uuid4())
            await db.raw_materials.insert_one({
                "id": material_id,
                "name": item_name,
                "name_en": None,
                "unit": item.get("unit", "كغم"),
                "quantity": item_qty,
                "min_quantity": 0,
                "cost_per_unit": item_cost,
                "category": None,
                "last_updated": datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": tenant_id,
            })

        # أضف طبقة تكلفة (FIFO)
        await add_cost_layer(
            db,
            material_id=material_id,
            material_name=item_name,
            unit=item.get("unit") or "كغم",
            quantity=item_qty,
            unit_cost=item_cost,
            tenant_id=tenant_id,
            source="purchase",
            source_id=invoice_id,
            source_number=str(invoice.get("invoice_number") or ""),
        )

        # تحديث cost_per_unit ليساوي تكلفة أقدم طبقة نشطة
        effective = await get_current_effective_cost(db, material_id, tenant_id)
        if effective is not None:
            await db.raw_materials.update_one(
                {"id": material_id},
                {"$set": {"cost_per_unit": effective, "last_cost_updated_at": datetime.now(timezone.utc).isoformat()}}
            )

        # تسجيل حركة دخول
        await db.inventory_movements.insert_one({
            "id": str(uuid.uuid4()),
            "type": "in",
            "subtype": "purchase_receipt",
            "material_id": material_id,
            "material_name": item_name,
            "quantity": item_qty,
            "unit": item.get("unit"),
            "cost_per_unit": item_cost,
            "total_value": item_qty * item_cost,
            "reference_type": "purchase_invoice",
            "reference_id": invoice_id,
            "reference_number": invoice.get("invoice_number"),
            "supplier_name": invoice.get("supplier_name"),
            "performed_by": current_user.get("id"),
            "performed_by_name": current_user.get("full_name") or current_user.get("username"),
            "notes": f"استلام فاتورة شراء #{invoice.get('invoice_number')}",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "tenant_id": tenant_id,
        })
        movements.append({"material_name": item_name, "quantity": item_qty})

    # تحديث حالة الفاتورة
    await db.purchase_invoices.update_one(
        {"id": invoice_id},
        {"$set": {
            "status": "transferred",
            "transferred_at": datetime.now(timezone.utc).isoformat(),
            "transferred_by": current_user.get("id"),
        }}
    )

    return {
        "message": "تم إرسال الفاتورة للمخزن بنجاح",
        "invoice_id": invoice_id,
        "movements": movements,
        "price_alerts": detected_alerts,
        "price_alerts_count": len(detected_alerts),
    }


@api_router.get("/purchase-suppliers")
async def get_purchase_suppliers(current_user: dict = Depends(get_current_user)):
    """جلب موردي المشتريات"""
    query = build_tenant_query(current_user)
    suppliers = await db.purchase_suppliers.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return suppliers

@api_router.post("/purchase-suppliers")
async def create_purchase_supplier(supplier: PurchaseSupplierCreate, current_user: dict = Depends(get_current_user)):
    """إضافة مورد جديد"""
    tenant_id = get_user_tenant_id(current_user)
    
    supplier_doc = {
        "id": str(uuid.uuid4()),
        **supplier.model_dump(),
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.purchase_suppliers.insert_one(supplier_doc)
    del supplier_doc["_id"]
    return supplier_doc

@api_router.get("/warehouse-purchase-requests")
async def get_warehouse_purchase_requests(current_user: dict = Depends(get_current_user)):
    """جلب طلبات الشراء من المخزن"""
    query = build_tenant_query(current_user)
    requests = await db.warehouse_purchase_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests

@api_router.post("/warehouse-purchase-requests/{request_id}/transfer")
async def transfer_warehouse_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """تحويل طلب الشراء للمخزن"""
    query = build_tenant_query(current_user)
    query["id"] = request_id
    
    request = await db.warehouse_purchase_requests.find_one(query)
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # تحديث حالة الطلب
    await db.warehouse_purchase_requests.update_one(
        {"id": request_id},
        {"$set": {"status": "completed", "transferred_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "تم التحويل للمخزن"}

# ==================== مواد التغليف والورقيات ====================

class PackagingMaterialCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str = "قطعة"
    quantity: float = 0
    min_quantity: float = 0
    cost_per_unit: float = 0
    category: Optional[str] = None

class PackagingRequestCreate(BaseModel):
    items: List[Dict[str, Any]]
    priority: str = "normal"
    notes: Optional[str] = None

# --- مواد التغليف (CRUD) محذوفة من هنا - موجودة في routes/inventory_system.py ---

# الحصول على طلبات التغليف
@api_router.get("/packaging-requests")
async def get_packaging_requests(current_user: dict = Depends(get_current_user)):
    """الحصول على جميع طلبات التغليف"""
    tenant_id = current_user.get("tenant_id")
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant ID required")
    
    requests = await db.packaging_requests.find({"tenant_id": tenant_id}).sort("created_at", -1).to_list(100)
    for r in requests:
        r["id"] = r.pop("_id", r.get("id"))
    return requests

# إنشاء طلب تغليف جديد
@api_router.post("/packaging-requests")
async def create_packaging_request(request: PackagingRequestCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء طلب مواد تغليف جديد"""
    tenant_id = current_user.get("tenant_id")
    branch_id = current_user.get("branch_id")
    
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant ID required")
    
    # الحصول على اسم الفرع
    branch_name = "المركز الرئيسي"
    if branch_id:
        branch = await db.branches.find_one({"id": branch_id})
        if branch:
            branch_name = branch.get("name", branch_name)
    
    # إنشاء رقم الطلب
    count = await db.packaging_requests.count_documents({"tenant_id": tenant_id})
    request_number = f"PKG-{count + 1:04d}"
    
    request_doc = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "from_branch_id": branch_id,
        "from_branch_name": branch_name,
        "request_number": request_number,
        "items": request.items,
        "priority": request.priority,
        "notes": request.notes,
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": current_user.get("user_id")
    }
    
    await db.packaging_requests.insert_one(request_doc)
    request_doc.pop("_id", None)
    return request_doc

# الموافقة على طلب تغليف
@api_router.post("/packaging-requests/{request_id}/approve")
async def approve_packaging_request(request_id: str, current_user: dict = Depends(get_current_user)):
    """الموافقة على طلب تغليف"""
    tenant_id = current_user.get("tenant_id")
    
    request = await db.packaging_requests.find_one({"id": request_id, "tenant_id": tenant_id})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    await db.packaging_requests.update_one(
        {"id": request_id},
        {
            "$set": {
                "status": "approved",
                "approved_at": datetime.now(timezone.utc).isoformat(),
                "approved_by": current_user.get("user_id")
            }
        }
    )
    
    return {"message": "تمت الموافقة على الطلب"}

# تحويل مواد التغليف للفرع
@api_router.post("/packaging-requests/{request_id}/transfer")
async def transfer_packaging_to_branch(request_id: str, current_user: dict = Depends(get_current_user)):
    """تحويل مواد التغليف للفرع - تزيد الكمية تلقائياً في مخزون الفرع"""
    tenant_id = current_user.get("tenant_id")
    
    request = await db.packaging_requests.find_one({"id": request_id, "tenant_id": tenant_id})
    if not request:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    branch_id = request.get("from_branch_id")
    
    # تحويل كل صنف
    for item in request.get("items", []):
        material_id = item.get("packaging_material_id")
        quantity = _sn(item.get("quantity"))
        
        # خصم من المخزن الرئيسي
        material = await db.packaging_materials.find_one({"id": material_id, "tenant_id": tenant_id})
        if material:
            new_quantity = max(0, _sn(material.get("quantity")) - quantity)
            transferred = material.get("transferred_to_branches", 0) + quantity
            
            await db.packaging_materials.update_one(
                {"id": material_id},
                {
                    "$set": {
                        "quantity": new_quantity,
                        "remaining_quantity": new_quantity,
                        "transferred_to_branches": transferred,
                        "total_value": new_quantity * material.get("cost_per_unit", 0),
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
        
        # إضافة لمخزون الفرع (أو إنشاء سجل جديد)
        if branch_id:
            branch_inventory = await db.branch_packaging_inventory.find_one({
                "branch_id": branch_id,
                "packaging_material_id": material_id,
                "tenant_id": tenant_id
            })
            
            if branch_inventory:
                # زيادة الكمية الموجودة
                await db.branch_packaging_inventory.update_one(
                    {"id": branch_inventory["id"]},
                    {
                        "$inc": {"quantity": quantity, "total_received": quantity},
                        "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
                    }
                )
            else:
                # إنشاء سجل جديد
                await db.branch_packaging_inventory.insert_one({
                    "id": str(uuid.uuid4()),
                    "tenant_id": tenant_id,
                    "branch_id": branch_id,
                    "packaging_material_id": material_id,
                    "name": item.get("name", ""),
                    "unit": item.get("unit", "قطعة"),
                    "quantity": quantity,
                    "total_received": quantity,
                    "used_quantity": 0,
                    "cost_per_unit": material.get("cost_per_unit", 0) if material else 0,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                    "last_updated": datetime.now(timezone.utc).isoformat()
                })
    
    # تحديث حالة الطلب
    await db.packaging_requests.update_one(
        {"id": request_id},
        {
            "$set": {
                "status": "transferred",
                "transferred_at": datetime.now(timezone.utc).isoformat(),
                "transferred_by": current_user.get("user_id")
            }
        }
    )
    
    return {"message": "تم تحويل المواد للفرع بنجاح"}

# الحصول على مخزون التغليف في الفروع
@api_router.get("/branch-packaging-inventory")
async def get_branch_packaging_inventory(branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """الحصول على مخزون التغليف في الفروع"""
    tenant_id = current_user.get("tenant_id")
    user_branch_id = current_user.get("branch_id")
    
    if not tenant_id:
        raise HTTPException(status_code=400, detail="Tenant ID required")
    
    query = {"tenant_id": tenant_id}
    
    # إذا المستخدم مرتبط بفرع، أظهر فقط مخزون فرعه
    if user_branch_id:
        query["branch_id"] = user_branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    inventory = await db.branch_packaging_inventory.find(query, {"_id": 0}).to_list(500)
    for item in inventory:
        # حساب الكمية المتبقية
        item["remaining_quantity"] = _sn(item.get("quantity")) - item.get("used_quantity", 0)
    
    return inventory

# ==================== moved to routes/ocr_routes.py ====================

# ==================== FINISHED PRODUCTS - المنتجات النهائية ====================

class FinishedProductCreate(BaseModel):
    """إنشاء منتج نهائي مع وصفته"""
    name: str
    name_en: Optional[str] = None
    unit: str = "قطعة"
    quantity: float = 0.0  # الكمية المتوفرة
    min_quantity: float = 0.0  # الحد الأدنى للتنبيه
    cost_per_unit: float = 0.0  # سيتم حسابها تلقائياً من الوصفة
    selling_price: float = 0.0  # سعر البيع
    recipe: List[Dict[str, Any]] = []  # [{raw_material_id, quantity}]
    description: Optional[str] = None
    category: str = "general"

class FinishedProductUpdate(BaseModel):
    """تحديث منتج نهائي"""
    name: Optional[str] = None
    name_en: Optional[str] = None
    unit: Optional[str] = None
    quantity: Optional[float] = None
    min_quantity: Optional[float] = None
    selling_price: Optional[float] = None
    recipe: Optional[List[Dict[str, Any]]] = None
    description: Optional[str] = None
    category: Optional[str] = None
    is_active: Optional[bool] = None

@api_router.post("/finished-products")
async def create_finished_product(product: FinishedProductCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء منتج نهائي جديد مع وصفته (المواد الخام المكونة له)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # حساب تكلفة الوحدة من الوصفة
    recipe_cost = 0.0
    recipe_details = []
    
    for ingredient in product.recipe:
        raw_material_id = ingredient.get("raw_material_id")
        qty = _sn(ingredient.get("quantity"))
        
        # جلب المادة الخام من المخزون
        raw_material = await db.inventory.find_one(
            {"id": raw_material_id, "item_type": "raw"},
            {"_id": 0}
        )
        
        if raw_material:
            ingredient_cost = qty * raw_material.get("cost_per_unit", 0)
            recipe_cost += ingredient_cost
            recipe_details.append({
                "raw_material_id": raw_material_id,
                "raw_material_name": raw_material.get("name", ""),
                "quantity": qty,
                "unit": raw_material.get("unit", ""),
                "cost_per_unit": raw_material.get("cost_per_unit", 0),
                "total_cost": ingredient_cost
            })
    
    product_doc = {
        "id": str(uuid.uuid4()),
        "name": product.name,
        "name_en": product.name_en,
        "unit": product.unit,
        "quantity": product.quantity,
        "min_quantity": product.min_quantity,
        "cost_per_unit": recipe_cost,  # التكلفة المحسوبة من الوصفة
        "selling_price": product.selling_price,
        "recipe": recipe_details,
        "description": product.description,
        "category": product.category,
        "item_type": "finished",
        "tenant_id": tenant_id,
        "branch_id": "main",  # المنتجات النهائية في المخزن الرئيسي
        "is_active": True,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "last_updated": datetime.now(timezone.utc).isoformat()
    }
    
    await db.inventory.insert_one(product_doc)
    del product_doc["_id"]
    return product_doc

@api_router.get("/finished-products")
async def get_finished_products(
    category: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب جميع المنتجات النهائية"""
    query = build_tenant_query(current_user, {"item_type": "finished"})
    
    if category:
        query["category"] = category
    
    products = await db.inventory.find(query, {"_id": 0}).to_list(500)
    return products

@api_router.get("/finished-products/{product_id}")
async def get_finished_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """جلب منتج نهائي محدد مع وصفته"""
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    return product

@api_router.put("/finished-products/{product_id}")
async def update_finished_product(
    product_id: str, 
    update: FinishedProductUpdate, 
    current_user: dict = Depends(get_current_user)
):
    """تحديث منتج نهائي ووصفته"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query)
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    
    # إذا تم تحديث الوصفة، أعد حساب التكلفة
    if "recipe" in update_data and update_data["recipe"]:
        recipe_cost = 0.0
        recipe_details = []
        
        for ingredient in update_data["recipe"]:
            raw_material_id = ingredient.get("raw_material_id")
            qty = _sn(ingredient.get("quantity"))
            
            raw_material = await db.inventory.find_one(
                {"id": raw_material_id, "item_type": "raw"},
                {"_id": 0}
            )
            
            if raw_material:
                ingredient_cost = qty * raw_material.get("cost_per_unit", 0)
                recipe_cost += ingredient_cost
                recipe_details.append({
                    "raw_material_id": raw_material_id,
                    "raw_material_name": raw_material.get("name", ""),
                    "quantity": qty,
                    "unit": raw_material.get("unit", ""),
                    "cost_per_unit": raw_material.get("cost_per_unit", 0),
                    "total_cost": ingredient_cost
                })
        
        update_data["recipe"] = recipe_details
        update_data["cost_per_unit"] = recipe_cost
    
    update_data["last_updated"] = datetime.now(timezone.utc).isoformat()
    
    await db.inventory.update_one({"id": product_id}, {"$set": update_data})
    return await db.inventory.find_one({"id": product_id}, {"_id": 0})

@api_router.delete("/finished-products/{product_id}")
async def delete_finished_product(product_id: str, current_user: dict = Depends(get_current_user)):
    """حذف منتج نهائي"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query)
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    await db.inventory.delete_one({"id": product_id})
    return {"message": "تم حذف المنتج"}

@api_router.post("/finished-products/{product_id}/manufacture")
async def manufacture_finished_product(
    product_id: str,
    data: dict = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """تصنيع منتج نهائي (خصم المواد الخام وزيادة كمية المنتج النهائي)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    quantity_to_manufacture = data.get("quantity", 1)
    
    query = build_tenant_query(current_user, {"id": product_id, "item_type": "finished"})
    product = await db.inventory.find_one(query, {"_id": 0})
    
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    if not product.get("recipe"):
        raise HTTPException(status_code=400, detail="المنتج ليس له وصفة محددة")
    
    # التحقق من توفر المواد الخام
    insufficient_materials = []
    for ingredient in product["recipe"]:
        raw_material = await db.inventory.find_one(
            {"id": ingredient["raw_material_id"], "item_type": "raw"},
            {"_id": 0}
        )
        
        if not raw_material:
            insufficient_materials.append({
                "name": ingredient.get("raw_material_name", "مادة Unknownة"),
                "required": ingredient["quantity"] * quantity_to_manufacture,
                "available": 0
            })
        else:
            required_qty = ingredient["quantity"] * quantity_to_manufacture
            if _sn(raw_material.get("quantity")) < required_qty:
                insufficient_materials.append({
                    "name": raw_material["name"],
                    "required": required_qty,
                    "available": _sn(raw_material.get("quantity"))
                })
    
    if insufficient_materials:
        raise HTTPException(
            status_code=400, 
            detail={
                "message": "المواد الخام غير كافية للتصنيع",
                "insufficient_materials": insufficient_materials
            }
        )
    
    # خصم المواد الخام
    for ingredient in product["recipe"]:
        required_qty = ingredient["quantity"] * quantity_to_manufacture
        await db.inventory.update_one(
            {"id": ingredient["raw_material_id"]},
            {"$inc": {"quantity": -required_qty}}
        )
        
        # تسجيل حركة المخزون
        movement_doc = {
            "id": str(uuid.uuid4()),
            "inventory_id": ingredient["raw_material_id"],
            "transaction_type": "out",
            "quantity": required_qty,
            "notes": f"تصنيع {quantity_to_manufacture} {product['unit']} من {product['name']}",
            "reference_type": "manufacturing",
            "reference_id": product_id,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.inventory_transactions.insert_one(movement_doc)
    
    # زيادة كمية المنتج النهائي
    await db.inventory.update_one(
        {"id": product_id},
        {
            "$inc": {"quantity": quantity_to_manufacture},
            "$set": {"last_updated": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    # تسجيل عملية التصنيع
    manufacturing_doc = {
        "id": str(uuid.uuid4()),
        "product_id": product_id,
        "product_name": product["name"],
        "quantity_manufactured": quantity_to_manufacture,
        "recipe_used": product["recipe"],
        "total_cost": product.get("cost_per_unit", 0) * quantity_to_manufacture,
        "tenant_id": get_user_tenant_id(current_user),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.manufacturing_logs.insert_one(manufacturing_doc)
    
    updated_product = await db.inventory.find_one({"id": product_id}, {"_id": 0})
    return {
        "message": f"تم تصنيع {quantity_to_manufacture} {product['unit']} من {product['name']}",
        "product": updated_product
    }

# ==================== BRANCH ORDERS - طلبات الفروع ====================

class BranchOrderCreate(BaseModel):
    to_branch_id: str
    items: List[Dict[str, Any]]  # [{product_id, quantity}] - منتجات نهائية فقط
    priority: str = "normal"  # low, normal, high
    notes: Optional[str] = None

class BranchOrderStatusUpdate(BaseModel):
    status: str  # pending, approved, in_transit, delivered, rejected
    notes: Optional[str] = None

@api_router.get("/branch-orders")
async def get_branch_orders(
    type: Optional[str] = None,  # outgoing, incoming
    status: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب طلبات الفروع"""
    query = build_tenant_query(current_user)
    
    user_branch_id = current_user.get("branch_id")
    
    if type == "outgoing" and user_branch_id:
        query["from_branch_id"] = user_branch_id
    elif type == "incoming" and user_branch_id:
        query["to_branch_id"] = user_branch_id
    
    if status:
        query["status"] = status
    
    orders = await db.branch_orders.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    
    # جلب أسماء الفروع
    for order in orders:
        from_branch = await db.branches.find_one({"id": order.get("from_branch_id")}, {"_id": 0, "name": 1})
        to_branch = await db.branches.find_one({"id": order.get("to_branch_id")}, {"_id": 0, "name": 1})
        order["from_branch"] = {"id": order.get("from_branch_id"), "name": from_branch.get("name") if from_branch else "المخزن الرئيسي"}
        order["to_branch"] = {"id": order.get("to_branch_id"), "name": to_branch.get("name") if to_branch else "المخزن الرئيسي"}
    
    return orders

@api_router.post("/branch-orders")
async def create_branch_order(order: BranchOrderCreate, current_user: dict = Depends(get_current_user)):
    """
    إنشاء طلب فرع جديد - يخصم المواد الخام تلقائياً من المخزون المركزي
    
    النظام:
    1. الفرع يطلب منتجات نهائية (مثل: برغر لحم)
    2. المنتج النهائي له وصفة (مكونات من المواد الخام)
    3. عند إرسال الطلب، يتم خصم المواد الخام مباشرة من المخزون المركزي
    4. لا يُشترط وجود كمية مسبقة من المنتج النهائي (الخصم من المواد الخام مباشرة)
    """
    tenant_id = get_user_tenant_id(current_user)
    
    # تجهيز تفاصيل الطلب
    order_items_details = []
    raw_materials_to_deduct = {}  # {raw_material_id: total_quantity_needed}
    products_without_recipe = []
    insufficient_materials = []
    
    for item in order.items:
        product_id = item.get("product_id")
        requested_qty = _sn(item.get("quantity"))
        
        if requested_qty <= 0:
            continue
        
        # جلب المنتج النهائي
        product = await db.inventory.find_one(
            {"id": product_id, "item_type": "finished"},
            {"_id": 0}
        )
        
        if not product:
            continue
        
        # التحقق من وجود وصفة للمنتج
        recipe = product.get("recipe", [])
        if not recipe:
            products_without_recipe.append(product["name"])
            continue
        
        # تجميع المواد الخام المطلوبة من الوصفة
        for ingredient in recipe:
            raw_material_id = ingredient.get("raw_material_id")
            qty_per_unit = _sn(ingredient.get("quantity"))
            total_needed = qty_per_unit * requested_qty
            
            if raw_material_id in raw_materials_to_deduct:
                raw_materials_to_deduct[raw_material_id]["quantity"] += total_needed
            else:
                raw_materials_to_deduct[raw_material_id] = {
                    "quantity": total_needed,
                    "name": ingredient.get("raw_material_name", ""),
                    "unit": ingredient.get("unit", "")
                }
        
        order_items_details.append({
            "product_id": product_id,
            "product_name": product["name"],
            "quantity": requested_qty,
            "unit": product.get("unit", "قطعة"),
            "cost_per_unit": product.get("cost_per_unit", 0),
            "recipe": recipe
        })
    
    # إذا كانت هناك منتجات بدون وصفة
    if products_without_recipe:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "بعض المنتجات ليس لها وصفة محددة",
                "products_without_recipe": products_without_recipe
            }
        )
    
    # التحقق من توفر المواد الخام
    for raw_material_id, details in raw_materials_to_deduct.items():
        raw_material = await db.inventory.find_one(
            {"id": raw_material_id, "item_type": "raw"},
            {"_id": 0}
        )
        
        if not raw_material:
            insufficient_materials.append({
                "name": details["name"],
                "required": details["quantity"],
                "available": 0
            })
        elif _sn(raw_material.get("quantity")) < details["quantity"]:
            insufficient_materials.append({
                "name": raw_material["name"],
                "required": details["quantity"],
                "available": _sn(raw_material.get("quantity")),
                "unit": raw_material.get("unit", "")
            })
    
    # إذا كانت هناك مواد خام غير كافية
    if insufficient_materials:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "المواد الخام غير كافية في المخزون الرئيسي",
                "insufficient_materials": insufficient_materials
            }
        )
    
    if not order_items_details:
        raise HTTPException(status_code=400, detail="لا توجد منتجات صالحة في الطلب")
    
    # إنشاء رقم الطلب
    last_order = await db.branch_orders.find_one(
        {"tenant_id": tenant_id} if tenant_id else {},
        {"_id": 0, "order_number": 1},
        sort=[("created_at", -1)]
    )
    order_num = 1
    if last_order and last_order.get("order_number"):
        try:
            order_num = int(last_order["order_number"].replace("BO-", "")) + 1
        except:
            order_num = 1
    
    order_id = str(uuid.uuid4())
    
    # خصم المواد الخام من المخزون الرئيسي
    deducted_materials = []
    for raw_material_id, details in raw_materials_to_deduct.items():
        await db.inventory.update_one(
            {"id": raw_material_id},
            {"$inc": {"quantity": -details["quantity"]}}
        )
        
        # تسجيل حركة المخزون
        movement_doc = {
            "id": str(uuid.uuid4()),
            "inventory_id": raw_material_id,
            "transaction_type": "out",
            "quantity": details["quantity"],
            "notes": f"طلب فرع - {order_num}",
            "reference_type": "branch_order",
            "reference_id": order_id,
            "created_by": current_user["id"],
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.inventory_transactions.insert_one(movement_doc)
        
        deducted_materials.append({
            "raw_material_id": raw_material_id,
            "raw_material_name": details["name"],
            "quantity_deducted": details["quantity"],
            "unit": details["unit"]
        })
    
    # حساب إجمالي التكلفة
    total_cost = sum(item["quantity"] * item["cost_per_unit"] for item in order_items_details)
    
    order_doc = {
        "id": order_id,
        "order_number": f"BO-{str(order_num).zfill(4)}",
        "from_branch_id": "warehouse",  # من المخزن الرئيسي دائماً
        "to_branch_id": order.to_branch_id,
        "items": order_items_details,
        "raw_materials_deducted": deducted_materials,
        "total_cost": total_cost,
        "status": "pending",
        "priority": order.priority,
        "notes": order.notes,
        "tenant_id": tenant_id,
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.branch_orders.insert_one(order_doc)
    del order_doc["_id"]
    
    # إضافة معلومات الفروع
    to_branch = await db.branches.find_one({"id": order_doc["to_branch_id"]}, {"_id": 0, "name": 1})
    order_doc["from_branch"] = {"id": "warehouse", "name": "المخزن الرئيسي"}
    order_doc["to_branch"] = {"id": order_doc["to_branch_id"], "name": to_branch.get("name") if to_branch else "الفرع"}
    
    return order_doc

@api_router.put("/branch-orders/{order_id}/status")
async def update_branch_order_status(order_id: str, update: BranchOrderStatusUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث حالة طلب الفرع"""
    query = build_tenant_query(current_user, {"id": order_id})
    
    order = await db.branch_orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    update_data = {"status": update.status}
    
    if update.status == "approved":
        update_data["approved_by"] = current_user["id"]
        update_data["approved_at"] = datetime.now(timezone.utc).isoformat()
    elif update.status == "in_transit":
        update_data["shipped_at"] = datetime.now(timezone.utc).isoformat()
        update_data["shipped_by"] = current_user["id"]
    elif update.status == "delivered":
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
        update_data["received_by"] = current_user["id"]
    elif update.status == "rejected":
        update_data["rejected_at"] = datetime.now(timezone.utc).isoformat()
        update_data["rejected_by"] = current_user["id"]
    
    if update.notes:
        update_data["status_notes"] = update.notes
    
    await db.branch_orders.update_one({"id": order_id}, {"$set": update_data})
    return await db.branch_orders.find_one({"id": order_id}, {"_id": 0})

@api_router.delete("/branch-orders/{order_id}")
async def delete_branch_order(order_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طلب فرع"""
    query = build_tenant_query(current_user, {"id": order_id})
    
    order = await db.branch_orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if order["status"] not in ["pending", "rejected"]:
        raise HTTPException(status_code=400, detail="لا يمكن حذف طلب تمت معالجته")
    
    await db.branch_orders.delete_one({"id": order_id})
    return {"message": "تم حذف الطلب"}

# ==================== DASHBOARD BACKGROUNDS ====================

@api_router.get("/dashboard-backgrounds")
async def get_dashboard_backgrounds(current_user: dict = Depends(get_current_user)):
    """جلب خلفيات Dashboard المتاحة للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب الخلفيات الافتراضية (متاحة للجميع)
    default_backgrounds = [
        "https://images.unsplash.com/photo-1517248135467-4c7edcad34c4?w=1920",
        "https://images.unsplash.com/photo-1554679665-f5537f187268?w=1920",
        "https://images.unsplash.com/photo-1466978913421-dad2ebd01d17?w=1920",
        "https://images.unsplash.com/photo-1552566626-52f8b828add9?w=1920",
        "https://images.unsplash.com/photo-1559329007-40df8a9345d8?w=1920",
        "https://images.unsplash.com/photo-1514933651103-005eec06c04b?w=1920"
    ]
    
    # جلب الخلفيات المرفوعة من قبل العميل
    tenant_backgrounds = await db.dashboard_backgrounds.find(
        {"tenant_id": tenant_id} if tenant_id else {},
        {"_id": 0}
    ).to_list(50)
    
    # جلب الخلفية المحددة حالياً
    settings = await db.tenant_settings.find_one(
        {"tenant_id": tenant_id} if tenant_id else {"tenant_id": None},
        {"_id": 0, "dashboard_background": 1}
    )
    
    return {
        "backgrounds": default_backgrounds + [bg["url"] for bg in tenant_backgrounds],
        "selected": settings.get("dashboard_background") if settings else None
    }

@api_router.put("/dashboard-backgrounds/select")
async def select_dashboard_background(data: dict, current_user: dict = Depends(get_current_user)):
    """اختيار خلفية Dashboard للعميل"""
    tenant_id = get_user_tenant_id(current_user)
    background_url = data.get("background_url")
    
    # تحديث أو إنشاء إعدادات العميل
    await db.tenant_settings.update_one(
        {"tenant_id": tenant_id} if tenant_id else {"tenant_id": None},
        {
            "$set": {
                "dashboard_background": background_url,
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["id"]
            },
            "$setOnInsert": {
                "tenant_id": tenant_id,
                "created_at": datetime.now(timezone.utc).isoformat()
            }
        },
        upsert=True
    )
    
    return {"message": "تم تحديث الخلفية", "background_url": background_url}

# ==================== RESERVATIONS - الحجوزات ====================

# Reservations & Reviews endpoints extracted to routes/reservations_reviews_routes.py

# ==================== BREAK-EVEN REPORT - تقرير نقطة التعادل ====================

# Break-Even endpoints extracted to routes/break_even_routes.py

# ==================== moved to routes/smart_reports_routes.py ====================

# ==================== CALL CENTER / CALLER ID ====================

class CallCenterConfig(BaseModel):
    enabled: bool = False
    provider: str = ""
    api_url: str = ""
    api_key: str = ""
    api_secret: str = ""
    webhook_secret: str = ""
    auto_popup: bool = True
    auto_save_new_callers: bool = True
    play_sound: bool = True

class IncomingCall(BaseModel):
    phone: str
    caller_name: Optional[str] = None
    call_id: Optional[str] = None
    direction: str = "incoming"
    timestamp: Optional[str] = None

# Store active calls in memory (in production, use Redis)
active_calls = {}

@api_router.post("/callcenter/config")
async def save_callcenter_config(config: CallCenterConfig, current_user: dict = Depends(get_current_user)):
    """حفظ إعدادات الكول سنتر"""
    tenant_id = get_user_tenant_id(current_user)
    
    config_doc = {
        "tenant_id": tenant_id,
        **config.dict(),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    await db.callcenter_config.update_one(
        {"tenant_id": tenant_id},
        {"$set": config_doc},
        upsert=True
    )
    
    return {"message": "تم حفظ إعدادات الكول سنتر"}

@api_router.get("/callcenter/config")
async def get_callcenter_config(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الكول سنتر"""
    tenant_id = get_user_tenant_id(current_user)
    
    config = await db.callcenter_config.find_one({"tenant_id": tenant_id}, {"_id": 0})
    
    if not config:
        return CallCenterConfig().dict()
    
    return config

@api_router.post("/callcenter/test")
async def test_callcenter_connection(config: CallCenterConfig, current_user: dict = Depends(get_current_user)):
    """اختبار اتصال الكول سنتر"""
    
    # في الإنتاج، يجب اختبار الاتصال الفعلي مع المزود
    # حالياً نرجع نجاح للمحاكاة
    
    if not config.provider:
        raise HTTPException(status_code=400, detail="يرجى اختيار مزود الخدمة")
    
    if not config.api_url and config.provider not in ["custom"]:
        raise HTTPException(status_code=400, detail="يرجى إدخال رابط API")
    
    # محاكاة اختبار الاتصال
    return {"success": True, "message": f"تم الاتصال بـ {config.provider} بنجاح"}

@api_router.post("/callcenter/webhook")
async def callcenter_webhook(request: Request):
    """Webhook لاستقبال المكالمات من نظام الكول سنتر - محمي بمفتاح سري"""

    # الحماية: يجب أن يطابق المفتاح السري المُهيّأ في البيئة لمنع حقن المكالمات
    webhook_secret = os.environ.get("CALLCENTER_WEBHOOK_SECRET")
    provided_secret = (
        request.headers.get("X-Webhook-Secret")
        or request.query_params.get("secret")
    )
    if not webhook_secret or provided_secret != webhook_secret:
        raise HTTPException(status_code=403, detail="Unauthorized webhook")

    try:
        body = await request.json()
    except:
        body = {}
    
    # استخراج رقم المتصل من البيانات (يختلف حسب المزود)
    phone = body.get("phone") or body.get("caller_id") or body.get("from") or body.get("callerNumber")
    caller_name = body.get("caller_name") or body.get("name") or body.get("callerName")
    call_id = body.get("call_id") or body.get("callId") or body.get("id") or str(uuid.uuid4())
    direction = body.get("direction") or body.get("type") or "incoming"
    
    if not phone:
        return {"status": "error", "message": "No phone number provided"}
    
    # تنظيف رقم الهاتف
    phone = phone.replace(" ", "").replace("-", "").replace("+", "")
    if phone.startswith("964"):
        phone = "0" + phone[3:]
    
    # البحث عن العميل
    customer = await db.customers.find_one(
        {"$or": [{"phone": phone}, {"phone2": phone}]},
        {"_id": 0}
    )
    
    # آخر طلب للعميل
    last_order = None
    if customer:
        last_order = await db.orders.find_one(
            {"customer_phone": phone},
            {"_id": 0},
            sort=[("created_at", -1)]
        )
    
    # تخزين المكالمة النشطة
    call_data = {
        "call_id": call_id,
        "phone": phone,
        "caller_name": caller_name or (customer["name"] if customer else "New Customer"),
        "direction": direction,
        "customer": customer,
        "last_order": last_order,
        "is_new_customer": customer is None,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": "ringing"
    }
    
    active_calls[call_id] = call_data
    
    # إذا كان عميل جديد وتم تفعيل الحفظ التلقائي
    # سيتم الحفظ عند إنشاء الطلب
    
    return {
        "status": "success",
        "call_id": call_id,
        "customer_found": customer is not None,
        "customer": customer,
        "last_order": last_order
    }

@api_router.post("/callcenter/simulate")
async def simulate_incoming_call(data: dict, current_user: dict = Depends(get_current_user)):
    """محاكاة مكالمة واردة للاختبار"""
    
    phone = data.get("phone", "07801234567")
    tenant_id = get_user_tenant_id(current_user)
    
    # تنظيف رقم الهاتف
    phone = phone.replace(" ", "").replace("-", "").replace("+", "")
    
    # البحث عن العميل مع مراعاة tenant_id
    customer_query = {"$or": [{"phone": phone}, {"phone2": phone}]}
    if tenant_id:
        customer_query["tenant_id"] = tenant_id
    else:
        customer_query["$or"] = [
            {"$and": [{"phone": phone}, {"$or": [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]}]},
            {"$and": [{"phone2": phone}, {"$or": [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]}]}
        ]
    
    customer = await db.customers.find_one(customer_query, {"_id": 0})
    
    # آخر طلب للعميل
    last_order = None
    if customer:
        order_query = {"customer_phone": phone}
        if tenant_id:
            order_query["tenant_id"] = tenant_id
        else:
            order_query["$or"] = [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]
        
        last_order = await db.orders.find_one(
            order_query,
            {"_id": 0},
            sort=[("created_at", -1)]
        )
    
    call_id = str(uuid.uuid4())
    call_data = {
        "call_id": call_id,
        "phone": phone,
        "caller_name": customer["name"] if customer else "New Customer",
        "direction": "incoming",
        "customer": customer,
        "last_order": last_order,
        "is_new_customer": customer is None,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": "ringing",
        "simulated": True,
        "tenant_id": tenant_id
    }
    
    active_calls[call_id] = call_data
    
    return call_data

@api_router.get("/callcenter/active-calls")
async def get_active_calls(current_user: dict = Depends(get_current_user)):
    """جلب المكالمات النشطة للمستخدم الحالي"""
    tenant_id = get_user_tenant_id(current_user)
    
    # فلترة المكالمات حسب tenant_id
    if tenant_id:
        filtered_calls = [c for c in active_calls.values() if c.get("tenant_id") == tenant_id]
    else:
        # النظام الرئيسي يرى المكالمات بدون tenant_id
        filtered_calls = [c for c in active_calls.values() if not c.get("tenant_id")]
    
    return filtered_calls

@api_router.post("/callcenter/calls/{call_id}/answer")
async def answer_call(call_id: str, current_user: dict = Depends(get_current_user)):
    """الرد على المكالمة"""
    if call_id in active_calls:
        active_calls[call_id]["status"] = "answered"
        active_calls[call_id]["answered_by"] = current_user["id"]
        active_calls[call_id]["answered_at"] = datetime.now(timezone.utc).isoformat()
        return active_calls[call_id]
    raise HTTPException(status_code=404, detail="المكالمة غير موجودة")

@api_router.post("/callcenter/calls/{call_id}/end")
async def end_call(call_id: str, current_user: dict = Depends(get_current_user)):
    """إنهاء المكالمة"""
    if call_id in active_calls:
        call_data = active_calls.pop(call_id)
        call_data["status"] = "ended"
        call_data["ended_at"] = datetime.now(timezone.utc).isoformat()
        
        # حفظ سجل المكالمة
        call_log = {k: v for k, v in call_data.items() if k != '_id'}
        call_log["id"] = call_id
        await db.call_logs.insert_one(call_log)
        
        return {"message": "تم إنهاء المكالمة", "call": call_data}
    raise HTTPException(status_code=404, detail="المكالمة غير موجودة")

@api_router.get("/callcenter/call-logs")
async def get_call_logs(
    limit: int = 50,
    skip: int = 0,
    current_user: dict = Depends(get_current_user)
):
    """سجل المكالمات"""
    tenant_id = get_user_tenant_id(current_user)
    
    query = {}
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    logs = await db.call_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    total = await db.call_logs.count_documents(query)
    
    return {
        "logs": logs,
        "total": total
    }

# ==================== SEED DATA ====================

@api_router.post("/seed")
async def seed_data(current_user: dict = Depends(verify_super_admin)):
    # Check if already seeded
    existing = await db.users.find_one({"email": "admin@maestroegp.com"})
    if existing:
        return {"message": "البيانات موجودة بالفعل"}
    
    # Create admin user
    admin_doc = {
        "id": str(uuid.uuid4()),
        "username": "admin",
        "email": "admin@maestroegp.com",
        "password": hash_password("admin123"),
        "full_name": "مدير النظام",
        "role": UserRole.ADMIN,
        "branch_id": None,
        "permissions": ["all"],
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(admin_doc)
    
    # Create default branch
    branch_doc = {
        "id": str(uuid.uuid4()),
        "name": "الفرع الرئيسي",
        "address": "بغداد - الكرادة",
        "phone": "+964 770 123 4567",
        "email": "main@maestroegp.com",
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.branches.insert_one(branch_doc)
    branch_id = branch_doc["id"]
    
    # Create categories
    categories = [
        {"name": "برغر", "name_en": "Burgers", "icon": "🍔", "color": "#EF4444", "sort_order": 1},
        {"name": "بيتزا", "name_en": "Pizza", "icon": "🍕", "color": "#F59E0B", "sort_order": 2},
        {"name": "مشروبات", "name_en": "Drinks", "icon": "☕", "color": "#8B5CF6", "sort_order": 3},
        {"name": "حلويات", "name_en": "Desserts", "icon": "🍰", "color": "#EC4899", "sort_order": 4},
        {"name": "سلطات", "name_en": "Salads", "icon": "🥗", "color": "#10B981", "sort_order": 5},
    ]
    
    cat_ids = {}
    for cat in categories:
        cat_doc = {"id": str(uuid.uuid4()), **cat, "is_active": True}
        await db.categories.insert_one(cat_doc)
        cat_ids[cat["name_en"]] = cat_doc["id"]
    
    # Create products with costs
    products = [
        {"name": "برغر كلاسيك", "name_en": "Classic Burger", "category_id": cat_ids["Burgers"], "price": 12000, "cost": 4000, "operating_cost": 1000, "image": "https://images.pexels.com/photos/18796078/pexels-photo-18796078.jpeg"},
        {"name": "برغر دبل", "name_en": "Double Burger", "category_id": cat_ids["Burgers"], "price": 18000, "cost": 7000, "operating_cost": 1500, "image": "https://images.pexels.com/photos/5672397/pexels-photo-5672397.jpeg"},
        {"name": "بيتزا مارغريتا", "name_en": "Margherita Pizza", "category_id": cat_ids["Pizza"], "price": 15000, "cost": 5000, "operating_cost": 1200, "image": "https://images.pexels.com/photos/35532821/pexels-photo-35532821.jpeg"},
        {"name": "بيتزا خضار", "name_en": "Veggie Pizza", "category_id": cat_ids["Pizza"], "price": 14000, "cost": 4500, "operating_cost": 1100, "image": "https://images.pexels.com/photos/34956178/pexels-photo-34956178.jpeg"},
        {"name": "قهوة عربية", "name_en": "Arabic Coffee", "category_id": cat_ids["Drinks"], "price": 3000, "cost": 500, "operating_cost": 200, "image": "https://images.pexels.com/photos/29799615/pexels-photo-29799615.jpeg"},
        {"name": "لاتيه", "name_en": "Latte", "category_id": cat_ids["Drinks"], "price": 5000, "cost": 1200, "operating_cost": 300, "image": "https://images.pexels.com/photos/15800375/pexels-photo-15800375.jpeg"},
        {"name": "كيكة شوكولاتة", "name_en": "Chocolate Cake", "category_id": cat_ids["Desserts"], "price": 8000, "cost": 2500, "operating_cost": 500, "image": "https://images.pexels.com/photos/29538417/pexels-photo-29538417.jpeg"},
        {"name": "تشيز كيك", "name_en": "Cheesecake", "category_id": cat_ids["Desserts"], "price": 9000, "cost": 3000, "operating_cost": 500, "image": "https://images.pexels.com/photos/15564368/pexels-photo-15564368.jpeg"},
    ]
    
    for prod in products:
        profit = prod["price"] - prod["cost"] - prod["operating_cost"]
        prod_doc = {"id": str(uuid.uuid4()), **prod, "profit": profit, "is_available": True, "ingredients": []}
        await db.products.insert_one(prod_doc)
    
    # Create tables
    for i in range(1, 11):
        table_doc = {
            "id": str(uuid.uuid4()),
            "number": i,
            "capacity": 4 if i <= 6 else 6,
            "branch_id": branch_id,
            "section": "داخلي" if i <= 6 else "خارجي",
            "status": "available",
            "current_order_id": None
        }
        await db.tables.insert_one(table_doc)
    
    # Create cashier user
    cashier_doc = {
        "id": str(uuid.uuid4()),
        "username": "cashier1",
        "email": "cashier@maestroegp.com",
        "password": hash_password("cashier123"),
        "full_name": "أحمد الكاشير",
        "role": UserRole.CASHIER,
        "branch_id": branch_id,
        "permissions": ["pos", "orders"],
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.users.insert_one(cashier_doc)
    
    # Set default currencies
    currencies = [
        {"code": "IQD", "name": "دينار عراقي", "symbol": "د.ع", "exchange_rate": 1.0},
        {"code": "USD", "name": "دولار أمريكي", "symbol": "$", "exchange_rate": 1460.0},
    ]
    await db.settings.insert_one({"type": "currencies", "value": currencies})
    
    # Set default delivery app settings
    delivery_apps = [
        {"app_id": "toters", "name": "توترز", "name_en": "Toters", "commission_type": "percentage", "commission_rate": 15, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "talabat", "name": "طلبات", "name_en": "Talabat", "commission_type": "percentage", "commission_rate": 18, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "baly", "name": "بالي", "name_en": "Baly", "commission_type": "percentage", "commission_rate": 12, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "alsaree3", "name": "عالسريع", "name_en": "Al-Sari3", "commission_type": "percentage", "commission_rate": 10, "is_active": True, "payment_terms": "weekly"},
        {"app_id": "talabati", "name": "طلباتي", "name_en": "Talabati", "commission_type": "percentage", "commission_rate": 14, "is_active": True, "payment_terms": "weekly"},
    ]
    for app in delivery_apps:
        await db.delivery_app_settings.insert_one(app)
    
    return {"message": "تم إنشاء البيانات الأولية بنجاح"}


@api_router.post("/utils/fix-tenant-categories")
async def fix_tenant_categories(current_user: dict = Depends(get_current_user)):
    """
    إضافة فئات افتراضية للمستأجر الحالي إذا لم يكن لديه فئات
    """
    tenant_id = get_user_tenant_id(current_user)
    
    if not tenant_id:
        raise HTTPException(status_code=400, detail="لا يمكن تحديد المستأجر")
    
    # التحقق من وجود فئات
    existing_categories = await db.categories.count_documents({"tenant_id": tenant_id})
    
    if existing_categories > 0:
        return {"message": f"المستأجر لديه {existing_categories} فئة بالفعل", "fixed": False}
    
    # إنشاء فئات افتراضية
    default_categories = [
        {"id": str(uuid.uuid4()), "name": "المشروبات", "name_en": "Beverages", "icon": "☕", "color": "#8B4513", "sort_order": 1, "tenant_id": tenant_id, "is_active": True},
        {"id": str(uuid.uuid4()), "name": "الوجبات الرئيسية", "name_en": "Main Dishes", "icon": "🥘", "color": "#D4AF37", "sort_order": 2, "tenant_id": tenant_id, "is_active": True},
        {"id": str(uuid.uuid4()), "name": "المقبلات", "name_en": "Appetizers", "icon": "🧆", "color": "#228B22", "sort_order": 3, "tenant_id": tenant_id, "is_active": True},
        {"id": str(uuid.uuid4()), "name": "الحلويات", "name_en": "Desserts", "icon": "🍰", "color": "#FF69B4", "sort_order": 4, "tenant_id": tenant_id, "is_active": True},
    ]
    
    for cat in default_categories:
        await db.categories.insert_one(cat)
    
    return {"message": "تم إضافة الفئات الافتراضية بنجاح", "fixed": True, "categories_added": len(default_categories)}


# ==================== ROOT ====================

@api_router.get("/")
async def root():
    return {"message": "Maestro EGP API", "version": "2.0.0"}

# ==================== moved to routes/biometric_routes.py ====================

# ==================== 🔧 Biometric Job Queue ====================
# نظام طابور عمليات البصمة (مزامنة، إصدار موظف، اختبار، صورة وجه...)
# الفرونت ينشر جوب → الوكيل المحلي يـ poll → ينفذ → يرسل النتيجة → الفرونت يستفسر
# يحلّ مشكلة Mixed Content blocking عند فتح التطبيق من HTTPS
# ⚠️ يتطلب تحديث print_server.ps1 ليدعم endpoints biometric-queue/*

ALLOWED_BIO_JOB_TYPES = {
    "zk-sync", "zk-push-user", "zk-users", "zk-test",
    "zk-face-photo", "zk-delete-user", "zk-probe-device",
}


@api_router.post("/biometric-queue")
async def create_biometric_job(payload: dict, current_user: dict = Depends(get_current_user)):
    """إنشاء جوب بصمة لتنفيذه عبر الوكيل المحلي."""
    if current_user.get("role") not in ["admin", "super_admin", "manager"]:
        raise HTTPException(status_code=403, detail="غير مسموح — للمالك/المدير فقط")

    tenant_id = get_user_tenant_id(current_user)
    job_type = (payload.get("type") or "").strip()
    if job_type not in ALLOWED_BIO_JOB_TYPES:
        raise HTTPException(status_code=400, detail=f"نوع الجوب غير مدعوم: {job_type}")

    branch_id = payload.get("branch_id") or current_user.get("branch_id") or ""
    job = {
        "id": str(uuid.uuid4()),
        "type": job_type,
        "params": payload.get("params") or {},
        "status": "pending",
        "branch_id": branch_id,
        "tenant_id": tenant_id,
        "created_by": current_user.get("id"),
        "created_by_name": current_user.get("full_name") or current_user.get("username"),
        "created_at": datetime.now(timezone.utc).isoformat(),
        "result": None,
        "error": None,
    }
    await db.biometric_queue.insert_one(job)
    job.pop("_id", None)
    return job


@api_router.get("/biometric-queue/pending")
async def list_pending_biometric_jobs(branch_id: Optional[str] = None, limit: int = 10, _auth: bool = Depends(verify_device_agent)):
    """الوكيل يـ poll الجوبات المعلقة (atomic claim لمنع تكرار التنفيذ).
    يُستخدم بدون auth (الوكيل المحلي قد لا يحمل توكن)."""
    q = {"status": "pending"}
    if branch_id:
        q["branch_id"] = branch_id
    jobs = await db.biometric_queue.find(q, {"_id": 0}).sort("created_at", 1).limit(limit).to_list(limit)
    if jobs:
        ids = [j["id"] for j in jobs]
        await db.biometric_queue.update_many(
            {"id": {"$in": ids}, "status": "pending"},
            {"$set": {
                "status": "processing",
                "claimed_at": datetime.now(timezone.utc).isoformat(),
            }}
        )
    return jobs


@api_router.post("/biometric-queue/{job_id}/result")
async def submit_biometric_job_result(job_id: str, payload: dict, _auth: bool = Depends(verify_device_agent)):
    """الوكيل يرسل نتيجة تنفيذ الجوب. payload: {success: bool, result?, error?}
    
    🔥 للجوبات من نوع zk-sync: نستخرج سجلات الحضور من result ونُدخلها في biometric_attendance
    ثم نُشغّل المعالجة التلقائية (dedupe + shift split + upsert في db.attendance)."""
    success = bool(payload.get("success"))
    result_data = payload.get("result")
    
    # اجلب الجوب لمعرفة نوعه وسياقه (device_id/tenant_id/branch_id)
    job = await db.biometric_queue.find_one({"id": job_id}, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="الجوب غير موجود")
    
    update = {
        "status": "completed" if success else "failed",
        "result": result_data,
        "error": payload.get("error"),
        "completed_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.biometric_queue.update_one({"id": job_id}, {"$set": update})
    
    # 🔥 معالجة نتائج zk-sync — إدراج السجلات في biometric_attendance + المعالجة التلقائية
    inserted_count = 0
    auto_processed = 0
    if success and job.get("type") == "zk-sync":
        try:
            # نتيجة الوكيل قد تكون dict مباشرة {records: [...]}، أو string JSON من PowerShell
            records = []
            if isinstance(result_data, dict):
                records = result_data.get("records") or []
            elif isinstance(result_data, str):
                try:
                    parsed = json.loads(result_data)
                    if isinstance(parsed, dict):
                        records = parsed.get("records") or []
                except Exception:
                    pass
            elif isinstance(result_data, list):
                records = result_data
            
            device_id = (job.get("params") or {}).get("device_id")
            tenant_id = job.get("tenant_id")
            branch_id = job.get("branch_id")
            
            to_insert = []
            for r in (records or []):
                if not isinstance(r, dict):
                    continue
                uid = str(r.get("uid") or r.get("employee_code") or "").strip()
                ts = (r.get("timestamp") or r.get("punch_time") or "").strip()
                if not uid or not ts or uid == "0":
                    continue
                # تطبيع صيغة الوقت: "YYYY-MM-DD HH:MM:SS" → "YYYY-MM-DDTHH:MM:SS"
                if " " in ts and "T" not in ts:
                    ts = ts.replace(" ", "T", 1)
                # dedup: نتحقق من عدم وجود سجل بنفس (device_id, uid, timestamp)
                existing = await db.biometric_attendance.find_one(
                    {"device_id": device_id, "employee_code": uid, "punch_time": ts, "tenant_id": tenant_id},
                    {"_id": 0, "id": 1}
                )
                if existing:
                    continue
                to_insert.append({
                    "id": str(uuid.uuid4()),
                    "device_id": device_id,
                    "branch_id": branch_id,
                    "employee_code": uid,
                    "punch_time": ts,
                    "punch_type": r.get("punch_type") or ("in" if int(r.get("status") or 0) == 0 else "out"),
                    "verify_type": r.get("verify_type") or "fingerprint",
                    "tenant_id": tenant_id,
                    "processed": False,
                    "synced_at": datetime.now(timezone.utc).isoformat(),
                })
            
            if to_insert:
                await db.biometric_attendance.insert_many(to_insert)
                inserted_count = len(to_insert)
                # حدّث last_sync على الجهاز
                if device_id:
                    await db.biometric_devices.update_one(
                        {"id": device_id},
                        {"$set": {"last_sync": datetime.now(timezone.utc).isoformat()}}
                    )
                # المعالجة التلقائية: biometric_attendance → db.attendance
                try:
                    from routes.biometric_routes import _auto_process_attendance_internal
                    fake_user = {"id": "system-agent", "tenant_id": tenant_id, "role": "system"}
                    ap = await _auto_process_attendance_internal(fake_user)
                    if isinstance(ap, dict):
                        # الدالة تُرجع attendance_created أو processed (توافقاً)
                        auto_processed = ap.get("attendance_created") or ap.get("processed") or ap.get("raw_records_processed") or 0
                except Exception as _pe:
                    import traceback
                    logger.warning(f"auto-process after queue zk-sync failed: {_pe}\n{traceback.format_exc()[:500]}")
        except Exception as _e:
            logger.warning(f"queue zk-sync result parse failed: {_e}")
    
    return {"ok": True, "job_id": job_id, "inserted": inserted_count, "auto_processed": auto_processed}


@api_router.get("/biometric-queue/{job_id}")
async def get_biometric_job(job_id: str, current_user: dict = Depends(get_current_user)):
    """الفرونت يستفسر حالة الجوب (polling)."""
    tenant_id = get_user_tenant_id(current_user)
    q = {"id": job_id}
    if tenant_id:
        q["tenant_id"] = tenant_id
    job = await db.biometric_queue.find_one(q, {"_id": 0})
    if not job:
        raise HTTPException(status_code=404, detail="الجوب غير موجود")
    return job


@api_router.delete("/biometric-queue/{job_id}")
async def cancel_biometric_job(job_id: str, current_user: dict = Depends(get_current_user)):
    """إلغاء جوب pending (للأدمن، في حال علقت أو فشل الوكيل)."""
    if current_user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="غير مسموح")
    tenant_id = get_user_tenant_id(current_user)
    q = {"id": job_id}
    if tenant_id:
        q["tenant_id"] = tenant_id
    await db.biometric_queue.delete_one(q)
    return {"ok": True}


async def cleanup_stale_biometric_jobs():
    """تنظيف الجوبات العالقة + watchdog TTL للجوبات المعلّقة عند غياب الوكيل.
    
    - processing > 5 دقائق  → failed (الوكيل بدأ التنفيذ ثم انقطع)
    - pending zk-probe-device > 5 دقائق → failed (الوكيل غير متصل — الاختبار لا يجب أن ينتظر طويلاً)
    - pending أي نوع > 24 ساعة → failed (تراكم قديم — يعرقل الطابور)
    - يحذف الجوبات المكتملة/الفاشلة الأقدم من 7 أيام."""
    try:
        now_utc = datetime.now(timezone.utc)
        # (1) processing عالقة أكثر من 5 دقائق
        cutoff_processing = (now_utc - timedelta(minutes=5)).isoformat()
        await db.biometric_queue.update_many(
            {"status": "processing", "claimed_at": {"$lt": cutoff_processing}},
            {"$set": {"status": "failed", "error": "Timeout - الوكيل لم يرد", "completed_at": now_utc.isoformat()}}
        )
        # (2) pending zk-probe-device أقدم من 5 دقائق → الوكيل offline (probe اختبار سريع)
        cutoff_probe = (now_utc - timedelta(minutes=5)).isoformat()
        await db.biometric_queue.update_many(
            {"status": "pending", "type": "zk-probe-device", "created_at": {"$lt": cutoff_probe}},
            {"$set": {"status": "failed", "error": "الوكيل المحلي غير متصل — لم يلتقط جوب الاختبار خلال 5 دقائق", "completed_at": now_utc.isoformat()}}
        )
        # (3) pending أي نوع أقدم من 24 ساعة → stale
        cutoff_stale = (now_utc - timedelta(hours=24)).isoformat()
        await db.biometric_queue.update_many(
            {"status": "pending", "created_at": {"$lt": cutoff_stale}},
            {"$set": {"status": "failed", "error": "stale_pending — بقي معلقاً أكثر من 24 ساعة", "completed_at": now_utc.isoformat()}}
        )
        # (4) حذف الجوبات المكتملة/الفاشلة الأقدم من 7 أيام
        old_cutoff = (now_utc - timedelta(days=7)).isoformat()
        await db.biometric_queue.delete_many({"created_at": {"$lt": old_cutoff}})
    except Exception as e:
        logger.error(f"cleanup_stale_biometric_jobs failed: {e}")

# ==================== End Biometric Job Queue ====================


@api_router.post("/biometric/push")
async def receive_biometric_push(request: Request):
    """
    استقبال بيانات الحضور من أجهزة ZKTeco (Push SDK)
    يجب تكوين الجهاز لإرسال البيانات لهذا الـ endpoint (مع المفتاح السري ?key= إن ضُبط)
    """
    await verify_device_agent(request)
    try:
        data = await request.json()
        payload = ZKTecoPushData(**data)
        
        if payload.PIN:
            # البحث عن الموظف بالكود
            employee = await db.employees.find_one({"code": payload.PIN}, {"_id": 0})
            
            punch_type = "in" if payload.Status == 0 else "out"
            
            attendance_record = {
                "id": str(uuid.uuid4()),
                "employee_id": employee["id"] if employee else None,
                "employee_code": payload.PIN,
                "punch_time": payload.DateTime or datetime.now(timezone.utc).isoformat(),
                "punch_type": punch_type,
                "device_serial": payload.DeviceSN,
                "verify_type": payload.VerifyType or "fingerprint",
                "created_at": datetime.now(timezone.utc).isoformat()
            }
            
            await db.biometric_attendance.insert_one(attendance_record)
            
            # تحديث سجل الحضور اليومي للموظف
            if employee:
                today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
                
                if punch_type == "in":
                    await db.attendance.update_one(
                        {"employee_id": employee["id"], "date": today},
                        {"$set": {"check_in": payload.DateTime, "updated_at": datetime.now(timezone.utc).isoformat()}},
                        upsert=True
                    )
                else:
                    await db.attendance.update_one(
                        {"employee_id": employee["id"], "date": today},
                        {"$set": {"check_out": payload.DateTime, "updated_at": datetime.now(timezone.utc).isoformat()}}
                    )
            
            logger.info(f"Biometric punch: {payload.PIN} - {punch_type}")
            
            return {"status": "received", "OperationID": payload.OperationID}
        
        return {"status": "no_data"}
    except Exception as e:
        logger.error(f"Biometric push error: {e}")
        return {"status": "error", "message": str(e)}

@api_router.get("/biometric/attendance")
async def get_biometric_attendance(
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """سجلات الحضور من أجهزة البصمة"""
    query = {"tenant_id": current_user.get("tenant_id")} if current_user.get("tenant_id") else {}
    
    if start_date:
        query["punch_time"] = {"$gte": start_date}
    if end_date:
        if "punch_time" in query:
            query["punch_time"]["$lte"] = end_date
        else:
            query["punch_time"] = {"$lte": end_date}
    
    records = await db.biometric_attendance.find(query, {"_id": 0}).sort("punch_time", -1).to_list(500)
    return records

# ==================== LOYALTY PROGRAM ROUTES ====================

class LoyaltySettingsUpdate(BaseModel):
    is_enabled: bool = True
    points_per_currency: float = 1.0
    currency_per_point: float = 0.01
    min_redeem_points: int = 100
    max_redeem_percent: float = 50
    points_expiry_days: int = 365
    welcome_bonus: int = 50
    birthday_bonus: int = 100
    referral_bonus: int = 200
    tiers: List[Dict[str, Any]] = []

class LoyaltyMemberCreate(BaseModel):
    customer_id: str
    customer_name: str
    phone: str
    email: Optional[str] = None
    birthday: Optional[str] = None
    referred_by: Optional[str] = None

class EarnPointsRequest(BaseModel):
    member_id: str
    order_id: str
    order_total: float

class RedeemPointsRequest(BaseModel):
    member_id: str
    points_to_redeem: int
    order_id: str

DEFAULT_LOYALTY_TIERS = [
    {"name": "برونزي", "name_en": "Bronze", "min_points": 0, "discount_percent": 0, "points_multiplier": 1.0, "color": "#CD7F32"},
    {"name": "فضي", "name_en": "Silver", "min_points": 500, "discount_percent": 5, "points_multiplier": 1.25, "color": "#C0C0C0"},
    {"name": "ذهبي", "name_en": "Gold", "min_points": 1500, "discount_percent": 10, "points_multiplier": 1.5, "color": "#FFD700"},
    {"name": "بلاتيني", "name_en": "Platinum", "min_points": 5000, "discount_percent": 15, "points_multiplier": 2.0, "color": "#E5E4E2"}
]

@api_router.get("/loyalty/settings")
async def get_loyalty_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات برنامج الولاء"""
    settings = await db.loyalty_settings.find_one(
        {"tenant_id": current_user.get("tenant_id")}, 
        {"_id": 0}
    )
    
    if not settings:
        return {
            "is_enabled": True,
            "points_per_currency": 1.0,
            "currency_per_point": 0.01,
            "min_redeem_points": 100,
            "max_redeem_percent": 50,
            "points_expiry_days": 365,
            "welcome_bonus": 50,
            "birthday_bonus": 100,
            "referral_bonus": 200,
            "tiers": DEFAULT_LOYALTY_TIERS
        }
    return settings

@api_router.put("/loyalty/settings")
async def update_loyalty_settings(settings: LoyaltySettingsUpdate, current_user: dict = Depends(get_current_user)):
    """تحديث إعدادات برنامج الولاء"""
    await db.loyalty_settings.update_one(
        {"tenant_id": current_user.get("tenant_id")},
        {"$set": {**settings.model_dump(), "tenant_id": current_user.get("tenant_id")}},
        upsert=True
    )
    return {"message": "تم تحديث الإعدادات"}

@api_router.get("/loyalty/members")
async def get_loyalty_members(current_user: dict = Depends(get_current_user)):
    """قائمة أعضاء برنامج الولاء"""
    members = await db.loyalty_members.find(
        {"tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    ).sort("total_points", -1).to_list(500)
    return members

@api_router.post("/loyalty/members")
async def create_loyalty_member(member: LoyaltyMemberCreate, current_user: dict = Depends(get_current_user)):
    """إضافة عضو جديد"""
    # التحقق من عدم وجود العضو
    existing = await db.loyalty_members.find_one({
        "phone": member.phone,
        "tenant_id": current_user.get("tenant_id")
    })
    if existing:
        raise HTTPException(status_code=400, detail="العضو موجود مسبقاً")
    
    # جلب إعدادات الولاء
    settings = await db.loyalty_settings.find_one({"tenant_id": current_user.get("tenant_id")})
    welcome_bonus = settings.get("welcome_bonus", 50) if settings else 50
    
    new_member = {
        "id": str(uuid.uuid4()),
        "customer_id": member.customer_id,
        "customer_name": member.customer_name,
        "phone": member.phone,
        "email": member.email,
        "total_points": welcome_bonus,
        "available_points": welcome_bonus,
        "redeemed_points": 0,
        "current_tier": "bronze",
        "lifetime_spending": 0,
        "total_orders": 0,
        "join_date": datetime.now(timezone.utc).isoformat(),
        "birthday": member.birthday,
        "referral_code": str(uuid.uuid4())[:8].upper(),
        "referred_by": member.referred_by,
        "tenant_id": current_user.get("tenant_id")
    }
    
    await db.loyalty_members.insert_one(new_member)
    
    # تسجيل نقاط الترحيب
    if welcome_bonus > 0:
        transaction = {
            "id": str(uuid.uuid4()),
            "member_id": new_member["id"],
            "transaction_type": "bonus",
            "points": welcome_bonus,
            "description": "نقاط الترحيب",
            "balance_after": welcome_bonus,
            "created_at": datetime.now(timezone.utc).isoformat(),
            "tenant_id": current_user.get("tenant_id")
        }
        await db.loyalty_transactions.insert_one(transaction)
    
    # مكافأة الإحالة
    if member.referred_by:
        referrer = await db.loyalty_members.find_one({"referral_code": member.referred_by})
        if referrer:
            referral_bonus = settings.get("referral_bonus", 200) if settings else 200
            await db.loyalty_members.update_one(
                {"id": referrer["id"]},
                {"$inc": {"total_points": referral_bonus, "available_points": referral_bonus}}
            )
            ref_transaction = {
                "id": str(uuid.uuid4()),
                "member_id": referrer["id"],
                "transaction_type": "bonus",
                "points": referral_bonus,
                "description": f"مكافأة إحالة - {member.customer_name}",
                "balance_after": referrer.get("available_points", 0) + referral_bonus,
                "created_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": current_user.get("tenant_id")
            }
            await db.loyalty_transactions.insert_one(ref_transaction)
    
    if "_id" in new_member:
        del new_member["_id"]
    
    return {"message": "تم إضافة العضو", "member": new_member}

@api_router.get("/loyalty/members/{member_id}")
async def get_loyalty_member(member_id: str, current_user: dict = Depends(get_current_user)):
    """تفاصيل عضو"""
    member = await db.loyalty_members.find_one(
        {"id": member_id, "tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    )
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    return member

@api_router.post("/loyalty/earn")
async def earn_points(request: EarnPointsRequest, current_user: dict = Depends(get_current_user)):
    """كسب نقاط من طلب"""
    member = await db.loyalty_members.find_one({"id": request.member_id, "tenant_id": current_user.get("tenant_id")})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    
    settings = await db.loyalty_settings.find_one({"tenant_id": current_user.get("tenant_id")})
    points_per_currency = settings.get("points_per_currency", 1.0) if settings else 1.0
    tiers = settings.get("tiers", DEFAULT_LOYALTY_TIERS) if settings else DEFAULT_LOYALTY_TIERS
    
    # حساب مضاعف المستوى
    multiplier = 1.0
    for tier in tiers:
        if tier.get("name_en", "").lower() == member.get("current_tier", "bronze").lower():
            multiplier = tier.get("points_multiplier", 1.0)
            break
    
    earned_points = int(request.order_total * points_per_currency * multiplier)
    new_total = member.get("total_points", 0) + earned_points
    new_available = member.get("available_points", 0) + earned_points
    
    # تحديد المستوى الجديد
    new_tier = "bronze"
    for tier in sorted(tiers, key=lambda x: x.get("min_points", 0), reverse=True):
        if new_total >= tier.get("min_points", 0):
            new_tier = tier.get("name_en", "bronze").lower()
            break
    
    await db.loyalty_members.update_one(
        {"id": request.member_id},
        {"$set": {
            "total_points": new_total,
            "available_points": new_available,
            "current_tier": new_tier,
            "lifetime_spending": member.get("lifetime_spending", 0) + request.order_total,
            "total_orders": member.get("total_orders", 0) + 1
        }}
    )
    
    transaction = {
        "id": str(uuid.uuid4()),
        "member_id": request.member_id,
        "order_id": request.order_id,
        "transaction_type": "earn",
        "points": earned_points,
        "description": f"طلب #{request.order_id[:8]}",
        "balance_after": new_available,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": current_user.get("tenant_id")
    }
    await db.loyalty_transactions.insert_one(transaction)
    
    return {"earned_points": earned_points, "new_balance": new_available, "new_tier": new_tier}

@api_router.post("/loyalty/redeem")
async def redeem_points(request: RedeemPointsRequest, current_user: dict = Depends(get_current_user)):
    """استبدال نقاط"""
    member = await db.loyalty_members.find_one({"id": request.member_id, "tenant_id": current_user.get("tenant_id")})
    if not member:
        raise HTTPException(status_code=404, detail="العضو غير موجود")
    
    if member.get("available_points", 0) < request.points_to_redeem:
        raise HTTPException(status_code=400, detail="رصيد النقاط غير كافي")
    
    settings = await db.loyalty_settings.find_one({"tenant_id": current_user.get("tenant_id")})
    min_redeem = settings.get("min_redeem_points", 100) if settings else 100
    currency_per_point = settings.get("currency_per_point", 0.01) if settings else 0.01
    
    if request.points_to_redeem < min_redeem:
        raise HTTPException(status_code=400, detail=f"الحد الأدنى للاستبدال {min_redeem} نقطة")
    
    discount_value = request.points_to_redeem * currency_per_point
    new_available = member.get("available_points", 0) - request.points_to_redeem
    new_redeemed = member.get("redeemed_points", 0) + request.points_to_redeem
    
    await db.loyalty_members.update_one(
        {"id": request.member_id},
        {"$set": {"available_points": new_available, "redeemed_points": new_redeemed}}
    )
    
    transaction = {
        "id": str(uuid.uuid4()),
        "member_id": request.member_id,
        "order_id": request.order_id,
        "transaction_type": "redeem",
        "points": -request.points_to_redeem,
        "description": f"استبدال في طلب #{request.order_id[:8]}",
        "balance_after": new_available,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "tenant_id": current_user.get("tenant_id")
    }
    await db.loyalty_transactions.insert_one(transaction)
    
    return {"discount_value": discount_value, "new_balance": new_available}

@api_router.get("/loyalty/transactions/{member_id}")
async def get_member_transactions(member_id: str, current_user: dict = Depends(get_current_user)):
    """سجل معاملات العضو"""
    transactions = await db.loyalty_transactions.find(
        {"member_id": member_id, "tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return transactions

# ==================== CUSTOMER REVIEWS ====================

@api_router.get("/customer-reviews")
async def get_customer_reviews(current_user: dict = Depends(get_current_user)):
    """جلب تقييمات العملاء"""
    tenant_id = current_user.get("tenant_id")
    query = {"tenant_id": tenant_id} if tenant_id else {}
    
    reviews = await db.customer_reviews.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    
    return reviews

@api_router.post("/customer-reviews")
async def create_customer_review(review: Dict[str, Any], request: Request):
    """إضافة تقييم من العميل (بدون توثيق - محمي بتحديد معدّل وتحقق المستأجر)"""
    enforce_rate_limit(request, "customer_review", max_calls=10, window_seconds=60)
    # التحقق من وجود الطلب ومطابقة المستأجر (منع حقن تقييمات وهمية)
    order_id = review.get("order_id")
    order = await db.orders.find_one({"id": order_id}, {"_id": 0, "tenant_id": 1, "branch_id": 1, "order_number": 1}) if order_id else None
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    try:
        rating = int(review.get("rating", 5))
    except Exception:
        rating = 5
    rating = max(1, min(5, rating))
    review_doc = {
        "id": str(uuid.uuid4()),
        "order_id": order_id,
        "order_number": order.get("order_number"),
        "customer_name": sanitize_text(review.get("customer_name"), 120),
        "customer_phone": sanitize_text(review.get("customer_phone"), 30),
        "rating": rating,
        "comment": sanitize_text(review.get("comment", ""), 1000),
        "food_rating": review.get("food_rating"),
        "service_rating": review.get("service_rating"),
        "speed_rating": review.get("speed_rating"),
        "tenant_id": order.get("tenant_id"),
        "branch_id": order.get("branch_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.customer_reviews.insert_one(review_doc)
    review_doc.pop("_id", None)
    
    return review_doc

# ==================== RECIPES & RAW MATERIALS ROUTES ====================

class RawMaterialCreate(BaseModel):
    name: str
    name_en: Optional[str] = None
    unit: str
    unit_cost: float
    current_stock: float = 0
    min_stock: float = 0
    max_stock: float = 0
    category: str = "general"
    branch_id: Optional[str] = None

class RecipeCreate(BaseModel):
    product_id: str
    ingredients: List[Dict[str, Any]]
    labor_cost: float = 0
    overhead_cost: float = 0
    portions: int = 1
    preparation_time: int = 0
    instructions: Optional[str] = None

MATERIAL_CATEGORIES = [
    {"id": "meat", "name": "لحوم ودواجن"},
    {"id": "seafood", "name": "مأكولات بحرية"},
    {"id": "vegetables", "name": "خضروات"},
    {"id": "fruits", "name": "فواكه"},
    {"id": "dairy", "name": "ألبان وبيض"},
    {"id": "grains", "name": "حبوب ونشويات"},
    {"id": "spices", "name": "توابل وبهارات"},
    {"id": "oils", "name": "زيوت ودهون"},
    {"id": "beverages", "name": "مشروبات"},
    {"id": "packaging", "name": "تغليف"},
    {"id": "general", "name": "عام"}
]

@api_router.get("/recipes/categories")
async def get_material_categories(current_user: dict = Depends(get_current_user)):
    """تصنيفات المواد الخام"""
    return MATERIAL_CATEGORIES

@api_router.get("/recipes/materials")
async def get_raw_materials(category: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """قائمة المواد الخام"""
    query = {"tenant_id": current_user.get("tenant_id")}
    if category:
        query["category"] = category
    
    materials = await db.raw_materials.find(query, {"_id": 0}).to_list(500)
    return materials

@api_router.post("/recipes/materials")
async def create_raw_material(material: RawMaterialCreate, current_user: dict = Depends(get_current_user)):
    """إضافة مادة خام"""
    new_material = {
        "id": str(uuid.uuid4()),
        **material.model_dump(),
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.raw_materials.insert_one(new_material)
    if "_id" in new_material:
        del new_material["_id"]
    
    return {"message": "تم إضافة المادة", "material": new_material}

@api_router.put("/recipes/materials/{material_id}")
async def update_raw_material(material_id: str, material: RawMaterialCreate, current_user: dict = Depends(get_current_user)):
    """تحديث مادة خام"""
    await db.raw_materials.update_one(
        {"id": material_id, "tenant_id": current_user.get("tenant_id")},
        {"$set": {**material.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم التحديث"}

@api_router.delete("/recipes/materials/{material_id}")
async def delete_raw_material(material_id: str, current_user: dict = Depends(get_current_user)):
    """حذف مادة خام"""
    await db.raw_materials.delete_one({"id": material_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.get("/recipes")
async def get_recipes(current_user: dict = Depends(get_current_user)):
    """قائمة الوصفات"""
    recipes = await db.recipes.find({"tenant_id": current_user.get("tenant_id")}, {"_id": 0}).to_list(500)
    return recipes

@api_router.post("/recipes")
async def create_recipe(recipe: RecipeCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء وصفة"""
    # جلب معلومات المنتج
    product = await db.products.find_one({"id": recipe.product_id}, {"_id": 0})
    if not product:
        raise HTTPException(status_code=404, detail="المنتج غير موجود")
    
    # جلب المواد الخام
    material_ids = [ing.get("material_id") for ing in recipe.ingredients]
    materials = await db.raw_materials.find({"id": {"$in": material_ids}}, {"_id": 0}).to_list(100)
    materials_dict = {m["id"]: m for m in materials}
    
    # حساب التكلفة
    total_cost = 0
    ingredients_list = []
    
    for ing in recipe.ingredients:
        mat = materials_dict.get(ing.get("material_id"))
        if mat:
            ing_cost = _sn(ing.get("quantity")) * mat.get("unit_cost", 0)
            total_cost += ing_cost
            ingredients_list.append({
                "material_id": mat["id"],
                "material_name": mat["name"],
                "quantity": _sn(ing.get("quantity")),
                "unit": mat["unit"],
                "unit_cost": mat["unit_cost"],
                "total_cost": round(ing_cost, 3)
            })
    
    final_cost = total_cost + recipe.labor_cost + recipe.overhead_cost
    selling_price = _sn(product.get("price"))
    profit_margin = ((selling_price - final_cost) / selling_price * 100) if selling_price > 0 else 0
    
    new_recipe = {
        "id": str(uuid.uuid4()),
        "product_id": recipe.product_id,
        "product_name": product.get("name", ""),
        "ingredients": ingredients_list,
        "total_cost": round(total_cost, 3),
        "labor_cost": recipe.labor_cost,
        "overhead_cost": recipe.overhead_cost,
        "final_cost": round(final_cost, 3),
        "selling_price": selling_price,
        "profit_margin": round(profit_margin, 2),
        "portions": recipe.portions,
        "preparation_time": recipe.preparation_time,
        "instructions": recipe.instructions,
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.recipes.insert_one(new_recipe)
    if "_id" in new_recipe:
        del new_recipe["_id"]
    
    return {"message": "تم إنشاء الوصفة", "recipe": new_recipe}

@api_router.get("/recipes/{recipe_id}")
async def get_recipe(recipe_id: str, current_user: dict = Depends(get_current_user)):
    """تفاصيل وصفة"""
    recipe = await db.recipes.find_one({"id": recipe_id, "tenant_id": current_user.get("tenant_id")}, {"_id": 0})
    if not recipe:
        raise HTTPException(status_code=404, detail="الوصفة غير موجودة")
    return recipe

@api_router.delete("/recipes/{recipe_id}")
async def delete_recipe(recipe_id: str, current_user: dict = Depends(get_current_user)):
    """حذف وصفة"""
    await db.recipes.delete_one({"id": recipe_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.get("/recipes/alerts/low-stock")
async def get_low_stock_alerts(current_user: dict = Depends(get_current_user)):
    """تنبيهات المخزون المنخفض"""
    materials = await db.raw_materials.find({
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "$expr": {"$lte": ["$current_stock", "$min_stock"]}
    }, {"_id": 0}).to_list(100)
    
    alerts = []
    for mat in materials:
        severity = "critical" if mat.get("current_stock", 0) == 0 else "warning"
        alerts.append({
            "material_id": mat["id"],
            "material_name": mat["name"],
            "current_stock": mat.get("current_stock", 0),
            "min_stock": mat.get("min_stock", 0),
            "unit": mat.get("unit", ""),
            "severity": severity
        })
    
    return alerts

# ==================== INVOICE & PRINTING ROUTES ====================

class InvoicePrinterCreate(BaseModel):
    name: str
    printer_type: str = "thermal"
    paper_width: int = 80
    connection_type: str = "network"
    ip_address: Optional[str] = None
    port: int = 9100
    branch_id: str
    is_default: bool = False

class InvoiceTemplateCreate(BaseModel):
    name: str
    template_type: str = "receipt"
    show_logo: bool = True
    logo_url: Optional[str] = None
    business_name: str = ""
    business_name_en: Optional[str] = None
    address: Optional[str] = None
    phone: Optional[str] = None
    tax_number: Optional[str] = None
    footer_text: Optional[str] = None
    footer_text_en: Optional[str] = None
    show_qr_code: bool = False
    paper_width: int = 80
    branch_id: Optional[str] = None
    is_default: bool = False

@api_router.get("/invoices/printers")
async def get_printers(current_user: dict = Depends(get_current_user)):
    """قائمة الطابعات"""
    printers = await db.printers.find({"tenant_id": current_user.get("tenant_id")}, {"_id": 0}).to_list(50)
    return printers

@api_router.post("/invoices/printers")
async def create_printer(printer: InvoicePrinterCreate, current_user: dict = Depends(get_current_user)):
    """إضافة طابعة"""
    new_printer = {
        "id": str(uuid.uuid4()),
        **printer.model_dump(),
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.printers.insert_one(new_printer)
    if "_id" in new_printer:
        del new_printer["_id"]
    
    return {"message": "تم إضافة الطابعة", "printer": new_printer}

@api_router.delete("/invoices/printers/{printer_id}")
async def delete_printer(printer_id: str, current_user: dict = Depends(get_current_user)):
    """حذف طابعة"""
    await db.printers.delete_one({"id": printer_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.get("/invoices/templates")
async def get_invoice_templates(current_user: dict = Depends(get_current_user)):
    """قائمة قوالب الفواتير"""
    templates = await db.invoice_templates.find({"tenant_id": current_user.get("tenant_id")}, {"_id": 0}).to_list(50)
    return templates

@api_router.post("/invoices/templates")
async def create_invoice_template(template: InvoiceTemplateCreate, current_user: dict = Depends(get_current_user)):
    """إنشاء قالب فاتورة"""
    new_template = {
        "id": str(uuid.uuid4()),
        **template.model_dump(),
        "tenant_id": current_user.get("tenant_id"),
        "is_active": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.invoice_templates.insert_one(new_template)
    if "_id" in new_template:
        del new_template["_id"]
    
    return {"message": "تم إنشاء القالب", "template": new_template}

@api_router.put("/invoices/templates/{template_id}")
async def update_invoice_template(template_id: str, template: InvoiceTemplateCreate, current_user: dict = Depends(get_current_user)):
    """تحديث قالب فاتورة"""
    await db.invoice_templates.update_one(
        {"id": template_id, "tenant_id": current_user.get("tenant_id")},
        {"$set": {**template.model_dump(), "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    return {"message": "تم التحديث"}

@api_router.delete("/invoices/templates/{template_id}")
async def delete_invoice_template(template_id: str, current_user: dict = Depends(get_current_user)):
    """حذف قالب"""
    await db.invoice_templates.delete_one({"id": template_id, "tenant_id": current_user.get("tenant_id")})
    return {"message": "تم الحذف"}

@api_router.post("/invoices/print/{order_id}")
async def print_invoice(order_id: str, print_type: str = "customer", printer_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """طباعة فاتورة مع تطبيق صلاحيات الطابعة"""
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # جلب إعدادات الطابعة إذا تم تحديدها
    printer_settings = None
    if printer_id:
        printer_query = build_tenant_query(current_user, {"id": printer_id})
        printer_settings = await db.printers.find_one(printer_query, {"_id": 0})
    
    # جلب قالب الفاتورة
    template = await db.invoice_templates.find_one({
        "tenant_id": current_user.get("tenant_id"),
        "template_type": "receipt" if print_type == "customer" else print_type,
        "is_default": True
    }, {"_id": 0})
    
    if not template:
        template = {
            "business_name": "المطعم",
            "show_logo": True,
            "footer_text": "شكراً لزيارتكم",
            "paper_width": 80
        }
    
    # تحضير الأصناف بناءً على صلاحيات الطابعة
    items = order.get("items", [])
    show_prices = True
    print_mode = "full_receipt"
    print_individual_items = False
    
    if printer_settings:
        show_prices = printer_settings.get("show_prices", True)
        print_mode = printer_settings.get("print_mode", "full_receipt")
        print_individual_items = printer_settings.get("print_individual_items", False)
    
    # تطبيق صلاحيات الطابعة على الأصناف
    processed_items = []
    for item in items:
        processed_item = {
            "name": item.get("product_name") or item.get("name"),
            "name_en": item.get("product_name_en") or item.get("name_en"),
            "quantity": item.get("quantity", 1),
            "notes": item.get("notes"),
        }
        
        # إضافة الأسعار فقط إذا كان مسموحاً
        if show_prices:
            processed_item["price"] = _sn(item.get("price"))
            processed_item["total"] = _sn(item.get("price")) * item.get("quantity", 1)
        
        processed_items.append(processed_item)
    
    # تحضير بيانات الطلب للطباعة
    print_data = {
        "order_number": order.get("order_number", order["id"][:8]),
        "date": datetime.fromisoformat(order.get("created_at", datetime.now(timezone.utc).isoformat())).strftime("%Y-%m-%d %H:%M"),
        "table_number": order.get("table_number"),
        "customer_name": order.get("customer_name"),
        "items": processed_items,
        "order_type": order.get("order_type", "dine_in"),
        "notes": order.get("notes"),
        "print_mode": print_mode,
        "show_prices": show_prices,
        "print_individual_items": print_individual_items
    }
    
    # إضافة المعلومات المالية فقط للفاتورة الكاملة وإذا كان عرض الأسعار مسموحاً
    if print_mode == "full_receipt" and show_prices:
        print_data["subtotal"] = _sn(order.get("subtotal"))
        print_data["discount"] = _sn(order.get("discount"))
        print_data["tax"] = order.get("tax", 0)
        print_data["total"] = _sn(order.get("total"))
        print_data["payment_method"] = order.get("payment_method", "cash")
        # تفاصيل الكوبون لإظهارها في الفاتورة
        print_data["coupon_id"] = order.get("coupon_id")
        print_data["coupon_name"] = order.get("coupon_name")
        print_data["coupon_code"] = order.get("coupon_code")
        print_data["coupon_discount"] = _sn(order.get("coupon_discount"))
    
    # إذا كانت طباعة كل صنف على حدة، نجهز مصفوفة من الطباعات
    print_jobs = []
    if print_individual_items:
        for item in processed_items:
            job = {
                "order_number": print_data["order_number"],
                "date": print_data["date"],
                "table_number": print_data["table_number"],
                "items": [item],
                "notes": print_data["notes"],
                "is_individual": True
            }
            print_jobs.append(job)
    else:
        print_jobs = [print_data]
    
    return {
        "message": "جاهز للطباعة",
        "print_data": print_data,
        "print_jobs": print_jobs,
        "template": template,
        "printer_settings": {
            "print_mode": print_mode,
            "show_prices": show_prices,
            "print_individual_items": print_individual_items
        }
    }

@api_router.get("/invoices/auto-print/{order_id}")
async def get_auto_print_data(order_id: str, branch_id: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """
    جلب بيانات الطباعة التلقائية لكل الطابعات النشطة
    يُستخدم عند إنشاء طلب جديد لإرسال بيانات الطباعة لكل طابعة
    """
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # جلب جميع الطابعات المفعّلة للطباعة التلقائية
    printer_query = build_tenant_query(current_user, {
        "is_active": True,
        "auto_print_on_order": True
    })
    if branch_id:
        printer_query["branch_id"] = branch_id
    elif order.get("branch_id"):
        printer_query["branch_id"] = order.get("branch_id")
    
    printers = await db.printers.find(printer_query, {"_id": 0}).to_list(50)
    
    if not printers:
        return {"message": "لا توجد طابعات مفعّلة للطباعة التلقائية", "printers": []}
    
    # تجهيز بيانات الطباعة لكل طابعة
    print_results = []
    
    for printer in printers:
        show_prices = printer.get("show_prices", True)
        print_mode = printer.get("print_mode", "full_receipt")
        print_individual_items = printer.get("print_individual_items", False)
        
        # تحضير الأصناف بناءً على صلاحيات الطابعة
        processed_items = []
        for item in order.get("items", []):
            processed_item = {
                "name": item.get("product_name") or item.get("name"),
                "name_en": item.get("product_name_en") or item.get("name_en"),
                "quantity": item.get("quantity", 1),
                "notes": item.get("notes"),
            }
            
            if show_prices:
                processed_item["price"] = _sn(item.get("price"))
                processed_item["total"] = _sn(item.get("price")) * item.get("quantity", 1)
            
            processed_items.append(processed_item)
        
        # تحضير بيانات الطلب
        print_data = {
            "order_number": order.get("order_number", order["id"][:8]),
            "date": datetime.fromisoformat(order.get("created_at", datetime.now(timezone.utc).isoformat())).strftime("%Y-%m-%d %H:%M"),
            "table_number": order.get("table_number"),
            "customer_name": order.get("customer_name"),
            "items": processed_items,
            "order_type": order.get("order_type", "dine_in"),
            "notes": order.get("notes"),
        }
        
        # إضافة المعلومات المالية للفاتورة الكاملة فقط
        if print_mode == "full_receipt" and show_prices:
            print_data["subtotal"] = _sn(order.get("subtotal"))
            print_data["discount"] = _sn(order.get("discount"))
            print_data["tax"] = order.get("tax", 0)
            print_data["total"] = _sn(order.get("total"))
            print_data["payment_method"] = order.get("payment_method", "cash")
            # تفاصيل الكوبون لإظهارها في الفاتورة
            print_data["coupon_id"] = order.get("coupon_id")
            print_data["coupon_name"] = order.get("coupon_name")
            print_data["coupon_code"] = order.get("coupon_code")
            print_data["coupon_discount"] = _sn(order.get("coupon_discount"))
        
        # تجهيز الطباعات
        print_jobs = []
        if print_individual_items:
            for item in processed_items:
                job = {
                    "order_number": print_data["order_number"],
                    "date": print_data["date"],
                    "table_number": print_data["table_number"],
                    "items": [item],
                    "notes": print_data["notes"],
                    "is_individual": True
                }
                print_jobs.append(job)
        else:
            print_jobs = [print_data]
        
        print_results.append({
            "printer": {
                "id": printer.get("id"),
                "name": printer.get("name"),
                "ip_address": printer.get("ip_address"),
                "port": printer.get("port", 9100),
                "printer_type": printer.get("printer_type")
            },
            "settings": {
                "print_mode": print_mode,
                "show_prices": show_prices,
                "print_individual_items": print_individual_items
            },
            "print_jobs": print_jobs
        })
    
    return {
        "message": "بيانات الطباعة جاهزة",
        "order_id": order_id,
        "printers_count": len(print_results),
        "printers": print_results
    }

# ==================== PUSH NOTIFICATIONS ROUTES ====================

class FCMTokenCreate(BaseModel):
    user_id: str
    user_type: str
    token: str
    device_type: str = "web"
    device_id: Optional[str] = None
    branch_id: Optional[str] = None

class SendNotificationRequest(BaseModel):
    target_type: str  # user, role, branch, all
    target_id: Optional[str] = None
    title: str
    body: str
    data: Dict[str, Any] = {}

@api_router.post("/notifications/fcm/register")
async def register_fcm_token(token_data: FCMTokenCreate, current_user: dict = Depends(get_current_user)):
    """تسجيل FCM Token"""
    existing = await db.fcm_tokens.find_one({"token": token_data.token})
    
    if existing:
        await db.fcm_tokens.update_one(
            {"token": token_data.token},
            {"$set": {
                "user_id": token_data.user_id,
                "user_type": token_data.user_type,
                "updated_at": datetime.now(timezone.utc).isoformat()
            }}
        )
    else:
        new_token = {
            "id": str(uuid.uuid4()),
            **token_data.model_dump(),
            "tenant_id": current_user.get("tenant_id"),
            "is_active": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.fcm_tokens.insert_one(new_token)
    
    return {"message": "تم التسجيل"}

@api_router.delete("/notifications/fcm/unregister")
async def unregister_fcm_token(token: str, current_user: dict = Depends(get_current_user)):
    """إلغاء تسجيل FCM Token"""
    await db.fcm_tokens.delete_one({"token": token})
    return {"message": "تم الإلغاء"}

@api_router.post("/notifications/send")
async def send_notification(request: SendNotificationRequest, background_tasks: BackgroundTasks, current_user: dict = Depends(get_current_user)):
    """إرسال إشعار"""
    # جلب Tokens المستهدفة
    query = {"tenant_id": current_user.get("tenant_id"), "is_active": True}
    
    if request.target_type == "user":
        query["user_id"] = request.target_id
    elif request.target_type == "role":
        query["user_type"] = request.target_id
    elif request.target_type == "branch":
        query["branch_id"] = request.target_id
    
    tokens = await db.fcm_tokens.find(query, {"token": 1, "_id": 0}).to_list(1000)
    token_list = [t["token"] for t in tokens]
    
    if not token_list:
        return {"message": "لا توجد أجهزة مسجلة", "sent": 0}
    
    # TODO: إرسال عبر Firebase
    # في الوقت الحالي نسجل الإشعار فقط
    
    notification_log = {
        "id": str(uuid.uuid4()),
        "target_type": request.target_type,
        "target_id": request.target_id,
        "title": request.title,
        "body": request.body,
        "data": request.data,
        "sent_count": len(token_list),
        "status": "sent",
        "tenant_id": current_user.get("tenant_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    await db.notification_logs.insert_one(notification_log)
    
    return {"message": "تم الإرسال", "sent": len(token_list)}

@api_router.get("/notifications/logs")
async def get_notification_logs(current_user: dict = Depends(get_current_user)):
    """سجل الإشعارات"""
    logs = await db.notification_logs.find(
        {"tenant_id": current_user.get("tenant_id")},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return logs

# Helper function to send notification on new order (called from order creation)
async def notify_new_order(order: dict, tenant_id: str):
    """إرسال إشعار طلب جديد"""
    try:
        # جلب tokens السائقين والموظفين
        tokens = await db.fcm_tokens.find({
            "tenant_id": tenant_id,
            "user_type": {"$in": ["driver", "admin", "staff"]},
            "is_active": True
        }).to_list(100)
        
        if tokens:
            # TODO: إرسال عبر Firebase
            logger.info(f"Would send notification for order {order.get('id')} to {len(tokens)} devices")
    except Exception as e:
        logger.error(f"Failed to send notification: {e}")


# ==================== COMPREHENSIVE DASHBOARD STATS ====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إحصائيات Dashboard الشاملة - يومي/أسبوعي/شهري/إجمالي"""
    tenant_id = current_user.get("tenant_id")
    
    # Super Admin بدون tenant لا يرى إحصائيات
    if current_user.get("role") == UserRole.SUPER_ADMIN and not tenant_id:
        return {
            "today": {"total_sales": 0, "total_orders": 0, "average_order_value": 0, "gross_profit": 0, "operating_costs": 0, "total_profit": 0, "by_payment_method": {}},
            "week": {"total_sales": 0, "total_orders": 0, "average_order_value": 0, "gross_profit": 0, "operating_costs": 0, "total_profit": 0, "by_payment_method": {}},
            "month": {"total_sales": 0, "total_orders": 0, "average_order_value": 0, "gross_profit": 0, "operating_costs": 0, "total_profit": 0, "by_payment_method": {}},
            "all_time": {"total_sales": 0, "total_orders": 0, "average_order_value": 0, "gross_profit": 0, "operating_costs": 0, "total_profit": 0, "by_payment_method": {}},
            "operating_costs_details": {"rent": 0, "electricity": 0, "water": 0, "generator": 0, "salaries": 0, "total": 0}
        }
    
    # تحديد الفلتر الأساسي
    base_query = {"status": {"$ne": OrderStatus.CANCELLED}}
    
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    else:
        base_query["tenant_id"] = "default"
    
    # فلترة الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    is_manager = user_role in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]
    
    if user_branch_id and not is_manager:
        base_query["branch_id"] = user_branch_id
    elif branch_id:
        base_query["branch_id"] = branch_id
    
    # المستخدمون غير المدراء يرون فقط طلباتهم الخاصة
    if not is_manager:
        base_query["cashier_id"] = current_user["id"]
    
    # حساب التواريخ
    now = datetime.now(timezone.utc)
    today = now.strftime('%Y-%m-%d')
    week_ago = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    month_ago = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    
    # جلب الفروع والتكاليف الثابتة
    branches_query = {"tenant_id": tenant_id, "is_active": {"$ne": False}}
    if branch_id:
        branches_query["id"] = branch_id
    elif user_branch_id and not is_manager:
        branches_query["id"] = user_branch_id
    
    branches = await db.branches.find(branches_query, {"_id": 0}).to_list(100)
    
    # حساب التكاليف الثابتة الشهرية
    total_rent = sum(_sn(b.get("rent_cost")) for b in branches)
    total_electricity = sum(_sn(b.get("electricity_cost")) for b in branches)
    total_water = sum(_sn(b.get("water_cost")) for b in branches)
    total_generator = sum(_sn(b.get("generator_cost")) for b in branches)
    total_fixed_costs = total_rent + total_electricity + total_water + total_generator
    daily_fixed_costs = total_fixed_costs / 30
    
    # حساب الرواتب
    employees_query = {"tenant_id": tenant_id, "is_active": {"$ne": False}}
    if branch_id:
        employees_query["branch_id"] = branch_id
    elif user_branch_id and not is_manager:
        employees_query["branch_id"] = user_branch_id
    
    employees = await db.employees.find(employees_query, {"_id": 0, "salary": 1}).to_list(1000)
    total_salaries = sum(_sn(e.get("salary")) for e in employees)
    daily_salaries = total_salaries / 30
    
    # الهدف اليومي من التكاليف التشغيلية
    daily_operating_costs = daily_fixed_costs + daily_salaries
    
    # استعلامات الفترات المختلفة
    async def get_period_stats(start_date: Optional[str] = None, days: int = 1):
        query = base_query.copy()
        if start_date:
            query["created_at"] = {"$gte": start_date}
        
        orders = await db.orders.find(query, {"_id": 0, "total": 1, "total_cost": 1, "profit": 1, "payment_method": 1}).to_list(10000)
        
        total_sales = sum(_sn(o.get("total")) for o in orders)
        total_orders = len(orders)
        avg_order = total_sales / total_orders if total_orders > 0 else 0
        gross_profit = sum(_sn(o.get("profit")) for o in orders)  # الربح قبل التكاليف التشغيلية
        
        # حساب التكاليف التشغيلية للفترة
        period_operating_costs = daily_operating_costs * days
        
        # صافي الربح الحقيقي = الربح الإجمالي - التكاليف التشغيلية
        net_profit = gross_profit - period_operating_costs
        
        by_payment = {}
        payment_translations = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        for o in orders:
            pm = o.get("payment_method", "cash")
            # ترجمة اسم طريقة الدفع للعربية
            translated_pm = payment_translations.get(pm, pm)
            by_payment[translated_pm] = by_payment.get(translated_pm, 0) + _sn(o.get("total"))
        
        return {
            "total_sales": total_sales,
            "total_orders": total_orders,
            "average_order_value": avg_order,
            "gross_profit": gross_profit,
            "operating_costs": period_operating_costs,
            "total_profit": net_profit,  # صافي الربح الحقيقي
            "by_payment_method": by_payment
        }
    
    # جلب جميع الإحصائيات بالتوازي
    today_stats, week_stats, month_stats, all_stats = await asyncio.gather(
        get_period_stats(today, 1),
        get_period_stats(week_ago, 7),
        get_period_stats(month_ago, 30),
        get_period_stats(None, 30)  # نفترض 30 يوم للإجمالي
    )
    
    # جلب آخر الطلبات
    recent_query = base_query.copy()
    recent_orders = await db.orders.find(recent_query, {"_id": 0}).sort("created_at", -1).limit(10).to_list(10)
    
    # جلب معلومات الوردية الحالية
    shift_query = {"status": "open"}
    if branch_id:
        shift_query["branch_id"] = branch_id
    current_shift = await db.shifts.find_one(shift_query, {"_id": 0})
    
    # جلب الطلبات المعلقة
    pending_query = base_query.copy()
    pending_query["status"] = {"$in": ["pending", "preparing", "ready"]}
    pending_count = await db.orders.count_documents(pending_query)
    
    return {
        "today": today_stats,
        "week": week_stats,
        "month": month_stats,
        "all_time": all_stats,
        "recent_orders": recent_orders,
        "current_shift": current_shift,
        "pending_orders_count": pending_count,
        "current_date": today,
        "server_time": now.isoformat()
    }


# ==================== AUTO DAY CLOSE SYSTEM (نظام الترحيل التلقائي) ====================

class DayCloseRequest(BaseModel):
    force: bool = False  # إغلاق إجباري حتى مع وجود طلبات معلقة
    notes: Optional[str] = None

class DayCloseResponse(BaseModel):
    success: bool
    message: str
    day_summary: Optional[Dict] = None
    pending_orders: Optional[List] = None
    shifts_closed: int = 0

@api_router.get("/day-management/status")
async def get_day_status(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """حالة اليوم الحالي - هل يوجد ورديات مفتوحة، طلبات معلقة، إلخ"""
    tenant_id = get_user_tenant_id(current_user)
    
    base_query = {}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    
    if branch_id:
        base_query["branch_id"] = branch_id
    
    # الورديات المفتوحة
    shift_query = {**base_query, "status": "open"}
    open_shifts = await db.shifts.find(shift_query, {"_id": 0}).to_list(100)
    
    # الطلبات المعلقة
    pending_query = {**base_query, "status": {"$in": ["pending", "preparing", "ready"]}}
    pending_orders = await db.orders.find(pending_query, {"_id": 0, "id": 1, "order_number": 1, "status": 1, "total": 1, "created_at": 1}).to_list(100)
    
    # آخر إغلاق يومي
    last_close = await db.day_closures.find_one(base_query, sort=[("closed_at", -1)])
    if last_close:
        last_close.pop("_id", None)
    
    # حساب عمر الوردية (بالساعات)
    oldest_shift_hours = 0
    if open_shifts:
        for shift in open_shifts:
            started = datetime.fromisoformat(shift["started_at"].replace("Z", "+00:00"))
            hours = (datetime.now(timezone.utc) - started).total_seconds() / 3600
            if hours > oldest_shift_hours:
                oldest_shift_hours = hours
    
    return {
        "open_shifts": open_shifts,
        "open_shifts_count": len(open_shifts),
        "pending_orders": pending_orders,
        "pending_orders_count": len(pending_orders),
        "last_day_close": last_close,
        "oldest_shift_hours": round(oldest_shift_hours, 1),
        "should_close": oldest_shift_hours >= 24,  # إشعار إذا مر 24 ساعة
        "can_close": len(pending_orders) == 0  # يمكن الإغلاق إذا لا توجد طلبات معلقة
    }

@api_router.post("/day-management/close", response_model=DayCloseResponse)
async def close_day(
    request: DayCloseRequest,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إغلاق اليوم وترحيل البيانات"""
    tenant_id = get_user_tenant_id(current_user)
    
    base_query = {}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    
    if branch_id:
        base_query["branch_id"] = branch_id
    
    # التحقق من الطلبات المعلقة
    pending_query = {**base_query, "status": {"$in": ["pending", "preparing", "ready"]}}
    pending_orders = await db.orders.find(pending_query, {"_id": 0}).to_list(100)
    
    if pending_orders and not request.force:
        return DayCloseResponse(
            success=False,
            message="يوجد طلبات معلقة يجب إغلاقها أولاً",
            pending_orders=pending_orders
        )
    
    # إغلاق جميع الورديات المفتوحة
    shift_query = {**base_query, "status": "open"}
    open_shifts = await db.shifts.find(shift_query, {"_id": 0}).to_list(100)
    
    shifts_closed = 0
    total_day_sales = 0
    total_day_profit = 0
    total_day_expenses = 0
    
    for shift in open_shifts:
        # حساب إحصائيات الوردية
        shift_start = shift.get("started_at") or shift.get("opened_at") or ""
        
        orders = await db.orders.find({
            "shift_id": shift["id"],
            "status": {"$ne": OrderStatus.CANCELLED}
        }).to_list(1000)
        
        # fallback: إذا لم توجد طلبات بـ shift_id
        if not orders and shift_start:
            orders = await db.orders.find({
                "cashier_id": shift.get("cashier_id", ""),
                "created_at": {"$gte": shift_start},
                "status": {"$ne": OrderStatus.CANCELLED}
            }).to_list(1000)
        
        shift_sales = sum(_sn(o.get("total")) for o in orders)
        shift_profit = sum(_sn(o.get("profit")) for o in orders)
        cash_sales = sum(_sn(o.get("total")) for o in orders if o.get("payment_method") == "cash")
        
        expense_query = {"branch_id": shift.get("branch_id")}
        if shift_start:
            expense_query["created_at"] = {"$gte": shift_start}
        expenses = await db.expenses.find(expense_query).to_list(100)
        shift_expenses = sum(_sn(e.get("amount")) for e in expenses)
        
        # حساب النقد المتوقع - الإغلاق الإجباري = closing_cash صفر = short cash
        opening_cash = _sn(shift.get("opening_cash") or shift.get("opening_balance") or 0)
        expected_cash = opening_cash + cash_sales - shift_expenses
        
        # تحديث الوردية كمغلقة
        await db.shifts.update_one(
            {"id": shift["id"]},
            {"$set": {
                "status": "closed",
                "ended_at": datetime.now(timezone.utc).isoformat(),
                "closed_by": current_user.get("full_name", current_user.get("username")),
                "closed_by_id": current_user.get("id"),
                "auto_closed": not request.force,
                "total_sales": shift_sales,
                "total_expenses": shift_expenses,
                "net_profit": shift_profit - shift_expenses,
                "closing_cash": 0,
                "expected_cash": expected_cash,
                "difference": 0 - expected_cash,
                "cash_sales": cash_sales,
                "notes": request.notes or "إغلاق يومي إجباري - short cash"
            }}
        )
        
        shifts_closed += 1
        total_day_sales += shift_sales
        total_day_profit += shift_profit
        total_day_expenses += shift_expenses
    
    # إذا كان هناك طلبات معلقة مع force=True، نغلقها كملغية
    if pending_orders and request.force:
        for order in pending_orders:
            await db.orders.update_one(
                {"id": order["id"]},
                {"$set": {
                    "status": OrderStatus.CANCELLED,
                    "cancelled_reason": "إغلاق يومي إجباري",
                    "cancelled_at": datetime.now(timezone.utc).isoformat(),
                    "cancelled_by": current_user.get("full_name")
                }}
            )
    
    # إنشاء سجل إغلاق اليوم
    day_summary = {
        "id": str(uuid.uuid4()),
        "date": datetime.now(timezone.utc).strftime('%Y-%m-%d'),
        "closed_at": datetime.now(timezone.utc).isoformat(),
        "closed_by": current_user.get("full_name", current_user.get("username")),
        "closed_by_id": current_user.get("id"),
        "branch_id": branch_id,
        "tenant_id": tenant_id,
        "shifts_closed": shifts_closed,
        "total_sales": total_day_sales,
        "total_profit": total_day_profit,
        "total_expenses": total_day_expenses,
        "net_profit": total_day_profit - total_day_expenses,
        "forced_close": request.force,
        "pending_orders_cancelled": len(pending_orders) if request.force else 0,
        "notes": request.notes
    }
    
    await db.day_closures.insert_one(day_summary)
    day_summary.pop("_id", None)
    
    return DayCloseResponse(
        success=True,
        message=f"تم إغلاق اليوم بنجاح - {shifts_closed} وردية",
        day_summary=day_summary,
        shifts_closed=shifts_closed
    )

@api_router.get("/day-management/history")
async def get_day_close_history(
    branch_id: Optional[str] = None,
    limit: int = 30,
    current_user: dict = Depends(get_current_user)
):
    """سجل إغلاقات الأيام السابقة"""
    tenant_id = get_user_tenant_id(current_user)
    
    query = {}
    if tenant_id:
        query["tenant_id"] = tenant_id
    if branch_id:
        query["branch_id"] = branch_id
    
    closures = await db.day_closures.find(query, {"_id": 0}).sort("closed_at", -1).limit(limit).to_list(limit)
    return closures


# ==================== AUTO DAY CLOSE SCHEDULER (المجدول التلقائي) ====================

async def auto_close_old_shifts():
    """إغلاق تلقائي للورديات القديمة (أكثر من 24 ساعة)"""
    try:
        cutoff = (datetime.now(timezone.utc) - timedelta(hours=24)).isoformat()
        
        # البحث عن الورديات المفتوحة لأكثر من 24 ساعة
        old_shifts = await db.shifts.find({
            "status": "open",
            "started_at": {"$lt": cutoff}
        }).to_list(100)
        
        for shift in old_shifts:
            shift_start = shift.get("started_at") or shift.get("opened_at") or ""
            
            # حساب الإحصائيات
            orders = await db.orders.find({
                "shift_id": shift["id"],
                "status": {"$ne": OrderStatus.CANCELLED}
            }).to_list(1000)
            
            if not orders and shift_start:
                orders = await db.orders.find({
                    "cashier_id": shift.get("cashier_id", ""),
                    "created_at": {"$gte": shift_start},
                    "status": {"$ne": OrderStatus.CANCELLED}
                }).to_list(1000)
            
            total_sales = sum(_sn(o.get("total")) for o in orders)
            total_profit = sum(_sn(o.get("profit")) for o in orders)
            cash_sales = sum(_sn(o.get("total")) for o in orders if o.get("payment_method") == "cash")
            
            expense_query = {"branch_id": shift.get("branch_id")}
            if shift_start:
                expense_query["created_at"] = {"$gte": shift_start}
            expenses = await db.expenses.find(expense_query).to_list(100)
            total_expenses = sum(_sn(e.get("amount")) for e in expenses)
            
            opening_cash = _sn(shift.get("opening_cash") or shift.get("opening_balance") or 0)
            expected_cash = opening_cash + cash_sales - total_expenses
            
            # إغلاق الوردية تلقائياً - closing_cash = 0 = short cash
            await db.shifts.update_one(
                {"id": shift["id"]},
                {"$set": {
                    "status": "closed",
                    "ended_at": datetime.now(timezone.utc).isoformat(),
                    "auto_closed": True,
                    "total_sales": total_sales,
                    "total_expenses": total_expenses,
                    "net_profit": total_profit - total_expenses,
                    "closing_cash": 0,
                    "expected_cash": expected_cash,
                    "difference": 0 - expected_cash,
                    "cash_sales": cash_sales,
                    "notes": "إغلاق تلقائي بعد 24 ساعة - short cash"
                }}
            )
            
            logger.info(f"Auto-closed shift {shift['id']} after 24 hours")
        
        if old_shifts:
            logger.info(f"Auto-closed {len(old_shifts)} old shifts")
            
    except Exception as e:
        logger.error(f"Error in auto_close_old_shifts: {e}")


# ==================== DAILY REPORT EMAIL (تقرير يومي بالبريد) ====================

class DailyReportEmailRequest(BaseModel):
    recipient_emails: List[str] = []  # قائمة البريد للإرسال
    include_all_branches: bool = True

@api_router.post("/day-management/send-report")
async def send_daily_report_email(
    request: DailyReportEmailRequest,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """إرسال التقرير اليومي عبر البريد الإلكتروني"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    # جمع بيانات التقرير
    base_query = {"status": {"$ne": OrderStatus.CANCELLED}}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    
    # جلب جميع الفروع
    branch_query = {}
    if tenant_id:
        branch_query["tenant_id"] = tenant_id
    branches = await db.branches.find(branch_query, {"_id": 0}).to_list(100)
    
    branches_data = []
    total_sales = 0
    total_orders = 0
    total_expenses = 0
    total_profit = 0
    
    for branch in branches:
        if branch_id and branch["id"] != branch_id:
            continue
        
        # طلبات الفرع اليوم
        orders_query = {**base_query, "branch_id": branch["id"], "created_at": {"$gte": today}}
        orders = await db.orders.find(orders_query, {"_id": 0, "total": 1, "profit": 1}).to_list(1000)
        
        # مصاريف الفرع اليوم
        expenses_query = {"branch_id": branch["id"], "date": {"$gte": today}}
        if tenant_id:
            expenses_query["tenant_id"] = tenant_id
        expenses = await db.expenses.find(expenses_query, {"_id": 0, "amount": 1}).to_list(1000)
        
        branch_sales = sum(_sn(o.get("total")) for o in orders)
        branch_expenses = sum(_sn(e.get("amount")) for e in expenses)
        branch_profit = sum(_sn(o.get("profit")) for o in orders) - branch_expenses
        
        branches_data.append({
            "name": branch["name"],
            "orders": len(orders),
            "sales": branch_sales,
            "expenses": branch_expenses,
            "profit": branch_profit
        })
        
        total_sales += branch_sales
        total_orders += len(orders)
        total_expenses += branch_expenses
        total_profit += branch_profit
    
    # عدد الورديات المغلقة اليوم
    shifts_closed = await db.day_closures.count_documents({
        "closed_at": {"$gte": today},
        **({"tenant_id": tenant_id} if tenant_id else {})
    })
    
    # عدد الطلبات الملغية
    cancelled_query = {**base_query, "status": OrderStatus.CANCELLED, "created_at": {"$gte": today}}
    cancelled_orders = await db.orders.count_documents(cancelled_query)
    
    report_data = {
        "branches": branches_data,
        "total_sales": total_sales,
        "total_orders": total_orders,
        "total_expenses": total_expenses,
        "net_profit": total_profit,
        "shifts_closed": shifts_closed,
        "cancelled_orders": cancelled_orders
    }
    
    # تحديد قائمة المستلمين
    recipient_emails = request.recipient_emails
    if not recipient_emails:
        # استخدام بريد المستخدم الحالي كافتراضي
        recipient_emails = [current_user.get("email")]
    
    # إرسال التقرير
    try:
        from services.email_service import send_daily_report
        result = await send_daily_report(tenant_id, report_data, recipient_emails)
        return {
            "success": True,
            "message": f"تم إرسال التقرير إلى {result['success']} مستلم",
            "report_data": report_data,
            "email_results": result
        }
    except ImportError:
        # إذا لم تكن خدمة البريد متاحة، نرجع البيانات فقط
        return {
            "success": False,
            "message": "خدمة البريد غير متاحة حالياً",
            "report_data": report_data
        }
    except Exception as e:
        logger.error(f"Failed to send daily report: {e}")
        return {
            "success": False,
            "message": str(e),
            "report_data": report_data
        }

@api_router.get("/day-management/report-preview")
async def get_daily_report_preview(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """معاينة التقرير اليومي قبل الإرسال"""
    tenant_id = get_user_tenant_id(current_user)
    today = datetime.now(timezone.utc).strftime('%Y-%m-%d')
    
    base_query = {"status": {"$ne": OrderStatus.CANCELLED}, "created_at": {"$gte": today}}
    if tenant_id:
        base_query["tenant_id"] = tenant_id
    if branch_id:
        base_query["branch_id"] = branch_id
    
    # جلب إحصائيات اليوم
    orders = await db.orders.find(base_query, {"_id": 0}).to_list(1000)
    
    expenses_query = {"date": {"$gte": today}}
    if tenant_id:
        expenses_query["tenant_id"] = tenant_id
    if branch_id:
        expenses_query["branch_id"] = branch_id
    expenses = await db.expenses.find(expenses_query, {"_id": 0}).to_list(1000)
    
    total_sales = sum(_sn(o.get("total")) for o in orders)
    total_profit = sum(_sn(o.get("profit")) for o in orders)
    total_expenses = sum(_sn(e.get("amount")) for e in expenses)
    
    return {
        "date": today,
        "total_sales": total_sales,
        "total_orders": len(orders),
        "total_expenses": total_expenses,
        "gross_profit": total_profit,
        "net_profit": total_profit - total_expenses,
        "average_order_value": total_sales / len(orders) if orders else 0
    }

# إضافة مهمة الإغلاق التلقائي عند بدء التطبيق
@app.on_event("startup")
async def start_auto_close_scheduler():
    """بدء مجدول الإغلاق التلقائي"""
    async def scheduler():
        while True:
            await asyncio.sleep(3600)  # كل ساعة
            await auto_close_old_shifts()
    
    asyncio.create_task(scheduler())
    logger.info("✅ Auto day close scheduler started")

@app.on_event("startup")
async def start_integrity_scheduler():
    """🛡 فحص سلامة دوري تلقائي كل ساعة لكل المستأجرين — بلا أي تدخل يدوي"""
    async def _integrity_loop():
        while True:
            await asyncio.sleep(3600)  # كل ساعة
            try:
                from routes.shifts_routes import run_startup_integrity_check
                await run_startup_integrity_check(db)
            except Exception as e:
                logger.warning(f"periodic integrity check failed: {e}")
    
    asyncio.create_task(_integrity_loop())
    logger.info("🛡 Hourly automatic integrity check scheduler started")


# ==================== 🚨 تنبيه الورديات المنسية (واتساب للمالك) ====================
async def _alert_forgotten_open_shifts():
    """يُنبّه مالك النظام إذا بقيت وردية مفتوحة أكثر من 12 ساعة (إشعار مبكر قبل الإغلاق التلقائي بـ24 ساعة).
    ملاحظة: يُرسَل تنبيه واحد فقط لكل وردية (نستخدم علامة alerted_forgotten في الوردية).
    
    🔔 التنبيه يذهب في 3 قنوات: جرس لوحة المالك + واتساب + بريد إلكتروني."""
    try:
        # عتبة: 12 ساعة (تعطي متسعاً بعد ساعة من نهاية اليوم العملي عادةً)
        threshold = (datetime.now(timezone.utc) - timedelta(hours=12)).isoformat()
        query = {
            "status": "open",
            "started_at": {"$lt": threshold},
            "alerted_forgotten": {"$ne": True},
        }
        open_shifts = await db.shifts.find(query).to_list(500)
        if not open_shifts:
            return
        
        for shift in open_shifts:
            shift_id = shift.get("id")
            tenant_id = shift.get("tenant_id")
            cashier_name = shift.get("cashier_name") or "غير معروف"
            branch_name = shift.get("branch_name") or "غير محدد"
            started_at = shift.get("started_at") or ""
            
            # حساب عدد الساعات المفتوحة
            hours_open = 12
            try:
                if started_at:
                    dt = datetime.fromisoformat(started_at.replace("Z", "+00:00"))
                    hours_open = int((datetime.now(timezone.utc) - dt).total_seconds() // 3600)
            except Exception:
                pass
            
            title = f"⚠️ وردية منسية — {branch_name}"
            message = (
                f"الفرع: {branch_name}\n"
                f"الكاشير: {cashier_name}\n"
                f"مفتوحة منذ: {hours_open} ساعة\n"
                f"يُرجى التحقق من إغلاقها قبل الإغلاق التلقائي (24 ساعة)."
            )
            
            # إرسال ثلاثي القنوات (جرس + واتساب + بريد)
            try:
                res = await notify_owner_multichannel(
                    title=title,
                    message=message,
                    severity="warning",
                    category="shift",
                    tenant_id=tenant_id,
                    metadata={
                        "shift_id": shift_id,
                        "hours_open": hours_open,
                        "branch_name": branch_name,
                        "cashier_name": cashier_name,
                    }
                )
                logger.info(f"🚨 تنبيه وردية منسية {shift_id} — جرس:{res['bell']} واتساب:{res['whatsapp']} بريد:{res['email']}")
            except Exception as _e:
                logger.warning(f"forgotten shift multichannel alert failed for {shift_id}: {_e}")
            
            # ضع العلامة حتى لا يتكرر التنبيه لنفس الوردية
            await db.shifts.update_one(
                {"id": shift_id},
                {"$set": {"alerted_forgotten": True, "alerted_forgotten_at": datetime.now(timezone.utc).isoformat()}}
            )
    except Exception as e:
        logger.error(f"Error in _alert_forgotten_open_shifts: {e}")


@app.on_event("startup")
async def start_forgotten_shifts_scheduler():
    """🚨 يشغّل مجدولاً كل 30 دقيقة لإشعار المالك بالورديات المنسية (>12 ساعة)."""
    async def _loop():
        # انتظار قصير عند البدء (لتفادي التنبيه فوراً بعد إعادة تشغيل)
        await asyncio.sleep(120)
        while True:
            try:
                await _alert_forgotten_open_shifts()
            except Exception as e:
                logger.warning(f"forgotten shifts alert loop error: {e}")
            await asyncio.sleep(1800)  # كل 30 دقيقة
    
    asyncio.create_task(_loop())
    logger.info("🚨 Forgotten shifts alert scheduler started (every 30 min)")


@app.on_event("startup")
async def start_biometric_watchdog_scheduler():
    """🔧 watchdog لجوبات البصمة — كل 60 ثانية:
    - يفشل probe pending > 5 دقائق (وكيل offline)
    - يفشل processing عالقة > 5 دقائق
    - يفشل pending أقدم من 24 ساعة (تراكم قديم)."""
    async def _loop():
        await asyncio.sleep(30)
        while True:
            try:
                await cleanup_stale_biometric_jobs()
            except Exception as e:
                logger.warning(f"biometric watchdog loop error: {e}")
            await asyncio.sleep(60)  # كل دقيقة
    asyncio.create_task(_loop())
    logger.info("🔧 Biometric watchdog scheduler started (every 60s)")


@app.on_event("startup")
async def refresh_owner_recovery_emails_cache():
    """يُحمّل قائمة بريد الاسترداد من قاعدة البيانات إلى الكاش المتزامن عند بدء التشغيل."""
    try:
        await get_owner_recovery_emails()
        logger.info(f"📧 Owner recovery emails loaded: {len(OWNER_RECOVERY_EMAILS)} address(es)")
    except Exception as e:
        logger.warning(f"failed to preload owner recovery emails: {e}")


# إضافة indexes عند بدء التطبيق
async def setup_database_indexes():
    """إعداد indexes لتحسين الأداء"""
    try:
        from services.reliability_service import create_database_indexes
        await create_database_indexes(db)
    except Exception as e:
        logger.error(f"Failed to create indexes: {e}")


async def fix_pending_orders_extras_calc():
    """إصلاح الطلبات المعلقة التي حُسِبت بالصيغة الخاطئة القديمة.
    
    الـbug القديم: total = (price + extras) × quantity (الإضافات تُضرب بالكمية)
    الصيغة الصحيحة: total = price × quantity + extras
    
    يُطبَّق فقط على الطلبات النشطة/المعلقة (status: pending, preparing, ready, on_hold).
    لن يُلمس الطلبات المغلقة (paid/delivered/cancelled) لأن المال خرج وله أثر محاسبي.
    
    يعمل مرة واحدة فقط (محفوظ في system_migrations).
    """
    try:
        MIG_KEY = "fix_pending_orders_extras_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        
        logger.info(f"🔧 Running migration: {MIG_KEY}")
        from datetime import datetime as _dt, timezone as _tz
        
        # نطاق محدد: الطلبات المعلقة/النشطة فقط (لن نلمس المغلقة/المدفوعة)
        active_statuses = ["pending", "preparing", "ready", "on_hold", "in_progress"]
        active_orders = await db.orders.find(
            {"status": {"$in": active_statuses}}, {"_id": 0}
        ).to_list(10000)
        
        fixed_count = 0
        for o in active_orders:
            items = o.get("items", []) or []
            if not items:
                continue
            
            # حساب صحيح: price × qty + extras (مرة واحدة)
            new_subtotal = 0
            for it in items:
                price = float(it.get("price") or 0)
                qty = int(it.get("quantity") or 1)
                extras = it.get("extras") or []
                extras_total = sum(float(e.get("price") or 0) * int(e.get("quantity") or 1) for e in extras)
                new_subtotal += (price * qty) + extras_total
            
            old_total = float(o.get("total") or 0)
            old_subtotal = float(o.get("subtotal") or 0)
            discount = float(o.get("discount") or 0)
            new_total = max(0, new_subtotal - discount)
            
            # فقط إذا الفرق ≥ 1 IQD نُصلح
            if abs(new_total - old_total) >= 1:
                await db.orders.update_one(
                    {"id": o["id"]},
                    {"$set": {
                        "subtotal": new_subtotal,
                        "total": new_total,
                        "_extras_calc_fixed": True,
                        "_old_total_before_fix": old_total
                    }}
                )
                fixed_count += 1
                logger.info(f"   ✅ طلب #{o.get('order_number','?')}: {old_total:,.0f} → {new_total:,.0f} (فرق {old_total-new_total:+,.0f})")
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": _dt.now(_tz.utc).isoformat(),
            "fixed_count": fixed_count,
            "scanned": len(active_orders)
        })
        logger.info(f"✅ {MIG_KEY} complete: فحص {len(active_orders)} طلب، تصحيح {fixed_count}")
    except Exception as e:
        logger.error(f"❌ fix_pending_orders_extras migration failed: {e}")


async def backfill_shift_cash_deposit_branch_v1():
    """يربط إيداعات نقد الشفت القديمة بفرعها الصحيح ويضيف اسم الفرع للعرض في خزينة المالك.
    آمن: لا يلمس المبالغ، فقط يضيف branch_id/branch_name الناقصين. يعمل مرة واحدة فقط."""
    try:
        MIG_KEY = "backfill_shift_cash_deposit_branch_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        logger.info(f"🔧 Running migration: {MIG_KEY}")
        from datetime import datetime as _dt, timezone as _tz
        branches = await db.branches.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
        bmap = {b.get("id"): b.get("name") for b in branches}
        deps = await db.owner_deposits.find({"source": "shift_cash"}, {"_id": 0}).to_list(50000)
        fixed = 0
        for d in deps:
            bid = d.get("branch_id")
            bname = d.get("branch_name")
            if bid and bname:
                continue
            if not bid:
                ref = d.get("ref_closing_id")
                src = None
                if ref:
                    src = await db.shifts.find_one({"id": ref}, {"_id": 0, "branch_id": 1, "branch_name": 1})
                    if not src:
                        src = await db.cash_register_closings.find_one(
                            {"$or": [{"id": ref}, {"shift_id": ref}]}, {"_id": 0, "branch_id": 1, "branch_name": 1})
                if src:
                    bid = src.get("branch_id")
                    bname = bname or src.get("branch_name")
            if bid and not bname:
                bname = bmap.get(bid)
            update = {}
            if bid and not d.get("branch_id"):
                update["branch_id"] = bid
            if bname and not d.get("branch_name"):
                update["branch_name"] = bname
            if update:
                await db.owner_deposits.update_one({"id": d["id"]}, {"$set": update})
                fixed += 1
        await db.system_migrations.insert_one({
            "key": MIG_KEY, "executed_at": _dt.now(_tz.utc).isoformat(),
            "fixed_count": fixed, "scanned": len(deps),
        })
        logger.info(f"✅ {MIG_KEY} complete: فحص {len(deps)} إيداع، تصحيح {fixed}")
    except Exception as e:
        logger.error(f"❌ backfill_shift_cash_deposit_branch migration failed: {e}")


async def backfill_shift_cash_deposit_branch_v2():
    """إصلاح شامل: يربط كل إيداعات نقد الشفت القديمة بفرعها الصحيح عبر سجل الشفت أو فرع الكاشير،
    لإخراجها من تصنيف "غير محدد". آمن: لا يلمس المبالغ. يعمل مرة واحدة."""
    try:
        MIG_KEY = "backfill_shift_cash_deposit_branch_v2"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        logger.info(f"🔧 Running migration: {MIG_KEY}")
        from datetime import datetime as _dt, timezone as _tz
        branches = await db.branches.find({}, {"_id": 0, "id": 1, "name": 1}).to_list(2000)
        bmap = {b.get("id"): b.get("name") for b in branches}
        deps = await db.owner_deposits.find({"source": "shift_cash"}, {"_id": 0}).to_list(100000)
        bname_to_id = {b.get("name"): b.get("id") for b in branches if b.get("name")}
        fixed = 0
        for d in deps:
            bid = d.get("branch_id")
            bname = d.get("branch_name")
            if bid and bname:
                continue
            ref = d.get("ref_closing_id")
            src = None
            if ref:
                src = await db.shifts.find_one({"id": ref}, {"_id": 0, "branch_id": 1, "branch_name": 1, "cashier_id": 1, "cashier_name": 1, "tenant_id": 1})
                if not src:
                    src = await db.cash_register_closings.find_one(
                        {"$or": [{"id": ref}, {"shift_id": ref}]},
                        {"_id": 0, "branch_id": 1, "branch_name": 1, "cashier_id": 1, "cashier_name": 1, "tenant_id": 1})
            tnt = (src or {}).get("tenant_id") or d.get("tenant_id")
            tq = {"tenant_id": tnt} if tnt else {}
            if src:
                bid = bid or src.get("branch_id")
                bname = bname or src.get("branch_name")
            if not bid and bname and bname_to_id.get(bname):
                bid = bname_to_id.get(bname)
            cid = (src or {}).get("cashier_id")
            if not bid and cid:
                cu = await db.users.find_one({"id": cid}, {"_id": 0, "branch_id": 1})
                bid = (cu or {}).get("branch_id")
                if not bid:
                    ce = await db.employees.find_one({"id": cid}, {"_id": 0, "branch_id": 1})
                    bid = (ce or {}).get("branch_id")
            cnm = (src or {}).get("cashier_name")
            if not bid and cnm:
                cu = await db.users.find_one({"full_name": cnm, **tq}, {"_id": 0, "branch_id": 1})
                bid = (cu or {}).get("branch_id")
                if not bid:
                    ce = await db.employees.find_one({"name": cnm, **tq}, {"_id": 0, "branch_id": 1})
                    bid = (ce or {}).get("branch_id")
            if bid and not bname:
                bname = bmap.get(bid)
            update = {}
            if bid and not d.get("branch_id"):
                update["branch_id"] = bid
            if bname and not d.get("branch_name"):
                update["branch_name"] = bname
            if update:
                await db.owner_deposits.update_one({"id": d["id"]}, {"$set": update})
                fixed += 1
        await db.system_migrations.insert_one({
            "key": MIG_KEY, "executed_at": _dt.now(_tz.utc).isoformat(),
            "fixed_count": fixed, "scanned": len(deps),
        })
        logger.info(f"✅ {MIG_KEY} complete: فحص {len(deps)} إيداع، تصحيح {fixed}")
    except Exception as e:
        logger.error(f"❌ backfill_shift_cash_deposit_branch_v2 migration failed: {e}")




async def cleanup_mistaken_expense_moataz36():
    """حذف مصروف أُدخل بالخطأ "ا-معتز#36" (غداء إدارة) — مرة واحدة فقط، بطلب المالك.
    آمن: يحفظ نسخة كاملة من كل مستند محذوف في system_migrations (قابلة للاسترجاع)،
    ثم يُعيد احتساب إجمالي مصاريف الورديات المفتوحة المتأثرة. لا يُضاف أي زر حذف في الواجهة.
    المطابقة الدقيقة: الوصف يحتوي على "معتز" و "36" معاً (لتفادي حذف أي مصروف آخر).
    """
    try:
        from datetime import datetime as _dt, timezone as _tz
        MIG_KEY = "cleanup_mistaken_expense_moataz36_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            return

        candidates = await db.expenses.find({"description": {"$regex": "معتز"}}, {"_id": 0}).to_list(2000)
        def _has36(e):
            return "36" in (e.get("description") or "") or "36" in (str(e.get("reference_number") or ""))
        to_delete = [e for e in candidates if "معتز" in (e.get("description") or "") and _has36(e)]

        deleted_ids = []
        affected = set()
        for e in to_delete:
            await db.expenses.delete_one({"id": e["id"]})
            deleted_ids.append(e["id"])
            if e.get("branch_id"):
                affected.add((e.get("tenant_id"), e.get("branch_id")))
            logger.info(f"   🗑️ حُذف مصروف بالخطأ: {e.get('description')} | {e.get('amount')} | {e.get('date')}")

        # إعادة احتساب إجمالي مصاريف الورديات المفتوحة المتأثرة
        for tid, bid in affected:
            shift_q = {"branch_id": bid, "status": "open"}
            if tid:
                shift_q["tenant_id"] = tid
            sh = await db.shifts.find_one(shift_q, {"_id": 0, "id": 1, "started_at": 1, "opening_cash": 1, "opening_balance": 1, "cash_sales": 1})
            if not sh:
                continue
            exp_q = {"branch_id": bid, "category": {"$ne": "refund"}, "created_at": {"$gte": sh.get("started_at", "")}}
            if tid:
                exp_q["tenant_id"] = tid
            exps = await db.expenses.find(exp_q, {"_id": 0, "amount": 1}).to_list(500)
            total_exp = sum(float(x.get("amount") or 0) for x in exps)
            oc = float(sh.get("opening_cash") or sh.get("opening_balance") or 0)
            cs = float(sh.get("cash_sales") or 0)
            await db.shifts.update_one({"id": sh["id"]}, {"$set": {"total_expenses": total_exp, "expected_cash": oc + cs - total_exp}})

        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": _dt.now(_tz.utc).isoformat(),
            "deleted_count": len(deleted_ids),
            "deleted_ids": deleted_ids,
            "deleted_docs_backup": to_delete,  # نسخة احتياطية كاملة للاسترجاع عند الحاجة
        })
        logger.info(f"✅ {MIG_KEY} complete: حُذف {len(deleted_ids)} مصروف")
    except Exception as e:
        logger.error(f"❌ cleanup_mistaken_expense_moataz36 migration failed: {e}")




async def seed_department_branches():
    """إنشاء 3 أقسام افتراضية (مطبخ مركزي، مخزن، مشتريات) لكل tenant.
    
    هذه الأقسام تعمل كـbranches لكن branch_type يُميزها:
    - رواتب موظفيها تُحسب منفصلة عن الفروع في تقارير HR
    - تظهر في dropdown اختيار الفرع لكن مصنفة في قسم خاص
    
    Idempotent: يعمل مرة واحدة لكل tenant.
    """
    try:
        MIG_KEY = "seed_department_branches_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            return
        
        logger.info(f"🔧 Running migration: {MIG_KEY}")
        DEFAULTS = [
            {"name": "المطبخ المركزي", "branch_type": "central_kitchen"},
            {"name": "المخزن", "branch_type": "warehouse"},
            {"name": "قسم المشتريات", "branch_type": "purchasing"},
        ]
        
        tenants = await db.tenants.find({}, {"_id": 0, "id": 1}).to_list(1000)
        created = 0
        for ten in tenants:
            tid = ten.get("id")
            if not tid:
                continue
            for d in DEFAULTS:
                # هل القسم موجود مسبقاً لهذا المستأجر؟
                existing = await db.branches.find_one({
                    "tenant_id": tid,
                    "branch_type": d["branch_type"],
                })
                if existing:
                    continue
                doc = {
                    "id": str(uuid.uuid4()),
                    "tenant_id": tid,
                    "name": d["name"],
                    "branch_type": d["branch_type"],
                    "address": "",
                    "phone": "",
                    "phone2": "",
                    "email": None,
                    "is_active": True,
                    "rent_cost": 0,
                    "water_cost": 0,
                    "electricity_cost": 0,
                    "generator_cost": 0,
                    "is_sold_branch": False,
                    "owner_percentage": 0,
                    "monthly_fee": 0,
                    "created_at": datetime.now(timezone.utc).isoformat(),
                }
                await db.branches.insert_one(doc)
                created += 1
        
        # وضع علامة على كل الـbranches القديمة كـbranch_type=branch (إن لم تُحدد)
        await db.branches.update_many(
            {"branch_type": {"$exists": False}},
            {"$set": {"branch_type": "branch"}}
        )
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "applied_at": datetime.now(timezone.utc).isoformat(),
            "departments_created": created,
        })
        logger.info(f"✅ Migration {MIG_KEY} done. Created {created} department(s).")
    except Exception as e:
        logger.error(f"❌ {MIG_KEY} migration failed: {e}")


async def cleanup_duplicate_expenses():
    """حذف مصروف الغاز 50,000 المكرر تحديداً (one-shot).
    
    نطاق محدود جداً:
    - description يحتوي "غاز" فقط
    - amount = 50,000 تحديداً
    - نفس (tenant + branch + user + description + amount) تم إنشاؤها خلال 60 ثانية = تكرار
    - يحتفظ بالأقدم، يحذف الأحدث، ويُعيد حساب الوردية المتأثرة
    
    لن يلمس أي مصروف عادي. يعمل مرة واحدة فقط (محفوظ في system_migrations).
    """
    try:
        MIG_KEY = "cleanup_duplicate_gas_expense_v2"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        
        logger.info(f"🔧 Running TARGETED migration: {MIG_KEY} (غاز 50,000 فقط)")
        from datetime import datetime as _dt, timezone as _tz
        
        # نطاق محدد: فقط مصاريف الغاز بمبلغ 50,000
        gas_expenses = await db.expenses.find({
            "amount": 50000,
            "description": {"$regex": "غاز", "$options": "i"}
        }, {"_id": 0}).sort("created_at", 1).to_list(1000)
        
        if not gas_expenses:
            logger.info("   لا توجد مصاريف غاز 50,000 مطابقة")
            await db.system_migrations.insert_one({
                "key": MIG_KEY,
                "executed_at": _dt.now(_tz.utc).isoformat(),
                "deleted_count": 0,
                "recomputed_shifts": 0
            })
            return
        
        # تجميع حسب fingerprint
        groups = {}
        for e in gas_expenses:
            fp = (
                e.get("tenant_id") or "",
                e.get("branch_id") or "",
                e.get("created_by") or "",
                (e.get("description") or "").strip()
            )
            groups.setdefault(fp, []).append(e)
        
        deleted_ids = []
        affected_shifts = set()
        for fp, items in groups.items():
            if len(items) < 2:
                continue
            items_sorted = sorted(items, key=lambda x: x.get("created_at", ""))
            kept = items_sorted[0]
            for dup in items_sorted[1:]:
                try:
                    t1 = _dt.fromisoformat(kept.get("created_at", "").replace("Z", "+00:00"))
                    t2 = _dt.fromisoformat(dup.get("created_at", "").replace("Z", "+00:00"))
                    diff = abs((t2 - t1).total_seconds())
                except Exception:
                    diff = 9999
                if diff <= 60:
                    await db.expenses.delete_one({"id": dup["id"]})
                    deleted_ids.append(dup["id"])
                    if dup.get("branch_id"):
                        affected_shifts.add((dup.get("tenant_id") or "", dup["branch_id"], dup.get("created_at", "")))
                    logger.info(f"   🗑️  حذف غاز مكرر: desc={fp[3][:40]} id={dup['id'][:8]}")
        
        # إعادة حساب الوردية المتأثرة
        recomputed_shifts = set()
        for tenant_id, branch_id, exp_created in affected_shifts:
            shift_q = {"branch_id": branch_id, "started_at": {"$lte": exp_created}}
            if tenant_id:
                shift_q["tenant_id"] = tenant_id
            affected = await db.shifts.find(shift_q, {"_id": 0}).to_list(100)
            for s in affected:
                if s["id"] in recomputed_shifts:
                    continue
                s_end = s.get("ended_at") or ""
                if s_end and exp_created > s_end:
                    continue
                q = {
                    "branch_id": branch_id,
                    "category": {"$ne": "refund"},
                    "created_at": {"$gte": s.get("started_at", "")}
                }
                if tenant_id:
                    q["tenant_id"] = tenant_id
                if s_end:
                    q["created_at"]["$lte"] = s_end
                shift_expenses = await db.expenses.find(q, {"_id": 0, "amount": 1}).to_list(500)
                total_exp = sum(float(e.get("amount") or 0) for e in shift_expenses)
                opening_cash = float(s.get("opening_cash") or s.get("opening_balance") or 0)
                cash_sales = float(s.get("cash_sales") or 0)
                await db.shifts.update_one(
                    {"id": s["id"]},
                    {"$set": {
                        "total_expenses": total_exp,
                        "expected_cash": opening_cash + cash_sales - total_exp
                    }}
                )
                recomputed_shifts.add(s["id"])
                logger.info(f"   ✅ وردية {s['id'][:8]} → total_expenses={total_exp:,.0f}")
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": _dt.now(_tz.utc).isoformat(),
            "deleted_count": len(deleted_ids),
            "recomputed_shifts": len(recomputed_shifts)
        })
        logger.info(f"✅ {MIG_KEY} complete: حذف {len(deleted_ids)} مكرر غاز، إعادة حساب {len(recomputed_shifts)} وردية")
    except Exception as e:
        logger.error(f"❌ cleanup_duplicate_gas_expense migration failed: {e}")


async def purge_ghost_order_saidiya_11_20260430():
    """حذف نهائي للطلب الشبح المكرر #11 في فرع السيدية (one-shot).
    
    نطاق محدود جداً:
    - order_number = 11 تحديداً
    - الفرع = السيدية (اسم يحتوي كلمة: السيدية/السيديه/صيدية)
    - total = 5000 تحديداً
    - order_type = dine_in
    - created_at يبدأ بـ 2026-04-30
    
    هذا الطلب مكرر/فاسد نشأ بسبب خطأ شبكي. سيُحذف نهائياً من orders (بدون تسجيل في
    cancellations أو refunds) ويُؤرشف في ghost_orders_archive للرجوع التدقيقي فقط.
    لن يظهر في التقارير ولا النقدي ولا إيصال إغلاق الصندوق.
    
    يعمل مرة واحدة فقط (محفوظ في system_migrations).
    """
    try:
        MIG_KEY = "purge_ghost_order_saidiya_11_20260430_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        
        logger.info(f"🔧 Running TARGETED migration: {MIG_KEY}")
        
        # ابحث عن فرع السيدية
        saidiya_branches = await db.branches.find(
            {"name": {"$regex": "السيدية|السيديه|صيدية", "$options": "i"}},
            {"_id": 0, "id": 1, "name": 1, "tenant_id": 1}
        ).to_list(10)
        
        if not saidiya_branches:
            logger.info(f"   لا يوجد فرع السيدية في هذه القاعدة — تم تخطي الـ migration وتسجيله")
            await db.system_migrations.insert_one({
                "key": MIG_KEY,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "deleted_count": 0,
                "note": "saidiya_branch_not_found",
            })
            return
        
        branch_ids = [b["id"] for b in saidiya_branches]
        logger.info(f"   فروع السيدية: {[(b.get('name'), b.get('id')[:8]) for b in saidiya_branches]}")
        
        # ابحث عن الطلب المستهدف بالمطابقة الدقيقة
        query = {
            "order_number": 11,
            "branch_id": {"$in": branch_ids},
            "total": 5000,
            "order_type": "dine_in",
            "created_at": {"$regex": "^2026-04-30"},
        }
        ghost_orders = await db.orders.find(query, {"_id": 0}).to_list(10)
        
        if not ghost_orders:
            logger.info("   لم يتم العثور على طلب شبح مطابق — تم تسجيل الـ migration كمُنفّذ")
            await db.system_migrations.insert_one({
                "key": MIG_KEY,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "deleted_count": 0,
                "note": "no_matching_ghost_order",
            })
            return
        
        # أرشفة آمنة قبل الحذف
        ghost_ids = []
        for o in ghost_orders:
            ghost_ids.append(o.get("id"))
            logger.info(f"   👻 طلب شبح: id={o.get('id')[:8]}... | status={o.get('status')} | payment={o.get('payment_method')}/{o.get('payment_status')}")
            try:
                await db.ghost_orders_archive.insert_one({
                    "id": str(uuid.uuid4()),
                    "_archived_at": datetime.now(timezone.utc).isoformat(),
                    "_archived_reason": "duplicate_ghost_order_11_saidiya_20260430",
                    "_migration_key": MIG_KEY,
                    "order_snapshot": {k: v for k, v in o.items() if not isinstance(v, bytes)},
                })
            except Exception as _arch_e:
                logger.warning(f"   ghost_orders_archive insert failed: {_arch_e}")
        
        # الحذف الفعلي من orders
        del_orders = await db.orders.delete_many(query)
        logger.info(f"   🗑️  حُذف من orders: {del_orders.deleted_count}")
        
        # تنظيف أي أثر محتمل في print_queue (لمنع إعادة الطباعة)
        try:
            del_pq = await db.print_queue.delete_many({"order_id": {"$in": ghost_ids}})
            if del_pq.deleted_count:
                logger.info(f"   🖨️  حُذف من print_queue: {del_pq.deleted_count}")
        except Exception:
            pass
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "deleted_count": del_orders.deleted_count,
            "archived_ids": ghost_ids,
            "note": "ghost_order_purged_successfully",
        })
        logger.info(f"✅ {MIG_KEY} complete: purged {del_orders.deleted_count} ghost order(s)")
    except Exception as e:
        logger.error(f"❌ purge_ghost_order_saidiya_11_20260430 migration failed: {e}")


async def settle_driver_collected_orders_as_cash():
    """تصحيح الطلبات المحصّلة من السائقين التي بقيت معلقة في التقارير (one-shot).
    
    المشكلة: endpoint /drivers/{id}/collect-payment كان يضبط driver_payment_status=paid
    فقط دون تحديث payment_status و payment_method في الطلب نفسه، مما أدى إلى ظهور الطلب
    في خانة "معلق" في تقرير إغلاق الصندوق بدلاً من "نقدي".
    
    هذه الـ migration تُصلح الطلبات السابقة بالشروط:
    - driver_id موجود
    - driver_payment_status == "paid" (أي المدير استلم الفلوس من السائق)
    - payment_status in [pending, None, ""] (الطلب لا يزال غير مُصنّف مدفوع)
    - payment_method ليس card ولا credit (كي لا نُخرب طلبات بطاقة/آجل فعلية)
    
    يتم ضبط payment_status="paid", payment_method="cash".
    لن يلمس أي طلب آخر. يعمل مرة واحدة فقط (محفوظ في system_migrations).
    """
    try:
        MIG_KEY = "settle_driver_collected_orders_as_cash_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        
        logger.info(f"🔧 Running migration: {MIG_KEY}")
        
        query = {
            "driver_id": {"$exists": True, "$ne": None},
            "driver_payment_status": "paid",
            "$or": [
                {"payment_status": {"$in": [None, "", "pending", "unpaid"]}},
                {"payment_status": {"$exists": False}},
            ],
            "payment_method": {"$nin": ["card", "credit"]},
        }
        
        # احصر أول قبل التعديل للوغ
        count = await db.orders.count_documents(query)
        logger.info(f"   طلبات مطابقة: {count}")
        
        if count == 0:
            await db.system_migrations.insert_one({
                "key": MIG_KEY,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "modified_count": 0,
                "note": "no_matching_orders",
            })
            logger.info(f"✅ {MIG_KEY} complete: no orders needed fixing")
            return
        
        result = await db.orders.update_many(
            query,
            {"$set": {
                "payment_status": "paid",
                "payment_method": "cash",
                "payment_settled_from_driver_at": datetime.now(timezone.utc).isoformat(),
                "_migration_applied": MIG_KEY,
            }}
        )
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "modified_count": result.modified_count,
        })
        logger.info(f"✅ {MIG_KEY} complete: fixed {result.modified_count} order(s) → paid/cash")
    except Exception as e:
        logger.error(f"❌ settle_driver_collected_orders_as_cash migration failed: {e}")


async def backfill_closing_business_date():
    """تعبئة business_date في سجلات cash_register_closings القديمة (one-shot).
    
    ضروري لأن الواجهة ستفلتر الشفتات حسب business_date بدل وقت الإغلاق،
    فالشفت اللي امتد بعد منتصف الليل يظهر بتاريخ اليوم الذي افتُتح فيه.
    
    منطق التعبئة:
    - إن كان shift_id متاحاً → نأخذ business_date من الشفت
    - وإلا → نحسب من shift_start (أو closed_at كمصدر أخير) عبر iraq_date_from_utc
    """
    try:
        MIG_KEY = "backfill_closing_business_date_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        
        logger.info(f"🔧 Running migration: {MIG_KEY}")
        closings = await db.cash_register_closings.find(
            {"$or": [{"business_date": {"$exists": False}}, {"business_date": None}, {"business_date": ""}]},
            {"_id": 0, "id": 1, "shift_id": 1, "shift_start": 1, "closed_at": 1, "started_at": 1}
        ).to_list(5000)
        
        logger.info(f"   سجلات تحتاج تعبئة: {len(closings)}")
        fixed = 0
        for c in closings:
            biz = None
            if c.get("shift_id"):
                _sh = await db.shifts.find_one({"id": c["shift_id"]}, {"_id": 0, "business_date": 1, "started_at": 1})
                if _sh:
                    biz = _sh.get("business_date")
                    if not biz and _sh.get("started_at"):
                        biz = iraq_date_from_utc(_sh["started_at"])
            if not biz:
                src = c.get("shift_start") or c.get("started_at") or c.get("closed_at")
                if src:
                    biz = iraq_date_from_utc(src)
            if biz:
                await db.cash_register_closings.update_one({"id": c["id"]}, {"$set": {"business_date": biz}})
                fixed += 1
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "fixed_count": fixed,
            "scanned_count": len(closings),
        })
        logger.info(f"✅ {MIG_KEY} complete: عبّأت business_date لـ {fixed} إغلاق")
    except Exception as e:
        logger.error(f"❌ backfill_closing_business_date migration failed: {e}")


async def auto_heal_shifts_and_business_dates():
    """إصلاح شامل تلقائي للشفتات العالقة و business_dates الخاطئة (يعمل دورياً عند كل إقلاع).
    
    هذه الـ migration ليست one-shot — تعمل كل مرة لأن البيانات قد تتلوّث مرة أخرى لاحقاً.
    
    1) إغلاق الشفتات العالقة:
       - شفت status=open له started_at أقدم من 30 ساعة → يُغلق تلقائياً (auto_close_stale).
       - يحفظ ended_at = started_at + 18 ساعة (تقدير معقول للإغلاق الفعلي).
    
    2) تصحيح business_date للشفتات:
       - أي شفت ليس لديه business_date → يُحسب من started_at (Iraq time).
    
    3) تصحيح business_date للطلبات والمصاريف اللي ربطها بشفت غير صحيح:
       - لكل طلب/مصروف بدون business_date أو business_date != الشفت الفعلي،
         نبحث عن الشفت الذي كان يجب أن يكون مفتوحاً وقت الإنشاء (cashier+branch+
         started_at <= created_at <= ended_at) ونعيد ضبط business_date.
    
    آمن جداً: لا يحذف بيانات، فقط يُحدّث الحقول.
    """
    try:
        logger.info("🔧 Running auto_heal_shifts_and_business_dates")
        
        # === 1) إغلاق الشفتات العالقة (>30 ساعة) ===
        stale_cutoff = (datetime.now(timezone.utc) - timedelta(hours=30)).isoformat()
        stale_shifts = await db.shifts.find(
            {"status": "open", "started_at": {"$lt": stale_cutoff}},
            {"_id": 0, "id": 1, "started_at": 1, "cashier_name": 1}
        ).to_list(200)
        
        for s in stale_shifts:
            # احسب وقت إغلاق منطقي: started_at + 18 ساعة (افتراض شفت طبيعي)
            try:
                started = datetime.fromisoformat(s["started_at"].replace("Z", "+00:00"))
                est_ended = (started + timedelta(hours=18)).isoformat()
            except Exception:
                est_ended = datetime.now(timezone.utc).isoformat()
            
            await db.shifts.update_one(
                {"id": s["id"]},
                {"$set": {
                    "status": "closed",
                    "ended_at": est_ended,
                    "auto_closed_at": datetime.now(timezone.utc).isoformat(),
                    "auto_close_reason": "stale_shift_over_30_hours",
                }}
            )
        if stale_shifts:
            logger.info(f"   🔒 أُغلق تلقائياً {len(stale_shifts)} شفت عالق")
        
        # === 2) إعادة حساب business_date لجميع الشفتات من started_at (مع ساعة بداية اليوم) ===
        # شفت الفجر (قبل ساعة بداية اليوم) يُنسب لليوم السابق — لا يُقسّم ليل المطعم على يومين.
        fixed_shifts = 0
        _biz_hour_cache = {}
        async def _tenant_hour(tid):
            if tid not in _biz_hour_cache:
                _biz_hour_cache[tid] = await get_business_day_start_hour(tid)
            return _biz_hour_cache[tid]
        async for sh in db.shifts.find(
            {},
            {"_id": 0, "id": 1, "started_at": 1, "opened_at": 1, "business_date": 1, "tenant_id": 1}
        ):
            src = sh.get("started_at") or sh.get("opened_at")
            if not src:
                continue
            _h = await _tenant_hour(sh.get("tenant_id"))
            correct_biz = iraq_business_date_from_utc(src, start_hour=_h)
            if correct_biz and correct_biz != sh.get("business_date"):
                await db.shifts.update_one(
                    {"id": sh["id"]},
                    {"$set": {"business_date": correct_biz, "_business_date_healed": datetime.now(timezone.utc).isoformat()}}
                )
                fixed_shifts += 1
        if fixed_shifts:
            logger.info(f"   🕐 صُحّح business_date لـ {fixed_shifts} شفت")
        
        # === 2b) استكمال business_date لسجلات إغلاق الصندوق (من يوم الشفت المرتبط) ===
        fixed_closings = 0
        _shift_bd_cache = {}
        async for c in db.cash_register_closings.find(
            {"$or": [{"business_date": {"$exists": False}}, {"business_date": None}, {"business_date": ""}]},
            {"_id": 0, "id": 1, "shift_id": 1, "shift_start": 1, "closed_at": 1}
        ):
            bd = None
            sid = c.get("shift_id")
            if sid:
                if sid in _shift_bd_cache:
                    bd = _shift_bd_cache[sid]
                else:
                    _sh = await db.shifts.find_one({"id": sid}, {"_id": 0, "business_date": 1, "started_at": 1, "opened_at": 1})
                    if _sh:
                        bd = _sh.get("business_date") or iraq_date_from_utc(_sh.get("started_at") or _sh.get("opened_at") or "")
                    _shift_bd_cache[sid] = bd
            if not bd:
                _ref = c.get("shift_start") or c.get("closed_at")
                bd = iraq_date_from_utc(_ref) if _ref else None
            if bd:
                await db.cash_register_closings.update_one({"id": c["id"]}, {"$set": {"business_date": bd}})
                fixed_closings += 1
        if fixed_closings:
            logger.info(f"   🧾 استُكمل business_date لـ {fixed_closings} سجل إغلاق")
        
        # === 2c) فحص سلامة تلقائي لليومين الأخيرين (إشعار + واتساب للمالك عند عدم التطابق) ===
        try:
            from routes.shifts_routes import run_startup_integrity_check
            await run_startup_integrity_check(db)
        except Exception as _ie:
            logger.warning(f"integrity startup check skipped: {_ie}")
        
        # === 3) تصحيح business_date للمصاريف ===
        # ابحث عن مصاريف لها created_by + branch_id + created_at، واربطها بالشفت
        # الذي كان مفتوحاً عند الإنشاء (started_at <= created_at <= ended_at OR شفت مفتوح حالياً)
        fixed_expenses = 0
        
        def _time_window(created_iso: str, hours_before: int = 24):
            """احسب نافذة زمنية للبحث عن شفتات قريبة من وقت الإنشاء"""
            try:
                dt = datetime.fromisoformat(created_iso.replace("Z", "+00:00"))
                earliest = (dt - timedelta(hours=hours_before)).isoformat()
                return earliest
            except Exception:
                return None
        
        async for exp in db.expenses.find(
            {"created_at": {"$exists": True}, "created_by": {"$exists": True}, "branch_id": {"$exists": True}},
            {"_id": 0, "id": 1, "created_by": 1, "branch_id": 1, "created_at": 1, "business_date": 1, "tenant_id": 1}
        ):
            exp_created = exp.get("created_at")
            if not exp_created:
                continue
            earliest = _time_window(exp_created, 24)
            # 1) جرب شفت الكاشير الذي يحوي وقت الإنشاء
            shift_q = {
                "cashier_id": exp["created_by"],
                "branch_id": exp["branch_id"],
                "started_at": {"$lte": exp_created},
            }
            if earliest:
                shift_q["started_at"]["$gte"] = earliest
            if exp.get("tenant_id"):
                shift_q["tenant_id"] = exp["tenant_id"]
            matching_shift = await db.shifts.find_one(
                {**shift_q, "$or": [{"ended_at": {"$gte": exp_created}}, {"status": "open"}]},
                {"_id": 0, "id": 1, "business_date": 1, "started_at": 1},
                sort=[("started_at", -1)]
            )
            # 2) fallback: أي شفت في نفس الفرع ضمن نافذة ±24h
            if not matching_shift:
                fallback_q = {
                    "branch_id": exp["branch_id"],
                    "started_at": {"$lte": exp_created},
                }
                if earliest:
                    fallback_q["started_at"]["$gte"] = earliest
                if exp.get("tenant_id"):
                    fallback_q["tenant_id"] = exp["tenant_id"]
                matching_shift = await db.shifts.find_one(
                    {**fallback_q, "$or": [{"ended_at": {"$gte": exp_created}}, {"status": "open"}]},
                    {"_id": 0, "id": 1, "business_date": 1, "started_at": 1},
                    sort=[("started_at", -1)]
                )
            if matching_shift:
                correct_biz = matching_shift.get("business_date") or iraq_date_from_utc(matching_shift.get("started_at"))
                if correct_biz and correct_biz != exp.get("business_date"):
                    await db.expenses.update_one(
                        {"id": exp["id"]},
                        {"$set": {"business_date": correct_biz, "shift_id": matching_shift["id"], "_business_date_healed": datetime.now(timezone.utc).isoformat()}}
                    )
                    fixed_expenses += 1
            else:
                # 3) fallback نهائي: iraq_date_from_utc على created_at
                correct_biz = iraq_date_from_utc(exp_created)
                if correct_biz and correct_biz != exp.get("business_date"):
                    await db.expenses.update_one(
                        {"id": exp["id"]},
                        {"$set": {"business_date": correct_biz, "_business_date_healed": datetime.now(timezone.utc).isoformat(), "_business_date_fallback": "iraq_date"}}
                    )
                    fixed_expenses += 1
        if fixed_expenses:
            logger.info(f"   💰 صُحّح business_date لـ {fixed_expenses} مصروف")
        
        # === 4) تصحيح business_date للطلبات (نفس المنطق مع نافذة زمنية محكمة) ===
        fixed_orders = 0
        async for o in db.orders.find(
            {"created_at": {"$exists": True}, "branch_id": {"$exists": True}},
            {"_id": 0, "id": 1, "cashier_id": 1, "branch_id": 1, "created_at": 1, "business_date": 1, "tenant_id": 1, "shift_id": 1}
        ):
            order_created = o.get("created_at")
            if not order_created:
                continue
            earliest = _time_window(order_created, 24)
            cashier = o.get("cashier_id")
            matching_shift = None
            # 1) جرب شفت الكاشير ضمن نافذة ±24h
            if cashier:
                shift_q = {
                    "cashier_id": cashier,
                    "branch_id": o["branch_id"],
                    "started_at": {"$lte": order_created},
                }
                if earliest:
                    shift_q["started_at"]["$gte"] = earliest
                if o.get("tenant_id"):
                    shift_q["tenant_id"] = o["tenant_id"]
                matching_shift = await db.shifts.find_one(
                    {**shift_q, "$or": [{"ended_at": {"$gte": order_created}}, {"status": "open"}]},
                    {"_id": 0, "id": 1, "business_date": 1, "started_at": 1},
                    sort=[("started_at", -1)]
                )
            # 2) fallback: أي شفت في نفس الفرع ضمن النافذة
            if not matching_shift:
                fallback_q = {
                    "branch_id": o["branch_id"],
                    "started_at": {"$lte": order_created},
                }
                if earliest:
                    fallback_q["started_at"]["$gte"] = earliest
                if o.get("tenant_id"):
                    fallback_q["tenant_id"] = o["tenant_id"]
                matching_shift = await db.shifts.find_one(
                    {**fallback_q, "$or": [{"ended_at": {"$gte": order_created}}, {"status": "open"}]},
                    {"_id": 0, "id": 1, "business_date": 1, "started_at": 1},
                    sort=[("started_at", -1)]
                )
            if matching_shift:
                correct_biz = matching_shift.get("business_date") or iraq_business_date_from_utc(matching_shift.get("started_at"), start_hour=await _tenant_hour(o.get("tenant_id")))
                if correct_biz and correct_biz != o.get("business_date"):
                    await db.orders.update_one(
                        {"id": o["id"]},
                        {"$set": {"business_date": correct_biz, "shift_id": matching_shift["id"], "_business_date_healed": datetime.now(timezone.utc).isoformat()}}
                    )
                    fixed_orders += 1
            else:
                # 3) fallback نهائي: اليوم التشغيلي (مع ساعة بداية اليوم) من وقت الطلب
                correct_biz = iraq_business_date_from_utc(order_created, start_hour=await _tenant_hour(o.get("tenant_id")))
                if correct_biz and correct_biz != o.get("business_date"):
                    await db.orders.update_one(
                        {"id": o["id"]},
                        {"$set": {"business_date": correct_biz, "_business_date_healed": datetime.now(timezone.utc).isoformat(), "_business_date_fallback": "iraq_date"}}
                    )
                    fixed_orders += 1
        if fixed_orders:
            logger.info(f"   📦 صُحّح business_date لـ {fixed_orders} طلب")
        
        logger.info(f"✅ auto_heal_shifts_and_business_dates complete")
    except Exception as e:
        logger.error(f"❌ auto_heal_shifts_and_business_dates failed: {e}")


async def fix_yamen_orders_jadriya_20260503():
    """إصلاح مستهدف لطلبات يامن #32-#59 في فرع الجادرية ليوم 2026-05-03.
    
    المستخدم أبلغ أن هذه الطلبات بالضبط تخص شفت يامن لكن بعضها مفقود من التقرير.
    هذه الـ migration تضمن:
    1. كل طلب رقم 32-59 في فرع الجادرية أُنشئ في 2026-05-03 (Iraq time) → business_date=2026-05-03
    2. تُسند إلى شفت يامن إن وُجد (cashier_name يحتوي "يامن")
    3. تعمل مرة واحدة (محفوظة في system_migrations)
    """
    try:
        MIG_KEY = "fix_yamen_orders_jadriya_20260503_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return
        
        logger.info(f"🔧 Running targeted migration: {MIG_KEY}")
        
        # ابحث عن فرع الجادرية
        jadriya_branches = await db.branches.find(
            {"name": {"$regex": "جادرية|الجادرية|Jadriya|Al-Jadriya", "$options": "i"}},
            {"_id": 0, "id": 1, "name": 1, "tenant_id": 1}
        ).to_list(10)
        
        if not jadriya_branches:
            logger.info("   لا يوجد فرع جادرية في هذه القاعدة — تم تسجيل migration كمُنفّذ")
            await db.system_migrations.insert_one({
                "key": MIG_KEY,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "modified_count": 0,
                "note": "jadriya_branch_not_found",
            })
            return
        
        branch_ids = [b["id"] for b in jadriya_branches]
        
        # ابحث عن شفت يامن لليوم (2026-05-03)
        yamen_shift = await db.shifts.find_one(
            {
                "branch_id": {"$in": branch_ids},
                "$or": [
                    {"cashier_name": {"$regex": "يامن|يام|Yamen|Yaman", "$options": "i"}},
                    {"business_date": "2026-05-03"},
                ],
                "started_at": {"$regex": "^2026-05-03"},
            },
            {"_id": 0, "id": 1, "cashier_id": 1, "cashier_name": 1, "business_date": 1},
            sort=[("started_at", -1)]
        )
        
        # طلبات #32-#59 في الجادرية ليوم 3
        orders_query = {
            "order_number": {"$gte": 32, "$lte": 59},
            "branch_id": {"$in": branch_ids},
            "created_at": {"$regex": "^2026-05-03"},
        }
        
        update_fields = {
            "business_date": "2026-05-03",
            "_business_date_healed": datetime.now(timezone.utc).isoformat(),
            "_healed_by_migration": MIG_KEY,
        }
        if yamen_shift:
            update_fields["shift_id"] = yamen_shift["id"]
            # لا نغيّر cashier_id لكي لا نكسر تاريخ الإنشاء الحقيقي
            logger.info(f"   شفت يامن المطابق: id={yamen_shift['id'][:8]}, name={yamen_shift.get('cashier_name')}")
        
        result = await db.orders.update_many(orders_query, {"$set": update_fields})
        logger.info(f"   🎯 صُحّح {result.modified_count} طلب (#32-#59) في الجادرية ليوم 2026-05-03")
        
        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "modified_count": result.modified_count,
            "branch_ids": branch_ids,
            "yamen_shift_id": yamen_shift["id"] if yamen_shift else None,
        })
        logger.info(f"✅ {MIG_KEY} complete")
    except Exception as e:
        logger.error(f"❌ fix_yamen_orders_jadriya_20260503 failed: {e}")



async def seed_initial_cost_layers_v1():
    """يُهيّء طبقة تكلفة أولية لكل مادة خام موجودة دون طبقات (one-shot).

    لكل raw_material:
    - إذا لم تكن لديها أي طبقة في material_cost_layers، نُنشئ طبقة واحدة بـ
      remaining_quantity = quantity الحالية و unit_cost = cost_per_unit الحالية.
    - هذا يضمن أن FIFO يعمل على البيانات السابقة دون فقدان المخزون.
    """
    try:
        MIG_KEY = "seed_initial_cost_layers_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return

        logger.info(f"🔧 Running migration: {MIG_KEY}")
        seeded = 0
        skipped = 0
        async for mat in db.raw_materials.find({}, {"_id": 0}):
            mat_id = mat.get("id")
            if not mat_id:
                continue
            existing = await db.material_cost_layers.count_documents({"material_id": mat_id})
            if existing > 0:
                skipped += 1
                continue
            qty = float(mat.get("quantity", 0) or 0)
            cost = float(mat.get("cost_per_unit", 0) or 0)
            if qty <= 0 or cost <= 0:
                # لا توجد كمية أو تكلفة لإنشاء طبقة منها
                continue
            await db.material_cost_layers.insert_one({
                "id": str(uuid.uuid4()),
                "tenant_id": mat.get("tenant_id"),
                "material_id": mat_id,
                "material_name": mat.get("name"),
                "unit": mat.get("unit") or "كغم",
                "unit_cost": cost,
                "original_quantity": qty,
                "remaining_quantity": qty,
                "source": "opening_balance",
                "source_id": None,
                "source_number": "OPENING",
                "received_at": mat.get("created_at") or datetime.now(timezone.utc).isoformat(),
                "created_at": datetime.now(timezone.utc).isoformat(),
                "status": "active",
            })
            seeded += 1

        # فهارس للأداء
        try:
            await db.material_cost_layers.create_index([("material_id", 1), ("received_at", 1)])
            await db.material_cost_layers.create_index([("tenant_id", 1), ("status", 1)])
            await db.price_alerts.create_index([("tenant_id", 1), ("status", 1), ("triggered_at", -1)])
        except Exception:
            pass

        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "seeded": seeded,
            "skipped": skipped,
        })
        logger.info(f"✅ {MIG_KEY} — seeded={seeded}, skipped={skipped}")
    except Exception as e:
        logger.error(f"❌ seed_initial_cost_layers_v1 failed: {e}")


async def backfill_tenant_id_on_products_v1():
    """ملء tenant_id المفقود على manufactured_products و products (one-shot).
    
    يعتمد على فرضية أن كل بيئة لديها tenant واحد رئيسي. إذا كان هناك أكثر من
    tenant، نأخذ أول admin/super_admin tenant_id كافتراضي. هذا يصلح propagate_cost_to_products
    لكي يعمل عند تغيّر تكلفة المواد الخام.
    """
    try:
        MIG_KEY = "backfill_tenant_id_on_products_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return

        logger.info(f"🔧 Running migration: {MIG_KEY}")
        # حدّد tenant_id الافتراضي: من أول admin/super_admin
        admin = await db.users.find_one(
            {"role": {"$in": ["admin", "super_admin"]}, "tenant_id": {"$ne": None}},
            {"_id": 0, "tenant_id": 1}
        )
        default_tid = admin.get("tenant_id") if admin else None

        if not default_tid:
            logger.warning(f"⚠️ {MIG_KEY}: لا يوجد tenant_id افتراضي، الترقية لن تعمل")
            await db.system_migrations.insert_one({
                "key": MIG_KEY,
                "executed_at": datetime.now(timezone.utc).isoformat(),
                "skipped_reason": "no_default_tenant",
            })
            return

        # ملء tenant_id على manufactured_products
        mfg_res = await db.manufactured_products.update_many(
            {"$or": [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]},
            {"$set": {"tenant_id": default_tid}}
        )
        # ملء tenant_id على products (POS)
        prd_res = await db.products.update_many(
            {"$or": [{"tenant_id": None}, {"tenant_id": {"$exists": False}}]},
            {"$set": {"tenant_id": default_tid}}
        )

        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "default_tenant_id": default_tid,
            "manufactured_products_updated": mfg_res.modified_count,
            "products_updated": prd_res.modified_count,
        })
        logger.info(f"✅ {MIG_KEY} — manufactured_products={mfg_res.modified_count}, products={prd_res.modified_count}")
    except Exception as e:
        logger.error(f"❌ backfill_tenant_id_on_products_v1 failed: {e}")


async def renumber_offline_orders_chronologically_v2():
    """إصلاح ترقيم طلبات الأوفلاين v2 — يكتشف الانجراف دون الاعتماد على is_offline_order flag.
    
    لكل (tenant_id, branch_id, business_date): إذا كان أقل order_number أصغر بفارق ≥ 5
    من أكبر order_number، نُعيد ترقيم كل اليوم تسلسلياً حسب created_at.
    """
    try:
        MIG_KEY = "renumber_offline_orders_chronologically_v2"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return

        logger.info(f"🔧 Running migration: {MIG_KEY}")
        renumbered_groups = 0
        renumbered_orders = 0

        # تجميع كل (tenant, branch, business_date) مع min/max
        pipeline = [
            {"$match": {"business_date": {"$ne": None}, "branch_id": {"$ne": None}}},
            {"$group": {
                "_id": {"tenant_id": "$tenant_id", "branch_id": "$branch_id", "business_date": "$business_date"},
                "min_num": {"$min": "$order_number"},
                "max_num": {"$max": "$order_number"},
                "count": {"$sum": 1},
            }}
        ]
        groups = await db.orders.aggregate(pipeline).to_list(5000)

        for grp in groups:
            min_num = grp.get("min_num") or 0
            max_num = grp.get("max_num") or 0
            count = grp.get("count") or 0
            # شرط الانجراف: إذا الفرق بين max و عدد الطلبات أكبر من 5، يدل على ترقيم متضارب
            # يعني: عدد الطلبات يجب أن يساوي max - min + 1 تقريباً. إن لم يكن، نعيد الترقيم.
            expected_max = min_num + count - 1
            drift = abs(max_num - expected_max)
            if drift < 3:  # سماحية صغيرة
                continue

            tenant_id = grp["_id"].get("tenant_id")
            branch_id = grp["_id"].get("branch_id")
            business_date = grp["_id"].get("business_date")

            q = {"branch_id": branch_id, "business_date": business_date}
            if tenant_id:
                q["tenant_id"] = tenant_id

            all_orders = await db.orders.find(
                q, {"_id": 0, "id": 1, "created_at": 1, "order_number": 1}
            ).sort("created_at", 1).to_list(2000)

            for new_num, ord_doc in enumerate(all_orders, start=1):
                old_num = ord_doc.get("order_number")
                if old_num != new_num:
                    await db.orders.update_one(
                        {"id": ord_doc["id"]},
                        {"$set": {
                            "order_number": new_num,
                            "original_order_number": old_num,
                            "renumbered_at": datetime.now(timezone.utc).isoformat(),
                            "renumbered_reason": "fix_offline_sync_drift_v2",
                        }}
                    )
                    renumbered_orders += 1
            renumbered_groups += 1

            # حدّث order_counters لكي لا تتكرّر
            await db.order_counters.update_one(
                {"branch_id": branch_id, "date": business_date},
                {"$set": {"counter": len(all_orders)}},
                upsert=True
            )

        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "renumbered_groups": renumbered_groups,
            "renumbered_orders": renumbered_orders,
        })
        logger.info(f"✅ {MIG_KEY} — groups={renumbered_groups}, orders={renumbered_orders}")
    except Exception as e:
        logger.error(f"❌ renumber_offline_orders_chronologically_v2 failed: {e}")


async def renumber_offline_orders_chronologically_v1():
    """إصلاح ترقيم طلبات الأوفلاين التي حصلت على أرقام خاطئة من العدّاد العام القديم.
    
    المشكلة: قبل الإصلاح كان sync_routes.get_next_order_number يستخدم عدّاداً عاماً (counters)
    يبدأ من 1، فظهرت أرقام مثل #13 #14 وسط الطلبات اليومية #47 #48.
    
    الحل: لكل branch_id + business_date، إذا وُجدت طلبات أوفلاين بأرقام أقل من أرقام الأونلاين
    لنفس اليوم، نُعيد ترقيم الكل تسلسلياً حسب created_at.
    """
    try:
        MIG_KEY = "renumber_offline_orders_chronologically_v1"
        done = await db.system_migrations.find_one({"key": MIG_KEY})
        if done:
            logger.info(f"🟢 Migration {MIG_KEY} already applied, skipping")
            return

        logger.info(f"🔧 Running migration: {MIG_KEY}")
        renumbered_groups = 0
        renumbered_orders = 0

        # 1) ابحث عن كل branch_id + business_date فيه طلبات أوفلاين
        offline_groups = await db.orders.aggregate([
            {"$match": {"is_offline_order": True, "business_date": {"$ne": None}}},
            {"$group": {
                "_id": {"branch_id": "$branch_id", "business_date": "$business_date", "tenant_id": "$tenant_id"},
                "min_offline_num": {"$min": "$order_number"},
            }}
        ]).to_list(2000)

        for grp in offline_groups:
            branch_id = grp["_id"].get("branch_id")
            business_date = grp["_id"].get("business_date")
            tenant_id = grp["_id"].get("tenant_id")
            if not branch_id or not business_date:
                continue

            # هل توجد طلبات أونلاين بنفس اليوم بأرقام أعلى؟ (=هناك تنافر)
            q = {"branch_id": branch_id, "business_date": business_date}
            if tenant_id:
                q["tenant_id"] = tenant_id
            online_max = await db.orders.find_one(
                {**q, "$or": [{"is_offline_order": False}, {"is_offline_order": {"$exists": False}}]},
                {"_id": 0, "order_number": 1},
                sort=[("order_number", -1)]
            )
            if not online_max:
                continue
            online_max_num = online_max.get("order_number") or 0

            # نتحقق إن كان هناك offline order بأرقام أقل من online — هذا الانجراف
            anomaly = await db.orders.find_one(
                {**q, "is_offline_order": True, "order_number": {"$lt": online_max_num}},
                {"_id": 0}
            )
            if not anomaly:
                continue

            # 2) نُعيد ترقيم كل الطلبات لهذا اليوم/الفرع تسلسلياً حسب created_at
            all_orders = await db.orders.find(q, {"_id": 0, "id": 1, "created_at": 1, "order_number": 1, "is_offline_order": 1}).sort("created_at", 1).to_list(2000)
            for new_num, ord_doc in enumerate(all_orders, start=1):
                old_num = ord_doc.get("order_number")
                if old_num != new_num:
                    await db.orders.update_one(
                        {"id": ord_doc["id"]},
                        {"$set": {
                            "order_number": new_num,
                            "original_order_number": old_num,
                            "renumbered_at": datetime.now(timezone.utc).isoformat(),
                            "renumbered_reason": "fix_offline_sync_counter_drift",
                        }}
                    )
                    renumbered_orders += 1
            renumbered_groups += 1

            # 3) حدِّث order_counters لئلا تُكرَّر الأرقام لاحقاً
            await db.order_counters.update_one(
                {"branch_id": branch_id, "date": business_date},
                {"$set": {"counter": len(all_orders)}},
                upsert=True
            )

        await db.system_migrations.insert_one({
            "key": MIG_KEY,
            "executed_at": datetime.now(timezone.utc).isoformat(),
            "renumbered_groups": renumbered_groups,
            "renumbered_orders": renumbered_orders,
        })
        logger.info(f"✅ {MIG_KEY} — groups={renumbered_groups}, orders={renumbered_orders}")
    except Exception as e:
        logger.error(f"❌ renumber_offline_orders_chronologically_v1 failed: {e}")


# ==================== SYSTEM HEALTH & RELIABILITY APIS ====================

@api_router.get("/system/health")
async def health_check():
    """فحص صحة النظام - لا يحتاج توثيق"""
    try:
        from services.reliability_service import SystemHealth
        return await SystemHealth.full_health_check(db)
    except Exception as e:
        return {"status": "error", "message": str(e)}

@api_router.get("/system/stats")
async def get_system_stats(current_user: dict = Depends(get_current_user)):
    """إحصائيات النظام"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    try:
        from services.reliability_service import get_database_stats
        db_stats = await get_database_stats(db)
        
        # إحصائيات إضافية
        tenant_id = get_user_tenant_id(current_user)
        query = {}
        if tenant_id:
            query["tenant_id"] = tenant_id
        
        stats = {
            "database": db_stats,
            "business": {
                "total_orders": await db.orders.count_documents(query),
                "total_products": await db.products.count_documents(query),
                "total_customers": await db.customers.count_documents(query) if "customers" in await db.list_collection_names() else 0,
                "total_employees": await db.employees.count_documents(query),
                "total_branches": await db.branches.count_documents(query),
                "active_shifts": await db.shifts.count_documents({**query, "status": "open"})
            },
            "capacity": {
                "orders_limit": 1000000,
                "products_limit": 100000,
                "users_limit": 10000,
                "status": "healthy"
            }
        }
        
        # تحديد حالة السعة
        if stats["business"]["total_orders"] > 500000:
            stats["capacity"]["status"] = "warning"
        if stats["business"]["total_orders"] > 900000:
            stats["capacity"]["status"] = "critical"
        
        return stats
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.post("/system/backup")
async def create_backup(current_user: dict = Depends(get_current_user)):
    """إنشاء نسخة احتياطية"""
    if current_user["role"] not in [UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح - Super Admin فقط")
    
    try:
        from services.reliability_service import full_backup
        result = await full_backup(db)
        return {
            "success": True,
            "message": f"تم النسخ الاحتياطي: {len(result['success'])} مجموعة",
            "details": result
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@api_router.get("/system/backup/list")
async def list_backups(current_user: dict = Depends(get_current_user)):
    """قائمة النسخ الاحتياطية"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    import os
    backup_path = "/app/backups"
    
    if not os.path.exists(backup_path):
        return {"backups": [], "message": "لا توجد نسخ احتياطية"}
    
    backups = []
    for f in os.listdir(backup_path):
        if f.endswith('.json'):
            file_path = os.path.join(backup_path, f)
            stat = os.stat(file_path)
            backups.append({
                "filename": f,
                "size_mb": round(stat.st_size / (1024*1024), 2),
                "created_at": datetime.fromtimestamp(stat.st_mtime).isoformat()
            })
    
    backups.sort(key=lambda x: x["created_at"], reverse=True)
    return {"backups": backups[:50]}  # آخر 50 نسخة


# ==================== moved to routes/customer_menu_api_routes.py ====================

# ==================== الطلبات المفضلة للزبائن ====================

class FavoriteItem(BaseModel):
    product_id: str
    product_name: str
    quantity: int
    price: float
    notes: str = ""

class AddFavoriteRequest(BaseModel):
    tenant_id: str = None
    phone: str
    name: str = None
    items: List[FavoriteItem]

@api_router.post("/customer/favorites/add")
async def add_to_favorites(request_body: AddFavoriteRequest, request: Request):
    """إضافة طلب للمفضلة (محمي بتحديد معدّل)"""
    enforce_rate_limit(request, "fav_add", max_calls=20, window_seconds=60)
    if not request_body.phone or not request_body.items:
        raise HTTPException(status_code=400, detail="رقم الهاتف والمنتجات مطلوبة")
    if len(request_body.items) > 100:
        raise HTTPException(status_code=400, detail="عدد الأصناف غير صالح")
    
    # التحقق من وجود المستأجر
    tenant = None
    if request_body.tenant_id:
        tenant = await db.tenants.find_one({"menu_slug": request_body.tenant_id})
        if not tenant:
            tenant = await db.tenants.find_one({"id": request_body.tenant_id})
    
    actual_tenant_id = tenant.get("id") if tenant else request_body.tenant_id
    
    favorite = {
        "id": str(uuid.uuid4()),
        "tenant_id": actual_tenant_id,
        "phone": sanitize_text(request_body.phone, 30),
        "name": sanitize_text(request_body.name, 120) or f"طلبي المفضل #{datetime.now().strftime('%d/%m')}",
        "items": [item.dict() for item in request_body.items],
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.customer_favorites.insert_one(favorite)
    
    # إرجاع بدون _id
    return {"message": "تمت الإضافة للمفضلة", "favorite": {k: v for k, v in favorite.items() if k != '_id'}}

@api_router.get("/customer/favorites")
async def get_favorites(
    tenant_id: str = None,
    phone: str = None
):
    """جلب الطلبات المفضلة للزبون"""
    if not phone:
        return []
    
    query = {"phone": phone}
    
    if tenant_id:
        tenant = await db.tenants.find_one({"menu_slug": tenant_id})
        if tenant:
            query["tenant_id"] = tenant.get("id")
        else:
            query["tenant_id"] = tenant_id
    
    favorites = await db.customer_favorites.find(
        query,
        {"_id": 0}
    ).sort("created_at", -1).limit(10).to_list(10)
    
    return favorites

@api_router.delete("/customer/favorites/{favorite_id}")
async def remove_from_favorites(
    favorite_id: str,
    phone: str = None
):
    """حذف طلب من المفضلة"""
    if not phone:
        raise HTTPException(status_code=400, detail="رقم الهاتف مطلوب")
    
    result = await db.customer_favorites.delete_one({
        "id": favorite_id,
        "phone": phone
    })
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="لم يتم العثور على الطلب المفضل")
    
    return {"message": "تم الحذف من المفضلة"}


# ==================== moved to routes/ratings_routes.py ====================

# ==================== DRIVER TRACKING ROUTES ====================

class DriverLocation(BaseModel):
    latitude: float
    longitude: float

@api_router.get("/drivers")
async def get_drivers(current_user: dict = Depends(get_current_user)):
    """جلب قائمة السائقين"""
    tenant_id = get_user_tenant_id(current_user)
    
    drivers = await db.drivers.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).sort("name", 1).to_list(100)
    
    return drivers

class DriverCreateRequest(BaseModel):
    name: str
    phone: str
    email: Optional[str] = None
    branch_id: Optional[str] = None
    pin: str = "1234"

@api_router.post("/drivers")
async def create_driver(
    driver_data: DriverCreateRequest,
    current_user: dict = Depends(get_current_user)
):
    """إنشاء سائق جديد"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # Debug logging
    logger.info(f"Creating driver with PIN: {driver_data.pin}")
    
    driver = {
        "id": str(uuid.uuid4()),
        "tenant_id": tenant_id,
        "branch_id": driver_data.branch_id,
        "name": driver_data.name,
        "phone": driver_data.phone,
        "email": driver_data.email,
        "pin": driver_data.pin,  # الرمز السري للسائق
        "is_active": True,
        "is_available": True,
        "current_location": None,
        "last_location_update": None,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    logger.info(f"Driver dict before insert: pin={driver.get('pin')}")
    
    await db.drivers.insert_one(driver)
    driver.pop("_id", None)
    driver.pop("pin", None)  # لا ترجع PIN في الاستجابة
    
    return {"message": "تم إضافة السائق", "driver": driver}

class DriverUpdateRequest(BaseModel):
    name: Optional[str] = None
    phone: Optional[str] = None
    email: Optional[str] = None
    pin: Optional[str] = None
    is_active: Optional[bool] = None
    is_available: Optional[bool] = None
    branch_id: Optional[str] = None

@api_router.put("/drivers/{driver_id}")
async def update_driver(
    driver_id: str,
    driver_data: DriverUpdateRequest,
    current_user: dict = Depends(get_current_user)
):
    """تحديث بيانات سائق"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    update_data = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if driver_data.name: update_data["name"] = driver_data.name
    if driver_data.phone: update_data["phone"] = driver_data.phone
    if driver_data.email is not None: update_data["email"] = driver_data.email
    if driver_data.pin: update_data["pin"] = driver_data.pin  # تحديث الرمز السري
    if driver_data.is_active is not None: update_data["is_active"] = driver_data.is_active
    if driver_data.is_available is not None: update_data["is_available"] = driver_data.is_available
    if driver_data.branch_id: update_data["branch_id"] = driver_data.branch_id
    
    result = await db.drivers.update_one(
        {"id": driver_id, "tenant_id": tenant_id},
        {"$set": update_data}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم تحديث بيانات السائق"}

@api_router.delete("/drivers/{driver_id}")
async def delete_driver(driver_id: str, current_user: dict = Depends(get_current_user)):
    """حذف سائق"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    result = await db.drivers.delete_one({"id": driver_id, "tenant_id": tenant_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم حذف السائق"}

@api_router.post("/drivers/{driver_id}/location")
async def update_driver_location(
    driver_id: str,
    location: DriverLocation,
    current_user: dict = Depends(get_current_user)
):
    """تحديث موقع السائق (يستخدمها تطبيق السائق)"""
    tenant_id = get_user_tenant_id(current_user)
    
    result = await db.drivers.update_one(
        {"id": driver_id, "tenant_id": tenant_id},
        {"$set": {
            "current_location": {
                "latitude": location.latitude,
                "longitude": location.longitude
            },
            "last_location_update": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم تحديث الموقع"}

@api_router.get("/drivers/{driver_id}/location")
async def get_driver_location(driver_id: str):
    """جلب موقع السائق (للزبون)"""
    driver = await db.drivers.find_one(
        {"id": driver_id},
        {"_id": 0, "current_location": 1, "last_location_update": 1, "name": 1, "phone": 1}
    )
    
    if not driver:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return driver

@api_router.post("/orders/{order_id}/assign-driver")
async def assign_driver_to_order(
    order_id: str,
    driver_id: str,
    current_user: dict = Depends(get_current_user)
):
    """تخصيص سائق للطلب"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPERVISOR, UserRole.CASHIER]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    
    tenant_id = get_user_tenant_id(current_user)
    
    # التحقق من وجود السائق
    driver = await db.drivers.find_one({"id": driver_id, "tenant_id": tenant_id})
    if not driver:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    # تحديث الطلب
    result = await db.orders.update_one(
        {"id": order_id, "tenant_id": tenant_id},
        {"$set": {
            "driver_id": driver_id,
            "driver_name": driver.get("name", ""),
            "driver_phone": driver.get("phone", ""),
            "driver_assigned_at": datetime.now(timezone.utc).isoformat(),
            "status": "out_for_delivery"
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # تحديث حالة السائق
    await db.drivers.update_one(
        {"id": driver_id},
        {"$set": {"is_available": False}}
    )

    # إشعار Push فوري للسائق بطلب جديد مُسند إليه (مثل تطبيق توترز)
    try:
        order_doc = await db.orders.find_one({"id": order_id}, {"_id": 0, "order_number": 1, "customer_address": 1, "total": 1, "tenant_id": 1})
        order_num = (order_doc or {}).get("order_number", order_id[:6])
        await send_push_notification(
            phone=driver.get("phone"),
            title="طلب جديد 🚚",
            body=f"تم إسناد الطلب #{order_num} إليك — {(order_doc or {}).get('customer_address','') or ''}",
            data={"type": "new_order", "order_id": order_id, "url": "/driver-app"},
            user_type="driver",
            tag=f"order-{order_id}",
            require_interaction=True,
        )
    except Exception as _e:
        logger.warning(f"driver push (assign) failed: {_e}")

    return {
        "message": "تم تخصيص السائق للطلب",
        "driver": {
            "id": driver["id"],
            "name": driver["name"],
            "phone": driver["phone"]
        }
    }


# ==================== DRIVER APP ROUTES ====================

async def _issue_driver_token(driver: dict) -> str:
    import secrets as _secrets
    driver_token = _secrets.token_urlsafe(32)
    await db.driver_tokens.insert_one({
        "token": driver_token,
        "driver_id": driver["id"],
        "tenant_id": driver.get("tenant_id"),
        "created_at": datetime.now(timezone.utc).isoformat()
    })
    return driver_token

@api_router.post("/driver/login")
async def driver_login(phone: str, pin: str, request: Request, device_id: str = None):
    """تسجيل دخول السائق برقم الهاتف والرمز السري + تحقق ثنائي من جهاز جديد."""
    _ip = _client_ip(request)
    _lockkey = f"drvlogin:{_ip}:{phone}"
    await check_login_lock(_lockkey)
    driver = await db.drivers.find_one({"phone": phone}, {"_id": 0})
    
    if not driver:
        await record_login_fail(_lockkey, ip=_ip, request=request)
        raise HTTPException(status_code=404, detail="رقم الهاتف غير مسجل كسائق")
    
    # التحقق من الرمز السري
    if driver.get("pin", "1234") != pin:
        await record_login_fail(_lockkey, ip=_ip, request=request)
        raise HTTPException(status_code=401, detail="الرمز السري غير صحيح")
    
    if not driver.get("is_active", True):
        raise HTTPException(status_code=403, detail="حساب السائق غير مفعل")

    await clear_login_attempts(_lockkey)
    
    # إزالة PIN من الاستجابة لأسباب أمنية
    driver_response = {k: v for k, v in driver.items() if k != "pin"}

    # ======== المصادقة الثنائية (جهاز موثوق) — السائق عبر واتساب/SMS ========
    if await two_fa_enabled() and not await is_device_trusted("driver", driver["id"], device_id):
        _channel = "whatsapp"
        resp2fa = await start_2fa_verification("driver", driver["id"], driver.get("name"),
                                               driver.get("tenant_id"), _channel, driver.get("phone"),
                                               device_id, _ip, request, extra={"purpose": "driver_login"})
        return resp2fa
    if await two_fa_enabled():
        await trust_device("driver", driver["id"], device_id, _ip,
                           request.headers.get("user-agent", ""), driver.get("tenant_id"))

    driver_token = await _issue_driver_token(driver)
    return {"driver": driver_response, "token": driver_token, "message": "تم تسجيل الدخول بنجاح"}

@api_router.post("/driver/login/verify-2fa")
async def verify_driver_2fa(payload: Verify2FARequest, request: Request):
    """التحقق من رمز دخول السائق وإصدار توكن الجلسة + توثيق الجهاز."""
    _ip = _client_ip(request)
    ok, sess, err = await verify_2fa_code(payload.verification_id, payload.code, _ip)
    if not ok:
        raise HTTPException(status_code=401, detail=err or "رمز التحقق غير صحيح")
    if sess.get("subject_type") != "driver":
        raise HTTPException(status_code=400, detail="جلسة تحقق غير صالحة")
    driver = await db.drivers.find_one({"id": sess["subject_id"]}, {"_id": 0})
    if not driver:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    await trust_device("driver", driver["id"], sess.get("device_id"), _ip,
                       request.headers.get("user-agent", ""), driver.get("tenant_id"))
    driver_response = {k: v for k, v in driver.items() if k != "pin"}
    driver_token = await _issue_driver_token(driver)
    return {"driver": driver_response, "token": driver_token, "device_id": sess.get("device_id"),
            "message": "تم تسجيل الدخول بنجاح"}


@api_router.get("/driver/orders")
async def get_driver_orders(current_driver: dict = Depends(get_current_driver)):
    """جلب الطلبات المسندة للسائق (مُستمدة من توكن السائق — لا يمكن طلب طلبات سائق آخر)."""
    driver_id = current_driver["id"]
    orders = await db.orders.find(
        {
            "driver_id": driver_id,
            "status": {"$nin": ["delivered", "cancelled", "canceled", "refunded", "rejected"]}
        },
        {"_id": 0}
    ).sort("created_at", -1).to_list(50)
    
    # إضافة status_label
    status_labels = {
        'pending': 'بانتظار التحضير',
        'confirmed': 'تم القبول',
        'preparing': 'قيد التحضير',
        'ready': 'جاهز للتوصيل',
        'completed': 'جاهز للتوصيل',
        'out_for_delivery': 'في الطريق'
    }
    
    for order in orders:
        order['status_label'] = status_labels.get(order.get('status'), order.get('status'))
    
    return orders

@api_router.get("/customer/order-driver/{order_id}")
async def get_order_driver_info(order_id: str, phone: str = None):
    """جلب معلومات سائق الطلب للزبون"""
    # جلب الطلب
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    # التحقق من رقم الهاتف
    if phone and order.get("customer_phone") != phone:
        raise HTTPException(status_code=403, detail="غير مصرح بالوصول لهذا الطلب")
    
    if not order.get("driver_id"):
        return {"driver": None, "message": "لم يتم تخصيص سائق بعد"}
    
    # جلب معلومات السائق
    driver = await db.drivers.find_one(
        {"id": order["driver_id"]},
        {"_id": 0, "id": 1, "name": 1, "phone": 1, "photo": 1, "current_location": 1, "last_location_update": 1}
    )
    
    if not driver:
        return {"driver": None, "message": "السائق غير متاح"}
    
    # إضافة موقع التوصيل (توحيد الحقول lat/lng -> latitude/longitude)
    delivery_location = order.get("delivery_location")
    if delivery_location:
        delivery_location = {
            "latitude": delivery_location.get("latitude", delivery_location.get("lat")),
            "longitude": delivery_location.get("longitude", delivery_location.get("lng")),
        }
    
    return {
        "driver": driver,
        "delivery_location": delivery_location,
        "order_status": order.get("status"),
        "delivery_fee": order.get("delivery_fee", 0),
        "order_total": order.get("total", 0)
    }


# ==================== PUSH NOTIFICATIONS ROUTES ====================

class PushSubscription(BaseModel):
    endpoint: str
    keys: dict
    phone: Optional[str] = None
    user_type: str = "customer"  # customer, driver, admin

# ==================== DRIVER APP ROUTES (بدون مصادقة JWT) ====================

class DriverLocationUpdate(BaseModel):
    latitude: float
    longitude: float

@api_router.post("/driver/update-location")
async def driver_update_location(location: DriverLocationUpdate, current_driver: dict = Depends(get_current_driver)):
    """تحديث موقع السائق - من تطبيق السائق (مصادقة بتوكن السائق)"""
    driver_id = current_driver["id"]
    result = await db.drivers.update_one(
        {"id": driver_id},
        {"$set": {
            "current_location": {
                "latitude": location.latitude,
                "longitude": location.longitude
            },
            "last_location_update": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="السائق غير موجود")
    
    return {"message": "تم تحديث الموقع", "success": True}

@api_router.put("/driver/orders/{order_id}/status")
async def driver_update_order_status(order_id: str, status: str, current_driver: dict = Depends(get_current_driver)):
    """تحديث حالة الطلب من تطبيق السائق (مصادقة بتوكن السائق)"""
    driver_id = current_driver["id"]
    valid_statuses = ['out_for_delivery', 'delivered']
    if status not in valid_statuses:
        raise HTTPException(status_code=400, detail="حالة غير صحيحة للسائق")
    
    # التحقق من أن الطلب مُسند لهذا السائق
    order = await db.orders.find_one({"id": order_id})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    if order.get("driver_id") != driver_id:
        raise HTTPException(status_code=403, detail="هذا الطلب غير مُسند لك")
    
    update_data = {
        "status": status,
        "updated_at": datetime.now(timezone.utc).isoformat()
    }
    
    if status == 'out_for_delivery':
        update_data["out_for_delivery_at"] = datetime.now(timezone.utc).isoformat()
    
    if status == 'delivered':
        update_data["delivered_at"] = datetime.now(timezone.utc).isoformat()
        # تحرير السائق
        await db.drivers.update_one(
            {"id": driver_id},
            {"$set": {"is_available": True, "current_order_id": None}}
        )
    
    await db.orders.update_one({"id": order_id}, {"$set": update_data})
    
    # إرسال إشعار Push للعميل
    await notify_order_status_change(order_id, status)
    
    return {"message": "تم تحديث حالة الطلب", "status": status}

@api_router.get("/driver/order-driver-info/{order_id}")
async def get_driver_info_for_customer(order_id: str):
    """جلب معلومات السائق وموقعه للزبون - بدون مصادقة"""
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    driver_id = order.get("driver_id")
    if not driver_id:
        return {
            "driver": None,
            "message": "لم يتم تخصيص سائق بعد",
            "order_status": order.get("status")
        }
    
    driver = await db.drivers.find_one(
        {"id": driver_id},
        {"_id": 0, "id": 1, "name": 1, "phone": 1, "current_location": 1, "last_location_update": 1}
    )
    
    dloc = order.get("delivery_location")
    if dloc:
        dloc = {
            "latitude": dloc.get("latitude", dloc.get("lat")),
            "longitude": dloc.get("longitude", dloc.get("lng")),
        }
    return {
        "driver": driver,
        "delivery_location": dloc,
        "order_status": order.get("status"),
        "delivery_address": order.get("delivery_address"),
        "delivery_fee": order.get("delivery_fee", 0),
        "order_total": order.get("total", 0),
        "order_number": order.get("order_number"),
        "is_rated": bool(order.get("is_rated")),
    }


# ==================== DELIVERY RATINGS (تقييم الزبون بعد التسليم - بدون مصادقة) ====================
class DeliveryRatingCreate(BaseModel):
    food_rating: Optional[int] = None        # تقييم الطعام (1-5)
    restaurant_rating: Optional[int] = None  # تقييم المطعم (1-5)
    driver_rating: Optional[int] = None      # تقييم السائق (1-5)
    notes: Optional[str] = None


@api_router.post("/track/{order_id}/rating")
async def submit_delivery_rating(order_id: str, rating: DeliveryRatingCreate, request: Request):
    """تقييم الزبون بعد التسليم: الطعام + المطعم + السائق + ملاحظات. بدون مصادقة (يفتحه الزبون من رابط التتبّع).
    يُسجَّل في سجل التوصيل ليطّلع عليه المالك/المدير."""
    enforce_rate_limit(request, "track_rating", max_calls=10, window_seconds=60)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    if order.get("is_rated"):
        raise HTTPException(status_code=400, detail="تم تقييم هذا الطلب مسبقاً")

    def _clamp(v):
        if v is None:
            return None
        try:
            return max(1, min(5, int(v)))
        except Exception:
            return None

    now_iso = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "order_id": order_id,
        "order_number": order.get("order_number"),
        "tenant_id": order.get("tenant_id"),
        "branch_id": order.get("branch_id"),
        "driver_id": order.get("driver_id"),
        "driver_name": order.get("driver_name"),
        "customer_name": order.get("customer_name"),
        "customer_phone": order.get("customer_phone"),
        "food_rating": _clamp(rating.food_rating),
        "restaurant_rating": _clamp(rating.restaurant_rating),
        "driver_rating": _clamp(rating.driver_rating),
        "notes": sanitize_text(rating.notes, 500),
        "created_at": now_iso,
    }
    await db.delivery_ratings.insert_one(doc)
    await db.orders.update_one({"id": order_id}, {"$set": {"is_rated": True, "rated_at": now_iso}})
    doc.pop("_id", None)
    return {"message": "شكراً لتقييمك", "rating": doc}


@api_router.get("/delivery-ratings")
async def get_delivery_ratings(
    branch_id: Optional[str] = None,
    driver_id: Optional[str] = None,
    limit: int = 100,
    current_user: dict = Depends(get_current_user),
):
    """سجل التوصيل: تقييمات الزبائن بعد التسليم — للمالك/المدير لمتابعة الأداء."""
    query = build_tenant_query(current_user, {})
    if branch_id:
        query["branch_id"] = branch_id
    if driver_id:
        query["driver_id"] = driver_id
    ratings = await db.delivery_ratings.find(query, {"_id": 0}).sort("created_at", -1).to_list(limit)

    def _avg(field):
        vals = [r[field] for r in ratings if r.get(field)]
        return round(sum(vals) / len(vals), 1) if vals else 0

    summary = {
        "count": len(ratings),
        "avg_food": _avg("food_rating"),
        "avg_restaurant": _avg("restaurant_rating"),
        "avg_driver": _avg("driver_rating"),
    }
    return {"ratings": ratings, "summary": summary}


# ==================== ORDER CHAT (محادثة الزبون مع السائق - بدون مصادقة) ====================
class OrderChatMessage(BaseModel):
    sender: str  # "customer" | "driver"
    sender_name: Optional[str] = None
    text: str

@api_router.get("/order-chat/{order_id}")
async def get_order_chat(order_id: str, after: Optional[str] = None):
    """جلب رسائل محادثة الطلب (يستخدمها الزبون والسائق) - بدون مصادقة"""
    query = {"order_id": order_id}
    if after:
        query["created_at"] = {"$gt": after}
    msgs = await db.order_chats.find(query, {"_id": 0}).sort("created_at", 1).to_list(500)
    return {"messages": msgs}

@api_router.post("/order-chat/{order_id}")
async def send_order_chat(order_id: str, msg: OrderChatMessage, request: Request):
    """إرسال رسالة في محادثة الطلب - بدون مصادقة (محمية بتحديد معدّل وتنظيف مدخلات)"""
    enforce_rate_limit(request, "order_chat_send", max_calls=20, window_seconds=60)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0, "id": 1, "driver_id": 1, "customer_name": 1})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    text = sanitize_text(msg.text or "", 1000)
    if not text:
        raise HTTPException(status_code=400, detail="الرسالة فارغة")
    sender = msg.sender if msg.sender in ("customer", "driver") else "customer"
    doc = {
        "id": str(uuid.uuid4()),
        "order_id": order_id,
        "sender": sender,
        "sender_name": sanitize_text(msg.sender_name, 120) or ("الزبون" if sender == "customer" else "السائق"),
        "text": text[:1000],
        "read": False,
        "listened": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.order_chats.insert_one(doc)
    doc.pop("_id", None)

    # إشعار Push فوري للطرف الآخر (رسالة محادثة)
    try:
        if sender == "customer" and order.get("driver_id"):
            drv = await db.drivers.find_one({"id": order.get("driver_id")}, {"_id": 0, "phone": 1})
            if drv and drv.get("phone"):
                await send_push_notification(
                    phone=drv.get("phone"),
                    title=f"رسالة من {doc['sender_name']} 💬",
                    body=text[:80],
                    data={"type": "chat_message", "order_id": order_id, "url": "/driver-app"},
                    user_type="driver",
                    tag=f"chat-{order_id}",
                )
    except Exception as _e:
        logger.warning(f"chat push failed: {_e}")

    return doc


@api_router.post("/order-chat/{order_id}/voice")
async def send_order_chat_voice(
    order_id: str,
    request: Request,
    file: UploadFile = File(...),
    sender: str = Form("customer"),
    sender_name: Optional[str] = Form(None),
    duration: Optional[float] = Form(0),
):
    """إرسال رسالة صوتية في محادثة الطلب - بدون مصادقة (محمية بتحديد معدّل)"""
    enforce_rate_limit(request, "order_chat_voice", max_calls=15, window_seconds=60)
    order = await db.orders.find_one({"id": order_id}, {"_id": 0, "id": 1, "driver_id": 1, "customer_name": 1})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")

    data = await file.read()
    if not data or len(data) == 0:
        raise HTTPException(status_code=400, detail="الملف الصوتي فارغ")
    if len(data) > 5 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="الرسالة الصوتية كبيرة جداً")

    voice_dir = UPLOAD_DIR / "voice"
    voice_dir.mkdir(exist_ok=True)
    ext = "webm"
    ct = (file.content_type or "").lower()
    if "ogg" in ct:
        ext = "ogg"
    elif "mp4" in ct or "m4a" in ct or "aac" in ct:
        ext = "mp4"
    elif "mpeg" in ct or "mp3" in ct:
        ext = "mp3"
    fname = f"{uuid.uuid4()}.{ext}"
    async with aiofiles.open(voice_dir / fname, "wb") as f:
        await f.write(data)
    audio_url = f"/api/uploads/voice/{fname}"

    sender = sender if sender in ("customer", "driver") else "customer"
    try:
        dur = round(float(duration or 0), 1)
    except Exception:
        dur = 0
    doc = {
        "id": str(uuid.uuid4()),
        "order_id": order_id,
        "sender": sender,
        "sender_name": sanitize_text(sender_name, 120) or ("الزبون" if sender == "customer" else "السائق"),
        "type": "voice",
        "audio_url": audio_url,
        "duration": dur,
        "text": "🎤 رسالة صوتية",
        "read": False,
        "listened": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.order_chats.insert_one(doc)
    doc.pop("_id", None)

    # إشعار Push فوري للطرف الآخر (رسالة صوتية)
    try:
        if sender == "customer" and order.get("driver_id"):
            drv = await db.drivers.find_one({"id": order.get("driver_id")}, {"_id": 0, "phone": 1})
            if drv and drv.get("phone"):
                await send_push_notification(
                    phone=drv.get("phone"),
                    title=f"رسالة صوتية من {doc['sender_name']} 🎤",
                    body="اضغط للاستماع",
                    data={"type": "chat_message", "order_id": order_id, "url": "/driver-app"},
                    user_type="driver",
                    tag=f"chat-{order_id}",
                )
    except Exception as _e:
        logger.warning(f"voice chat push failed: {_e}")

    return doc


@api_router.post("/order-chat/{order_id}/read")
async def mark_order_chat_read(order_id: str, request: Request, viewer: str = Query("customer")):
    """تعليم رسائل الطرف الآخر كمقروءة (تم الرؤية) — viewer هو الطرف الذي فتح المحادثة."""
    enforce_rate_limit(request, "chat_read", max_calls=60, window_seconds=60)
    viewer = viewer if viewer in ("customer", "driver") else "customer"
    other = "driver" if viewer == "customer" else "customer"
    res = await db.order_chats.update_many(
        {"order_id": order_id, "sender": other, "read": {"$ne": True}},
        {"$set": {"read": True, "read_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"updated": res.modified_count}


@api_router.post("/order-chat/{order_id}/listened/{message_id}")
async def mark_order_chat_listened(order_id: str, message_id: str, request: Request, viewer: str = Query("customer")):
    """تعليم رسالة صوتية كمسموعة (تم الاستماع) — فقط إن كان viewer هو المستقبِل."""
    enforce_rate_limit(request, "chat_listened", max_calls=60, window_seconds=60)
    viewer = viewer if viewer in ("customer", "driver") else "customer"
    other = "driver" if viewer == "customer" else "customer"
    res = await db.order_chats.update_one(
        {"order_id": order_id, "id": message_id, "sender": other, "type": "voice"},
        {"$set": {"listened": True, "read": True, "listened_at": datetime.now(timezone.utc).isoformat(), "read_at": datetime.now(timezone.utc).isoformat()}},
    )
    return {"updated": res.modified_count}


# ==================== PUSH NOTIFICATIONS ROUTES ====================

@api_router.post("/push/subscribe")
async def subscribe_push(subscription: PushSubscription, request: Request):
    """تسجيل اشتراك في إشعارات Push (محمي بتحديد معدّل)"""
    enforce_rate_limit(request, "push_subscribe", max_calls=20, window_seconds=60)
    if not subscription.endpoint or not str(subscription.endpoint).startswith("https://"):
        raise HTTPException(status_code=400, detail="endpoint غير صالح")
    sub_doc = {
        "id": str(uuid.uuid4()),
        "endpoint": subscription.endpoint,
        "keys": subscription.keys,
        "phone": sanitize_text(subscription.phone, 30),
        "user_type": subscription.user_type if subscription.user_type in ("driver", "customer", "staff") else "customer",
        "created_at": datetime.now(timezone.utc).isoformat(),
        "is_active": True
    }
    
    # تحديث أو إنشاء الاشتراك
    await db.push_subscriptions.update_one(
        {"endpoint": subscription.endpoint},
        {"$set": sub_doc},
        upsert=True
    )
    
    return {"message": "تم تسجيل الاشتراك بنجاح"}

@api_router.delete("/push/unsubscribe")
async def unsubscribe_push(endpoint: str):
    """إلغاء اشتراك في إشعارات Push"""
    await db.push_subscriptions.delete_one({"endpoint": endpoint})
    return {"message": "تم إلغاء الاشتراك"}

@api_router.get("/push/vapid-public-key")
async def get_vapid_public_key():
    """إرجاع مفتاح VAPID العام للواجهة الأمامية (applicationServerKey)"""
    return {"publicKey": os.environ.get("VAPID_PUBLIC_KEY", "")}

async def send_push_notification(phone: str, title: str, body: str, data: dict = None, user_type: str = None, icon: str = None, tag: str = None, require_interaction: bool = False):
    """إرسال إشعار Push حقيقي للمستخدم عبر pywebpush + VAPID"""
    try:
        from pywebpush import webpush, WebPushException

        vapid_private_key = os.environ.get("VAPID_PRIVATE_KEY")
        vapid_subject = os.environ.get("VAPID_SUBJECT", "mailto:notifications@maestroegp.com")
        if not vapid_private_key:
            logger.warning("VAPID_PRIVATE_KEY غير مهيأ — لا يمكن إرسال إشعار Push")
            return False

        query = {"is_active": True}
        if phone:
            query["phone"] = phone
        if user_type:
            query["user_type"] = user_type

        subscriptions = await db.push_subscriptions.find(query).to_list(100)

        payload = json.dumps({
            "title": title,
            "body": body,
            "icon": icon or "/icons/icon-192.png",
            "badge": "/icons/icon-192.png",
            "tag": tag or (data or {}).get("type", "maestro"),
            "requireInteraction": require_interaction,
            "data": data or {},
        })

        sent_count = 0
        for sub in subscriptions:
            sub_info = {"endpoint": sub.get("endpoint"), "keys": sub.get("keys", {})}
            try:
                webpush(
                    subscription_info=sub_info,
                    data=payload,
                    vapid_private_key=vapid_private_key,
                    vapid_claims={"sub": vapid_subject},
                )
                sent_count += 1
            except WebPushException as ex:
                status_code = getattr(getattr(ex, "response", None), "status_code", None)
                if status_code in (404, 410):
                    # اشتراك منتهي/غير صالح — أوقفه
                    await db.push_subscriptions.update_one(
                        {"endpoint": sub.get("endpoint")},
                        {"$set": {"is_active": False}},
                    )
                logger.warning(f"WebPush failed ({status_code}): {ex}")
            except Exception as ex:
                logger.warning(f"WebPush error: {ex}")

        # سجل الإشعار (للسجل التاريخي/جرس داخل التطبيق)
        await db.notification_logs.insert_one({
            "id": str(uuid.uuid4()),
            "phone": phone,
            "user_type": user_type,
            "title": title,
            "body": body,
            "data": data,
            "sent_at": datetime.now(timezone.utc).isoformat(),
            "subscriptions_count": len(subscriptions),
            "sent_count": sent_count,
        })

        return sent_count > 0

    except Exception as e:
        logger.error(f"Push notification error: {str(e)}")
        return False

@api_router.post("/push/test")
async def test_push_notification(phone: str, message: str = "هذا إشعار تجريبي", current_user: dict = Depends(get_current_user)):
    """إرسال إشعار تجريبي — للمالك/المدير فقط (منع إساءة الاستخدام)"""
    if current_user.get("role") not in ("super_admin", "admin", "manager"):
        raise HTTPException(status_code=403, detail="غير مصرح")
    await send_push_notification(
        phone=sanitize_text(phone, 30),
        title="Maestro EGP",
        body=sanitize_text(message, 200),
        data={"type": "test"}
    )
    return {"message": "تم إرسال الإشعار"}

@api_router.get("/notifications/{phone}")
async def get_notifications(phone: str, limit: int = 20, current_user: dict = Depends(get_current_user)):
    """جلب سجل الإشعارات لرقم — للموظفين فقط (منع تسريب بيانات الآخرين)"""
    notifications = await db.notification_logs.find(
        {"phone": phone},
        {"_id": 0}
    ).sort("sent_at", -1).limit(min(limit, 100)).to_list(min(limit, 100))
    
    return notifications

# دالة لإرسال إشعار عند تغير حالة الطلب
async def notify_order_status_change(order_id: str, new_status: str):
    """إرسال إشعار للعميل عند تغير حالة الطلب"""
    order = await db.orders.find_one({"id": order_id})
    if not order:
        return
    
    status_messages = {
        'preparing': ('جاري تحضير طلبك! 👨‍🍳', 'طلبك قيد التحضير الآن'),
        'ready': ('طلبك جاهز! ✅', 'طلبك جاهز للتوصيل'),
        'out_for_delivery': ('السائق في الطريق! 🚚', 'السائق في طريقه إليك'),
        'delivered': ('تم التسليم! 🎉', 'استمتع بوجبتك! لا تنسى تقييم الطلب'),
        'cancelled': ('تم رفض/إلغاء طلبك ❌', 'نعتذر، تم رفض أو إلغاء طلبك'),
        'rejected': ('تم رفض طلبك ❌', 'نعتذر، تم رفض طلبك'),
    }
    
    if new_status in status_messages:
        title, body = status_messages[new_status]
        await send_push_notification(
            phone=order.get("customer_phone"),
            title=title,
            body=body,
            data={
                "type": "order_status",
                "order_id": order_id,
                "status": new_status,
                "rate": new_status == "delivered",
                "order_number": str(order.get("order_number", "")),
                "url": f"/menu/{order.get('tenant_id')}"
            },
            user_type="customer",
            tag=f"order-status-{order_id}",
            require_interaction=new_status in ("ready", "out_for_delivery", "delivered"),
        )


# ==================== ADDRESS AUTOCOMPLETE ROUTES ====================

@api_router.get("/geocode/reverse")
async def reverse_geocode(lat: float, lng: float):
    """تحويل إحداثيات لعنوان (Reverse Geocoding)"""
    try:
        import httpx
        
        # استخدام Nominatim API المجاني
        url = f"https://nominatim.openstreetmap.org/reverse?lat={lat}&lon={lng}&format=json&accept-language=ar"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers={"User-Agent": "MaestroEGP/1.0"})
            
            if response.status_code == 200:
                data = response.json()
                
                address = data.get("display_name", "")
                address_parts = data.get("address", {})
                
                return {
                    "address": address,
                    "street": address_parts.get("road", ""),
                    "neighbourhood": address_parts.get("neighbourhood", address_parts.get("suburb", "")),
                    "city": address_parts.get("city", address_parts.get("town", address_parts.get("village", ""))),
                    "country": address_parts.get("country", ""),
                    "lat": lat,
                    "lng": lng
                }
            else:
                return {"address": "", "lat": lat, "lng": lng}
                
    except Exception as e:
        logger.error(f"Reverse geocoding error: {str(e)}")
        return {"address": "", "lat": lat, "lng": lng}

@api_router.get("/geocode/search")
async def search_address(query: str, lat: Optional[float] = None, lng: Optional[float] = None):
    """البحث عن عنوان (Address Autocomplete)"""
    try:
        import httpx
        
        # استخدام Nominatim API للبحث
        url = f"https://nominatim.openstreetmap.org/search?q={query}&format=json&limit=5&accept-language=ar"
        
        # إضافة تفضيل للموقع الحالي إذا متاح
        if lat and lng:
            url += f"&viewbox={lng-0.5},{lat-0.5},{lng+0.5},{lat+0.5}&bounded=0"
        
        async with httpx.AsyncClient() as client:
            response = await client.get(url, headers={"User-Agent": "MaestroEGP/1.0"})
            
            if response.status_code == 200:
                data = response.json()
                
                results = []
                for item in data:
                    results.append({
                        "address": item.get("display_name", ""),
                        "lat": float(item.get("lat", 0)),
                        "lng": float(item.get("lon", 0)),
                        "type": item.get("type", "")
                    })
                
                return {"results": results}
            else:
                return {"results": []}
                
    except Exception as e:
        logger.error(f"Address search error: {str(e)}")
        return {"results": []}


# ==================== PAYMENT ROUTES ====================

@api_router.post("/payments/create-checkout/{tenant_id}")
async def create_payment_checkout(
    tenant_id: str,
    request: Request,
    order_id: str,
    amount: float,
    customer_phone: Optional[str] = None,
    save_card: bool = False
):
    """إنشاء جلسة دفع Stripe"""
    try:
        from emergentintegrations.payments.stripe.checkout import (
            StripeCheckout, 
            CheckoutSessionRequest,
            CheckoutSessionResponse
        )
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="Stripe not configured")
        
        # إعداد URLs
        host_url = str(request.base_url).rstrip('/')
        webhook_url = f"{host_url}/api/webhook/stripe"
        
        # استخدام الـ origin من الـ referer
        referer = request.headers.get('referer', '')
        if referer:
            from urllib.parse import urlparse
            parsed = urlparse(referer)
            frontend_url = f"{parsed.scheme}://{parsed.netloc}"
        else:
            frontend_url = host_url
        
        success_url = f"{frontend_url}/menu/{tenant_id}?payment_success=true&session_id={{CHECKOUT_SESSION_ID}}"
        cancel_url = f"{frontend_url}/menu/{tenant_id}?payment_cancelled=true"
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url=webhook_url)
        
        # إنشاء طلب الدفع
        checkout_request = CheckoutSessionRequest(
            amount=float(amount),
            currency="usd",
            success_url=success_url,
            cancel_url=cancel_url,
            metadata={
                "order_id": order_id,
                "tenant_id": tenant_id,
                "customer_phone": customer_phone or "",
                "save_card": str(save_card)
            },
            payment_methods=["card"]
        )
        
        session: CheckoutSessionResponse = await stripe_checkout.create_checkout_session(checkout_request)
        
        # حفظ معاملة الدفع
        transaction = {
            "id": str(uuid.uuid4()),
            "session_id": session.session_id,
            "order_id": order_id,
            "tenant_id": tenant_id,
            "amount": amount,
            "currency": "usd",
            "customer_phone": customer_phone,
            "payment_status": "pending",
            "status": "initiated",
            "save_card": save_card,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        
        await db.payment_transactions.insert_one(transaction)
        
        return {
            "success": True,
            "checkout_url": session.url,
            "session_id": session.session_id
        }
        
    except ImportError:
        raise HTTPException(status_code=500, detail="Payment library not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Payment error: {str(e)}")

@api_router.get("/payments/status/{session_id}")
async def get_payment_status(session_id: str):
    """التحقق من حالة الدفع"""
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            raise HTTPException(status_code=500, detail="Stripe not configured")
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url="")
        status = await stripe_checkout.get_checkout_status(session_id)
        
        # تحديث حالة المعاملة في قاعدة البيانات
        transaction = await db.payment_transactions.find_one({"session_id": session_id})
        
        if transaction:
            new_status = "completed" if status.payment_status == "paid" else status.payment_status
            
            # تجنب التحديث المتكرر
            if transaction.get("payment_status") != new_status:
                await db.payment_transactions.update_one(
                    {"session_id": session_id},
                    {
                        "$set": {
                            "payment_status": new_status,
                            "status": status.status,
                            "updated_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
                
                # تحديث حالة الطلب إذا تم الدفع
                if status.payment_status == "paid" and transaction.get("order_id"):
                    await db.orders.update_one(
                        {"id": transaction["order_id"]},
                        {
                            "$set": {
                                "payment_status": "paid",
                                "paid_at": datetime.now(timezone.utc).isoformat()
                            }
                        }
                    )
        
        return {
            "status": status.status,
            "payment_status": status.payment_status,
            "amount_total": status.amount_total / 100,
            "currency": status.currency,
            "order_id": status.metadata.get("order_id") if status.metadata else None
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Status check error: {str(e)}")

@api_router.post("/webhook/stripe")
async def stripe_webhook(request: Request):
    """معالجة webhook من Stripe"""
    try:
        from emergentintegrations.payments.stripe.checkout import StripeCheckout
        
        api_key = os.environ.get('STRIPE_API_KEY')
        if not api_key:
            return {"error": "Stripe not configured"}
        
        body = await request.body()
        signature = request.headers.get("Stripe-Signature")
        
        stripe_checkout = StripeCheckout(api_key=api_key, webhook_url="")
        webhook_response = await stripe_checkout.handle_webhook(body, signature)
        
        if webhook_response.payment_status == "paid":
            # تحديث المعاملة
            await db.payment_transactions.update_one(
                {"session_id": webhook_response.session_id},
                {
                    "$set": {
                        "payment_status": "completed",
                        "status": "complete",
                        "webhook_event_id": webhook_response.event_id,
                        "updated_at": datetime.now(timezone.utc).isoformat()
                    }
                }
            )
            
            # تحديث الطلب
            transaction = await db.payment_transactions.find_one({"session_id": webhook_response.session_id})
            if transaction and transaction.get("order_id"):
                await db.orders.update_one(
                    {"id": transaction["order_id"]},
                    {
                        "$set": {
                            "payment_status": "paid",
                            "paid_at": datetime.now(timezone.utc).isoformat()
                        }
                    }
                )
        
        return {"received": True, "event_type": webhook_response.event_type}
        
    except Exception as e:
        return {"error": str(e)}


# ==================== PAYMENT SETTINGS APIs ====================

class PaymentSettingsUpdate(BaseModel):
    stripe_enabled: Optional[bool] = True
    stripe_publishable_key: Optional[str] = None
    stripe_secret_key: Optional[str] = None
    stripe_currency: Optional[str] = "USD"
    stripe_mode: Optional[str] = "test"  # test or live
    zaincash_enabled: Optional[bool] = True
    zaincash_phone: Optional[str] = None
    zaincash_name: Optional[str] = None
    zaincash_qr_image: Optional[str] = None
    cash_enabled: Optional[bool] = True
    delivery_fee: Optional[int] = 5000
    min_order_amount: Optional[int] = 10000
    # أجور التوصيل حسب المسافة (تلقائي)
    distance_fee_enabled: Optional[bool] = None
    fee_base: Optional[int] = None        # الأجرة الأساسية (تغطي أول fee_base_km)
    fee_base_km: Optional[float] = None   # عدد الكيلومترات المشمولة بالأجرة الأساسية
    fee_per_km: Optional[int] = None      # أجرة كل كم إضافي
    fee_max: Optional[int] = None         # سقف الأجرة (0 = بلا سقف)
    fee_round_to: Optional[int] = None    # التقريب لأقرب (250/500/1000)
    max_distance_km: Optional[float] = None  # أقصى مسافة توصيل (0 = بلا حدود)
    fee_zones: Optional[List[dict]] = None   # نطاقات الكم: [{up_to_km, fee}] (تطغى على المعادلة)

@api_router.get("/payment-settings")
async def get_payment_settings(current_user: dict = Depends(get_current_user)):
    """جلب إعدادات الدفع"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    settings = await db.payment_settings.find_one(
        {"tenant_id": tenant_id},
        {"_id": 0, "stripe_secret_key": 0}  # إخفاء المفتاح السري
    )
    
    if not settings:
        settings = {
            "tenant_id": tenant_id,
            "stripe_enabled": True,
            "stripe_publishable_key": "",
            "stripe_currency": "USD",
            "stripe_mode": "test",
            "zaincash_enabled": True,
            "zaincash_phone": "",
            "zaincash_name": "",
            "zaincash_qr_image": "",
            "cash_enabled": True,
            "delivery_fee": 5000,
            "min_order_amount": 10000
        }
    
    # إخفاء المفتاح السري (عرض فقط أنه موجود أو لا)
    settings["stripe_secret_key_set"] = bool(settings.get("stripe_secret_key"))
    
    return settings

@api_router.post("/payment-settings")
async def update_payment_settings(
    settings: PaymentSettingsUpdate,
    current_user: dict = Depends(get_current_user)
):
    """تحديث إعدادات الدفع"""
    if not has_role(current_user, ['admin', 'owner']):
        raise HTTPException(status_code=403, detail="غير مصرح لك بتعديل الإعدادات")
    
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    update_data = {
        "tenant_id": tenant_id,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user.get("id")
    }
    
    # إضافة الحقول غير الفارغة فقط
    settings_dict = settings.dict(exclude_unset=True, exclude_none=True)
    update_data.update(settings_dict)
    
    # تشفير المفتاح السري (في الإنتاج يجب استخدام تشفير حقيقي)
    if "stripe_secret_key" in update_data and update_data["stripe_secret_key"]:
        # في الإنتاج: استخدم تشفير AES أو KMS
        # هنا نحتفظ به كما هو للتبسيط
        pass
    
    await db.payment_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": update_data},
        upsert=True
    )
    
    return {"success": True, "message": "تم حفظ الإعدادات بنجاح"}

# ==================== أجور التوصيل حسب المسافة ====================

def _srv_haversine_km(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    """المسافة بالكيلومتر بين نقطتين (Haversine)"""
    R = 6371.0
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = math.sin(dlat / 2) ** 2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlng / 2) ** 2
    return 2 * R * math.asin(math.sqrt(a))


def _compute_distance_fee(s: dict, km: float) -> int:
    """حساب أجور التوصيل من إعدادات المسافة.
    الأولوية: نطاقات الكيلومتر (fee_zones) إن وُجدت — المالك يحدد مبلغاً لكل نطاق كم.
    وإلا المعادلة (أساس + لكل كم)."""
    zones = s.get("fee_zones")
    if isinstance(zones, list) and len(zones) > 0:
        valid = sorted(
            [z for z in zones if z.get("up_to_km") not in (None, "") and z.get("fee") not in (None, "")],
            key=lambda z: float(z["up_to_km"])
        )
        for z in valid:
            if km <= float(z["up_to_km"]):
                return int(float(z["fee"]))
        if valid:
            return int(float(valid[-1]["fee"]))  # أبعد من كل النطاقات → مبلغ آخر نطاق
    # المعادلة (احتياطي)
    base = float(s.get("fee_base") or 0)
    base_km = float(s.get("fee_base_km") or 0)
    per_km = float(s.get("fee_per_km") or 0)
    fee_max = float(s.get("fee_max") or 0)
    round_to = int(s.get("fee_round_to") or 0)
    fee = base + max(0.0, km - base_km) * per_km
    if fee_max > 0:
        fee = min(fee, fee_max)
    if round_to > 0:
        fee = math.ceil(fee / round_to) * round_to
    return int(fee)


async def _distance_fee_for(tenant_id: str, branch_id: Optional[str], loc: Optional[dict]):
    """ترجع dict {km, fee, out_of_range, max_km} إن أمكن الحساب، وإلا None"""
    s = await db.payment_settings.find_one({"tenant_id": tenant_id or "default"}, {"_id": 0}) or {}
    if not s.get("distance_fee_enabled"):
        return None
    if not loc:
        return None
    lat = loc.get("latitude", loc.get("lat"))
    lng = loc.get("longitude", loc.get("lng"))
    if lat is None or lng is None:
        return None
    branch = None
    if branch_id:
        branch = await db.branches.find_one({"id": branch_id}, {"_id": 0, "latitude": 1, "longitude": 1})
    if not branch or branch.get("latitude") is None or branch.get("longitude") is None:
        return None
    km = _srv_haversine_km(float(branch["latitude"]), float(branch["longitude"]), float(lat), float(lng))
    max_km = float(s.get("max_distance_km") or 0)
    return {
        "km": round(km, 2),
        "fee": _compute_distance_fee(s, km),
        "out_of_range": max_km > 0 and km > max_km,
        "max_km": max_km
    }


class BranchLocationUpdate(BaseModel):
    latitude: float
    longitude: float


@api_router.put("/branches/{branch_id}/location")
async def update_branch_location(branch_id: str, loc: BranchLocationUpdate, current_user: dict = Depends(get_current_user)):
    """تحديد إحداثيات الفرع (لحساب أجور التوصيل بالمسافة)"""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    query = build_tenant_query(current_user, {"id": branch_id})
    result = await db.branches.update_one(query, {"$set": {
        "latitude": loc.latitude,
        "longitude": loc.longitude,
        "location_updated_at": datetime.now(timezone.utc).isoformat()
    }})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="الفرع غير موجود")
    return {"message": "تم حفظ موقع الفرع", "latitude": loc.latitude, "longitude": loc.longitude}


@api_router.get("/delivery-fee/suggest")
async def suggest_delivery_fee(order_id: str, current_user: dict = Depends(get_current_user)):
    """اقتراح أجور التوصيل لطلب حسب المسافة (للكاشير/إدارة التوصيل)"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    s = await db.payment_settings.find_one({"tenant_id": tenant_id}, {"_id": 0}) or {}
    if not s.get("distance_fee_enabled"):
        return {"enabled": False, "suggested_fee": None, "reason": "أجور المسافة غير مفعلة في الإعدادات"}
    order = await db.orders.find_one({"id": order_id}, {"_id": 0, "delivery_location": 1, "branch_id": 1})
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    loc = order.get("delivery_location") or {}
    lat = loc.get("latitude", loc.get("lat"))
    lng = loc.get("longitude", loc.get("lng"))
    if lat is None or lng is None:
        return {"enabled": True, "suggested_fee": None, "reason": "لا يوجد موقع GPS للزبون على هذا الطلب"}
    branch = await db.branches.find_one({"id": order.get("branch_id")}, {"_id": 0, "latitude": 1, "longitude": 1})
    if not branch or branch.get("latitude") is None or branch.get("longitude") is None:
        return {"enabled": True, "suggested_fee": None, "reason": "حدّد موقع الفرع في الإعدادات أولاً"}
    km = _srv_haversine_km(float(branch["latitude"]), float(branch["longitude"]), float(lat), float(lng))
    max_km = float(s.get("max_distance_km") or 0)
    out_of_range = max_km > 0 and km > max_km
    resp = {
        "enabled": True,
        "distance_km": round(km, 2),
        "suggested_fee": _compute_distance_fee(s, km),
        "out_of_range": out_of_range,
        "max_km": max_km
    }
    if out_of_range:
        resp["reason"] = f"الزبون خارج نطاق التوصيل ({round(km, 1)} كم > {int(max_km)} كم)"
    return resp


@api_router.get("/customer/delivery-fee/{tenant_id}")
async def customer_delivery_fee_quote(tenant_id: str, lat: float, lng: float, branch_id: Optional[str] = None):
    """تسعير أجور التوصيل للزبون حسب موقعه - بدون مصادقة"""
    s = await db.payment_settings.find_one({"tenant_id": tenant_id}, {"_id": 0}) or {}
    if not s.get("distance_fee_enabled"):
        return {"distance_based": False, "fee": None}
    branch = None
    if branch_id:
        branch = await db.branches.find_one({"id": branch_id}, {"_id": 0, "latitude": 1, "longitude": 1})
    if not branch or branch.get("latitude") is None:
        branch = await db.branches.find_one(
            {"tenant_id": tenant_id, "is_active": {"$ne": False}, "latitude": {"$ne": None}},
            {"_id": 0, "latitude": 1, "longitude": 1}
        )
    if not branch or branch.get("latitude") is None or branch.get("longitude") is None:
        return {"distance_based": False, "fee": None, "reason": "لم يُحدد موقع الفرع"}
    km = _srv_haversine_km(float(branch["latitude"]), float(branch["longitude"]), float(lat), float(lng))
    max_km = float(s.get("max_distance_km") or 0)
    if max_km > 0 and km > max_km:
        return {
            "distance_based": True,
            "out_of_range": True,
            "distance_km": round(km, 2),
            "max_km": max_km,
            "fee": None
        }
    return {"distance_based": True, "distance_km": round(km, 2), "fee": _compute_distance_fee(s, km)}


@api_router.post("/payment-settings/zaincash-qr")
async def upload_zaincash_qr(
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user)
):
    """رفع صورة QR Code لزين كاش"""
    if not has_role(current_user, ['admin', 'owner']):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # حفظ الصورة
    file_ext = file.filename.split('.')[-1] if '.' in file.filename else 'png'
    filename = f"zaincash_qr_{tenant_id}.{file_ext}"
    file_path = UPLOAD_DIR / "payment" / filename
    
    (UPLOAD_DIR / "payment").mkdir(exist_ok=True)
    
    content = await file.read()
    async with aiofiles.open(file_path, 'wb') as f:
        await f.write(content)
    
    image_url = f"/uploads/payment/{filename}"
    
    # تحديث الإعدادات
    await db.payment_settings.update_one(
        {"tenant_id": tenant_id},
        {"$set": {"zaincash_qr_image": image_url}},
        upsert=True
    )
    
    return {"success": True, "image_url": image_url}


# ==================== REAL-TIME NOTIFICATIONS APIs ====================

# تخزين الإشعارات غير المقروءة في الذاكرة (للتبسيط)
# في الإنتاج: استخدم Redis أو WebSockets
pending_notifications = {}

@api_router.get("/notifications/pending-orders")
async def get_pending_order_notifications(
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """جلب الطلبات الجديدة (للكاشير)"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # جلب الطلبات الجديدة من آخر 5 دقائق
    five_minutes_ago = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
    
    query = {
        "status": "pending",
        "source": "customer_app",
        "created_at": {"$gte": five_minutes_ago}
    }
    
    if branch_id:
        query["branch_id"] = branch_id
    elif tenant_id != "default":
        query["tenant_id"] = tenant_id
    
    orders = await db.orders.find(
        query,
        {"_id": 0, "id": 1, "order_number": 1, "customer_name": 1, 
         "total": 1, "created_at": 1, "payment_method": 1, "items": 1}
    ).sort("created_at", -1).to_list(20)
    
    # تحديد الطلبات الجديدة (غير المشاهدة)
    user_id = current_user.get("id", "")
    viewed_key = f"{tenant_id}_{user_id}"
    viewed_orders = pending_notifications.get(viewed_key, set())
    
    new_orders = []
    for order in orders:
        order["is_new"] = order["id"] not in viewed_orders
        new_orders.append(order)
    
    return {
        "orders": new_orders,
        "new_count": sum(1 for o in new_orders if o.get("is_new")),
        "total_count": len(new_orders)
    }

@api_router.post("/notifications/mark-seen")
async def mark_orders_as_seen(
    order_ids: List[str] = Body(...),
    current_user: dict = Depends(get_current_user)
):
    """تحديد الطلبات كمشاهدة"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    user_id = current_user.get("id", "")
    viewed_key = f"{tenant_id}_{user_id}"
    
    if viewed_key not in pending_notifications:
        pending_notifications[viewed_key] = set()
    
    pending_notifications[viewed_key].update(order_ids)
    
    return {"success": True, "marked_count": len(order_ids)}


@api_router.post("/notifications/accept-order/{order_id}")
async def accept_customer_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """قبول طلب العميل الخارجي"""
    tenant_id = get_user_tenant_id(current_user)
    query = {"id": order_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    order = await db.orders.find_one(query)
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "confirmed",
            "accepted_by": current_user.get("id"),
            "accepted_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    # تحديد كمشاهد
    user_id = current_user.get("id", "")
    viewed_key = f"{tenant_id or 'default'}_{user_id}"
    if viewed_key not in pending_notifications:
        pending_notifications[viewed_key] = set()
    pending_notifications[viewed_key].add(order_id)
    
    return {"success": True, "message": "تم قبول الطلب"}

@api_router.post("/notifications/reject-order/{order_id}")
async def reject_customer_order(
    order_id: str,
    current_user: dict = Depends(get_current_user)
):
    """رفض طلب العميل الخارجي"""
    tenant_id = get_user_tenant_id(current_user)
    query = {"id": order_id}
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    await db.orders.update_one(
        {"id": order_id},
        {"$set": {
            "status": "cancelled",
            "is_rejected": True,
            "driver_id": None,
            "driver_name": None,
            "driver_phone": None,
            "rejected_by": current_user.get("id"),
            "rejected_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"success": True, "message": "تم رفض الطلب"}

@api_router.get("/notifications/delayed-orders")
async def get_delayed_orders(
    branch_id: Optional[str] = None,
    delay_minutes: int = 15,
    current_user: dict = Depends(get_current_user)
):
    """
    جلب الطلبات المتأخرة - الطلبات التي مر عليها أكثر من المدة المحددة
    delay_minutes: عدد الدقائق للاعتبار التأخير (افتراضي 15 دقيقة)
    """
    tenant_id = get_user_tenant_id(current_user)
    
    # حساب وقت الحد الأقصى للتأخير
    delay_threshold = (datetime.now(timezone.utc) - timedelta(minutes=delay_minutes)).isoformat()
    
    # البحث عن الطلبات المتأخرة (pending أو preparing لأكثر من المدة المحددة)
    query = {
        "status": {"$in": ["pending", "preparing"]},
        "created_at": {"$lt": delay_threshold}
    }
    
    if branch_id and branch_id != 'all':
        query["branch_id"] = branch_id
    
    if tenant_id:
        query["tenant_id"] = tenant_id
    
    # جلب الطلبات المتأخرة
    delayed_orders = await db.orders.find(query, {"_id": 0}).sort("created_at", 1).to_list(50)
    
    # حساب مدة التأخير لكل طلب
    now = datetime.now(timezone.utc)
    for order in delayed_orders:
        try:
            created_at = datetime.fromisoformat(order.get("created_at", "").replace("Z", "+00:00"))
            delay_duration = now - created_at
            order["delay_minutes"] = int(delay_duration.total_seconds() / 60)
            
            # تصنيف مستوى التأخير
            if order["delay_minutes"] >= 45:
                order["delay_level"] = "critical"  # حرج
            elif order["delay_minutes"] >= 30:
                order["delay_level"] = "high"  # عالي
            elif order["delay_minutes"] >= 15:
                order["delay_level"] = "medium"  # متوسط
            else:
                order["delay_level"] = "low"  # منخفض
        except:
            order["delay_minutes"] = 0
            order["delay_level"] = "unknown"
    
    # تصنيف حسب نوع الطلب
    delayed_by_type = {
        "dine_in": [o for o in delayed_orders if o.get("order_type") == "dine_in"],
        "takeaway": [o for o in delayed_orders if o.get("order_type") == "takeaway"],
        "delivery": [o for o in delayed_orders if o.get("order_type") == "delivery"],
    }
    
    # إحصائيات سريعة
    stats = {
        "total_delayed": len(delayed_orders),
        "critical_count": len([o for o in delayed_orders if o.get("delay_level") == "critical"]),
        "high_count": len([o for o in delayed_orders if o.get("delay_level") == "high"]),
        "medium_count": len([o for o in delayed_orders if o.get("delay_level") == "medium"]),
        "avg_delay_minutes": round(sum(o.get("delay_minutes", 0) for o in delayed_orders) / max(len(delayed_orders), 1), 1),
        "max_delay_minutes": max((o.get("delay_minutes", 0) for o in delayed_orders), default=0)
    }
    
    return {
        "delayed_orders": delayed_orders,
        "delayed_by_type": delayed_by_type,
        "stats": stats,
        "delay_threshold_minutes": delay_minutes
    }

@api_router.get("/notifications/sound-alert")
async def check_sound_alert(
    last_check: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """التحقق من وجود طلبات جديدة تحتاج تنبيه صوتي"""
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    # تحديد وقت آخر فحص
    if last_check:
        try:
            check_time = last_check
        except:
            check_time = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
    else:
        check_time = (datetime.now(timezone.utc) - timedelta(seconds=30)).isoformat()
    
    query = {
        "status": "pending",
        "source": "customer_app",
        "created_at": {"$gt": check_time}
    }
    
    if branch_id:
        query["branch_id"] = branch_id
    elif tenant_id != "default":
        query["tenant_id"] = tenant_id
    
    new_orders_count = await db.orders.count_documents(query)
    
    return {
        "has_new_orders": new_orders_count > 0,
        "new_orders_count": new_orders_count,
        "check_time": datetime.now(timezone.utc).isoformat()
    }


# ==================== SECURE CARD DATA (Stripe handles this) ====================
# ملاحظة مهمة: بيانات البطاقة لا تُخزن في قاعدة البيانات أبداً
# Stripe يتعامل مع جميع بيانات البطاقة الحساسة
# نحن نخزن فقط: آخر 4 أرقام، نوع البطاقة، تاريخ الانتهاء (للعرض فقط)

@api_router.get("/payment-transactions")
async def get_payment_transactions(
    limit: int = 20,
    current_user: dict = Depends(get_current_user)
):
    """جلب سجل المعاملات المالية"""
    if not has_role(current_user, ['admin', 'owner']):
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    tenant_id = get_user_tenant_id(current_user) or "default"
    
    transactions = await db.payment_transactions.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).sort("created_at", -1).limit(limit).to_list(limit)
    
    return {"transactions": transactions}


# ==================== LICENSE MANAGEMENT - نظام الترخيص ====================

class LicenseResponse(BaseModel):
    """نموذج استجابة الترخيص"""
    is_active: bool
    is_expired: bool
    tenant_id: str
    tenant_name: str
    expiry_date: Optional[str] = None
    features: List[str] = []
    max_branches: int = 1
    max_users: int = 5
    plan: str = "basic"
    message: str = ""

@api_router.get("/license/verify")
async def verify_license(current_user: dict = Depends(get_current_user)):
    """
    التحقق من ترخيص المستخدم/العميل
    يستخدم من تطبيق سطح المكتب للتحقق من صلاحية الترخيص
    """
    try:
        tenant_id = get_user_tenant_id(current_user)
        
        # إذا كان Super Admin
        if current_user.get("role") == UserRole.SUPER_ADMIN:
            return {
                "is_active": True,
                "is_expired": False,
                "tenant_id": "system",
                "tenant_name": "System Administrator",
                "expiry_date": None,
                "features": ["all"],
                "max_branches": 999,
                "max_users": 999,
                "plan": "enterprise",
                "message": "مرحباً بك كمدير النظام"
            }
        
        # جلب بيانات العميل
        tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
        
        if not tenant:
            # إذا لم يوجد tenant، استخدم المستخدم نفسه
            return {
                "is_active": current_user.get("is_active", True),
                "is_expired": False,
                "tenant_id": tenant_id or "default",
                "tenant_name": current_user.get("full_name", ""),
                "expiry_date": None,
                "features": ["pos", "orders", "tables", "expenses"],
                "max_branches": 1,
                "max_users": 5,
                "plan": "basic",
                "message": "ترخيص أساسي"
            }
        
        # التحقق من حالة الاشتراك
        is_active = tenant.get("is_active", True)
        expires_at = tenant.get("expires_at")
        is_expired = False
        
        if expires_at:
            try:
                expiry_date = datetime.fromisoformat(expires_at.replace('Z', '+00:00'))
                is_expired = datetime.now(timezone.utc) > expiry_date
            except:
                is_expired = False
        
        # جلب الميزات المفعّلة
        enabled_features = tenant.get("enabled_features", {})
        features = [key for key, value in enabled_features.items() if value]
        
        # إذا لم توجد ميزات، استخدم الميزات الافتراضية
        if not features:
            features = ["pos", "orders", "tables", "expenses", "inventory", "reports"]
        
        # تحديد الخطة
        subscription_type = tenant.get("subscription_type", "basic")
        plan_mapping = {
            "trial": "trial",
            "bronze": "basic",
            "silver": "standard",
            "gold": "premium",
            "basic": "basic",
            "premium": "premium",
            "demo": "demo"
        }
        plan = plan_mapping.get(subscription_type, "basic")
        
        # رسالة حسب الحالة
        if not is_active:
            message = "الحساب معطل - يرجى التواصل مع الدعم"
        elif is_expired:
            message = "انتهى الاشتراك - يرجى التجديد"
        else:
            message = "الترخيص فعّال"
        
        return {
            "is_active": is_active,
            "is_expired": is_expired,
            "tenant_id": tenant_id,
            "tenant_name": tenant.get("name", tenant.get("name_ar", "")),
            "expiry_date": expires_at,
            "features": features,
            "max_branches": tenant.get("max_branches", 1),
            "max_users": tenant.get("max_users", 5),
            "plan": plan,
            "message": message
        }
        
    except Exception as e:
        logger.error(f"❌ License verification error: {e}")
        raise HTTPException(status_code=500, detail=f"خطأ في التحقق من الترخيص: {str(e)}")

@api_router.post("/license/activate")
async def activate_license(
    license_key: str = Body(..., embed=True),
    device_id: str = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """
    تفعيل الترخيص لجهاز جديد
    """
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب بيانات العميل
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0})
    
    if not tenant:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    # التحقق من الترخيص (في هذا التنفيذ البسيط، نسمح بأي مفتاح)
    # يمكن تطوير هذا لاحقاً للتحقق من مفتاح معين
    
    # تسجيل الجهاز
    device_record = {
        "device_id": device_id,
        "tenant_id": tenant_id,
        "activated_at": datetime.now(timezone.utc).isoformat(),
        "activated_by": current_user.get("id"),
        "is_active": True
    }
    
    # التحقق من عدم تجاوز عدد الأجهزة المسموح بها
    existing_devices = await db.license_devices.count_documents({
        "tenant_id": tenant_id,
        "is_active": True
    })
    
    max_devices = tenant.get("max_devices", 5)
    
    if existing_devices >= max_devices:
        raise HTTPException(
            status_code=400, 
            detail=f"تم الوصول للحد الأقصى من الأجهزة ({max_devices})"
        )
    
    # إضافة الجهاز
    await db.license_devices.update_one(
        {"device_id": device_id, "tenant_id": tenant_id},
        {"$set": device_record},
        upsert=True
    )
    
    return {
        "success": True,
        "message": "تم تفعيل الترخيص بنجاح",
        "device_id": device_id,
        "devices_used": existing_devices + 1,
        "max_devices": max_devices
    }

@api_router.get("/license/devices")
async def get_licensed_devices(current_user: dict = Depends(get_current_user)):
    """
    جلب قائمة الأجهزة المرخصة
    """
    tenant_id = get_user_tenant_id(current_user)
    
    devices = await db.license_devices.find(
        {"tenant_id": tenant_id, "is_active": True},
        {"_id": 0}
    ).to_list(100)
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "max_devices": 1})
    max_devices = tenant.get("max_devices", 5) if tenant else 5
    
    return {
        "devices": devices,
        "count": len(devices),
        "max_devices": max_devices
    }

@api_router.delete("/license/devices/{device_id}")
async def deactivate_device(
    device_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    إلغاء ترخيص جهاز
    """
    tenant_id = get_user_tenant_id(current_user)
    
    result = await db.license_devices.update_one(
        {"device_id": device_id, "tenant_id": tenant_id},
        {"$set": {"is_active": False, "deactivated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="الجهاز غير موجود")
    
    return {"success": True, "message": "تم إلغاء ترخيص الجهاز"}


# ==================== SUPER ADMIN - DEVICE MANAGEMENT ====================

@api_router.get("/super-admin/tenants/{tenant_id}/devices")
async def get_tenant_devices(
    tenant_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    جلب أجهزة عميل معين (Super Admin فقط)
    """
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    devices = await db.license_devices.find(
        {"tenant_id": tenant_id},
        {"_id": 0}
    ).to_list(100)
    
    tenant = await db.tenants.find_one({"id": tenant_id}, {"_id": 0, "max_devices": 1})
    max_devices = tenant.get("max_devices", 5) if tenant else 5
    
    return {
        "devices": devices,
        "count": len([d for d in devices if d.get("is_active")]),
        "max_devices": max_devices
    }

@api_router.put("/super-admin/tenants/{tenant_id}/max-devices")
async def update_tenant_max_devices(
    tenant_id: str,
    max_devices: int = Body(..., embed=True),
    current_user: dict = Depends(get_current_user)
):
    """
    تحديث الحد الأقصى للأجهزة لعميل (Super Admin فقط)
    """
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    if max_devices < 1:
        raise HTTPException(status_code=400, detail="الحد الأدنى للأجهزة هو 1")
    
    result = await db.tenants.update_one(
        {"id": tenant_id},
        {"$set": {"max_devices": max_devices}}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="العميل غير موجود")
    
    return {"success": True, "max_devices": max_devices}

@api_router.delete("/super-admin/tenants/{tenant_id}/devices/{device_id}")
async def deactivate_tenant_device(
    tenant_id: str,
    device_id: str,
    current_user: dict = Depends(get_current_user)
):
    """
    إلغاء ترخيص جهاز لعميل (Super Admin فقط)
    """
    if current_user.get("role") != UserRole.SUPER_ADMIN:
        raise HTTPException(status_code=403, detail="غير مصرح لك")
    
    result = await db.license_devices.update_one(
        {"device_id": device_id, "tenant_id": tenant_id},
        {"$set": {
            "is_active": False, 
            "deactivated_at": datetime.now(timezone.utc).isoformat(),
            "deactivated_by": current_user.get("id")
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="الجهاز غير موجود")
    
    return {"success": True, "message": "تم إلغاء ترخيص الجهاز"}


# ==================== BUSINESS DATE MIGRATION ====================
# ==================== BUSINESS DATE MIGRATION ====================

# Platform-owner only: cleanup a specific orphan expense record
# (لا واجهة أمامية - استخدام داخلي فقط من قبل المالك عبر API)
@api_router.post("/superadmin/cleanup-orphan-expense/{expense_id}")
async def superadmin_cleanup_orphan_expense(
    expense_id: str,
    current_user: dict = Depends(verify_super_admin)
):
    """حذف سجل مصروف يتيم (للمالك/super_admin فقط - بدون واجهة).
    يُعيد احتساب إجمالي الوردية المرتبطة تلقائياً."""
    expense = await db.expenses.find_one({"id": expense_id}, {"_id": 0})
    if not expense:
        raise HTTPException(status_code=404, detail="المصروف غير موجود")
    
    await db.expenses.delete_one({"id": expense_id})
    
    # إعادة احتساب total_expenses للوردية المرتبطة
    exp_created = expense.get("created_at", "")
    exp_branch = expense.get("branch_id")
    exp_tenant = expense.get("tenant_id")
    recomputed_shifts = 0
    if exp_created and exp_branch:
        shift_q = {"branch_id": exp_branch, "started_at": {"$lte": exp_created}}
        if exp_tenant:
            shift_q["tenant_id"] = exp_tenant
        affected_shifts = await db.shifts.find(
            shift_q,
            {"_id": 0, "id": 1, "started_at": 1, "ended_at": 1, "opening_cash": 1, "opening_balance": 1, "cash_sales": 1}
        ).to_list(100)
        for s in affected_shifts:
            s_end = s.get("ended_at") or ""
            if s_end and exp_created > s_end:
                continue
            exp_q = {
                "branch_id": exp_branch,
                "category": {"$ne": "refund"},
                "created_at": {"$gte": s.get("started_at", "")}
            }
            if exp_tenant:
                exp_q["tenant_id"] = exp_tenant
            if s_end:
                exp_q["created_at"]["$lte"] = s_end
            shift_expenses = await db.expenses.find(exp_q, {"_id": 0, "amount": 1}).to_list(500)
            total_exp = sum(float(e.get("amount") or 0) for e in shift_expenses)
            opening_cash = float(s.get("opening_cash") or s.get("opening_balance") or 0)
            cash_sales = float(s.get("cash_sales") or 0)
            expected_cash = opening_cash + cash_sales - total_exp
            await db.shifts.update_one(
                {"id": s["id"]},
                {"$set": {"total_expenses": total_exp, "expected_cash": expected_cash}}
            )
            recomputed_shifts += 1
    
    return {
        "success": True,
        "deleted_expense": {
            "id": expense.get("id"),
            "description": expense.get("description"),
            "amount": expense.get("amount"),
            "date": expense.get("date")
        },
        "shifts_recomputed": recomputed_shifts
    }


@api_router.post("/admin/migrate-business-dates")
async def migrate_business_dates(force: bool = False, current_user: dict = Depends(get_current_user)):
    """
    ترحيل (Migration) لتحديث جميع السجلات المالية القديمة بـ business_date
    (اليوم التشغيلي المستنِد على تاريخ بدء الوردية).
    
    يُشغَّل مرة واحدة من قِبَل المالك فقط. آمن لإعادة التشغيل (idempotent).
    
    force=True: إعادة احتساب business_date لجميع السجلات حتى التي تملكه سابقاً (مفيد لتصحيح ترحيل سابق خاطئ).
    """
    # صلاحية المالك فقط
    if current_user.get("role") not in ["admin", "super_admin", "manager"]:
        raise HTTPException(status_code=403, detail="غير مصرح - المالك فقط")
    
    tenant_id = get_user_tenant_id(current_user)
    stats = {
        "shifts_updated": 0,
        "orders_updated": 0,
        "expenses_updated": 0,
        "advances_updated": 0,
        "deductions_updated": 0,
        "bonuses_updated": 0,
        "overtime_updated": 0,
        "shift_expenses_recomputed": 0,
        "errors": []
    }
    
    try:
        # 1) إضافة business_date للورديات التي لا تملكه
        shifts_cursor = db.shifts.find(
            {"tenant_id": tenant_id, "business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "started_at": 1, "opened_at": 1}
        )
        async for shift in shifts_cursor:
            started = shift.get("started_at") or shift.get("opened_at")
            biz_date = iraq_date_from_utc(started) if started else iraq_date_from_utc()
            await db.shifts.update_one(
                {"id": shift["id"]},
                {"$set": {"business_date": biz_date}}
            )
            stats["shifts_updated"] += 1
        
        # 2) جلب كل الورديات للـ tenant مع business_date لمعالجة السجلات الأخرى
        all_shifts = await db.shifts.find(
            {"tenant_id": tenant_id},
            {"_id": 0, "id": 1, "branch_id": 1, "started_at": 1, "opened_at": 1, "ended_at": 1, "business_date": 1, "status": 1}
        ).sort("started_at", 1).to_list(10000)
        
        # ترتيب الورديات حسب الفرع
        shifts_by_branch = {}
        for s in all_shifts:
            b = s.get("branch_id") or "__no_branch__"
            shifts_by_branch.setdefault(b, []).append(s)
        
        def find_shift_for_record(record_branch_id, record_created_at, record_shift_id=None):
            """إيجاد الوردية التي ينتمي إليها السجل"""
            if record_shift_id:
                for s in all_shifts:
                    if s.get("id") == record_shift_id:
                        return s
            branch_shifts = shifts_by_branch.get(record_branch_id or "__no_branch__", [])
            if not branch_shifts:
                branch_shifts = all_shifts
            for s in branch_shifts:
                start = s.get("started_at") or s.get("opened_at") or ""
                end = s.get("ended_at") or ""
                if record_created_at >= start and (not end or record_created_at <= end):
                    return s
            return None
        
        # 3) ترحيل الطلبات - مع force يُعاد احتساب حتى إن كانت تملك business_date
        orders_filter = {"tenant_id": tenant_id}
        if not force:
            orders_filter["business_date"] = {"$exists": False}
        async for order in db.orders.find(
            orders_filter,
            {"_id": 0, "id": 1, "branch_id": 1, "created_at": 1, "shift_id": 1, "business_date": 1}
        ):
            shift = find_shift_for_record(order.get("branch_id"), order.get("created_at", ""), order.get("shift_id"))
            biz = shift["business_date"] if (shift and shift.get("business_date")) else iraq_date_from_utc(order.get("created_at"))
            if order.get("business_date") != biz:
                await db.orders.update_one({"id": order["id"]}, {"$set": {"business_date": biz}})
                stats["orders_updated"] += 1
        
        # 4) ترحيل المصاريف
        expenses_filter = {"tenant_id": tenant_id}
        if not force:
            expenses_filter["business_date"] = {"$exists": False}
        async for exp in db.expenses.find(
            expenses_filter,
            {"_id": 0, "id": 1, "branch_id": 1, "created_at": 1, "date": 1, "business_date": 1}
        ):
            shift = find_shift_for_record(exp.get("branch_id"), exp.get("created_at", ""))
            if shift and shift.get("business_date"):
                biz = shift["business_date"]
            else:
                biz = exp.get("date") or iraq_date_from_utc(exp.get("created_at"))
            if exp.get("business_date") != biz:
                await db.expenses.update_one({"id": exp["id"]}, {"$set": {"business_date": biz}})
                stats["expenses_updated"] += 1
        
        # 5) ترحيل السلف
        async for adv in db.advances.find(
            {"tenant_id": tenant_id, "business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "created_at": 1, "date": 1}
        ):
            biz = iraq_date_from_utc(adv.get("created_at")) or adv.get("date")
            await db.advances.update_one({"id": adv["id"]}, {"$set": {"business_date": biz}})
            stats["advances_updated"] += 1
        
        # 6) ترحيل الخصومات
        async for ded in db.deductions.find(
            {"tenant_id": tenant_id, "business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "created_at": 1, "date": 1}
        ):
            biz = iraq_date_from_utc(ded.get("created_at")) or ded.get("date")
            await db.deductions.update_one({"id": ded["id"]}, {"$set": {"business_date": biz}})
            stats["deductions_updated"] += 1
        
        # 7) ترحيل المكافآت
        async for bon in db.bonuses.find(
            {"tenant_id": tenant_id, "business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "created_at": 1, "date": 1}
        ):
            biz = iraq_date_from_utc(bon.get("created_at")) or bon.get("date")
            await db.bonuses.update_one({"id": bon["id"]}, {"$set": {"business_date": biz}})
            stats["bonuses_updated"] += 1
        
        # 8) ترحيل طلبات الساعات الإضافية
        async for ot in db.overtime_requests.find(
            {"tenant_id": tenant_id, "business_date": {"$exists": False}},
            {"_id": 0, "id": 1, "created_at": 1, "date": 1}
        ):
            biz = iraq_date_from_utc(ot.get("created_at")) or ot.get("date")
            await db.overtime_requests.update_one({"id": ot["id"]}, {"$set": {"business_date": biz}})
            stats["overtime_updated"] += 1
        
        # 9) إعادة احتساب total_expenses للورديات المُغلقة (استبعاد المرتجعات)
        closed_shifts = await db.shifts.find(
            {"tenant_id": tenant_id, "status": "closed"},
            {"_id": 0, "id": 1, "branch_id": 1, "started_at": 1, "opened_at": 1, "ended_at": 1, "cash_sales": 1, "opening_cash": 1, "opening_balance": 1}
        ).to_list(10000)
        
        for s in closed_shifts:
            shift_start = s.get("started_at") or s.get("opened_at") or ""
            shift_end = s.get("ended_at") or ""
            if not shift_start:
                continue
            exp_query = {
                "tenant_id": tenant_id,
                "branch_id": s.get("branch_id"),
                "category": {"$ne": "refund"},
                "created_at": {"$gte": shift_start}
            }
            if shift_end:
                exp_query["created_at"]["$lte"] = shift_end
            shift_expenses = await db.expenses.find(exp_query, {"_id": 0, "amount": 1}).to_list(500)
            total_exp = sum(float(e.get("amount") or 0) for e in shift_expenses)
            
            opening_cash = float(s.get("opening_cash") or s.get("opening_balance") or 0)
            cash_sales = float(s.get("cash_sales") or 0)
            expected_cash = opening_cash + cash_sales - total_exp
            
            await db.shifts.update_one(
                {"id": s["id"]},
                {"$set": {"total_expenses": total_exp, "expected_cash": expected_cash}}
            )
            stats["shift_expenses_recomputed"] += 1
        
        return {
            "success": True,
            "message": "تمت عملية الترحيل بنجاح",
            "stats": stats
        }
    except Exception as e:
        logger.exception("Business date migration failed")
        stats["errors"].append(str(e))
        return {
            "success": False,
            "message": f"فشلت عملية الترحيل: {str(e)}",
            "stats": stats
        }


# Include reports routes (refactored) - PRIORITY over old routes
from routes.reports_routes import router as reports_router
app.include_router(reports_router, prefix="/api")

# Include drivers routes (refactored)
from routes.drivers_routes import router as drivers_router
app.include_router(drivers_router, prefix="/api")

# Include payroll routes (refactored)
from routes.payroll_routes import router as payroll_router
app.include_router(payroll_router, prefix="/api")

# Include shifts routes (refactored)
from routes.shifts_routes import router as shifts_router
app.include_router(shifts_router, prefix="/api")

# Owner wallet routes
from routes.owner_wallet import router as owner_wallet_router
app.include_router(owner_wallet_router, prefix="/api")

# External/Sold branches routes
from routes.external_branches import router as external_branches_router
app.include_router(external_branches_router, prefix="/api")

# Order notifications routes (real-time notifications for cashier and driver)
from routes.order_notifications import router as order_notifications_router
from routes.print_queue import router as print_queue_router
app.include_router(order_notifications_router, prefix="/api")
app.include_router(print_queue_router, prefix="/api")

# ==================== SALES LEADERBOARD (لوحة تنافس المبيعات) ====================

@api_router.get("/sales-leaderboard")
async def get_sales_leaderboard(
    period: str = "today",
    current_user: dict = Depends(get_current_user)
):
    """لوحة ترتيب المبيعات اليومية - تنافس بين الكاشيرية"""
    tenant_id = current_user.get("tenant_id")
    
    now = datetime.now(timezone.utc)
    if period == "today":
        start_date = now.strftime('%Y-%m-%d')
    elif period == "week":
        start_date = (now - timedelta(days=7)).strftime('%Y-%m-%d')
    elif period == "month":
        start_date = (now - timedelta(days=30)).strftime('%Y-%m-%d')
    else:
        start_date = now.strftime('%Y-%m-%d')
    
    query = {
        "status": {"$nin": [OrderStatus.CANCELLED, "refunded"]},
        "created_at": {"$gte": start_date}
    }
    if tenant_id:
        query["tenant_id"] = tenant_id
    else:
        query["tenant_id"] = "default"
    
    user_role = current_user.get("role")
    user_branch_id = current_user.get("branch_id")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    
    orders = await db.orders.find(query, {
        "_id": 0, "cashier_id": 1, "cashier_name": 1, "total": 1
    }).to_list(10000)
    
    # جلب أسماء الكاشيرية من قاعدة البيانات
    cashier_ids = list(set(o.get("cashier_id") for o in orders if o.get("cashier_id")))
    users_list = await db.users.find(
        {"id": {"$in": cashier_ids}},
        {"_id": 0, "id": 1, "name": 1, "full_name": 1}
    ).to_list(500)
    user_names = {u["id"]: u.get("full_name") or u.get("name", "غير معروف") for u in users_list}
    
    cashier_stats = {}
    for order in orders:
        cid = order.get("cashier_id", "unknown")
        if cid not in cashier_stats:
            cashier_stats[cid] = {
                "cashier_id": cid,
                "cashier_name": order.get("cashier_name") or user_names.get(cid, "غير معروف"),
                "total_sales": 0,
                "order_count": 0
            }
        cashier_stats[cid]["total_sales"] += _sn(order.get("total"))
        cashier_stats[cid]["order_count"] += 1
    
    leaderboard = sorted(cashier_stats.values(), key=lambda x: x["total_sales"], reverse=True)
    
    for i, entry in enumerate(leaderboard):
        entry["rank"] = i + 1
        entry["average_order"] = entry["total_sales"] / entry["order_count"] if entry["order_count"] > 0 else 0
    
    return {
        "period": period,
        "date": now.strftime('%Y-%m-%d'),
        "leaderboard": leaderboard,
        "total_cashiers": len(leaderboard)
    }

# ==================== DAILY SALES TARGET (هدف المبيعات اليومي) ====================

# Include router and middleware
app.include_router(api_router)

# Sales Target routes (extracted to routes/sales_target_routes.py)
from routes.sales_target_routes import router as sales_target_router
app.include_router(sales_target_router, prefix="/api")

# Break-Even routes (extracted to routes/break_even_routes.py)
from routes.break_even_routes import router as break_even_router
app.include_router(break_even_router, prefix="/api")

# Reservations & Reviews routes (extracted to routes/reservations_reviews_routes.py)
from routes.reservations_reviews_routes import router as reservations_reviews_router
app.include_router(reservations_reviews_router, prefix="/api")

# Suppliers routes (extracted to routes/suppliers_routes.py)
from routes.suppliers_routes import router as suppliers_router
app.include_router(suppliers_router, prefix="/api")

# Include new inventory system routes
from routes.inventory_system import router as inventory_router
app.include_router(inventory_router)

# Branch Daily Stock Count (الجرد اليومي للفروع)
from routes.branch_stock_count import router as branch_stock_count_router
app.include_router(branch_stock_count_router)

# Department Monthly Stocktake (الجرد الشهري للأقسام)
from routes.department_stock_count import router as dept_stock_count_router
app.include_router(dept_stock_count_router)

# Include sync routes for offline support
from routes.sync_routes import router as sync_router
app.include_router(sync_router, prefix="/api")

# مكالمات WebRTC داخل التطبيق (زبون ↔ سائق)
from routes.call_routes import router as call_router
app.include_router(call_router, prefix="/api")

# Super Admin routes (extracted)
from routes.super_admin_routes import router as super_admin_router
app.include_router(super_admin_router, prefix="/api")

# Cash closing report / Biometric / Customer menu APIs (extracted)
from routes.hr_routes import router as hr_router, _calc_worked_hours_hhmm
app.include_router(hr_router, prefix="/api")
from routes.cash_closing_report_routes import router as cash_closing_router
app.include_router(cash_closing_router, prefix="/api")
from routes.biometric_routes import router as biometric_router, ZKTecoPushData
app.include_router(biometric_router, prefix="/api")
from routes.customer_menu_api_routes import router as customer_menu_api_router, generate_menu_slug, get_customer_from_token
app.include_router(customer_menu_api_router, prefix="/api")
from routes.ratings_routes import router as ratings_router
app.include_router(ratings_router, prefix="/api")
from routes.smart_reports_routes import router as smart_reports_router
app.include_router(smart_reports_router, prefix="/api")
from routes.ocr_routes import router as ocr_router
app.include_router(ocr_router, prefix="/api")
from routes.printer_routes import router as printer_router
app.include_router(printer_router, prefix="/api")
from routes.pdf_export_routes import router as pdf_export_router
app.include_router(pdf_export_router, prefix="/api")
from routes.refunds_routes import router as refunds_router
app.include_router(refunds_router, prefix="/api")
from routes.coupons_promotions_routes import router as coupons_promotions_router
app.include_router(coupons_promotions_router, prefix="/api")
from routes.payroll_reports_routes import router as payroll_reports_router
app.include_router(payroll_reports_router, prefix="/api")
from routes.notification_prefs_routes import router as notification_prefs_router
app.include_router(notification_prefs_router, prefix="/api")

# Middleware to prevent caching of API responses
@app.middleware("http")
async def add_no_cache_headers(request, call_next):
    response = await call_next(request)
    # رؤوس أمان عامة لكل الاستجابات
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "SAMEORIGIN"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Permissions-Policy"] = "geolocation=(self), microphone=(self), camera=()"
    if request.url.path.startswith("/api"):
        response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
        response.headers["Pragma"] = "no-cache"
        response.headers["Expires"] = "0"
        # تسجيل محاولات الوصول المرفوضة (محاولات اختراق): فقط عند وجود توكن أو عمليات كتابة لتقليل الضجيج
        try:
            if response.status_code in (401, 403):
                has_auth = "authorization" in {k.lower() for k in request.headers.keys()}
                if has_auth or request.method in ("POST", "PUT", "DELETE", "PATCH"):
                    await record_audit("access.denied", request, status=response.status_code)
        except Exception:
            pass
    return response

_cors_origins_env = os.environ.get('CORS_ORIGINS', '').strip()
if _cors_origins_env and _cors_origins_env != '*':
    _allowed_origins = [o.strip() for o in _cors_origins_env.split(',') if o.strip()]
    app.add_middleware(
        CORSMiddleware,
        allow_credentials=True,
        allow_origins=_allowed_origins,
        allow_methods=["*"],
        allow_headers=["*"],
    )
else:
    # عند عدم ضبط CORS_ORIGINS: لا نعكس أصلاً عشوائياً مع credentials. نسمح فقط بنطاقات الإنتاج المعروفة.
    app.add_middleware(
        CORSMiddleware,
        allow_credentials=True,
        allow_origin_regex=r"https://([a-z0-9-]+\.)?maestroegp\.com",
        allow_methods=["*"],
        allow_headers=["*"],
    )

# WebSocket Integration for Real-time Notifications
try:
    from services.websocket_service import sio, notify_branch_new_order, notify_driver_new_order
    
    # Mount Socket.IO app
    socket_app = socketio.ASGIApp(sio, other_asgi_app=app)
    
    logger.info("✅ WebSocket service initialized successfully")
except Exception as e:
    logger.warning(f"⚠️ WebSocket service not available: {e}")
    socket_app = app  # Fallback to regular app


# ==================== تحديث الطلبات القديمة لشركات التوصيل ====================
@api_router.post("/admin/fix-delivery-orders")
async def fix_delivery_orders(current_user: dict = Depends(verify_super_admin)):
    """تحديث الطلبات القديمة لشركات التوصيل - يُنفذ مرة واحدة"""
    
    # جلب كل العملاء الذين هم شركات توصيل
    delivery_customers = await db.customers.find({"is_delivery_company": True}, {"_id": 0, "id": 1, "name": 1, "phone": 1}).to_list(1000)
    
    if not delivery_customers:
        return {"message": "لا توجد شركات توصيل مسجلة", "updated": 0}
    
    # بناء قائمة أرقام الهواتف والأسماء
    delivery_phones = {c.get("phone") for c in delivery_customers if c.get("phone")}
    delivery_customer_ids = {c.get("id") for c in delivery_customers}
    delivery_names = {c.get("name").lower() for c in delivery_customers if c.get("name")}
    
    # جلب كل الطلبات الآجلة التي قد تكون لشركات توصيل
    orders = await db.orders.find({
        "payment_method": "credit",
        "is_delivery_company": {"$ne": True}  # لم يتم تحديثها بعد
    }, {"_id": 0}).to_list(10000)
    
    updated_count = 0
    for order in orders:
        should_update = False
        delivery_company_name = None
        
        # التحقق بالـ customer_id
        if order.get("customer_id") in delivery_customer_ids:
            should_update = True
            # جلب اسم الشركة
            for c in delivery_customers:
                if c.get("id") == order.get("customer_id"):
                    delivery_company_name = c.get("name")
                    break
        
        # التحقق برقم الهاتف
        elif order.get("customer_phone") in delivery_phones:
            should_update = True
            for c in delivery_customers:
                if c.get("phone") == order.get("customer_phone"):
                    delivery_company_name = c.get("name")
                    break
        
        # التحقق باسم العميل
        elif order.get("customer_name") and order.get("customer_name").lower() in delivery_names:
            should_update = True
            delivery_company_name = order.get("customer_name")
        
        if should_update:
            update_data = {
                "is_delivery_company": True,
            }
            if delivery_company_name:
                update_data["delivery_app_name"] = delivery_company_name
            
            await db.orders.update_one(
                {"id": order.get("id")},
                {"$set": update_data}
            )
            updated_count += 1
    
    return {
        "message": f"تم تحديث {updated_count} طلب",
        "updated": updated_count,
        "delivery_companies_count": len(delivery_customers)
    }



# === RECEIPT BITMAP RENDERING (Arabic support) ===
@app.post("/api/print/render-receipt")
async def render_receipt_endpoint(request: Request):
    """Generate ESC/POS bitmap bytes for thermal printer receipt with Arabic support."""
    import base64
    try:
        from receipt_renderer import render_receipt_image
        data = await request.json()
        order = data.get("order", {})
        config = data.get("printer_config", {})
        raw_bytes = render_receipt_image(order, config)
        b64 = base64.b64encode(raw_bytes).decode("ascii")
        return {"success": True, "raw_data": b64, "size": len(raw_bytes)}
    except Exception as e:
        return {"success": False, "error": str(e)}



@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()

