"""Smart Reports + Export (extracted from server.py)"""
from fastapi import APIRouter
from server import *  # noqa: F401,F403
from server import (_sn)

router = APIRouter()

# ==================== SMART REPORTS - التقارير الذكية ====================


@router.get("/smart-reports/sales")
async def get_sales_report(
    period: str = "today",  # today, week, month, year, custom
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير المبيعات - النقدي فقط الكاش المحصّل باليد (بدون التوصيل وبدون البطاقة)"""
    query = build_tenant_query(current_user)
    
    if branch_id:
        query["branch_id"] = branch_id
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    end_dt = None
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "yesterday":
        y = now - timedelta(days=1)
        start = y.replace(hour=0, minute=0, second=0, microsecond=0)
        end_dt = y.replace(hour=23, minute=59, second=59, microsecond=999999)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    elif period == "last_month":
        # الشهر السابق كاملاً (من اليوم الأول للشهر الماضي إلى آخر يوم)
        first_of_this_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        end_dt = first_of_this_month - timedelta(microseconds=1)
        start = end_dt.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif period == "six_months":
        start = now - timedelta(days=180)
    elif period == "year":
        start = now - timedelta(days=365)
    elif period == "custom" and start_date:
        start = datetime.fromisoformat(start_date)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    query["created_at"] = {"$gte": start.isoformat()}
    if end_dt:
        query["created_at"]["$lte"] = end_dt.isoformat()
    elif end_date:
        query["created_at"]["$lte"] = end_date
    
    # استبعاد الطلبات الملغية والمرتجعة من التقارير
    query["status"] = {"$nin": ["cancelled", "refunded"]}
    query["is_refunded"] = {"$ne": True}
    
    orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
    
    # فصل الطلبات المدفوعة عن المعلقة
    paid_orders = [o for o in orders if o.get("payment_status") in ["paid", "credit", None]]
    pending_orders = [o for o in orders if o.get("payment_status") == "pending"]
    
    # إجمالي المبيعات (فقط الطلبات المدفوعة/الآجلة - بدون المعلقة)
    total_sales = sum(_sn(o.get("total")) for o in paid_orders)
    total_orders = len(paid_orders)
    avg_order_value = total_sales / total_orders if total_orders > 0 else 0
    
    # المبيعات النقدية (فقط الكاش المحصّل باليد - بدون التوصيل وبدون بطاقة)
    cash_amount = sum(
        _sn(o.get("total")) for o in paid_orders 
        if o.get("payment_method") == "cash" and o.get("order_type") != "delivery"
    )
    
    # نقدي السائقين (كاش محصّل من سائقي التوصيل — منفصل عن نقدي المطعم المباشر)
    # هذه طلبات توصيل COD تم استلام فلوسها من السائق (payment_settled_from_driver_at)
    driver_cash_amount = sum(
        _sn(o.get("total")) for o in paid_orders
        if o.get("payment_method") == "cash"
        and o.get("order_type") == "delivery"
        and o.get("driver_id")
        and not o.get("is_delivery_company")
        and not o.get("delivery_app")
        and not o.get("delivery_app_name")
    )
    
    # مبيعات البطاقة (منفصلة - ليست نقدي)
    card_amount = sum(
        _sn(o.get("total")) for o in paid_orders 
        if o.get("payment_method") == "card"
    )
    
    # الآجل العادي (بدون شركة توصيل وبدون سائقين)
    credit_amount = sum(
        _sn(o.get("total")) for o in paid_orders 
        if o.get("payment_method") == "credit" 
        and not o.get("delivery_app") 
        and not o.get("delivery_app_name")
        and not o.get("delivery_app_id")
        and not o.get("is_delivery_company")
        and not (o.get("delivery_commission") and float(o.get("delivery_commission", 0)) > 0)
        and not (o.get("order_type") == "delivery" and o.get("driver_id") and not o.get("is_delivery_company"))
    )
    
    # تقسيم حسب نوع الطلب
    dine_in_amount = sum(_sn(o.get("total")) for o in paid_orders if o.get("order_type") == "dine_in")
    takeaway_amount = sum(_sn(o.get("total")) for o in paid_orders if o.get("order_type") == "takeaway")
    delivery_amount = sum(_sn(o.get("total")) for o in paid_orders if o.get("order_type") == "delivery" or o.get("delivery_app") or o.get("delivery_app_name") or o.get("is_delivery_company"))
    
    # عدد الطلبات حسب طريقة الدفع
    cash_orders_count = len([o for o in paid_orders if o.get("payment_method") == "cash" and o.get("order_type") != "delivery" and not o.get("delivery_app")])
    card_orders_count = len([o for o in paid_orders if o.get("payment_method") == "card"])
    credit_orders_count = len([o for o in paid_orders if o.get("payment_method") == "credit" and not o.get("delivery_app") and not o.get("delivery_app_name") and not o.get("delivery_app_id") and not o.get("is_delivery_company") and not (o.get("delivery_commission") and float(o.get("delivery_commission", 0)) > 0)])
    
    # شركات التوصيل الافتراضية (للتحويل من المعرف للاسم)
    default_delivery_apps_names = {
        "toters": "توترز",
        "talabat": "طلبات",
        "baly": "بالي",
        "alsaree3": "عالسريع",
        "talabati": "طلباتي",
    }
    
    # تجميع شركات التوصيل والسائقين حسب الاسم (في حسب طريقة الدفع)
    delivery_apps_amounts = {}
    driver_delivery_amount = 0  # مجموع توصيل السائقين (ليس شركات)
    
    for o in paid_orders:
        # فقط طلبات الآجل/التوصيل
        if o.get("payment_method") != "credit":
            continue
        
        # التحقق هل هي شركة توصيل أم سائق عادي
        is_company = o.get("is_delivery_company", False)
        app_name = o.get("delivery_app_name")
        
        # إذا لم يكن هناك اسم، نحول المعرف للاسم
        if not app_name and o.get("delivery_app"):
            app_id = o.get("delivery_app")
            app_name = default_delivery_apps_names.get(app_id, app_id)
            is_company = True  # إذا له delivery_app فهو شركة توصيل
        
        # إذا كان العميل شركة توصيل
        if not app_name and is_company:
            app_name = o.get("customer_name") or "شركة توصيل"
        
        # إذا كان لديه عمولة توصيل ولكن بدون اسم
        if not app_name and o.get("delivery_commission") and float(o.get("delivery_commission", 0)) > 0:
            app_name = o.get("customer_name") or "شركة توصيل"
            is_company = True
        
        if app_name and is_company:
            # شركة توصيل - يظهر باسم الشركة
            if app_name not in delivery_apps_amounts:
                delivery_apps_amounts[app_name] = 0
            delivery_apps_amounts[app_name] += _sn(o.get("total"))
        elif o.get("order_type") == "delivery" and o.get("driver_id") and not is_company:
            # توصيل بسائق عادي (ليس شركة توصيل) - يجمع تحت "توصيل سائقين"
            driver_delivery_amount += _sn(o.get("total"))
    
    # بناء by_payment_method مع أسماء شركات التوصيل
    by_payment_method = {}
    
    # إضافة النقدي فقط إذا كان أكبر من 0
    if cash_amount > 0:
        by_payment_method["نقدي"] = cash_amount
    
    # توصيل داخلي (كاش محصّل من سائقي المطعم) — سطر منفصل للتدقيق
    if driver_cash_amount > 0:
        by_payment_method["توصيل داخلي"] = driver_cash_amount
    
    # إضافة البطاقة فقط إذا كانت أكبر من 0
    if card_amount > 0:
        by_payment_method["بطاقة"] = card_amount
    
    # إضافة الآجل فقط إذا كان أكبر من 0
    if credit_amount > 0:
        by_payment_method["آجل"] = credit_amount
    
    # إضافة كل شركة توصيل باسمها
    for app_name, amount in delivery_apps_amounts.items():
        if amount > 0:
            by_payment_method[app_name] = amount
    
    # إضافة توصيل السائقين غير المحصّل (إذا وجد)
    if driver_delivery_amount > 0:
        by_payment_method["توصيل داخلي (بذمة السائقين)"] = driver_delivery_amount

    # ⚠️ الطلبات المعلقة (pending) لا تُحتسب كمبيعات/طريقة دفع — هي طلبات لم تُدفع بعد
    # نعرضها كملخّص منفصل لتطابق تقرير إغلاق الصندوق (الذي يستثني المعلق)
    pending_amount = sum(_sn(o.get("total")) for o in pending_orders)
    pending_count = len(pending_orders)
    
    # خدمة التوصيل الداخلية: أجور التوصيل المحصلة لسائقي المطعم (قيمها ضمن إجمالي المبيعات والنقدي)
    _internal_delivery_orders = [
        o for o in paid_orders
        if o.get("order_type") == "delivery"
        and o.get("driver_id")
        and not o.get("is_delivery_company")
        and not o.get("delivery_app")
        and not o.get("delivery_app_name")
        and _sn(o.get("delivery_fee")) > 0
    ]
    internal_delivery_fees = sum(_sn(o.get("delivery_fee")) for o in _internal_delivery_orders)
    internal_delivery_orders_count = len(_internal_delivery_orders)
    
    return {
        "period": period,
        "total_sales": total_sales,  # إجمالي كل المبيعات (بدون المعلقة)
        "total_orders": total_orders,
        "average_order_value": round(avg_order_value, 2),
        "by_type": {
            "dine_in": dine_in_amount,
            "takeaway": takeaway_amount,
            "delivery": delivery_amount
        },
        "by_payment_method": by_payment_method,
        # خدمة التوصيل الداخلية (أجور التوصيل) — قيمها مدمجة ضمن النقدي/توصيل داخلي وإجمالي المبيعات
        "internal_delivery_fees": internal_delivery_fees,
        "internal_delivery_orders_count": internal_delivery_orders_count,
        "by_payment": {
            "cash": cash_orders_count,
            "card": card_orders_count,
            "credit": credit_orders_count
        },
        # ⭐ ملخّص الطلبات المعلقة منفصل (لا يدخل في طرق الدفع لتطابق تقرير إغلاق الصندوق)
        "pending_orders_summary": {
            "count": pending_count,
            "amount": pending_amount,
        }
    }


@router.get("/reports/credit")
async def get_credit_report(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير الآجل - فقط الآجل العادي (بدون شركات التوصيل)"""
    query = build_tenant_query(current_user)
    
    # فقط الآجل العادي
    query["payment_method"] = "credit"
    
    if branch_id:
        query["branch_id"] = branch_id
    
    if start_date:
        query["created_at"] = {"$gte": start_date}
    if end_date:
        if "created_at" not in query:
            query["created_at"] = {}
        query["created_at"]["$lte"] = end_date
    
    # جلب الطلبات
    all_orders = await db.orders.find(query, {"_id": 0}).to_list(1000)
    
    # جلب قائمة العملاء الذين هم شركات توصيل
    tenant_id = current_user.get("tenant_id")
    delivery_customers_query = {"is_delivery_company": True}
    if tenant_id:
        delivery_customers_query["tenant_id"] = tenant_id
    else:
        delivery_customers_query["$or"] = [{"tenant_id": {"$exists": False}}, {"tenant_id": None}]
    
    delivery_customers = await db.customers.find(delivery_customers_query, {"id": 1}).to_list(1000)
    delivery_customer_ids = {c.get("id") for c in delivery_customers}
    
    # فلترة يدوية - استبعاد طلبات شركات التوصيل والمرتجعات بكل الطرق الممكنة
    orders = []
    for o in all_orders:
        # استبعاد المرتجعات والملغية
        if o.get("status") in ("refunded", "cancelled", "canceled"):
            continue
        # استبعاد إذا كان له شركة توصيل بأي شكل
        if o.get("delivery_app") or o.get("delivery_app_name") or o.get("delivery_app_id"):
            continue
        # استبعاد إذا كان العميل شركة توصيل (من الـ flag)
        if o.get("is_delivery_company") or o.get("is_delivery_app"):
            continue
        # استبعاد إذا كان لديه عمولة توصيل (يعني شركة توصيل)
        if o.get("delivery_commission") and float(o.get("delivery_commission", 0)) > 0:
            continue
        # استبعاد إذا كان الـ customer_id مرتبط بشركة توصيل
        if o.get("customer_id") in delivery_customer_ids:
            continue
        # استبعاد إذا كان اسم العميل يحتوي على كلمات شركات التوصيل المعروفة
        customer_name = (o.get("customer_name") or "").lower()
        delivery_keywords = ["toters", "تترز", "baly", "بالي", "talabat", "طلبات", "carriage", "كاريدج", "hungerstation", "هنقرستيشن", "jahez", "جاهز", "marsool", "مرسول"]
        if any(keyword in customer_name for keyword in delivery_keywords):
            continue
        # الطلب ليس له شركة توصيل - يُضاف للآجل العادي
        orders.append(o)
    
    total_credit = sum(_sn(o.get("total")) for o in orders)
    collected_amount = sum(o.get("collected_amount", 0) for o in orders)
    remaining_amount = total_credit - collected_amount
    
    # تحضير بيانات الطلبات
    order_list = []
    for order in orders:
        collected = order.get("collected_amount", 0)
        remaining = _sn(order.get("total")) - collected
        order_list.append({
            "id": order.get("id"),
            "order_number": order.get("order_number"),
            "created_at": order.get("created_at"),
            "customer_name": order.get("customer_name"),
            "customer_phone": order.get("customer_phone"),
            "total": _sn(order.get("total")),
            "collected_amount": collected,
            "remaining_amount": remaining,
            "is_fully_collected": remaining <= 0
        })
    
    return {
        "total_credit": total_credit,
        "collected_amount": collected_amount,
        "remaining_amount": remaining_amount,
        "total_orders": len(orders),
        "orders": order_list
    }

@router.post("/reports/credit/collect")
async def collect_credit_payment(
    data: dict,
    current_user: dict = Depends(get_current_user)
):
    """تسجيل تحصيل آجل"""
    order_id = data.get("order_id")
    amount = _sn(data.get("amount"))
    collected_by = data.get("collected_by", "")
    notes = data.get("notes", "")
    
    query = build_tenant_query(current_user, {"id": order_id})
    order = await db.orders.find_one(query)
    
    if not order:
        raise HTTPException(status_code=404, detail="الطلب غير موجود")
    
    current_collected = order.get("collected_amount", 0)
    new_collected = current_collected + amount
    
    await db.orders.update_one(
        {"id": order_id},
        {
            "$set": {
                "collected_amount": new_collected,
                "last_collection_date": datetime.now(timezone.utc).isoformat(),
                "last_collected_by": collected_by,
                "collection_notes": notes
            }
        }
    )
    
    return {"message": "تم تسجيل التحصيل بنجاح", "new_collected_amount": new_collected}



# ==================== moved to routes/cash_closing_report_routes.py ====================

# ==================== SMART REPORTS EXPORT ====================

@router.get("/smart-reports/export/excel")
async def export_smart_report_excel(
    report_type: str = "sales",  # sales, products, hourly
    period: str = "month",
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير الذكية إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    tenant_id = get_user_tenant_id(current_user)
    
    wb = Workbook()
    ws = wb.active
    
    # التنسيق
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    elif period == "year":
        start = now - timedelta(days=365)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Build query
    query = build_tenant_query(current_user)
    
    # فلترة الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.GENERAL_MANAGER, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    query["created_at"] = {"$gte": start.isoformat()}
    
    if report_type == "sales":
        ws.title = "تقرير المبيعات الذكي"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        # العنوان
        ws.merge_cells('A1:F1')
        ws['A1'] = f"تقرير المبيعات الذكي - {period}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # الرؤوس
        headers = ['#', 'رقم الطلب', 'التاريخ', 'النوع', 'طريقة الدفع', 'المبلغ']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        
        row_num = 4
        total = 0
        for idx, order in enumerate(orders, 1):
            total += _sn(order.get("total"))
            data = [
                idx,
                order.get("order_number", ""),
                order.get("created_at", "")[:10],
                order_types.get(order.get("order_type"), order.get("order_type", "")),
                payment_methods.get(order.get("payment_method"), order.get("payment_method", "")),
                _sn(order.get("total"))
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 6:
                    cell.number_format = '#,##0'
            row_num += 1
        
        # الإجمالي
        ws.cell(row=row_num, column=5, value="الإجمالي:").font = Font(bold=True)
        ws.cell(row=row_num, column=6, value=total).font = Font(bold=True)
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        
    elif report_type == "products":
        ws.title = "المنتجات الأكثر مبيعاً"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        # حساب مبيعات المنتجات
        product_sales = {}
        for order in orders:
            for item in order.get("items", []):
                pid = item.get("product_id")
                name = item.get("name", "Unknown")
                if pid not in product_sales:
                    product_sales[pid] = {"name": name, "qty": 0, "revenue": 0}
                product_sales[pid]["qty"] += _sn(item.get("quantity"))
                product_sales[pid]["revenue"] += _sn(item.get("price")) * _sn(item.get("quantity"))
        
        # ترتيب حسب الكمية
        sorted_products = sorted(product_sales.values(), key=lambda x: x["qty"], reverse=True)[:20]
        
        ws.merge_cells('A1:D1')
        ws['A1'] = "المنتجات الأكثر مبيعاً"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['#', 'المنتج', 'الكمية', 'الإيرادات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row_num = 4
        for idx, prod in enumerate(sorted_products, 1):
            data = [idx, prod["name"], prod["qty"], prod["revenue"]]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 4:
                    cell.number_format = '#,##0'
            row_num += 1
    
    elif report_type == "hourly":
        ws.title = "التقرير الساعي"
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["created_at"] = {"$regex": f"^{today}"}
        
        orders = await db.orders.find(query, {"_id": 0, "created_at": 1, "total": 1}).to_list(10000)
        
        hourly_data = {str(h).zfill(2): {"orders": 0, "sales": 0} for h in range(24)}
        for order in orders:
            try:
                created_at = order.get("created_at", "")
                if "T" in created_at:
                    hour = created_at.split("T")[1][:2]
                else:
                    hour = "00"
                if hour in hourly_data:
                    hourly_data[hour]["orders"] += 1
                    hourly_data[hour]["sales"] += _sn(order.get("total"))
            except:
                pass
        
        ws.merge_cells('A1:D1')
        ws['A1'] = f"التقرير الساعي - {today}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        headers = ['الساعة', 'عدد الطلبات', 'المبيعات']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        row_num = 4
        for hour in sorted(hourly_data.keys()):
            data = [f"{hour}:00", hourly_data[hour]["orders"], hourly_data[hour]["sales"]]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 3:
                    cell.number_format = '#,##0'
            row_num += 1
    
    # ضبط عرض الأعمدة
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max(max_length + 2, 12)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=smart_report_{report_type}_{period}.xlsx"}
    )

@router.get("/smart-reports/export/pdf")
async def export_smart_report_pdf(
    report_type: str = "sales",
    period: str = "month",
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير التقارير الذكية إلى PDF"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    
    tenant_id = get_user_tenant_id(current_user)
    
    # تحديد الفترة
    now = datetime.now(timezone.utc)
    if period == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    elif period == "week":
        start = now - timedelta(days=7)
    elif period == "month":
        start = now - timedelta(days=30)
    elif period == "year":
        start = now - timedelta(days=365)
    else:
        start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    # Build query
    query = build_tenant_query(current_user)
    
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.GENERAL_MANAGER, UserRole.MANAGER]:
        query["branch_id"] = user_branch_id
    elif branch_id:
        query["branch_id"] = branch_id
    
    query["created_at"] = {"$gte": start.isoformat()}
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), rightMargin=1*cm, leftMargin=1*cm, topMargin=1*cm, bottomMargin=1*cm)
    
    styles = getSampleStyleSheet()
    elements = []
    
    title_style = styles['Heading1']
    title_style.alignment = 1
    
    headers = []
    data_rows = []
    totals_row = None
    title_text = ""
    
    if report_type == "sales":
        title_text = f"تقرير المبيعات الذكي - {period}"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        headers = ["#", "رقم الطلب", "التاريخ", "النوع", "طريقة الدفع", "المبلغ"]
        order_types = {"dine_in": "محلي", "takeaway": "سفري", "delivery": "توصيل"}
        payment_methods = {"cash": "نقدي", "card": "بطاقة", "credit": "آجل"}
        
        total = 0
        for idx, order in enumerate(orders[:100], 1):  # Limit to 100 for PDF
            total += _sn(order.get("total"))
            data_rows.append([
                str(idx),
                order.get("order_number", ""),
                order.get("created_at", "")[:10],
                order_types.get(order.get("order_type"), order.get("order_type", "")),
                payment_methods.get(order.get("payment_method"), order.get("payment_method", "")),
                f"{order.get('total', 0):,.0f}"
            ])
        
        totals_row = ["", "", "", "", "الإجمالي:", f"{total:,.0f}"]
        
    elif report_type == "products":
        title_text = "المنتجات الأكثر مبيعاً"
        orders = await db.orders.find(query, {"_id": 0}).to_list(10000)
        
        product_sales = {}
        for order in orders:
            for item in order.get("items", []):
                pid = item.get("product_id")
                name = item.get("name", "Unknown")
                if pid not in product_sales:
                    product_sales[pid] = {"name": name, "qty": 0, "revenue": 0}
                product_sales[pid]["qty"] += _sn(item.get("quantity"))
                product_sales[pid]["revenue"] += _sn(item.get("price")) * _sn(item.get("quantity"))
        
        sorted_products = sorted(product_sales.values(), key=lambda x: x["qty"], reverse=True)[:20]
        
        headers = ["#", "المنتج", "الكمية", "الإيرادات"]
        for idx, prod in enumerate(sorted_products, 1):
            data_rows.append([str(idx), prod["name"], str(prod["qty"]), f"{prod['revenue']:,.0f}"])
    
    elif report_type == "hourly":
        title_text = f"التقرير الساعي - {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
        today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
        query["created_at"] = {"$regex": f"^{today}"}
        
        orders = await db.orders.find(query, {"_id": 0, "created_at": 1, "total": 1}).to_list(10000)
        
        hourly_data = {str(h).zfill(2): {"orders": 0, "sales": 0} for h in range(24)}
        for order in orders:
            try:
                created_at = order.get("created_at", "")
                if "T" in created_at:
                    hour = created_at.split("T")[1][:2]
                else:
                    hour = "00"
                if hour in hourly_data:
                    hourly_data[hour]["orders"] += 1
                    hourly_data[hour]["sales"] += _sn(order.get("total"))
            except:
                pass
        
        headers = ["الساعة", "عدد الطلبات", "المبيعات"]
        for hour in sorted(hourly_data.keys()):
            data_rows.append([f"{hour}:00", str(hourly_data[hour]["orders"]), f"{hourly_data[hour]['sales']:,.0f}"])
    
    # Build PDF
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 20))
    
    if headers and data_rows:
        table_data = [headers] + data_rows
        if totals_row:
            table_data.append(totals_row)
        
        col_widths = [doc.width / len(headers)] * len(headers)
        table = Table(table_data, colWidths=col_widths)
        
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F5F5')]),
        ])
        
        if totals_row:
            style.add('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#E6E6E6'))
            style.add('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold')
        
        table.setStyle(style)
        elements.append(table)
    
    doc.build(elements)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=smart_report_{report_type}_{period}.pdf"}
    )

