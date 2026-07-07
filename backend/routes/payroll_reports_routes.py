"""Payroll Reports + Salary Payments (extracted from server.py)"""
from fastapi import APIRouter
from server import *  # noqa: F401,F403
from server import (_sn)

router = APIRouter()

# ==================== PAYROLL REPORTS & EXPORT - تقارير الرواتب والتصدير ====================

@router.get("/reports/payroll-summary")
async def get_payroll_summary_report(
    month: str,  # YYYY-MM or YYYY for yearly
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تقرير شامل للرواتب - يدعم شهر/سنة/فترة مخصصة"""
    tenant_id = get_user_tenant_id(current_user)
    
    # بناء استعلام الموظفين
    emp_query = {"is_active": True}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    
    # فلترة حسب الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        emp_query["branch_id"] = user_branch_id
    elif branch_id:
        emp_query["branch_id"] = branch_id
    
    # استعلام محسّن مع limit معقول
    employees = await db.employees.find(emp_query, {"_id": 0}).limit(200).to_list(200)
    
    if not employees:
        return {
            "month": month,
            "employee_count": 0,
            "employees": [],
            "totals": {"basic_salary": 0, "total_deductions": 0, "total_bonuses": 0, "total_advances": 0, "overtime_pay": 0, "net_payable": 0}
        }
    
    # حساب تواريخ البداية والنهاية
    if start_date and end_date:
        q_start = start_date
        q_end = end_date
    elif len(month) == 4:  # سنة فقط YYYY
        q_start = f"{month}-01-01"
        q_end = f"{month}-12-31"
    else:
        q_start = f"{month}-01"
        q_end = f"{month}-31"
    
    # === تحسين الأداء: Batch fetch بدلاً من N+1 queries ===
    employee_ids = [emp["id"] for emp in employees]
    branch_ids = list(set([emp.get("branch_id") for emp in employees if emp.get("branch_id")]))
    
    # جلب جميع الخصومات دفعة واحدة
    all_deductions = await db.deductions.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": q_start, "$lte": q_end}
    }, {"_id": 0}).to_list(5000)
    
    # جلب جميع المكافآت دفعة واحدة
    all_bonuses = await db.bonuses.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": q_start, "$lte": q_end}
    }, {"_id": 0}).to_list(5000)
    
    # جلب جميع السلف دفعة واحدة
    all_advances = await db.advances.find({
        "employee_id": {"$in": employee_ids},
        "status": "approved",
        "remaining_amount": {"$gt": 0}
    }, {"_id": 0}).to_list(5000)

    # 💰 أقساط السلف المُحتسبة يدوياً لهذا الشهر (للسلف السابقة — المالك يحدّد المبلغ)
    all_advance_installments = await db.advance_installments.find({
        "employee_id": {"$in": employee_ids},
        "month": month
    }, {"_id": 0}).to_list(5000)
    manual_installments_by_emp = {}
    for ins in all_advance_installments:
        eid = ins.get("employee_id")
        manual_installments_by_emp[eid] = manual_installments_by_emp.get(eid, 0) + _sn(ins.get("amount"))
    
    # جلب جميع الحضور دفعة واحدة
    all_attendance = await db.attendance.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": q_start, "$lte": q_end}
    }, {"_id": 0}).to_list(10000)
    
    # جلب الوقت الإضافي الموافق عليه
    all_overtime = await db.overtime_requests.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": q_start, "$lte": q_end},
        "status": "approved"
    }, {"_id": 0}).to_list(5000)

    # 💰 جلب دفعات الرواتب النقدية (المالك يصرف من النقدي الفعلي)
    all_payments = await db.salary_payments.find({
        "employee_id": {"$in": employee_ids},
        "payment_date": {"$gte": q_start, "$lte": q_end}
    }, {"_id": 0}).to_list(5000)
    payments_by_emp = {}
    for p in all_payments:
        eid = p.get("employee_id")
        payments_by_emp[eid] = payments_by_emp.get(eid, 0) + (p.get("amount") or 0)

    # جلب جميع الفروع دفعة واحدة
    all_branches = await db.branches.find(
        {"id": {"$in": branch_ids}},
        {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    
    # تجميع البيانات
    deductions_by_emp = {}
    for d in all_deductions:
        emp_id = d.get("employee_id")
        if emp_id not in deductions_by_emp:
            deductions_by_emp[emp_id] = []
        deductions_by_emp[emp_id].append(d)
    
    bonuses_by_emp = {}
    for b in all_bonuses:
        emp_id = b.get("employee_id")
        if emp_id not in bonuses_by_emp:
            bonuses_by_emp[emp_id] = []
        bonuses_by_emp[emp_id].append(b)
    
    advances_by_emp = {}
    for a in all_advances:
        emp_id = a.get("employee_id")
        if emp_id not in advances_by_emp:
            advances_by_emp[emp_id] = []
        advances_by_emp[emp_id].append(a)
    
    attendance_by_emp = {}
    for a in all_attendance:
        emp_id = a.get("employee_id")
        if emp_id not in attendance_by_emp:
            attendance_by_emp[emp_id] = []
        attendance_by_emp[emp_id].append(a)
    
    overtime_by_emp = {}
    for o in all_overtime:
        emp_id = o.get("employee_id")
        if emp_id not in overtime_by_emp:
            overtime_by_emp[emp_id] = []
        overtime_by_emp[emp_id].append(o)
    
    branches_by_id = {b["id"]: b for b in all_branches}
    
    # بناء بيانات التقرير لكل موظف
    employee_data = []
    totals = {
        "basic_salary": 0,
        "total_deductions": 0,
        "total_bonuses": 0,
        "total_advances": 0,
        "overtime_pay": 0,
        "net_payable": 0
    }
    
    for emp in employees:
        emp_id = emp["id"]
        basic_salary = _sn(emp.get("salary"))
        salary_type = emp.get("salary_type", "monthly")
        work_hours_per_day = _sn(emp.get("work_hours_per_day", 8))
        hourly_rate = basic_salary / (30 * work_hours_per_day) if work_hours_per_day > 0 else 0
        daily_rate = basic_salary / 30 if basic_salary > 0 else 0
        
        # الحضور
        emp_attendance = attendance_by_emp.get(emp_id, [])
        present_days = len([a for a in emp_attendance if a.get("status") == "present"])
        late_days = len([a for a in emp_attendance if a.get("status") == "late"])
        early_leave_days = len([a for a in emp_attendance if a.get("status") == "early_leave"])
        worked_days = present_days + late_days
        absent_days = len([a for a in emp_attendance if a.get("status") == "absent"])
        total_worked_hours = sum(_sn(a.get("worked_hours")) for a in emp_attendance)
        
        # حساب الراتب المستحق حسب النوع - pro-rata حسب أيام العمل الفعلية
        # مثال: راتب 600,000 / 30 = 20,000 يومي. عمل 10 أيام → مستحق = 200,000
        # الصيغة: earned = daily_rate × worked_days
        # صافي الراتب يمكن أن يكون سالباً (الموظف مدين للشركة) وهذا مقصود للدقة
        worked_days_count = present_days + late_days + early_leave_days
        if salary_type == "hourly":
            earned_salary = round(total_worked_hours * hourly_rate, 2)
        elif salary_type == "daily":
            earned_salary = round(worked_days_count * daily_rate, 2)
        else:
            # monthly: pro-rata حسب أيام العمل الفعلية
            earned_salary = round(daily_rate * worked_days_count, 2) if basic_salary else 0
        
        # الخصومات
        deductions = deductions_by_emp.get(emp_id, [])
        emp_deductions = sum(_sn(d.get("amount")) for d in deductions)
        
        # المكافآت
        bonuses = bonuses_by_emp.get(emp_id, [])
        emp_bonuses = sum(_sn(b.get("amount")) for b in bonuses)
        
        # السلف:
        #  - سلفة الشهر الحالي (تاريخها ضمن الشهر): تُخصم تلقائياً (القسط الشهري المحدّد)
        #  - سلفة شهر سابق: لا تُخصم تلقائياً — تظهر كإشعار، ويُخصم فقط ما يحتسبه المالك يدوياً (أقساط)
        advances = advances_by_emp.get(emp_id, [])
        current_month_advances = [a for a in advances if (a.get("date") or "")[:7] == month]
        previous_advances = [a for a in advances if (a.get("date") or "")[:7] < month]
        emp_advances_auto = sum(_sn(a.get("monthly_deduction", 0)) for a in current_month_advances)
        emp_advances_manual = round(manual_installments_by_emp.get(emp_id, 0), 2)
        emp_advances = round(emp_advances_auto + emp_advances_manual, 2)
        # الإشعار: رصيد السلف السابقة المتبقّي فقط (يحتاج قراراً يدوياً من المالك)
        pending_advances = sum(_sn(a.get("remaining_amount", 0)) for a in previous_advances)
        
        # 👑 المدير العام/الأونر: لا يُحتسب عليه الحضور/السلف/الخصومات (راتب كامل بلا أي خصم)
        if bool(emp.get("is_general_manager")):
            emp_deductions = 0
            emp_advances = 0
            pending_advances = 0
            absent_days = 0
            deductions = []
            earned_salary = basic_salary
            worked_days = present_days + late_days
        
        # الوقت الإضافي الموافق عليه
        emp_overtime = overtime_by_emp.get(emp_id, [])
        approved_ot_hours = sum(_sn(o.get("hours")) for o in emp_overtime)
        emp_overtime_pay = round(approved_ot_hours * hourly_rate * 1.5, 2)
        
        net_payable = round(earned_salary + emp_overtime_pay + emp_bonuses - emp_deductions - emp_advances, 2)
        # 💰 المدفوع نقداً والمتبقي
        emp_paid = round(payments_by_emp.get(emp_id, 0), 2)
        emp_remaining = round(net_payable - emp_paid, 2)

        branch = branches_by_id.get(emp.get("branch_id"), {})
        
        employee_data.append({
            "id": emp["id"],
            "name": emp.get("name"),
            "position": emp.get("position"),
            "branch_id": emp.get("branch_id"),
            "branch_name": branch.get("name", "-"),
            "salary_type": salary_type,
            "basic_salary": basic_salary,
            "earned_salary": earned_salary,
            "worked_days": worked_days,
            "absent_days": absent_days,
            "total_worked_hours": total_worked_hours,
            "deductions": emp_deductions,
            "deductions_details": deductions,
            "bonuses": emp_bonuses,
            "bonuses_details": bonuses,
            "advances_deduction": emp_advances,
            "pending_advances": pending_advances,
            "overtime_hours": approved_ot_hours,
            "overtime_pay": emp_overtime_pay,
            "net_payable": net_payable,
            "paid_amount": emp_paid,
            "remaining": emp_remaining,
        })
        
        totals["basic_salary"] += basic_salary
        totals["total_deductions"] += emp_deductions
        totals["total_bonuses"] += emp_bonuses
        totals["total_advances"] += emp_advances
        totals["overtime_pay"] += emp_overtime_pay
        totals["net_payable"] += net_payable
        totals["paid_amount"] = totals.get("paid_amount", 0) + emp_paid
        totals["remaining"] = totals.get("remaining", 0) + emp_remaining
    
    return {
        "month": month,
        "employee_count": len(employees),
        "employees": employee_data,
        "totals": totals
    }


@router.post("/advances/{advance_id}/deduct-installment")
async def deduct_advance_installment(advance_id: str, payload: dict, current_user: dict = Depends(get_current_user)):
    """احتساب قسط من سلفة سابقة يدوياً في شهر محدّد (المالك يحدّد المبلغ).
    يُنشئ سجل قسط (advance_installments) ويُنقص الرصيد المتبقّي للسلفة."""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    tenant_id = get_user_tenant_id(current_user)
    advance = await db.advances.find_one({"id": advance_id}, {"_id": 0})
    if not advance:
        raise HTTPException(status_code=404, detail="السلفة غير موجودة")
    month = (payload.get("month") or "").strip()
    if not month or len(month) != 7:
        raise HTTPException(status_code=400, detail="الشهر مطلوب بصيغة YYYY-MM")
    amount = _sn(payload.get("amount"))
    remaining = _sn(advance.get("remaining_amount"))
    if amount <= 0:
        raise HTTPException(status_code=400, detail="المبلغ يجب أن يكون أكبر من صفر")
    if amount > remaining:
        raise HTTPException(status_code=400, detail=f"المبلغ ({amount:,.0f}) أكبر من الرصيد المتبقّي ({remaining:,.0f})")
    installment = {
        "id": str(uuid.uuid4()),
        "advance_id": advance_id,
        "employee_id": advance.get("employee_id"),
        "employee_name": advance.get("employee_name"),
        "tenant_id": advance.get("tenant_id") or tenant_id,
        "month": month,
        "amount": round(amount, 2),
        "notes": payload.get("notes", ""),
        "date": f"{month}-28",
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.advance_installments.insert_one(installment)
    new_remaining = round(remaining - amount, 2)
    update = {
        "remaining_amount": new_remaining,
        "deducted_amount": round(_sn(advance.get("deducted_amount")) + amount, 2),
    }
    if new_remaining <= 0:
        update["status"] = "paid"
    await db.advances.update_one({"id": advance_id}, {"$set": update})
    installment.pop("_id", None)
    return {"success": True, "installment": installment, "remaining_amount": new_remaining}


@router.get("/advances/installments")
async def list_advance_installments(month: str, employee_id: str = None, current_user: dict = Depends(get_current_user)):
    """قائمة أقساط السلف المُحتسبة يدوياً لشهر/موظف معيّن."""
    q = {"month": month}
    if employee_id:
        q["employee_id"] = employee_id
    items = await db.advance_installments.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return {"installments": items}


@router.delete("/advances/installments/{installment_id}")
async def delete_advance_installment(installment_id: str, current_user: dict = Depends(get_current_user)):
    """التراجع عن احتساب قسط سلفة (يُعيد المبلغ للرصيد المتبقّي)."""
    if current_user["role"] not in [UserRole.ADMIN, UserRole.MANAGER, UserRole.SUPER_ADMIN]:
        raise HTTPException(status_code=403, detail="غير مصرح")
    ins = await db.advance_installments.find_one({"id": installment_id}, {"_id": 0})
    if not ins:
        raise HTTPException(status_code=404, detail="القسط غير موجود")
    await db.advance_installments.delete_one({"id": installment_id})
    adv = await db.advances.find_one({"id": ins.get("advance_id")}, {"_id": 0})
    if adv:
        restored = round(_sn(adv.get("remaining_amount")) + _sn(ins.get("amount")), 2)
        upd = {
            "remaining_amount": restored,
            "deducted_amount": max(0, round(_sn(adv.get("deducted_amount")) - _sn(ins.get("amount")), 2)),
        }
        if restored > 0 and adv.get("status") == "paid":
            upd["status"] = "approved"
        await db.advances.update_one({"id": ins.get("advance_id")}, {"$set": upd})
    return {"success": True}

# ==================== 💰 Salary Payments + Daily Payroll Summary ====================

class SalaryPaymentCreate(BaseModel):
    employee_id: str
    amount: float
    payment_date: Optional[str] = None  # YYYY-MM-DD (defaults to today)
    notes: Optional[str] = None
    payment_method: Optional[str] = "cash"  # cash | bank | other
    salary_month: Optional[str] = None  # YYYY-MM — الشهر المستحق؛ يُؤرَّخ السحب من خزينة المالك بآخر يوم منه


@router.post("/payroll/payments")
async def create_salary_payment(
    payload: SalaryPaymentCreate,
    current_user: dict = Depends(get_current_user)
):
    """صرف دفعة من راتب موظف.
    تُسجَّل في `salary_payments` (سلفة على الراتب) + تُسحَب تلقائياً من **خزينة المالك**
    (`owner_withdrawals` بفئة `salary_payment`) لأن المالك يدفعها من ماله الخاص.
    لا تؤثر على الشفت، المصاريف، أو نقدي المبيعات."""
    if current_user.get("role") not in ["admin", "super_admin", "manager"]:
        raise HTTPException(status_code=403, detail="غير مسموح — للمالك/المدير فقط")

    if payload.amount is None or payload.amount <= 0:
        raise HTTPException(status_code=400, detail="المبلغ يجب أن يكون أكبر من صفر")

    tenant_id = get_user_tenant_id(current_user)
    emp_q = {"id": payload.employee_id}
    if tenant_id:
        emp_q["tenant_id"] = tenant_id
    emp = await db.employees.find_one(emp_q, {"_id": 0})
    if not emp:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")

    payment_id = str(uuid.uuid4())
    actual_date = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    # تحديد تاريخ القيد: إذا حُدِّد الشهر المستحق (salary_month) نُؤرّخ بآخر يوم منه
    # (مثال: صرف في 1/7 لراتب شهر 6 → القيد بتاريخ 2026-06-30 ويُسحب من إيداعات شهر 6)
    salary_month = (payload.salary_month or "").strip() or None
    if salary_month:
        import calendar as _cal
        try:
            _y, _m = int(salary_month[:4]), int(salary_month[5:7])
            _last = _cal.monthrange(_y, _m)[1]
            pay_date = f"{salary_month}-{_last:02d}"
        except Exception:
            raise HTTPException(status_code=400, detail="صيغة الشهر المستحق غير صحيحة (المتوقع YYYY-MM)")
    else:
        pay_date = payload.payment_date or actual_date
    paid_by_name = current_user.get("full_name") or current_user.get("username")

    # جلب اسم الفرع من جدول branches
    emp_branch_id = emp.get("branch_id")
    emp_branch_name = None
    if emp_branch_id:
        br = await db.branches.find_one({"id": emp_branch_id}, {"_id": 0, "name": 1})
        emp_branch_name = (br or {}).get("name")

    # تحقق من رصيد الفرع المتاح في خزينة المالك (إيداعات الفرع - سحوباته)
    if emp_branch_id:
        branch_q = {"branch_id": emp_branch_id}
        if tenant_id:
            branch_q["tenant_id"] = tenant_id
        br_deposits = await db.owner_deposits.find(branch_q, {"_id": 0}).to_list(5000)
        br_withdrawals = await db.owner_withdrawals.find(branch_q, {"_id": 0}).to_list(5000)
        br_transfers = await db.owner_profit_transfers.find(branch_q, {"_id": 0}).to_list(5000)
        if salary_month:
            # السحب من إيداعات الشهر المستحق فقط لهذا الفرع
            _m = lambda x: str((x or {}).get("date", "")).startswith(salary_month)
            br_total_dep = sum(d.get("amount", 0) for d in br_deposits if _m(d))
            br_total_wd = sum(w.get("amount", 0) for w in br_withdrawals if _m(w))
            br_total_tf = sum(t.get("amount", 0) for t in br_transfers if _m(t))
            scope_label = f" لشهر {salary_month}"
        else:
            br_total_dep = sum(d.get("amount", 0) for d in br_deposits)
            br_total_wd = sum(w.get("amount", 0) for w in br_withdrawals)
            br_total_tf = sum(t.get("amount", 0) for t in br_transfers)
            scope_label = ""
        branch_balance = br_total_dep - br_total_wd - br_total_tf
        if float(payload.amount) > branch_balance:
            raise HTTPException(
                status_code=400,
                detail=f"رصيد فرع \"{emp_branch_name or emp_branch_id}\"{scope_label} غير كافٍ في خزينة المالك. المتاح: {branch_balance:,.0f} IQD، المطلوب: {float(payload.amount):,.0f} IQD"
            )

    # 1) سجل الدفعة في salary_payments
    pay_doc = {
        "id": payment_id,
        "employee_id": payload.employee_id,
        "employee_name": emp.get("name"),
        "branch_id": emp_branch_id,
        "branch_name": emp_branch_name,
        "amount": round(float(payload.amount), 2),
        "payment_date": pay_date,
        "salary_month": salary_month,
        "actual_payment_date": actual_date,
        "payment_method": payload.payment_method or "cash",
        "notes": payload.notes,
        "paid_by": current_user.get("id"),
        "paid_by_name": paid_by_name,
        "tenant_id": tenant_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    # 2) سحب موازي من خزينة المالك (مرتبط بالفرع لاستقطاع من إيداعاته)
    withdrawal_id = str(uuid.uuid4())
    withdrawal_doc = {
        "id": withdrawal_id,
        "tenant_id": tenant_id,
        "amount": round(float(payload.amount), 2),
        "date": pay_date,
        "salary_month": salary_month,
        "actual_payment_date": actual_date,
        "beneficiary": f"راتب: {emp.get('name')}",
        "description": (payload.notes or "دفعة على راتب الموظف") + (f" (مستحق شهر {salary_month})" if salary_month else ""),
        "category": "salary_payment",
        "branch_id": emp_branch_id,
        "branch_name": emp_branch_name,
        "linked_salary_payment_id": payment_id,
        "created_by": paid_by_name,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

    pay_doc["linked_owner_withdrawal_id"] = withdrawal_id

    await db.salary_payments.insert_one(pay_doc)
    await db.owner_withdrawals.insert_one(withdrawal_doc)
    pay_doc.pop("_id", None)
    return pay_doc


@router.get("/payroll/payments")
async def list_salary_payments(
    employee_id: Optional[str] = None,
    branch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    salary_month: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """عرض دفعات الرواتب - فلترة بالموظف/الفرع/الفترة/الشهر المستحق.
    salary_month (YYYY-MM): يعرض كل الدفعات المخصومة من إيداعات ذلك الشهر بصرف النظر عن تاريخ الصرف الفعلي."""
    tenant_id = get_user_tenant_id(current_user)
    q = {}
    if tenant_id:
        q["tenant_id"] = tenant_id
    if employee_id:
        q["employee_id"] = employee_id
    if branch_id:
        q["branch_id"] = branch_id
    if salary_month:
        q["salary_month"] = salary_month
    if start_date or end_date:
        date_q = {}
        if start_date:
            date_q["$gte"] = start_date
        if end_date:
            date_q["$lte"] = end_date
        q["payment_date"] = date_q
    payments = await db.salary_payments.find(q, {"_id": 0}).sort("payment_date", -1).to_list(2000)
    return payments


@router.delete("/payroll/payments/{payment_id}")
async def delete_salary_payment(
    payment_id: str,
    current_user: dict = Depends(get_current_user)
):
    """حذف دفعة - للمالك فقط (في حال خطأ تسجيل).
    يحذف أيضاً السحب المرتبط من خزينة المالك."""
    if current_user.get("role") not in ["admin", "super_admin"]:
        raise HTTPException(status_code=403, detail="غير مسموح — للمالك فقط")

    tenant_id = get_user_tenant_id(current_user)
    q = {"id": payment_id}
    if tenant_id:
        q["tenant_id"] = tenant_id
    existing = await db.salary_payments.find_one(q, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="الدفعة غير موجودة")

    # حذف السحب المرتبط من خزينة المالك (إن وُجد)
    linked_wid = existing.get("linked_owner_withdrawal_id")
    if linked_wid:
        await db.owner_withdrawals.delete_one({"id": linked_wid})

    await db.salary_payments.delete_one({"id": payment_id})
    return {"message": "تم حذف الدفعة + السحب المرتبط"}


@router.get("/payroll/daily-summary")
async def get_daily_payroll_summary(
    date: str,  # YYYY-MM-DD
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """كشف رواتب يومي — يعرض راتب اليوم لكل موظف مع حضوره وما تم صرفه نقداً."""
    tenant_id = get_user_tenant_id(current_user)
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")

    emp_query = {"is_active": True}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        emp_query["branch_id"] = user_branch_id
    elif branch_id:
        emp_query["branch_id"] = branch_id

    employees = await db.employees.find(emp_query, {"_id": 0}).limit(500).to_list(500)
    if not employees:
        return {"date": date, "rows": [], "totals": {}}

    employee_ids = [e["id"] for e in employees]
    branch_ids = list(set([e.get("branch_id") for e in employees if e.get("branch_id")]))
    month = date[:7]
    month_start = f"{month}-01"

    today_attendance = await db.attendance.find({
        "employee_id": {"$in": employee_ids},
        "date": date
    }, {"_id": 0}).to_list(500)
    att_by_emp = {a["employee_id"]: a for a in today_attendance}

    mtd_attendance = await db.attendance.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": month_start, "$lte": date}
    }, {"_id": 0}).to_list(20000)
    mtd_days_by_emp = {}
    for a in mtd_attendance:
        if a.get("status") in ["present", "late", "early_leave"]:
            mtd_days_by_emp[a["employee_id"]] = mtd_days_by_emp.get(a["employee_id"], 0) + 1

    today_ded = await db.deductions.find({
        "employee_id": {"$in": employee_ids},
        "date": date
    }, {"_id": 0}).to_list(500)
    ded_today_by_emp = {}
    for d in today_ded:
        ded_today_by_emp[d["employee_id"]] = ded_today_by_emp.get(d["employee_id"], 0) + (d.get("amount") or 0)

    mtd_ded = await db.deductions.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": month_start, "$lte": date}
    }, {"_id": 0}).to_list(5000)
    mtd_ded_by_emp = {}
    for d in mtd_ded:
        mtd_ded_by_emp[d["employee_id"]] = mtd_ded_by_emp.get(d["employee_id"], 0) + (d.get("amount") or 0)

    today_bon = await db.bonuses.find({
        "employee_id": {"$in": employee_ids},
        "date": date
    }, {"_id": 0}).to_list(500)
    bon_today_by_emp = {}
    for b in today_bon:
        bon_today_by_emp[b["employee_id"]] = bon_today_by_emp.get(b["employee_id"], 0) + (b.get("amount") or 0)

    mtd_bon = await db.bonuses.find({
        "employee_id": {"$in": employee_ids},
        "date": {"$gte": month_start, "$lte": date}
    }, {"_id": 0}).to_list(5000)
    mtd_bon_by_emp = {}
    for b in mtd_bon:
        mtd_bon_by_emp[b["employee_id"]] = mtd_bon_by_emp.get(b["employee_id"], 0) + (b.get("amount") or 0)

    advances_q = {
        "employee_id": {"$in": employee_ids},
        "status": "approved",
        "remaining_amount": {"$gt": 0}
    }
    advances = await db.advances.find(advances_q, {"_id": 0}).to_list(500)
    adv_by_emp = {}
    for a in advances:
        adv_by_emp[a["employee_id"]] = adv_by_emp.get(a["employee_id"], 0) + (a.get("remaining_amount") or 0)

    payments_this_month = await db.salary_payments.find({
        "employee_id": {"$in": employee_ids},
        "payment_date": {"$gte": month_start, "$lte": f"{month}-31"}
    }, {"_id": 0}).to_list(5000)
    paid_by_emp = {}
    for p in payments_this_month:
        paid_by_emp[p["employee_id"]] = paid_by_emp.get(p["employee_id"], 0) + (p.get("amount") or 0)

    all_branches = await db.branches.find(
        {"id": {"$in": branch_ids}}, {"_id": 0, "id": 1, "name": 1}
    ).to_list(100)
    branch_name_by_id = {b["id"]: b.get("name", "-") for b in all_branches}

    rows = []
    totals = {
        "daily_earned": 0.0, "deductions_today": 0.0, "bonuses_today": 0.0,
        "mtd_earned": 0.0, "mtd_deductions": 0.0, "mtd_bonuses": 0.0,
        "pending_advances": 0.0, "paid_this_month": 0.0, "remaining_this_month": 0.0,
        "present_count": 0, "absent_count": 0,
    }

    for emp in employees:
        emp_id = emp["id"]
        basic = float(emp.get("salary") or 0)
        daily_rate = round(basic / 30, 2) if basic else 0
        att = att_by_emp.get(emp_id)
        present = bool(att and att.get("status") in ["present", "late"])
        worked_hours = float(att.get("worked_hours") or 0) if att else 0
        check_in = att.get("check_in") if att else None
        check_out = att.get("check_out") if att else None

        earned_today = daily_rate if present else 0
        ded_today = ded_today_by_emp.get(emp_id, 0)
        bon_today = bon_today_by_emp.get(emp_id, 0)

        mtd_days = mtd_days_by_emp.get(emp_id, 0)
        mtd_earned = round(daily_rate * mtd_days, 2)
        mtd_ded_amt = mtd_ded_by_emp.get(emp_id, 0)
        mtd_bon_amt = mtd_bon_by_emp.get(emp_id, 0)
        paid_amt = paid_by_emp.get(emp_id, 0)
        pending_adv = adv_by_emp.get(emp_id, 0)

        # المتبقي للشهر = (مكتسب حتى اليوم + مكافآت) - خصومات - مدفوع نقداً
        remaining = round(mtd_earned + mtd_bon_amt - mtd_ded_amt - paid_amt, 2)

        rows.append({
            "employee_id": emp_id,
            "employee_name": emp.get("name"),
            "branch_id": emp.get("branch_id"),
            "branch_name": branch_name_by_id.get(emp.get("branch_id"), "-"),
            "position": emp.get("position"),
            "basic_salary": basic,
            "daily_rate": daily_rate,
            "present": present,
            "status": (att.get("status") if att else "absent"),
            "check_in": check_in,
            "check_out": check_out,
            "worked_hours": worked_hours,
            "earned_today": earned_today,
            "deductions_today": ded_today,
            "bonuses_today": bon_today,
            "mtd_days": mtd_days,
            "mtd_earned": mtd_earned,
            "mtd_deductions": mtd_ded_amt,
            "mtd_bonuses": mtd_bon_amt,
            "pending_advances": pending_adv,
            "paid_this_month": paid_amt,
            "remaining_this_month": remaining,
        })

        totals["daily_earned"] += earned_today
        totals["deductions_today"] += ded_today
        totals["bonuses_today"] += bon_today
        totals["mtd_earned"] += mtd_earned
        totals["mtd_deductions"] += mtd_ded_amt
        totals["mtd_bonuses"] += mtd_bon_amt
        totals["pending_advances"] += pending_adv
        totals["paid_this_month"] += paid_amt
        totals["remaining_this_month"] += remaining
        if present:
            totals["present_count"] += 1
        else:
            totals["absent_count"] += 1

    totals = {k: (round(v, 2) if isinstance(v, float) else v) for k, v in totals.items()}
    return {"date": date, "month": month, "rows": rows, "totals": totals}

# ==================== End Salary Payments + Daily Payroll ====================


@router.get("/reports/employee-salary-slip/{employee_id}")
async def get_employee_salary_slip(
    employee_id: str,
    month: str,  # YYYY-MM
    current_user: dict = Depends(get_current_user)
):
    """مفردات مرتب موظف واحد"""
    tenant_id = get_user_tenant_id(current_user)
    
    # جلب الموظف
    emp_query = {"id": employee_id}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    
    employee = await db.employees.find_one(emp_query, {"_id": 0})
    if not employee:
        raise HTTPException(status_code=404, detail="الموظف غير موجود")
    
    # التحقق من صلاحية الفرع
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        if employee.get("branch_id") != user_branch_id:
            raise HTTPException(status_code=403, detail="غير مصرح")
    
    start_date = f"{month}-01"
    end_date = f"{month}-31"
    
    # جلب الفرع
    branch = await db.branches.find_one({"id": employee.get("branch_id")}, {"_id": 0})
    
    # جلب معلومات العميل
    tenant_info = await db.tenants.find_one({"id": tenant_id}, {"_id": 0}) if tenant_id else None
    
    # الخصومات التفصيلية
    deductions = await db.deductions.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(100)
    
    # تصنيف الخصومات
    deductions_by_type = {}
    for d in deductions:
        dtype = d.get("deduction_type", "other")
        if dtype not in deductions_by_type:
            deductions_by_type[dtype] = {"items": [], "total": 0}
        deductions_by_type[dtype]["items"].append(d)
        deductions_by_type[dtype]["total"] += _sn(d.get("amount"))
    
    # المكافآت التفصيلية
    bonuses = await db.bonuses.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(100)
    
    # تصنيف المكافآت
    bonuses_by_type = {}
    for b in bonuses:
        btype = b.get("bonus_type", "other")
        if btype not in bonuses_by_type:
            bonuses_by_type[btype] = {"items": [], "total": 0}
        bonuses_by_type[btype]["items"].append(b)
        bonuses_by_type[btype]["total"] += _sn(b.get("amount"))
    
    # السلف
    advances = await db.advances.find({
        "employee_id": employee_id,
        "status": {"$in": ["approved", "paid"]}
    }, {"_id": 0}).to_list(100)
    
    # الحضور
    attendance = await db.attendance.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date}
    }, {"_id": 0}).to_list(31)
    
    # إحصائيات الحضور
    attendance_stats = {
        "present": len([a for a in attendance if a.get("status") == "present"]),
        "absent": len([a for a in attendance if a.get("status") == "absent"]),
        "late": len([a for a in attendance if a.get("status") == "late"]),
        "early_leave": len([a for a in attendance if a.get("status") == "early_leave"]),
        "holiday": len([a for a in attendance if a.get("status") == "holiday"])
    }
    
    # حساب الإجماليات
    total_deductions = sum(_sn(d.get("amount")) for d in deductions)
    total_bonuses = sum(_sn(b.get("amount")) for b in bonuses)
    advance_deduction = sum(a.get("monthly_deduction", 0) for a in advances if a.get("status") == "approved" and _sn(a.get("remaining_amount", 0)) > 0)
    pending_advances = sum(a.get("remaining_amount", 0) for a in advances if a.get("status") == "approved")
    
    basic_salary = _sn(employee.get("salary"))
    salary_type = employee.get("salary_type", "monthly")
    
    # حساب الراتب المستحق حسب النوع (pro-rata للراتب الشهري حسب أيام العمل الفعلية)
    worked_days = attendance_stats["present"] + attendance_stats["late"] + attendance_stats["early_leave"]
    daily_rate = round(basic_salary / 30, 2) if basic_salary else 0
    
    if salary_type == "monthly":
        # الراتب الشهري: يُحسب بالتناسب مع أيام العمل الفعلية
        # مثال: راتب 600, يومي = 20, عمل 10 أيام → مستحق = 200
        earned_salary = round(daily_rate * worked_days, 2)
    elif salary_type == "daily":
        earned_salary = round(daily_rate * worked_days, 2)
    elif salary_type == "hourly":
        total_worked_hours = sum(_sn(a.get("worked_hours")) for a in attendance)
        hourly_rate = _sn(employee.get("hourly_rate")) or (daily_rate / (_sn(employee.get("work_hours_per_day")) or 8))
        earned_salary = round(total_worked_hours * hourly_rate, 2)
    else:
        earned_salary = basic_salary
    
    # الوقت الإضافي الموافق عليه (موحّد مع تقرير الرواتب)
    work_hours_per_day = _sn(employee.get("work_hours_per_day", 8)) or 8
    hourly_rate_ot = (daily_rate / work_hours_per_day) if work_hours_per_day else 0
    approved_overtime = await db.overtime_requests.find({
        "employee_id": employee_id,
        "date": {"$gte": start_date, "$lte": end_date},
        "status": "approved"
    }, {"_id": 0}).to_list(100)
    approved_ot_hours = sum(_sn(o.get("hours")) for o in approved_overtime)
    overtime_pay = round(approved_ot_hours * hourly_rate_ot * 1.5, 2)

    # صافي الراتب = المستحق + الوقت الإضافي + المكافآت - الخصومات - السلف
    # يمكن أن يكون سالباً (الموظف مدين للشركة) وهذا متوقّع ودقيق
    net_salary = round(earned_salary + overtime_pay + total_bonuses - total_deductions - advance_deduction, 2)
    
    return {
        "employee": employee,
        "branch": branch,
        "tenant": tenant_info,
        "month": month,
        "salary_details": {
            "basic_salary": basic_salary,
            "salary_type": salary_type,
            "daily_rate": daily_rate,
            "worked_days": worked_days,
            "earned_salary": earned_salary,
            "work_hours_per_day": employee.get("work_hours_per_day", 8)
        },
        "deductions": {
            "items": deductions,
            "by_type": deductions_by_type,
            "total": total_deductions
        },
        "bonuses": {
            "items": bonuses,
            "by_type": bonuses_by_type,
            "total": total_bonuses
        },
        "advances": {
            "items": advances,
            "deduction_this_month": advance_deduction,
            "pending_total": pending_advances
        },
        "attendance": {
            "records": attendance,
            "stats": attendance_stats
        },
        "summary": {
            "basic_salary": basic_salary,
            "worked_days": worked_days,
            "daily_rate": daily_rate,
            "earned_salary": earned_salary,
            "overtime_pay": overtime_pay,
            "total_additions": total_bonuses + overtime_pay,
            "total_deductions": total_deductions + advance_deduction,
            "net_salary": net_salary
        },
        "generated_at": datetime.now(timezone.utc).isoformat()
    }

@router.get("/reports/payroll/export/excel")
async def export_payroll_excel(
    month: str,
    branch_id: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    """تصدير تقرير الرواتب إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # جلب بيانات التقرير
    tenant_id = get_user_tenant_id(current_user)
    
    emp_query = {"is_active": True}
    if tenant_id:
        emp_query["tenant_id"] = tenant_id
    
    user_branch_id = current_user.get("branch_id")
    user_role = current_user.get("role")
    
    if user_branch_id and user_role not in [UserRole.SUPER_ADMIN, UserRole.ADMIN, UserRole.MANAGER]:
        emp_query["branch_id"] = user_branch_id
    elif branch_id:
        emp_query["branch_id"] = branch_id
    
    employees = await db.employees.find(emp_query, {"_id": 0}).to_list(500)
    
    start_date = f"{month}-01"
    end_date = f"{month}-31"
    
    # إنشاء ملف Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "تقرير الرواتب"
    
    # التنسيق
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:H1')
    ws['A1'] = f"تقرير الرواتب - {month}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # الرؤوس
    headers = ['#', 'الموظف', 'الفرع', 'الراتب الأساسي', 'المكافآت', 'الخصومات', 'السلف', 'صافي الراتب']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # البيانات
    row_num = 4
    totals = [0, 0, 0, 0, 0]
    
    for idx, emp in enumerate(employees, 1):
        # الخصومات
        deductions = await db.deductions.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(100)
        emp_deductions = sum(_sn(d.get("amount")) for d in deductions)
        
        # المكافآت
        bonuses = await db.bonuses.find({
            "employee_id": emp["id"],
            "date": {"$gte": start_date, "$lte": end_date}
        }, {"_id": 0}).to_list(100)
        emp_bonuses = sum(_sn(b.get("amount")) for b in bonuses)
        
        # السلف
        advances = await db.advances.find({
            "employee_id": emp["id"],
            "status": "approved",
            "remaining_amount": {"$gt": 0}
        }, {"_id": 0}).to_list(100)
        emp_advances = sum(a.get("monthly_deduction", 0) for a in advances)
        
        basic_salary = _sn(emp.get("salary"))
        net_salary = basic_salary + emp_bonuses - emp_deductions - emp_advances
        
        # جلب اسم الفرع
        branch = await db.branches.find_one({"id": emp.get("branch_id")}, {"_id": 0, "name": 1})
        
        data = [
            idx,
            emp.get("name", ""),
            branch.get("name", "-") if branch else "-",
            basic_salary,
            emp_bonuses,
            emp_deductions,
            emp_advances,
            net_salary
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            if col >= 4:
                cell.number_format = '#,##0'
        
        totals[0] += basic_salary
        totals[1] += emp_bonuses
        totals[2] += emp_deductions
        totals[3] += emp_advances
        totals[4] += net_salary
        
        row_num += 1
    
    # الإجماليات
    ws.cell(row=row_num, column=2, value="الإجمالي").font = Font(bold=True)
    ws.cell(row=row_num, column=2).border = thin_border
    for col, total in enumerate(totals, 4):
        cell = ws.cell(row=row_num, column=col, value=total)
        cell.font = Font(bold=True)
        cell.border = thin_border
        cell.number_format = '#,##0'
    
    # عرض الأعمدة
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 15
    
    # حفظ الملف
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=payroll_report_{month}.xlsx"}
    )

@router.get("/reports/employee-salary-slip/{employee_id}/export/excel")
async def export_employee_salary_slip_excel(
    employee_id: str,
    month: str,
    current_user: dict = Depends(get_current_user)
):
    """تصدير مفردات مرتب موظف إلى Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from io import BytesIO
    
    # جلب بيانات الموظف
    slip_data = await get_employee_salary_slip(employee_id, month, current_user)
    employee = slip_data["employee"]
    
    wb = Workbook()
    ws = wb.active
    ws.title = "مفردات المرتب"
    
    # التنسيق
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # العنوان
    ws.merge_cells('A1:D1')
    ws['A1'] = "مفردات المرتب"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # معلومات الموظف
    ws['A3'] = "اسم الموظف:"
    ws['B3'] = employee.get("name", "")
    ws['C3'] = "الشهر:"
    ws['D3'] = month
    
    ws['A4'] = "الوظيفة:"
    ws['B4'] = employee.get("position", "")
    ws['C4'] = "الفرع:"
    ws['D4'] = slip_data["branch"].get("name", "-") if slip_data["branch"] else "-"
    
    # الراتب الأساسي
    ws['A6'] = "الراتب الأساسي"
    ws['A6'].font = Font(bold=True)
    ws['B6'] = slip_data["salary_details"]["basic_salary"]
    ws['B6'].number_format = '#,##0'
    
    # المكافآت
    row = 8
    ws[f'A{row}'] = "المكافآت"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    row += 1
    
    for bonus in slip_data["bonuses"]["items"]:
        ws[f'A{row}'] = bonus.get("reason", bonus.get("bonus_type", ""))
        ws[f'B{row}'] = _sn(bonus.get("amount"))
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
    
    ws[f'A{row}'] = "إجمالي المكافآت"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = slip_data["bonuses"]["total"]
    ws[f'B{row}'].number_format = '#,##0'
    row += 2
    
    # الخصومات
    ws[f'A{row}'] = "الخصومات"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    row += 1
    
    for deduction in slip_data["deductions"]["items"]:
        ws[f'A{row}'] = deduction.get("reason", deduction.get("deduction_type", ""))
        ws[f'B{row}'] = _sn(deduction.get("amount"))
        ws[f'B{row}'].number_format = '#,##0'
        row += 1
    
    ws[f'A{row}'] = "إجمالي الخصومات"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = slip_data["deductions"]["total"]
    ws[f'B{row}'].number_format = '#,##0'
    row += 2
    
    # السلف
    ws[f'A{row}'] = "خصم السلف"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = slip_data["advances"]["deduction_this_month"]
    ws[f'B{row}'].number_format = '#,##0'
    row += 2
    
    # صافي الراتب
    ws[f'A{row}'] = "صافي الراتب"
    ws[f'A{row}'].font = Font(bold=True, size=14)
    ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws[f'A{row}'].font = Font(bold=True, size=14, color="FFFFFF")
    ws[f'B{row}'] = slip_data["summary"]["net_salary"]
    ws[f'B{row}'].number_format = '#,##0'
    ws[f'B{row}'].font = Font(bold=True, size=14)
    
    # عرض الأعمدة
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=salary_slip_{employee.get('name', '')}_{month}.xlsx"}
    )

